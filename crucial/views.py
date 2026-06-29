from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .forms import *
from .models import *
from exam.models import Examname
from miscellaneous.models import Institute,WeekendDay
from django.contrib import messages
from django.db.models import Q, Count, Case, When, IntegerField, Sum, DecimalField, Sum, ExpressionWrapper, DecimalField, F, Subquery, OuterRef
from user.models import CustomUser
from itertools import chain
from django.contrib import messages
import json
from core.models import StudentClass, SubjectAssign
from user.decorators import *
from django.contrib.auth.decorators import user_passes_test, login_required
from django.forms import inlineformset_factory
from datetime import datetime
from datetime import date
from collections import defaultdict
from attendance.models import StudentAttendance
from exam.views import result_sms
from sms.utils import *
from accounting.models import *
from .forms import *
from crucial.models import FeeHead, FeeSubHead, FeeSubHeadConfig
from django.views import View
import time
from django.db.models import Prefetch
from django.db import transaction 
from decimal import Decimal, InvalidOperation
import pandas as pd
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from django.utils import timezone
from reportlab.lib.pagesizes import letter, landscape
from xhtml2pdf import pisa
from io import BytesIO

# -------------------------------Routine--------------------------


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'HR']))
def list_routine(request):
    classroutinelist = ClassConfig.objects.all()
    context = {
        'heading': 'Routine',
        'subheading': 'List Routine',
        'classroutinelist': classroutinelist
    }
    return render(request, 'crucial/routine/list_routine.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'HR']))
def create_routine(request, class_config_id):
    current_year = str(datetime.now().year)
    admission_year = Admission_Year.objects.get(name=current_year)

    DAYS_OF_WEEK = [
        "Saturday",
        "Sunday",
        "Monday",
        "Tuesday",
        "Wednesday",
        "Thursday",
        "Friday",
    ]

    DAYS_MAPPING = {
        "sat": 0,
        "sun": 1,
        "mon": 2,
        "tue": 3,
        "wed": 4,
        "thu": 5,
        "fri": 6,
    }

    # Define weekends
    # WEEKENDS = ["Saturday", "Friday"]
    weekends_qs = WeekendDay.objects.filter(academic_year=admission_year)
    WEEKENDS = [DAYS_OF_WEEK[DAYS_MAPPING[weekend.day]] for weekend in weekends_qs if weekend.day in DAYS_MAPPING]



    class_instance = get_object_or_404(ClassConfig, pk=class_config_id)

    # Filter out weekends to get working days
    working_days = [day for day in DAYS_OF_WEEK if day not in WEEKENDS]

    if request.method == 'POST':
        try:
            routine_data = json.loads(request.POST.get('routine_data', '[]'))
            previousRoutineForTheClass = Routine.objects.filter(class_id=class_instance)
            if previousRoutineForTheClass.exists():
                previousRoutineForTheClass.delete()
            for entry in routine_data:
                period_id = entry.get('period_id')
                class_id = entry.get('class_id')
                subject_id = entry.get('subject_id')
                teacher_name = entry.get('teacher_name')
                day_name = entry.get('day_name')

                # Use update_or_create to save or update routine
                Routine.objects.update_or_create(
                    class_id=ClassConfig.objects.get(id=class_id),
                    period_id=PeriodConfig.objects.get(id=period_id),
                    day_name=day_name,
                    defaults={
                        'subject_id': Subject.objects.get(id=subject_id),
                        'teacher_name': StaffProfile.objects.get(id=teacher_name),
                    }
                )
            
            messages.success(request, 'Routine has been Updated!')

            return redirect(request.META['HTTP_REFERER'])

        except Exception as e:
            print({'status': 'error', 'message': str(e)})
            return redirect(request.META['HTTP_REFERER'])
  
  
    # period_configs = PeriodConfig.objects.select_related('class_id', 'period_id').filter(class_id=class_config_id).order_by('start_time')

    period_configs = PeriodConfig.objects.filter(
        class_id=class_config_id
    ).order_by('start_time')

    subject_assign_for_class = SubjectAssign.objects.filter(
        class_id=class_instance.class_group_id
    ).prefetch_related('subjects')

    subject_ids_for_class = subject_assign_for_class.values_list('subjects__id', flat=True)


    # classToFilter = ClassConfig.objects.get(id=class_config_id)

    # Query to create multiple rows for each teacher-subject combination
    # teacher_subjects = TeacherSubjectAssign.objects.filter(
    #     class_assigns=classToFilter
    # ).prefetch_related(
    #     Prefetch('subject_assigns')
    # ).select_related('teacher_id')

    teacher_subjects = TeacherSubjectAssign.objects.filter(
        class_assigns=class_instance,
        subject_assigns__id__in=subject_ids_for_class
    ).prefetch_related(
        Prefetch('subject_assigns', queryset=Subject.objects.filter(id__in=subject_ids_for_class))
    ).select_related('teacher_id').distinct()

    print("teacher_subjects",teacher_subjects)
    # Flatten results into rows for each teacher-subject pair
    flattened_teacher_subjects = [
        {"teacher": obj.teacher_id, "subject": subject}
        for obj in teacher_subjects
        for subject in obj.subject_assigns.all()
    ]

    routines = Routine.objects.filter(class_id=class_config_id).select_related(
        'class_id', 'period_id', 'subject_id', 'teacher_name'
    )

    formattedRoutines = {}
    for item in routines:
        day_name = item.day_name

        if item.period_id is None:
            print(f"Warning: Routine ID {item.id} has no period assigned.")  # Debugging
            continue

        if day_name not in formattedRoutines:
            formattedRoutines[day_name] = {}
        formattedRoutines[day_name][item.period_id.period_id.name] = {
        'subject_name': item.subject_id.name,
        'teacher_name': item.teacher_name.staff_field.name,
        'subject_id': item.subject_id.id,
        'teacher_id': item.teacher_name.id,
        }
    # Pass working days to the template
    context = {
        "working_days": working_days,
        'routines': formattedRoutines, 
        "working_days_json": json.dumps(working_days),  
        "period_configs": period_configs, 
        "teacher_subjects": flattened_teacher_subjects,
        'class_id': class_config_id,
        'class_instance':class_instance,
        'heading': 'Routine',
        'subheading': 'Create Routine',
        }
    return render(request, 'crucial/routine/create_routine.html', context)

@csrf_exempt
def get_existing_routines(request):
    routines = Routine.objects.select_related('period_id', 'teacher_name').all()

    routine_data = []
    for routine in routines: 
        routine_data.append({
            "teacher_id": routine.teacher_name.id if routine.teacher_name else None,
            "teacher_name": routine.teacher_name.staff_field.name if routine.teacher_name else None,
            "day_name": routine.day_name,
            "start_time": routine.period_id.start_time.strftime('%H:%M') if routine.period_id else None,
            "end_time": routine.period_id.end_time.strftime('%H:%M') if routine.period_id else None,
        })

    return JsonResponse(routine_data, safe=False)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'HR']))
def del_routine(request, pk):
    routine = get_object_or_404(Routine, pk=pk)
    routine.delete()
    return redirect(request.META['HTTP_REFERER'])

from django.http import HttpResponseBadRequest

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, "teacher", roles=['Manager', 'HR']))
def search_teacher(request):
    current_year = str(datetime.now().year)
    admission_year = Admission_Year.objects.get(name=current_year)

    teacherList = StaffProfile.objects.filter(Q(staff_field__status="Active") & Q(role__name="Teacher") )

    teacher_id = request.GET.get('teacher_id')
    if not teacher_id:
        teacher = None
        routines = []
    else:
        try:
            teacher = StaffProfile.objects.get(id=teacher_id)
            routines = Routine.objects.filter(teacher_name=teacher).select_related(
                'class_id', 'period_id', 'subject_id', 'teacher_name'
            ).order_by('period_id__start_time')
        except StaffProfile.DoesNotExist:
            return HttpResponseBadRequest("Invalid teacher selected.")
    
    DAYS_OF_WEEK = [
        "Saturday",
        "Sunday",
        "Monday",
        "Tuesday",
        "Wednesday",
        "Thursday",
        "Friday",
    ]

    # Define weekends
    DAYS_MAPPING = {
        "sat": 0,
        "sun": 1,
        "mon": 2,
        "tue": 3,
        "wed": 4,
        "thu": 5,
        "fri": 6,
    }

    # Define weekends
    # WEEKENDS = ["Saturday", "Friday"]
    weekends_qs = WeekendDay.objects.filter(academic_year=admission_year)
    WEEKENDS = [DAYS_OF_WEEK[DAYS_MAPPING[weekend.day]] for weekend in weekends_qs if weekend.day in DAYS_MAPPING]


    # Filter out weekends to get working days
    working_days = [day for day in DAYS_OF_WEEK if day not in WEEKENDS]

    unique_periods = {}
    for routine in routines:
        period_id = routine.period_id.id
        if period_id not in unique_periods:
            unique_periods[period_id] = {
                "class_name": routine.class_id,
                "period_name": routine.period_id.period_id.name,
                "start_time": routine.period_id.start_time,
                "end_time": routine.period_id.end_time,
                "break_time": routine.period_id.break_time,
                "subject": routine.subject_id.name,
                "note": routine.note,
            }

    # Convert the unique periods to a list for the response
    period_configs = list(unique_periods.values())

    # print(period_configs)

    formattedRoutines = {}
    for item in routines:
        day_name = item.day_name
        if day_name not in formattedRoutines:
            formattedRoutines[day_name] = {}
        formattedRoutines[day_name][item.period_id.period_id.name] = { 
        'subject_name': item.subject_id.name,
        'teacher_name': item.teacher_name.staff_field.name,
        'subject_id': item.subject_id.id,
        'teacher_id': item.teacher_name.id,
        'class_name': item.class_id,
        }

    today = date.today()

    context={
        "working_days": working_days, 
        "teacher": teacher,
        "teacherList":teacherList,
        "today": today, 
        "period_configs": period_configs, 
        "routines": formattedRoutines, 
        'heading': 'Routine',
        'subheading': 'Teacher Routine'
        }
    # return render(request, 'routine/teacher_routine.html', context=context)
    return render(request, 'crucial/routine/search_teacher.html', context) 


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, "teacher", roles=['Manager', 'HR']))
@csrf_exempt
def teacher_routinelist(request, teacher_id):
    routineList = Routine.objects.filter(Q(teacher_name=teacher_id))
    teacherRoutineList = list(
        routineList.values(
            'class_id__class_group_id__class_id__name',  # Class name
            'class_id__shift_id__name',                 # Shift name
            'class_id__section_id__name',               # Section name
            'period_id__period_id__name',               # Corrected Period name
            'day_name',                                 # Day of the week
            'subject_id__name',                         # Subject name
            'teacher_name__staff_field__name',                       # Teacher name
        ).order_by('period_id__name')
    )
    return JsonResponse({"teacherRoutineList": teacherRoutineList})

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, "student", "parent", roles=['Manager', 'HR']))
def search_class(request):

    classList = ClassConfig.objects.all()
    context = {
        'classList': classList,
        'heading': 'Routine',
        'subheading': 'Class Routine',
    }
    return render(request, 'crucial/routine/search_class.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, "student", "parent", roles=['Manager', 'HR']))
@csrf_exempt
def get_routinelist(request, class_id):
    routineList = Routine.objects.filter(class_id=class_id)
    classRoutineList = list(routineList.values('class_id__class_group_id__class_id__name', 'class_id__section_id__name',
                            'period_id__name', 'day_name', 'subject_id__name', 'teacher_name__name').order_by('period_id__name'))
    return JsonResponse({"classRoutineList": classRoutineList})


# -------------------------------Notice--------------------------

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'HR']))
def notice(request):
    context = {
        'heading': 'Comunication',
        'subheading': 'Notice',
    }
    return render(request, 'crucial/notice/notice.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'HR']))
def list_notice(request):
    noticelist = Notice.objects.all()
    context = {
        'noticelist': noticelist
    }
    return render(request, 'crucial/notice/list_notice.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'HR']))
def add_notice(request):
    if request.method == 'POST':
        form = Noticeform(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Notice has been Saved ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "noticeListChanged": "noticeListChanged"
            })})
    else:
        form = Noticeform()

    context = {
        'form': form
    }

    return render(request, 'crucial/notice/add_notice.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'HR']))
def edit_notice(request, pk):
    notice = get_object_or_404(Notice, pk=pk)
    if request.method == 'POST':
        form = Noticeform(request.POST, instance=notice)
        if form.is_valid():
            form.save()
            messages.success(request, 'Notice has been Edited ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "noticeListChanged": "noticeListChanged"
            })})
    else:
        form = Noticeform(instance=notice)

    context = {
        'form': form
    }

    return render(request, 'crucial/notice/add_notice.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'HR']))
def del_notice(request, pk):
    notice = get_object_or_404(Notice, pk=pk)
    notice.delete()
    return HttpResponse(status=204, headers={'HX-Trigger': 'noticeListChanged'})


# -------------------------------Notification--------------------------

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'HR']))
def notification(request):
    context = {
        'heading': 'Comunication',
        'subheading': 'Notification',
    }
    return render(request, 'crucial/notification/notification.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'HR']))
def list_notification(request):
    notificationlist = Notification.objects.all()
    context = {
        'notificationlist': notificationlist
    }
    return render(request, 'crucial/notification/list_notification.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'HR']))
def add_notification(request):
    if request.method == 'POST':
        form = Notificationform(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Notification has been Saved ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "notificationListChanged": "notificationListChanged"
            })})
    else:
        form = Notificationform()

    context = {
        'form': form
    }

    return render(request, 'crucial/notification/add_notification.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'HR']))
def edit_notification(request, pk):
    notification = get_object_or_404(Notification, pk=pk)
    if request.method == 'POST':
        form = Notificationform(request.POST, instance=notification)
        if form.is_valid():
            form.save()
            messages.success(request, 'Notification has been Edited ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "notificationListChanged": "notificationListChanged"
            })})
    else:
        form = Notificationform(instance=notification)

    context = {
        'form': form
    }

    return render(request, 'crucial/notification/add_notification.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'HR']))
def del_notification(request, pk):
    notification = get_object_or_404(Notificationform, pk=pk)
    notification.delete()
    return HttpResponse(status=204, headers={'HX-Trigger': 'notificationListChanged'})



@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'HR', 'Accountant']))
def sms(request):
    current_year = str(datetime.now().year)
    receiver = None
    admission_year_list=Admission_Year.objects.all()
    academic_session_list=AcademicSession.objects.all()
    admission_year = Admission_Year.objects.get(name=current_year)
    sms_template_list = SMSTemplate.objects.all()
    class_list = StudentClass.objects.all()
    section_list = ClassConfig.objects.all()
    group_list = ClassGroupConfig.objects.filter(group_id__isnull=False)
    exam_list = Examname.objects.filter(academic_year=admission_year)
    template_count = sms_template_list.count()
    student_list = StudentProfile.objects.filter(
    Q(
        academic_session_year__start_year__lte=admission_year.name,
        academic_session_year__end_year__gte=admission_year.name
    ) |
    Q(admission_year_id=admission_year), student_field__status="Active")

    employee_list = StaffProfile.objects.filter(Q(staff_field__status="Active"))
    phone_numbers = []
    sms_count = None

    if request.method == 'POST':
        current_user = request.user
        title = request.POST.get('title')
        body = request.POST.get('body')
        class_id_list = [int(class_id)
                         for class_id in request.POST.getlist('class_id')]
        section_id_list = request.POST.getlist('section')
        wise_id = request.POST.get('wise')
        admission_year_id = request.POST.get('admission_year')
        academic_session_id = request.POST.get('academic_session')
        sms_body = request.POST.get('body')
        if sms_body:
            sms_count = count_sms(sms_body)

        if class_id_list:
            class_instance = StudentClass.objects.filter(
                id__in=class_id_list)

            phone_numbers = []
            for student_class in class_instance:
                studentList = StudentProfile.objects.filter(
                    Q(
                        academic_session_year__start_year__lte=admission_year_id,
                        academic_session_year__end_year__gte=admission_year_id
                    )|Q(admission_year_id=admission_year_id) & Q(class_id__class_group_id__class_id=student_class) & Q(student_field__status="Active"))
                
                parent_numbers = [{'phone_number': student.parent_id.phone_number} for student in studentList if
                                  student.parent_id and student.parent_id.phone_number]
                phone_numbers.extend([parent['phone_number']
                                     for parent in parent_numbers])
            parent_phone_numbers = list(phone_numbers)
            phone_numbers_string = ','.join(
                [f"880{number.lstrip('0')}" for number in parent_phone_numbers])
        
            total_number = (len(parent_phone_numbers))
            total_sms = total_number*sms_count
            sms_limit_obj = SMSUsage.objects.filter(
                Msg_type='NONMASKING').first()

            if sms_limit_obj.total_sms < total_sms:
                messages.error(request, 'SMS LIMIT OVER !!!')
            else:
                sms_limit_obj.total_sms -= total_sms
                sms_limit_obj.save()
                receiver = f'{phone_numbers_string}'
                send_sms(receiver, sms_body)

                for phone_number in parent_phone_numbers:
                    SMS.objects.create(
                        mobile=phone_number, title=title, msg=sms_body, created_by=current_user)

            return redirect('sms')

        if section_id_list:
            section_instance = ClassConfig.objects.filter(
                id__in=section_id_list)

            phone_numbers = []
            for student_section in section_instance:
                studentList = StudentProfile.objects.filter(
                    Q(
                        academic_session_year__start_year__lte=admission_year_id,
                        academic_session_year__end_year__gte=admission_year_id
                    )|Q(admission_year_id=admission_year_id) & Q(
                        class_id=student_section) & Q(student_field__status="Active")
                )

                parent_numbers = [{'phone_number': student.parent_id.phone_number} for student in studentList if
                                  student.parent_id and student.parent_id.phone_number]
                phone_numbers.extend([parent['phone_number']
                                     for parent in parent_numbers])
            parent_phone_numbers = list(phone_numbers)
            phone_numbers_string = ','.join(
                [f"880{number.lstrip('0')}" for number in parent_phone_numbers])
            total_number = (len(parent_phone_numbers))
            total_sms = total_number*sms_count
            sms_limit_obj = SMSUsage.objects.filter(
                Msg_type='NONMASKING').first()
            if sms_limit_obj.total_sms < total_sms:
                messages.error(request, 'SMS LIMIT OVER !!!')
            else:
                sms_limit_obj.total_sms -= total_sms
                sms_limit_obj.save()
                receiver = f'{phone_numbers_string}'
                send_sms(receiver, sms_body)
                for phone_number in parent_phone_numbers:
                    SMS.objects.create(
                        mobile=phone_number, title=title, msg=sms_body, created_by=current_user)
            return redirect('sms')

        if wise_id == "th":
            teacher_list = StaffProfile.objects.filter(
                    Q(staff_field__status="Active") & Q(role__name="Teacher")
                ).values_list('staff_field__phone_number', flat=True)
            phone_numbers_string = ','.join(
                [f"880{number.lstrip('0')}" for number in teacher_list])
            total_number = (len(teacher_list))
            total_sms = total_number*sms_count
            sms_limit_obj = SMSUsage.objects.filter(
                Msg_type='NONMASKING').first()
            if sms_limit_obj.total_sms < total_sms:
                messages.error(request, 'SMS LIMIT OVER !!!')
            else:
                sms_limit_obj.total_sms -= total_sms
                sms_limit_obj.save()
                receiver = f'{phone_numbers_string}'
                send_sms(receiver, sms_body)
                for phone_number in teacher_list:
                    SMS.objects.create(mobile=phone_number,
                                       title=title, msg=sms_body)
            return redirect('sms')
        if wise_id == "sf":
            staff_list = StaffProfile.objects.filter(
                Q(staff_field__status="Active")
            ).exclude(
                role__name="Teacher"
            ).values_list('staff_field__phone_number', flat=True)
            phone_numbers_string = ','.join(
                [f"880{number.lstrip('0')}" for number in staff_list])
            total_number = (len(staff_list))
            total_sms = total_number*sms_count
            sms_limit_obj = SMSUsage.objects.filter(
                Msg_type='NONMASKING').first()
            if sms_limit_obj.total_sms < total_sms:
                messages.error(request, 'SMS LIMIT OVER !!!')
            else:
                sms_limit_obj.total_sms -= total_sms
                sms_limit_obj.save()
                receiver = f'{phone_numbers_string}'
                send_sms(receiver, sms_body)
                for phone_number in staff_list:
                    SMS.objects.create(
                        mobile=phone_number, title=title, msg=sms_body, created_by=current_user)
            return redirect('sms')
        if wise_id == "ins":
            teacher_list = StaffProfile.objects.values_list(
                'staff_field__phone_number', flat=True)
            
            studentList = StudentProfile.objects.filter(
                Q(academic_session_year__start_year__lte=admission_year_id,
                        academic_session_year__end_year__gte=admission_year_id
                    )|Q(admission_year_id=admission_year_id) & Q(student_field__status="Active"))
            parent_numbers = [
                student.parent_id.phone_number for student in studentList if student.parent_id and student.parent_id.phone_number]

            phone_numbers = list(teacher_list) + parent_numbers

            phone_numbers_string = ','.join(
                [f"880{number.lstrip('0')}" for number in phone_numbers])

            total_number = (len(phone_numbers))
            total_sms = total_number*sms_count
            sms_limit_obj = SMSUsage.objects.filter(
                Msg_type='NONMASKING').first()
            if sms_limit_obj.total_sms < total_sms:
                messages.error(request, 'SMS LIMIT OVER !!!')
            else:
                sms_limit_obj.total_sms -= total_sms
                sms_limit_obj.save()
                receiver = f'{phone_numbers_string}'
                send_sms(receiver, sms_body)
                for phone_number in phone_numbers:
                    SMS.objects.create(
                        mobile=phone_number, title=title, msg=sms_body, created_by=current_user)

            return redirect('sms')

        if wise_id == "sp":

            selected_student_ids = request.POST.getlist('student')
            selected_student_ids = [int(student_id)
                                    for student_id in selected_student_ids]
            studentList = StudentProfile.objects.filter(
                id__in=selected_student_ids)
            parent_numbers = [{'phone_number': student.parent_id.phone_number} for student in studentList if
                              student.parent_id and student.parent_id.phone_number]
            phone_numbers.extend([parent['phone_number']
                                 for parent in parent_numbers])
            parent_phone_numbers = list(phone_numbers)
            phone_numbers_string = ','.join(
                [f"880{number.lstrip('0')}" for number in parent_phone_numbers])
            total_number = (len(parent_phone_numbers))
            total_sms = total_number*sms_count
            sms_limit_obj = SMSUsage.objects.filter(
                Msg_type='NONMASKING').first()
            if sms_limit_obj.total_sms < total_sms:
                messages.error(request, 'SMS LIMIT OVER !!!')
            else:
                sms_limit_obj.total_sms -= total_sms
                sms_limit_obj.save()
                receiver = f'{phone_numbers_string}'
                send_sms(receiver, sms_body)
                for phone_number in parent_phone_numbers:
                    SMS.objects.create(
                        mobile=phone_number, title=title, msg=sms_body, created_by=current_user)
            return redirect('sms')

        if wise_id == "se":
            selected_employee_ids = request.POST.getlist('employee')
            selected_employee = StaffProfile.objects.filter(
                id__in=selected_employee_ids)
            
            selected_employees = list(selected_employee) 
            phone_numbers = [
                employee.staff_field.phone_number for employee in selected_employees
                if employee.staff_field and employee.staff_field.phone_number
            ]
            phone_numbers_string = ','.join(
                [f"880{number.lstrip('0')}" for number in phone_numbers])
            total_number = (len(phone_numbers))
            total_sms = total_number*sms_count
            sms_limit_obj = SMSUsage.objects.filter(
                Msg_type='NONMASKING').first()
            if sms_limit_obj.total_sms < total_sms:
                messages.error(request, 'SMS LIMIT OVER !!!')
            else:
                sms_limit_obj.total_sms -= total_sms
                sms_limit_obj.save()
                receiver = f'{phone_numbers_string}'
                send_sms(receiver ,sms_body)
                for phone_number in phone_numbers:
                    SMS.objects.create(
                        mobile=phone_number, title=title, msg=sms_body, created_by=current_user)
            return redirect('sms')

        if title:
            SMSTemplate.objects.create(
                title=title,
                body=body
            )

            return redirect('sms')

    context = {
        'student_list': student_list,
        'employee_list': employee_list,
        'class_list': class_list,
        'section_list': section_list,
        'group_list': group_list,
        'sms_template_list': sms_template_list,
        'exam_list': exam_list,
        'template_count': template_count,
        'admission_year_list':admission_year_list,
        'academic_session_list':academic_session_list,
        'heading': 'Comunication',
        'subheading': 'General SMS',
    }
    return render(request, 'crucial/comunication/sms.html', context)




@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def update_notification_status(request, sms_id):
    if request.method == 'POST':
        sms = get_object_or_404(SMSTemplateNotification, id=sms_id)
        status = request.POST.get('notification_status')
        if status in ['Active', 'Inactive']:
            sms.notification_status = status
            sms.save()
            return JsonResponse({'success': True})
        return JsonResponse({'success': False, 'error': 'Invalid status'}, status=400)
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
@csrf_exempt
def update_notification_body(request):
    if request.method == 'POST':
        sms_id = request.POST.get('sms_id')
        body = request.POST.get('body')
        
        try:
            sms = SMSTemplateNotification.objects.get(id=sms_id)
            sms.body = body
            sms.save()
            return JsonResponse({'success': True})
        except SMSTemplateNotification.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'SMS not found'}, status=404)
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'HR', 'Accountant']))
def notification_sms(request): 
    current_year = str(datetime.now().year)
    admission_year = Admission_Year.objects.get(name=current_year)
    fee_month_list=Fee_month.objects.all()
    feetype = Feetype.objects.all() 
    class_list = StudentClass.objects.all()
    section_list = ClassConfig.objects.all()
    exam_list = Examname.objects.filter(academic_year=admission_year)
    sms_template_list=SMSTemplateNotification.objects.all()
    sms_count = None
    if request.method == 'POST':
        current_user = request.user
        class_id_list = [int(class_id)
                         for class_id in request.POST.getlist('class_id')]
        
        if class_id_list:
            class_instance = StudentClass.objects.filter(
                id__in=class_id_list, academic_year=admission_year)
        section_id_list = request.POST.getlist('section')
        if section_id_list:
            section_instance = ClassConfig.objects.filter(
                id__in=section_id_list, academic_year=admission_year)
        exam_id = request.POST.get('exam_id')
        if exam_id:
            exam_instance = get_object_or_404(Examname, pk=exam_id)

        fee_month = request.POST.get('feemonth')
        aistatus = request.POST.get('aistatus')
        title = request.POST.get('title')
        notification_type = request.POST.get('notification_type')
        notification_body = request.POST.get('body')

        if fee_month:
            fee_month_instance = get_object_or_404(Fee_month, pk=fee_month)
        notify = request.POST.get('notification')
        
        if notify == "er":
            return result_sms(request, exam_instance.id, class_id_list)
        if notify == "psi":
            if fee_month_instance:
                fees = Fees.objects.filter(
                    month_id=fee_month_instance, academic_year=admission_year)
            if class_id_list:
                template = SMSTemplateNotification.objects.filter(
                        notification_type='Monthly Due Info',
                        notification_status='Active'
                        ).first()
                
                for student_class in class_instance:
                    student_profiles = StudentProfile.objects.filter(
                        Q(admission_year_id=admission_year) & Q(class_id__class_group_id__class_id=student_class) & Q(student_field__status="Active"))
                    
                    for student_profile in student_profiles:
                        student_fees = fees.filter(Q(student_id=student_profile) & Q(status__in=["unpaid", "partial"]))
                        total_amount = int(sum(fee.total_fee_after_partial_payments() for fee in student_fees))
                        
                        for fee in student_fees:
                            if fee.amount > 0:
                                name = student_profile.student_field.name
                                name = name.title()
                                amount = fee.amount
                                student_number = student_profile.parent_id.phone_number
                                formatted_number = '880' + \
                                    student_number.lstrip('0')
                                if template:
                                    sms_body = template.body.format(
                                    name=name,
                                    amount=total_amount,
                                    fees_month=fee_month_instance
                                    )
                                    sms_count = count_sms(sms_body)
                                    sms_limit_obj = SMSUsage.objects.filter(
                                        Msg_type='NONMASKING').first()
                                    if sms_limit_obj.total_sms < 1:
                                        print('SMS LIMIT OVER')
                                    else:
                                        
                                        receiver = f'{formatted_number}'
                                        send_sms(receiver , sms_body)
                                        sms_limit_obj.total_sms -= sms_count
                                        sms_limit_obj.save()
                                        SMS.objects.create(
                                            mobile=student_number, title='Monthly Due Info', msg=sms_body, created_by=current_user)

                messages.success(request, 'Messages sent successfully ! ! !')

            elif section_id_list:
                template = SMSTemplateNotification.objects.filter(
                        notification_type='Monthly Due Info',
                        notification_status='Active'
                        ).first()
                
                for student_section in section_instance:
                    studentList = StudentProfile.objects.filter(
                        Q(admission_year_id=admission_year) & Q(
                            class_id=student_section) & Q(student_field__status="Active")
                    )
                    for student in studentList:
                        student_fees = fees.filter(Q(student_id=student) & Q(status__in=["unpaid", "partial"]))
                        total_amount = int(sum(fee.total_fee_after_partial_payments() for fee in student_fees))
                        
                        for fee in stu_fees:
                            if fee.amount > 0:
                                name = student.student_field.name
                                student_number = student.parent_id.phone_number
                                formatted_number = '880' + \
                                    student_number.lstrip('0')
                                if template:
                                    sms_body = template.body.format(
                                    name=name,
                                    amount=total_amount,
                                    fees_month=fee_month_instance
                                    )
                                    
                                    sms_count = count_sms(sms_body)
                                    sms_limit_obj = SMSUsage.objects.filter(
                                        Msg_type='NONMASKING').first()
                                    if sms_limit_obj.total_sms < 1:
                                        print('SMS LIMIT OVER') 
                                    else:
                                        receiver = f'{formatted_number}"'
                                        send_sms(receiver, sms_body)
                                        sms_limit_obj.total_sms -= sms_count
                                        sms_limit_obj.save()
                                        SMS.objects.create(
                                            mobile=student_number, title='Monthly Due Info', msg=sms_body, created_by=current_user)

                messages.success(request, 'Messages sent successfully ! ! !')

            else:
                studentList = StudentProfile.objects.filter(
                    Q(admission_year_id=admission_year) & Q(
                        student_field__status="Active")
                )

                template = SMSTemplateNotification.objects.filter(
                        notification_type='Monthly Due Info',
                        notification_status='Active'
                        ).first()
                
                for student in studentList:
                    stu_fees = fees.filter(Q(student_id=student) & Q(status__in=["unpaid", "partial"]))
                    total_amount = int(sum(fee.total_fee_after_partial_payments() for fee in stu_fees))
                    for fee in stu_fees:
                        if fee.amount > 0:
                            name = student.student_field.name
                            
                            student_number = student.parent_id.phone_number
                            formatted_number = '880' + student_number.lstrip('0')
                            sms_body = template.body.format(
                                name=name,
                                amount=total_amount,
                                fees_month=fee_month_instance
                                )
                            
                            sms_count = count_sms(sms_body)
                            sms_limit_obj = SMSUsage.objects.filter(
                                Msg_type='NONMASKING').first()
                            if sms_limit_obj.total_sms < 1:
                                print('SMS LIMIT OVER')
                            else:
                                receiver = f'{formatted_number}'
                                send_sms(receiver, sms_body)
                                sms_limit_obj.total_sms -= sms_count
                                sms_limit_obj.save()
                                SMS.objects.create(
                                    mobile=student_number, title='Monthly Due Info', msg=sms_body, created_by=current_user)

                messages.success(request, 'Messages sent successfully ! ! !')

            return redirect('notification_sms')

        if notify == "di":
            fees = Fees.objects.filter(
                Q(status__in=["unpaid", "partial"]), Q(academic_year=admission_year))
            if class_id_list:
                for student_class in class_instance:
                    student_profiles = StudentProfile.objects.filter(
                        Q(admission_year_id=admission_year) & Q(
                            class_id__class_group_id__class_id=student_class) & Q(student_field__status="Active")
                    )
                    template = SMSTemplateNotification.objects.filter(
                        notification_type='Total Due Info',
                        notification_status='Active'
                        ).first()
                    
                    for student_profile in student_profiles:
                        student_fees = fees.filter(student_id=student_profile)
                        total_unpaid_fees = int(sum(fee.total_fee_after_partial_payments() for fee in student_fees))
                        name = student_profile.student_field.name
                        student_number = student_profile.parent_id.phone_number
                        formatted_number = '880' + student_number.lstrip('0')
                        if total_unpaid_fees:
                            current_date = datetime.now().date()
                            today_date = current_date.strftime("%d-%m-%Y")
                            if template:
                                sms_body = template.body.format(
                                    name=name,
                                    today_date=today_date,
                                    total_unpaid_fees=total_unpaid_fees
                                    )
                                sms_count = count_sms(sms_body)
                                sms_limit_obj = SMSUsage.objects.filter(
                                    Msg_type='NONMASKING').first()
                                if sms_limit_obj.total_sms < 1:
                                    print('SMS LIMIT OVER')
                                else:
                                    receiver = f'{formatted_number}'
                                    send_sms(receiver, sms_body)
                                    sms_limit_obj.total_sms -= sms_count
                                    sms_limit_obj.save()
                                    SMS.objects.create(
                                        mobile=student_number, title='Due Fee', msg=sms_body, created_by=current_user)

                messages.success(request, 'Messages sent successfully ! ! !')

            elif section_id_list:
                for student_section in section_instance:
                    studentList = StudentProfile.objects.filter(
                        Q(admission_year_id=admission_year) & Q(
                            class_id=student_section) & Q(student_field__status="Active")
                    )
                    template = SMSTemplateNotification.objects.filter(
                        notification_type='Total Due Info',
                        notification_status='Active'
                        ).first()
                    
                    for student in studentList:
                        student_fees = fees.filter(student_id=student)
                        total_unpaid_fees = int(sum(fee.total_fee_after_partial_payments() for fee in student_fees))
                        name = student.student_field.name
                        student_number = student.parent_id.phone_number
                        formatted_number = '880' + student_number.lstrip('0')
                        if total_unpaid_fees:
                            current_date = datetime.now().date()
                            today_date = current_date.strftime("%d-%m-%Y")
                            if template:
                                sms_body = template.body.format(
                                    name=name,
                                    today_date=today_date,
                                    total_unpaid_fees=total_unpaid_fees
                                    )
                                sms_count = count_sms(sms_body)
                                sms_limit_obj = SMSUsage.objects.filter(
                                    Msg_type='NONMASKING').first()
                                if sms_limit_obj.total_sms < 1:
                                    print('SMS LIMIT OVER')
                                else:
                                    receiver = f'{formatted_number}'
                                    send_sms(receiver, sms_body)
                                    sms_limit_obj.total_sms -= sms_count
                                    sms_limit_obj.save()
                                    SMS.objects.create(
                                        mobile=student_number, title='Due Fee', msg=sms_body, created_by=current_user)

                messages.success(request, 'Messages sent successfully ! ! !')

            else:
                studentList = StudentProfile.objects.filter(
                    Q(admission_year_id=admission_year) & Q(
                        student_field__status="Active")
                )
                template = SMSTemplateNotification.objects.filter(
                        notification_type='Total Due Info',
                        notification_status='Active'
                        ).first()
                
                for student in studentList:
                    stu_fees = fees.filter(student_id=student)
                    total_unpaid_fees = int(sum(fee.total_fee_after_partial_payments() for fee in stu_fees))
                    name = student.student_field.name
                    student_number = student.parent_id.phone_number
                    formatted_number = '880' + student_number.lstrip('0')
                    if total_unpaid_fees:
                        current_date = datetime.now().date()
                        today_date = current_date.strftime("%d-%m-%Y")
                        
                        if template:
                            sms_body = template.body.format(
                            name=name,
                            today_date=today_date,
                            total_unpaid_fees=total_unpaid_fees
                            )
                            
                            sms_count = count_sms(sms_body)
                            sms_limit_obj = SMSUsage.objects.filter(
                                Msg_type='NONMASKING').first()
                            if sms_limit_obj.total_sms < 1:
                                print('SMS LIMIT OVER')
                            else:
                                receiver = f'{formatted_number}'
                                send_sms(receiver, sms_body)
                                sms_limit_obj.total_sms -= sms_count
                                sms_limit_obj.save()
                                SMS.objects.create(
                                    mobile=student_number, title='Due Fee', msg=sms_body, created_by=current_user)
                messages.success(request, 'Messages sent successfully ! ! !')

            return redirect('notification_sms')
        
        if notify == "ai":
            today = date.today()
            day_name = today.strftime('%A')
            today_date = today.strftime('%B %d, %Y')

            if aistatus == 'present':
                
                students_with_present = StudentAttendance.objects.filter(
                    Q(attendance_date=today) &  
                    Q(status=True)  
                ).values('name')
                # print(students_with_present)
                present_students = StudentProfile.objects.filter(
                    admission_year_id=admission_year,
                    id__in=students_with_present,
                    student_field__status="Active"
                )
                template = SMSTemplateNotification.objects.filter(
                        notification_type='Attendance Present',
                        notification_status='Active'
                    ).first()
                
                for student_profile in present_students:
                    student_attendance = StudentAttendance.objects.filter(Q(name=student_profile) & Q(attendance_date=today) ).first()
                    if student_attendance:
                        in_time = student_attendance.created_at.strftime('%H:%M')
                        name = student_profile.student_field.name
                        student_number = student_profile.parent_id.phone_number
                        formatted_number = '880' + student_number.lstrip('0')
                    
                    if template:
                        sms_body = template.body.format(
                        name=name,
                        today_date=today_date,
                        day_name=day_name,
                        in_time=in_time
                        )

                        sms_count = count_sms(sms_body)
                        sms_limit_obj = SMSUsage.objects.filter(
                            Msg_type='NONMASKING').first()
                        if sms_limit_obj.total_sms < 1:
                            print('SMS LIMIT OVER')
                        else:
                            receiver = f'{formatted_number}'
                            send_sms(receiver, sms_body)
                            sms_limit_obj.total_sms -= sms_count
                            sms_limit_obj.save()
                            SMS.objects.create(
                                mobile=student_number, title='Attendance Info', msg=sms_body, created_by=current_user)
                messages.success(request, 'Messages sent successfully ! ! !')

                return redirect('notification_sms')
            if aistatus == 'absent':
                students_with_present = StudentAttendance.objects.filter(
                    Q(attendance_date=today) &  
                    Q(status=True)  
                ).values('name')
                
                absent_students = StudentProfile.objects.filter(
                    admission_year_id=admission_year,
                    student_field__status="Active"
                ).exclude(id__in=Subquery(students_with_present.values('name')))
                template = SMSTemplateNotification.objects.filter(
                        notification_type='Attendance Absent',
                        notification_status='Active'
                    ).first()
                
                
                for student_profile in absent_students:    
                    name = student_profile.student_field.name
                    student_number = student_profile.parent_id.phone_number
                    formatted_number = '880' + student_number.lstrip('0')
                    
                    if template:
                        sms_body = template.body.format( 
                        name=name,
                        today_date=today_date,
                        day_name=day_name
                        )
                        sms_count = count_sms(sms_body)
                        sms_limit_obj = SMSUsage.objects.filter(
                            Msg_type='NONMASKING').first()
                        if sms_limit_obj.total_sms < 1:
                            print('SMS LIMIT OVER')
                        else:
                            receiver = f'{formatted_number}'
                            send_sms(receiver, sms_body)
                            sms_limit_obj.total_sms -= sms_count
                            sms_limit_obj.save()
                            SMS.objects.create(
                                mobile=student_number, title='Attendance Info', msg=sms_body, created_by=current_user)
                messages.success(request, 'Messages sent successfully ! ! !')
                return redirect('notification_sms')
            
        if title:
            SMSTemplateNotification.objects.create(
                title=title,
                notification_type=notification_type,
                body = notification_body,
                notification_status='Active'
            )

            return redirect('notification_sms')

    context = {
        'exam_list': exam_list,
        'class_list': class_list,
        'section_list': section_list,
        'fee_month_list':fee_month_list,
        'feetype': feetype,
        'sms_template_list':sms_template_list,
        'heading': 'Comunication',
        'subheading': 'Notification SMS',
    }
    return render(request, 'crucial/comunication/notification_sms.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def del_sms(request, pk):
    sms_template = get_object_or_404(SMSTemplate, pk=pk)
    sms_template.delete()
    return redirect('sms')


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant', 'HR']))
def send_summary(request):
    num_queries = None
    sms_summary = None
    if request.method == 'POST':
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')

        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

            sms_summary = SMS.objects.filter(
                created_at__date__range=(start_date, end_date)
            ).values('created_at__date').annotate(num_sms=Count('id'))

            sms_summary_list = list(sms_summary)
            num_queries = len(sms_summary_list)

    context = {
        'sms_summary': sms_summary,
        'num_queries': num_queries,
        'heading': 'SMS',
        'subheading': 'Send Sms Summary',
    }
    return render(request, 'crucial/comunication/send_summary.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant', 'HR']))
def purchase_history(request):
    smslimit = SMSLimit.objects.all()
    smslimit_count = smslimit.count()
    if request.method == 'POST':
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')

        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

            smslimit = SMSLimit.objects.filter(
                created_at__date__range=(start_date, end_date)
            ).order_by("-id")

    context = {
        'smslimit': smslimit,
        'smslimit_count': smslimit_count,
        'heading': 'SMS',
        'subheading': 'Purchase History',
    }
    return render(request, 'crucial/comunication/purchase_history.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant', 'HR']))
def sms_delivery(request):
    smsdelivery = None
    userlist = CustomUser.objects.all()
    if request.method == 'POST':
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')

        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

            smsdelivery = SMS.objects.filter(
                created_at__date__range=(start_date, end_date)
            ).order_by("-created_at")
    context = {
        'smsdelivery': smsdelivery,
        'userlist': userlist,
        'heading': 'SMS',
        'subheading': 'Teacher sms delivery',
    }
    sms_message = request.GET.get('message', '')
    if sms_message:
        context['sms_message'] = sms_message
    return render(request, 'crucial/comunication/sms_delivery.html', context)





# ------------------------------ Fee_package --------------------------


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def feepackage(request):
    context = {
        'heading': 'Fees',
        'subheading': 'Fee Amount',
    }
    return render(request, 'crucial/finance/fees_package.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def list_feepackage(request):
    current_year = str(datetime.now().year)
    admission_year = Admission_Year.objects.get(name=current_year)
    feepackagelist = Fee_package.objects.filter(
        academic_year=admission_year).order_by("student_class")
    context = {
        'feepackagelist': feepackagelist
    }
    return render(request, 'crucial/finance/listfee_package.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def add_feepackage(request):
    current_year = str(datetime.now().year)
    admission_year = Admission_Year.objects.get(name=current_year)
    form = Fee_package_form(initial={'academic_year': admission_year})
    
    if request.method == 'POST':
        form = Fee_package_form(request.POST)
        if form.is_valid():
            # Check if the fee package already exists
            student_class = form.cleaned_data['student_class']
            fees_type = form.cleaned_data['fees_type']
            academic_year = form.cleaned_data['academic_year']
            
            existing_fee_package = Fee_package.objects.filter(
                student_class=student_class,
                fees_type=fees_type,
                academic_year=academic_year
            ).exists()
            
            if existing_fee_package:
                messages.error(request, 'Fee package already exists for the selected class, fee type, and session year.')
                return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                    "messages": 'Success',
                    "feePackageChanged": "feePackageChanged"
                })})
            else:
                form.save()
                messages.success(request, 'Fee Amount has been saved successfully!')
                return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                    "messages": 'Success',
                    "feePackageChanged": "feePackageChanged"
                })})
        else:
            # Form is invalid, re-render with errors
            messages.error(request, 'Form validation error! Please correct the errors.')
    
    context = {
        'form': form
    }
    
    return render(request, 'crucial/finance/addfee_package.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def edit_feepackage(request, pk):
    feeamount = get_object_or_404(Fee_package, pk=pk)
    if request.method == 'POST':
        form = Fee_package_form(request.POST, instance=feeamount)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fee Amount has been Updated ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "feePackageChanged": "feePackageChanged"
            })})
    else:
        form = Fee_package_form(instance=feeamount)

    context = {
        'form': form
    }

    return render(request, 'crucial/finance/addfee_package.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def del_feepackage(request, pk):
    feeamount = get_object_or_404(Fee_package, pk=pk)
    feeamount.delete()
    return HttpResponse(status=204, headers={'HX-Trigger': 'feePackageChanged'})



# ------------------------------fee type view --------------------------


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def feetype(request):
    context = {
        'heading': 'Fees',
        'subheading': 'Fee Type',
    }
    return render(request, 'crucial/finance/feetype.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def list_feetype(request):
    
    feetypelist = Feetype.objects.all().order_by("id")

    context = {
        'feetypelist': feetypelist
    }
    return render(request, 'crucial/finance/listfeetype.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def add_feetype(request):
    form = Feetypeform(initial={'status': 'Active'})

    if request.method == 'POST':
        form = Feetypeform(request.POST)
        if form.is_valid():
            fee_head = form.cleaned_data['fee_head'] 
            fee_Schedule = form.cleaned_data['fee_Schedule']
            version = form.cleaned_data['version']

            existing_fee_type = Feetype.objects.filter( 
                fee_head=fee_head,
                fee_Schedule=fee_Schedule,
                version=version,

            ).exists()

            if existing_fee_type:
                messages.error(request, 'Fee type already exists for the selected fee head, schedule, and session year.')
                return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                    "messages": 'Error',
                    "feeListChanged": "feeListChanged"
                })})
            else:
                form.save()
                messages.success(request, 'Fee type has been saved successfully!')
                return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                    "messages": 'Success',
                    "feeListChanged": "feeListChanged"
                })})
        else:
            messages.error(request, 'Form validation error! Please correct the errors.')

    context = {
        'form': form
    }

    return render(request, 'crucial/finance/addfeetype.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def edit_feetype(request, pk):
    feetype = get_object_or_404(Feetype, pk=pk)
    if request.method == 'POST':
        form = Feetypeform(request.POST, instance=feetype)
        if form.is_valid():
            form.save()
            messages.success(request, 'Feetype has been Updated ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({ 
                "messages": 'Success',
                            "feeListChanged": "feeListChanged"
            })})
    else:
        form = Feetypeform(instance=feetype)

    context = {
        'form': form
    }

    return render(request, 'crucial/finance/addfeetype.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def del_feetype(request, pk):
    feetype = get_object_or_404(Feetype, pk=pk)
    feetype.delete()
    return HttpResponse(status=204, headers={'HX-Trigger': 'feeListChanged'})

# ------------------------------Waiver--------------------------



@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'HR', 'Accountant']))
def waiver(request):
    update_academic_year = Admission_Year.objects.latest('updated_at')
    stulist = StudentProfile.objects.filter(
        Q(admission_year_id=update_academic_year) & Q(student_field__status="Active")
    )
    waiver_list = Waiver.objects.filter(academic_year=update_academic_year).order_by('-created_at')

    if request.method == 'POST':
        current_user = request.user
        student_id = request.POST.get('student_id')
        student_instance = get_object_or_404(StudentProfile, pk=student_id)
        waiver_amount = request.POST.get('waiver_amount')
        description = request.POST.get('description')
        fee_type_ids = request.POST.getlist('fee_types') 

        existing_waivers = Waiver.objects.filter(student_id=student_instance, academic_year=update_academic_year)

        if not existing_waivers.exists():
            waiver = Waiver.objects.create(
                student_id=student_instance,
                waiver_amount=waiver_amount,
                academic_year=update_academic_year,
                description=description,
                created_by=current_user
            )
            waiver.fee_types.set(fee_type_ids)
            
            ledger_name = f"waiver_{student_instance.student_field.user_id}"
            ledger_category, _ = LedgerCategory.objects.get_or_create(name="Waivers", defaults={"description": "Tracking student fee waivers"})
            if not Ledger.objects.filter(name=ledger_name).exists():
                Ledger.objects.create(
                    name=ledger_name,
                    category=ledger_category,
                    note=f"Waiver ledger for student ID {student_instance.student_field.user_id}",
                    created_at=timezone.now()
                )
            
            messages.success(request, 'Waiver has been created successfully')
            return redirect('waiver')
        else:
            messages.error(request, 'Selected Student Waiver already exists!')

    context = {
        'stulist': stulist,
        'waiver_list': waiver_list,
        'fee_types': Feetype.objects.all(),
        'heading': 'Student',
        'subheading': 'Waiver',
    }
    return render(request, 'crucial/finance/student_waiver.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'HR', 'Accountant']))
def edit_waiver(request, id):
    waiver_instance = get_object_or_404(Waiver, id=id)
    update_academic_year = Admission_Year.objects.latest('updated_at')

    if request.method == 'POST':
        form = WaiverForm(request.POST, instance=waiver_instance)
        if form.is_valid():
            current_user = request.user
            waiver_instance = form.save(commit=False)
            waiver_instance.updated_by = current_user
            waiver_instance.academic_year = update_academic_year
            waiver_instance.save()
            form.save_m2m()  # Save the many-to-many relationship
            messages.success(request, 'Waiver has been updated successfully')
            return redirect('waiver')
    else:
        form = WaiverForm(instance=waiver_instance)

    context = {
        'form': form,
        'heading': 'Student',
        'subheading': 'Update Waiver',
    }
    return render(request, 'crucial/finance/update_waiver.html', context)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'HR', 'Accountant']))
def del_waiver(request, pk):
    waiver = get_object_or_404(Waiver, pk=pk)
    waiver.delete()
    messages.error(request, 'Selected Student Waiver has been deleted!')
    return redirect('waiver')

# ------------------------------Fee Master --------------------------
def get_number_of_months(request):
    feetype_id = request.GET.get('feetype_id')
    if feetype_id:
        try:
            feetype = Feetype.objects.get(id=feetype_id)
            months = feetype.get_number_of_months()
            return JsonResponse({'months': months})
        except Feetype.DoesNotExist:
            return JsonResponse({'error': 'Feetype not found'}, status=404)
    return JsonResponse({'error': 'Invalid request'}, status=400)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'HR', 'Accountant']))
def get_existing_fees(request):
    feetype_id = request.GET.get('feetype_id')
    
    existing_fees = Fees_name.objects.filter(fees_type_id=feetype_id)
    
    ordered_fees = sorted(existing_fees, key=lambda fee: (
            fee.fee_amount_id.student_class.class_id.name if fee.fee_amount_id and fee.fee_amount_id.student_class.class_id else '',
            fee.fee_amount_id.student_class.group_id.name if fee.fee_amount_id and fee.fee_amount_id.student_class.group_id else ''
        ))

    fees_data = []
    for fee in ordered_fees:
        fees_data.append({
            'fees_title': fee.fees_title,
            'class': fee.fee_amount_id.student_class.class_id.name if fee.fee_amount_id else 'N/A',
            'group': fee.fee_amount_id.student_class.group_id.name if fee.fee_amount_id and fee.fee_amount_id.student_class.group_id else '',
            'startdate': fee.startdate,
            'enddate': fee.enddate
        })
    return JsonResponse({'fees': fees_data})

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'HR', 'Accountant']))
def fees_master(request):
    current_year = str(datetime.now().year)
    admission_year = Admission_Year.objects.get(name=current_year)
    stuclass = ClassGroupConfig.objects.all()
    stufeetype = Feetype.objects.all()
    monthlist = Fee_month.objects.all()
    admission_year_list = Admission_Year.objects.all()
    if request.method == 'POST':
        feetype_id = request.POST.get('feetype_id')
        admission_id = request.POST.get('admission_id')
        print("feetype_id",feetype_id,"admission_id",admission_id)
        feetype_instance = get_object_or_404(Feetype, pk=feetype_id)
        admission_year_instance = get_object_or_404(Admission_Year, pk=admission_id)
        additional_class_ids = request.POST.getlist('additional_class_ids[]')
        additional_month_ids = request.POST.getlist('additional_month_ids[]')
        additional_start_dates = request.POST.getlist('additional_start_dates[]')
        additional_end_dates = request.POST.getlist('additional_end_dates[]')
        print("additional_class_ids",additional_class_ids)
        success_messages = []
        error_messages = []

        if feetype_instance.is_hostel_fee or feetype_instance.is_transport_fee or feetype_instance.is_coaching_fee:
            for month_id, start, end in zip(additional_month_ids, additional_start_dates, additional_end_dates):
                if month_id and start and end:  # Only process complete sets of data
                    try:
                        month_instance = Fee_month.objects.get(pk=month_id)
                        
                        existing_fees = Fees_name.objects.filter(
                            fees_type=feetype_instance,
                            month=month_instance,
                        ).exists()
                        
                        if not existing_fees:
                            if feetype_instance.is_hostel_fee:
                                Fees_name.create_from_fee_amount( 
                                    fees_type=feetype_instance,
                                    month=month_instance,
                                    start_date=start,
                                    end_date=end,
                                    academic_year=admission_year_instance,
                                    user=request.user
                                )
                                success_messages.append('Successfully created Hostel Fee!')

                            if feetype_instance.is_transport_fee:
                                print("Transport Print")
                                Fees_name.create_from_fee_amount(
                                    fees_type=feetype_instance,
                                    month=month_instance,
                                    start_date=start,
                                    end_date=end,
                                    academic_year=admission_year_instance,
                                    user=request.user
                                )
                                success_messages.append('Successfully created Transport Fee!')

                            if feetype_instance.is_coaching_fee:
                                print("Coching Print")
                                Fees_name.create_from_fee_amount(
                                    fees_type=feetype_instance,
                                    month=month_instance,
                                    start_date=start,
                                    end_date=end,
                                    academic_year=admission_year_instance,
                                    user=request.user
                                )
                                success_messages.append('Successfully created Coaching Fee!')
                        else:
                            error_messages.append('Fee already exists for the selected class, fee type!')
                    except Fee_month.DoesNotExist:
                        error_messages.append(f"Fee_month with id {month_id} does not exist")
        else:
            for class_id in additional_class_ids:
                additional_fee_amount = Fee_package.objects.filter(fees_type_id=feetype_id, student_class_id=class_id,academic_year=admission_year).first()
                print("additional_fee_amount",additional_fee_amount)
                if additional_fee_amount:
                    for month_id, start, end in zip(additional_month_ids, additional_start_dates, additional_end_dates):
                        if month_id and start and end: 
                            try:
                                month_instance = Fee_month.objects.get(pk=month_id)
                                
                                existing_fees = Fees_name.objects.filter(
                                    fee_amount_id=additional_fee_amount.id,
                                    month=month_instance,
                                    academic_year=admission_year_instance,
                                    fees_type=additional_fee_amount.fees_type,
                                ).exists()
                                
                                if not existing_fees:
                                    Fees_name.create_from_fee_amount(
                                        fees_type=additional_fee_amount.fees_type,
                                        month=month_instance,
                                        start_date=start,
                                        end_date=end,
                                        academic_year=admission_year_instance,
                                        user=request.user,
                                        fee_amount=additional_fee_amount
                                    )
                                    success_messages.append('Successfully created Selected Fee!')
                                else:
                                    error_messages.append('Fee already exists for the selected class, fee type!')
                            except Fee_month.DoesNotExist:
                                error_messages.append(f"Fee_month with id {month_id} does not exist")
                else:
                    error_messages.append(f"No Fee_amount found for class_id: {class_id} and fees_type_id: {feetype_id}")

        
        if success_messages:
            messages.success(request, success_messages[-1])
        if error_messages:
            messages.error(request, error_messages[-1])

        return redirect('fees_master')

    context = {
        'stufeetype': stufeetype,
        'stuclass': stuclass,
        'monthlist': monthlist,
        'admission_year_list':admission_year_list,
        'heading': 'Fees',
        'subheading': 'Fees Master',
    }
    return render(request, 'crucial/finance/fees_master.html', context)


# ------------------------------Fee view --------------------------

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def add_fee(request):
    current_year = str(datetime.now().year)
    current_admission_year = Admission_Year.objects.get(name=current_year)
    fee_months = Fee_month.objects.all()
    fee_name_list = Fees_name.objects.all()  
    stulist = None
    admission_year_list = Admission_Year.objects.all().order_by("id")
    admission_year_instance = None

    if request.method == 'POST':
        current_user = request.user
        invoice = request.POST.get('invoice_type')
        version = request.POST.get('version') 
        fee_month_id = request.POST.get('fee_month_id')
        fee_name_id = request.POST.get('fee_name')  
        academic_year = request.POST.get('academic_year')
        individual_student = request.POST.get('student_id')

        # Validate required fields
        if not fee_name_id:
            messages.error(request, "Please select a Fee Name.")
            return redirect('addfee')
        if not version:
            messages.error(request, "Please select a version (Bangla or English).")
            return redirect('addfee')
        if not academic_year:
            messages.error(request, "Please select an Academic Year.")
            return redirect('addfee')
        if not fee_month_id:
            messages.error(request, "Please select a Fee Month.")
            return redirect('addfee')
        if invoice == 'individual' and not individual_student:
            messages.error(request, "Please select a student for individual invoice.")
            return redirect('addfee')

        admission_year_instance = get_object_or_404(Admission_Year, id=academic_year)
        fee_month_instance = get_object_or_404(Fee_month, pk=fee_month_id)
        fee_name_instance = get_object_or_404(Fees_name, pk=fee_name_id)

        description = request.POST.get('description')

        success_count = 0  
        error_count = 0 

        if invoice == 'individual' and individual_student:
            students = StudentProfile.objects.filter(pk=individual_student, version=version) 
        else:
            students = StudentProfile.objects.filter(Q(student_field__status="Active"), version=version)

        for student in students:
            fee_amount = None
            waiver_percentage = 0

            # Check for waiver
            waiver = Waiver.objects.filter(student_id=student, fee_types=fee_name_instance.fees_type).first()
            if waiver:
                waiver_percentage = waiver.waiver_amount

            # Determine fee amount based on fee type
            if fee_name_instance.fees_type.is_hostel_fee:
                hostel_package = Hostel.objects.filter(student_id=student, academic_year=admission_year_instance).first()
                if hostel_package and hostel_package.hostel_package:
                    fee_amount = hostel_package.hostel_package.amount
                else:
                    error_count += 1
            elif fee_name_instance.fees_type.is_transport_fee:
                transport_package = Transport.objects.filter(student_id=student, academic_year=admission_year_instance).first()
                if transport_package and transport_package.transport_package:
                    fee_amount = transport_package.transport_package.amount
                else:
                    error_count += 1
            elif fee_name_instance.fees_type.is_coaching_fee:
                coaching_package = Tution.objects.filter(student_id=student, academic_year=admission_year_instance).first()
                if coaching_package and coaching_package.tution_package:
                    fee_amount = coaching_package.tution_package.amount
                else:
                    error_count += 1
            elif (student.class_id and 
                  fee_name_instance.fee_amount_id and 
                  student.class_id.class_group_id == fee_name_instance.fee_amount_id.student_class):
                fee_amount = fee_name_instance.fee_amount_id.amount if fee_name_instance.fee_amount_id else 0

            if fee_amount is not None:
                fee_amount_after_waiver = fee_amount * (Decimal(1) - waiver_percentage / Decimal(100))

                # Check for existing fee entry
                existing_fee = Fees.objects.filter(
                    student_id=student,
                    feetype_id=fee_name_instance,
                    month_id=fee_month_instance,
                    academic_year=admission_year_instance
                ).exists()

                if not existing_fee:
                    Fees.objects.create(
                        student_id=student,
                        feetype_id=fee_name_instance,
                        amount=fee_amount_after_waiver,
                        status='unpaid',
                        month_id=fee_month_instance,
                        academic_year=admission_year_instance,
                        created_by=current_user,
                        description=description
                    )
                    success_count += 1
                else:
                    error_count += 1
            else:
                error_count += 1

        # Handle success and error messages
        if success_count > 0:
            messages.success(request, f"Successfully allocated fees for {success_count} student(s).")
        if error_count > 0:
            messages.error(request, f"Failed to allocate fees for {error_count} student(s). Please check student eligibility or missing packages.")

        return redirect('addfee')

    context = {
        'stulist': stulist,
        'fee_months': fee_months,
        'admission_year_list': admission_year_list,
        'fee_types': Feetype.objects.all(),
        'fee_name_list': fee_name_list,
        'heading': 'Fees',
        'subheading': 'Allocation Fee',
    }
    return render(request, 'crucial/finance/addfee.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def generate_fee_report(request, fee_id):
    institute = Institute.objects.latest('id')
    fee = Fees.objects.get(id=fee_id)
    context = {
        'institute': institute,
        'fee': fee
    }
    return render(request, 'report/fees.html', context)



@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def listfeestatus(request):
    academic_year = Admission_Year.objects.latest('updated_at')
    fee_months = Fee_month.objects.all()
    classlist = ClassConfig.objects.all()
    feelist, grouped_fees = None, defaultdict(list)
    total_amount = 0
    total_due = 0
    total_discount = 0
    total_paid = 0

    if request.method == 'POST':
        month_id = request.POST.get('month_id')
        class_id = request.POST.get('class_name_id')
        feestatus_id = request.POST.get('status_check')
        selected_month = Fee_month.objects.get(id=month_id)
        selected_class = ClassConfig.objects.get(id=class_id)

        if feestatus_id == 'paid':
            feestatus='paid'
        else:
            feestatus='unpaid'


        if feestatus == 'paid':
            feelist = Fees.objects.filter(
                Q(student_id__class_id=selected_class) &
                Q(month_id=selected_month) &
                (Q(status='paid') | Q(status='partial'))
            )
        elif feestatus == 'unpaid':
            feelist = Fees.objects.filter(
                Q(student_id__class_id=selected_class) &
                Q(month_id=selected_month) &
                (Q(status='unpaid') | Q(status='partial'))
            )
        else:
            feelist = Fees.objects.filter(
                Q(student_id__class_id=selected_class) &
                Q(month_id=selected_month) &
                Q(status=feestatus)
            )
        class_feelist = Fees.objects.filter(
                Q(student_id__class_id=selected_class) &
                Q(month_id=selected_month)
            )
        
        if class_feelist.exists():
            total_amount = class_feelist.aggregate(total_amount=Sum('amount'))['total_amount'] or 0
            total_discount = class_feelist.aggregate(total_discount=Sum('discount_amount'))['total_discount'] or 0
            total_paid = sum(fee.total_netTotal() for fee in class_feelist) or 0
            total_due = total_amount - total_discount - total_paid

            for fee in feelist:
                grouped_fees[fee.student_id].append(fee)

    context = {
        'fee_months': fee_months,
        'classlist': classlist,
        'grouped_fees': dict(grouped_fees), 
        'total_amount': total_amount,
        'total_discount': total_discount,
        'total_paid': total_paid,
        'total_due': total_due,
        'heading': 'Fees',
        'subheading': 'Fees Collection',
    }

    return render(request, 'crucial/finance/listfee.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
@csrf_exempt

def feestatus(request):
    session_year = Admission_Year.objects.latest('updated_at')
    admission_year_instance = get_object_or_404(Admission_Year, name=session_year)
    current_user = request.user
    fee_id = request.POST.get('id')
    
    try:
        fees_id = Fees.objects.get(id=fee_id)
    except Fees.DoesNotExist:
        return JsonResponse({'error': 'Fees matching query does not exist.'}, status=404)

    discount = Decimal(request.POST.get('discount') or '0.00')
    method = request.POST.get('paymentMethod')
    payment_amount = Decimal(request.POST.get('paymentAmount') or '0.00')
    latefee = fees_id.calculate_late_fee() or fees_id.late_amount
    
    if method == 'full':
        status = 'paid'
        fees_id.status = status
        fees_id.late_amount = latefee
        fees_id.discount_amount = discount
        f_amount = fees_id.amount + latefee - discount
        fees_id.created_by = current_user

        # Correctly pass the `fees_id` object to the function
        create_receive_and_journal(fees_id, f_amount, current_user)
        print("create_receive_and_journal")
        fees_id.save()

    elif method == 'partial':
        fees_id.late_amount = latefee
        fees_id.discount_amount = discount
        fees_id.created_by = current_user
        f_amount = payment_amount

        # Update fields for partial payment and save
        fees_id.save(update_fields=['late_amount', 'discount_amount', 'updated_by'])
        fees_id.record_payment(payment_amount, current_user)
        
        create_receive_and_journal(fees_id, f_amount, current_user)
        fees_id.save()

    student_id = fees_id.student_id.id
    student = StudentProfile.objects.get(id=student_id)
    total_amount = float(f_amount)
    fee_month = fees_id.month_id.name
    phonenumber = student.parent_id.phone_number
    name = student.student_field.name
    class_name = student.class_id
    roll = student.roll_no
    formatted_number = '880' + phonenumber.lstrip('0')

    # Sending SMS logic
    template = SMSTemplateNotification.objects.filter(
        notification_type='Pay Slip Info',
        notification_status='Active'
    ).first()
    if template:
        sms_body = template.body.format(
            name=name,
            fee_month=fee_month,
            total_amount=total_amount
        )
        sms_count = count_sms(sms_body)
        sms_limit_obj = SMSUsage.objects.filter(Msg_type='NONMASKING').first()
        if sms_limit_obj.total_sms < 1:
            print('SMS LIMIT OVER')
        else:
            receiver = f'{formatted_number}'
            send_sms(receiver, sms_body)
            sms_limit_obj.total_sms -= sms_count
            sms_limit_obj.save()
            SMS.objects.create(mobile=phonenumber, title='Paid Msg', msg=sms_body, created_by=current_user)

    if class_name:
        class_config_data = {
            'class_id': getattr(class_name.class_group_id.class_id, 'name', 'Unknown Class'),
            'section_id': getattr(class_name.section_id, 'name', 'Unknown Section')
        }

        print(class_config_data)


    fees = Fees.objects.filter(student_id=student_id)
    total_net_total = sum(fee.total_netTotal() for fee in fees)
    feesList = []

    for fee in fees:
        latefee = fee.calculate_late_fee() or fee.late_amount
        total_fee_after_discount = fee.total_fee_after_discount()
        waiver_amount = Waiver.objects.filter(student_id=student_id).aggregate(
            total_waiver=Sum('waiver_amount'))['total_waiver'] or 0
        total_paid_amount = fee.total_paid_amount()
        total_fee_after_partial_payments = fee.total_fee_after_partial_payments()
        fee_data = {
            'id': fee.id,
            'feetype': fee.feetype_id.fees_title,
            'amount': fee.amount,
            'startdate': fee.feetype_id.startdate,
            'enddate': fee.feetype_id.enddate,
            'discount_amount': fee.discount_amount,
            'status': fee.status,
            'late_fee': latefee,
            'total_fee_after_discount': total_fee_after_discount,
            'waiver_amount': waiver_amount,
            'total_paid_amount': total_paid_amount,
            'total_netTotal': total_net_total,
            'total_fee_after_partial_payments': total_fee_after_partial_payments,
            'total_fee': fee.total_fee()
        }
        feesList.append(fee_data)

    student_data = [
        {
            'name': name,
            'class': class_config_data['class_id'] + " " + class_config_data['section_id'],
            'roll': roll,
            'phonenumber': phonenumber,
            'feesList': feesList,
            'id': fees_id.id,
            'status': fees_id.status,
            'total_paid_amount': fees_id.total_paid_amount(),
            'total_fee_after_partial_payments': fees_id.total_fee_after_partial_payments()
        }
    ]

    return JsonResponse(student_data, safe=False)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
# @csrf_exempt
# def student_search(request):

#     if request.method == 'POST':
#         student_id = request.POST.get('studentId')
#         student = StudentProfile.objects.get(id=student_id)
#         name = student.student_field.name
#         class_name = student.class_id
#         roll = student.roll_no
#         parent=student.parent_id.name

#         if class_name:
#             class_config_data = {
#             'class_id': getattr(class_name.class_group_id.class_id, 'name', 'Unknown Class'),
#             'group_id': getattr(class_name.class_group_id.group_id, 'name', ''),
#             'section_id': getattr(class_name.section_id, 'name', '')
#         }

#         fees = Fees.objects.filter(student_id=student_id)

#         total_net_total = sum(fee.total_netTotal() for fee in fees)

#         feesList = []

#         for fee in fees:
#             latefee = fee.calculate_late_fee() or fee.late_amount
#             total_fee_after_discount = fee.total_fee_after_discount()
#             waiver_amount = Waiver.objects.filter(student_id=student_id).aggregate(
#                 total_waiver=Sum('waiver_amount'))['total_waiver'] or 0
#             total_paid_amount = fee.total_paid_amount()
#             total_fee_after_partial_payments = fee.total_fee_after_partial_payments()

#             fee_data = {
#                 'id': fee.id,
#                 'feetype': fee.feetype_id.fees_title,
#                 'amount': fee.amount,
#                 'startdate': fee.feetype_id.startdate if fee.feetype_id else None,
#                 'enddate': fee.feetype_id.enddate if fee.feetype_id else None,
#                 'discount_amount': fee.discount_amount,
#                 'status': fee.status,
#                 'total_fee_after_discount': total_fee_after_discount,
#                 'late_fee': latefee,
#                 'waiver_amount': waiver_amount,
#                 'total_paid_amount': total_paid_amount,
#                 'total_netTotal':total_net_total,
#                 'total_fee_after_partial_payments': total_fee_after_partial_payments,
#                 'total_fee':fee.total_fee()
                
#             }
#             feesList.append(fee_data)

#         student_data = [
#             {'name': name, 'parent' : parent, 'class': class_config_data['class_id'] + " " + class_config_data['group_id'] + " " +
#                 class_config_data['section_id'], 'roll': roll, 'feesList': feesList}
#         ]

        
#         return JsonResponse(student_data, safe=False)

#     return JsonResponse({'error': 'Invalid request'})


@csrf_exempt
def student_search(request):
    if request.method == 'POST':
        try:
            student_id = request.POST.get('studentId')
            if not student_id:
                return JsonResponse({'error': 'Missing studentId'}, status=400)

            # Retrieve student object
            student = StudentProfile.objects.select_related('student_field', 'class_id__class_group_id__class_id', 'class_id__class_group_id__group_id', 'class_id__section_id', 'parent_id').get(id=student_id)

            # Basic student info
            name = student.student_field.name if student.student_field else "N/A"
            roll = student.roll_no
            parent = student.parent_id.name if student.parent_id else "N/A"

            # Class config data
            class_info = student.class_id
            class_config_data = {
                'class_id': getattr(class_info.class_group_id.class_id, 'name', 'Unknown Class') if class_info else 'Unknown Class',
                'group_id': getattr(class_info.class_group_id.group_id, 'name', '') if class_info else '',
                'section_id': getattr(class_info.section_id, 'name', '') if class_info else '',
            }

            # Fetch and calculate fee-related data
            fees = Fees.objects.filter(student_id=student_id)
            total_net_total = sum(fee.total_netTotal() for fee in fees)

            feesList = []
            for fee in fees:
                latefee = fee.calculate_late_fee() or fee.late_amount
                total_fee_after_discount = fee.total_fee_after_discount()
                waiver_amount = Waiver.objects.filter(student_id=student_id).aggregate(
                    total_waiver=Sum('waiver_amount'))['total_waiver'] or 0
                total_paid_amount = fee.total_paid_amount()
                total_fee_after_partial_payments = fee.total_fee_after_partial_payments()
                total_fee = fee.total_fee()

                fee_data = {
                    'id': fee.id,
                    'feetype': fee.feetype_id.fees_title if fee.feetype_id else "N/A",
                    'amount': fee.amount,
                    'startdate': fee.feetype_id.startdate if fee.feetype_id else None,
                    'enddate': fee.feetype_id.enddate if fee.feetype_id else None,
                    'discount_amount': fee.discount_amount,
                    'status': fee.status,
                    'late_fee': latefee,
                    'waiver_amount': waiver_amount,
                    'total_paid_amount': total_paid_amount,
                    'total_fee_after_discount': total_fee_after_discount,
                    'total_fee_after_partial_payments': total_fee_after_partial_payments,
                    'total_fee': total_fee,
                }
                feesList.append(fee_data)

            # Final student data
            student_data = {
                'name': name,
                'parent': parent,
                'class': f"{class_config_data['class_id']} {class_config_data['group_id']} {class_config_data['section_id']}",
                'roll': roll,
                'total_netTotal': total_net_total,
            }

            return JsonResponse({'student': student_data, 'fees': feesList}, status=200)

        except StudentProfile.DoesNotExist:
            return JsonResponse({'error': 'Student not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=400)



import logging

from decimal import Decimal, InvalidOperation
logger = logging.getLogger(__name__)
@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))

# @csrf_exempt
# # def collect_fees(request):
# #     if request.method == 'POST':
# #         data = json.loads(request.body.decode('utf-8'))
# #         student_id = data.get('studentId')
# #         fees_status = data.get('status')
# #         selected_fee_ids = data.get('selectedFeeIds')
# #         discounts = data.get('discountAmounts')
# #         input_data = data.get('inputData', {})

# #         current_user = request.user
# #         total_fees_amount = 0
# #         total_late_fees = 0
# #         total_discount_amount = 0
# #         total_paid_fees_amount = 0
# #         collected_fees = []

# #         try:
# #             student_profile = StudentProfile.objects.get(id=student_id)

# #             for fee_id, discount in zip(selected_fee_ids, discounts):
# #                 fee = Fees.objects.get(id=fee_id)
# #                 latefee = fee.calculate_late_fee() or fee.late_amount
# #                 payment_amount = Decimal(input_data.get(str(fee_id), '0.00'))
# #                 discount_amount = Decimal(discount or '0.00')

# #                 if fees_status == 'collectfees':
# #                     if payment_amount > 0:
# #                         fee.late_amount = latefee
# #                         fee.discount_amount = discount_amount
# #                         fee.created_by = current_user
# #                         fee.save(update_fields=['late_amount', 'discount_amount', 'updated_by'])
# #                         fee.record_payment(payment_amount, current_user)
# #                     else:
# #                         fee.status = 'paid'
# #                         fee.late_amount = latefee
# #                         fee.discount_amount = discount_amount
# #                         fee.created_by = current_user
# #                         fee.save()

# #                 # Create Receive and Journal Transactions for multiple fees
# #                 create_receive_and_journal(fee, payment_amount, current_user)

# #                 collected_fees.append({
# #                     'fee_id': fee_id,
# #                     'fees_name': fee.feetype_id.fees_title,
# #                     'fees_amount': fee.amount,
# #                     'discount_amount': fee.discount_amount,
# #                     'late_amount': fee.late_amount,
# #                     'total_fee': fee.total_netTotal(),
# #                     'status': fee.status
# #                 })

# #                 total_paid_fees_amount += fee.total_netTotal()
# #                 total_fees_amount += fee.amount
# #                 total_late_fees += fee.late_amount
# #                 total_discount_amount += fee.discount_amount

# #             due_total = (total_fees_amount + total_late_fees) - (total_discount_amount + total_paid_fees_amount)

# #             return JsonResponse({
# #                 'fees': collected_fees,
# #                 'total_fees_amount': total_fees_amount,
# #                 'total_late_fees': total_late_fees,
# #                 'total_paid_fees_amount': total_paid_fees_amount,
# #                 'total_discount_amount': total_discount_amount,
# #                 'due_total': due_total
# #             }, status=200)

# #         except StudentProfile.DoesNotExist:
# #             return JsonResponse({'error': 'Student profile not found'}, status=404)

# #         except Fees.DoesNotExist:
# #             return JsonResponse({'error': 'Fee not found'}, status=404)

# #         except Exception as e:
# #             print(f"Error in collect_fees: {e}")
# #             return JsonResponse({'error': str(e)}, status=500)

# #     return JsonResponse({'msg': 'Invalid request'}, status=400)


#fee_collect via zihad
@csrf_exempt
def collect_fees(request):
    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        print("data", data)
        student_id = data.get('studentId')
        fees_status = data.get('status')
        selected_fee_ids = data.get('selectedFeeIds')
        discounts = data.get('discountAmounts')
        input_data = data.get('inputData', {})  
        current_user = request.user
        
        try:
            fees_collected = []
            total_fees_amount = 0
            total_late_fees = 0
            total_discount_amount = 0
            total_paid_fees_amount = 0
            total_partial_paid = 0
            partial_paid = 0
            
            student_profile = StudentProfile.objects.get(id=student_id)
            name = student_profile.student_field.name
            class_id = student_profile.class_id.class_group_id.class_id.name
            section_id = student_profile.class_id.section_id.name if student_profile.class_id and student_profile.class_id.section_id else "No Section"
            phone_number = student_profile.parent_id.phone_number if student_profile.parent_id else "N/A"
            formatted_number = f'880{phone_number.lstrip("0")}' if phone_number != "N/A" else "N/A"
            total_fee_paid = None
            
            for fee_id, discount in zip(selected_fee_ids, discounts):
                fee = Fees.objects.get(id=fee_id)
                fee_month = fee.month_id.name
                latefee = fee.calculate_late_fee() or fee.late_amount
                print("input", input_data.get(str(fee_id)))
                
                if fees_status == 'collectfees' and input_data.get(str(fee_id)):
                    print("check1")
                    payment_value = input_data.get(str(fee_id), '0.00')
                    
                    if not payment_value or str(payment_value).strip() == '':
                        payment_value = '0.00'
                    
                    try:
                        payment_amount = Decimal(payment_value)
                    except Exception as e:
                        return JsonResponse({'error': f'Invalid payment amount ({payment_value}) for fee ID {fee_id}: {str(e)}'}, status=400)
                    
                    fee.late_amount = latefee
                    fee.discount_amount = discount
                    fee.created_by = current_user
                    fee.save(update_fields=['late_amount', 'discount_amount', 'updated_by'])
                    fee.record_payment(payment_amount, current_user)
                    
                    partial_instance = PartialPayment.objects.filter(fee=fee.id).first()
                    payment_amount = partial_instance.amount if partial_instance else Decimal('0.00')
                    create_receive_and_journal(fee, payment_amount, current_user)
                
                elif fees_status == 'collectfees':
                    print("check2")
                    status_checking = fee.status
                    if status_checking == 'unpaid':
                        fee.status = 'paid'
                        fee.late_amount = latefee
                        fee.discount_amount = discount
                        fee.created_by = current_user
                        fee.save()
                        create_receive_and_journal(fee, fee.total_netTotal(), current_user)
                    elif status_checking == 'partial':
                        print("check3")
                        fee.status = 'paid'
                        fee.late_amount = latefee
                        fee.discount_amount = discount or fee.discount_amount
                        previous_amount = fee.total_paid_amount()
                        payment_amount = fee.total_fee() - previous_amount - fee.discount_amount
                        fee.created_by = current_user
                        fee.save(update_fields=['late_amount', 'discount_amount', 'updated_by'])
                        fee.record_payment(payment_amount, current_user)
                        

                        create_receive_and_journal(fee, payment_amount, current_user)
                
                total_fee_amount = fee.amount  
                total_late_fee = fee.late_amount
                
                if fee.status == 'paid':
                    total_fee_paid = fee.total_netTotal()
                elif fee.status == 'partial':
                    partial_instance = PartialPayment.objects.filter(fee=fee.id).first()
                    total_fee_paid = partial_instance.amount if partial_instance else Decimal('0.00')
                
                partial_payments = PartialPayment.objects.filter(fee=fee.id).order_by('payment_date')
                if partial_payments.exists():
                    previous_partial_payments = partial_payments.exclude(id=partial_payments.last().id)
                    partial_paid = previous_partial_payments.aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
                    total_fee_paid = partial_payments.last().amount
                
                total_discount_amount += fee.discount_amount
                
                fees_collected.append({
                    'fee_id': fee_id,
                    'fees_name': fee.feetype_id.fees_title,
                    'fees_amount': fee.amount,
                    'discount_amount': fee.discount_amount,
                    'late_amount': fee.late_amount,
                    'total_fee': total_fee_paid,
                    'partial_paid': partial_paid,
                    'status': fee.status
                })
                
                if total_fee_paid is not None:
                    total_paid_fees_amount += total_fee_paid
                if total_fee_amount is not None:
                    total_fees_amount += total_fee_amount   
                if total_late_fee is not None:
                    total_late_fees += total_late_fee   
                if partial_paid is not None:
                    total_partial_paid += partial_paid   
            
            due_total = (total_fees_amount + total_late_fees) - (total_discount_amount + total_paid_fees_amount + total_partial_paid)
            full_and_final = total_fee_paid if total_fee_paid is not None else Decimal('0.00')

            student_info = {
                'name': name,
                'class': class_id,
                'section_id': section_id,
                'phone_number': phone_number
            }

            return JsonResponse({
                'fees': fees_collected,
                'total_fees_amount': total_fees_amount,
                'total_late_fees': total_late_fees,
                'total_paid_fees_amount': total_paid_fees_amount,
                'total_discount_amount': total_discount_amount,
                'due_total': due_total,
                'full_and_final': full_and_final,
                'student_info': student_info
            }, status=200)
 
        except StudentProfile.DoesNotExist:
            return JsonResponse({'error': 'Student profile not found'}, status=404)

        except Fees.DoesNotExist:
            return JsonResponse({'error': 'Fee not found'}, status=404)

        except Exception as e:
            logger.error(f"Error in collect_fees: {str(e)}", exc_info=True)
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'msg': 'Invalid request'}, status=400)

def create_receive_and_journal(fee, payment_amount, current_user):
    try:
        if not fee or payment_amount is None or payment_amount <= Decimal('0.00'):
            raise ValueError("Invalid fee or payment amount")

        try:
            fee_head = fee.feetype_id.fees_type.fee_head
            if not fee_head:
                raise AttributeError()
        except AttributeError:
            raise ValueError("Fee structure is missing FeeHead association")

        try:
            income_ledger = Ledger.objects.get(
                name__iexact=f"{fee_head.name} Income",
                category__name='Income'
            )
            cash_ledger = Ledger.objects.get(
                name__iexact="Cash in Hand",
                category__name='Asset'
            )
        except Ledger.DoesNotExist as e:
            raise ValueError(f"Missing ledger: {str(e)}")

        timestamp = timezone.now().strftime('%Y%m%d%H%M%S')
        voucher_no = f"FEE-{fee.id}-{timestamp}"
        transaction_date = timezone.now().date()

        # Create the receive entry
        Receive.objects.create(
            voucher_no=voucher_no,
            date=transaction_date,
            amount=payment_amount,
            student=fee.student_id,
            fee_head=fee_head,
            cash_ledger=cash_ledger,
            income_ledger=income_ledger,
            description=f"Payment received for {fee.feetype_id.fees_title}",
        )

        # Force balance calculation
        cash_ledger.refresh_from_db()
        income_ledger.refresh_from_db()
        _ = cash_ledger.current_balance
        _ = income_ledger.current_balance

    except ValueError as ve:
        logger.error(f"Validation Error: {str(ve)}")
        raise
    except Exception as e:
        logger.error(f"Unexpected Error: {str(e)}", exc_info=True)
        raise ValueError("Transaction processing failed due to system error")

   
@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
@csrf_exempt
def generate_multiple_invoice(request):
    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        print(data)
        student_id = data.get('studentId')
        fees_status = data.get('status')
        selected_fee_ids = data.get('selectedFeeIds')
        discounts = data.get('discountAmounts')
        input_data = data.get('inputData', {})  
        current_user = request.user

        try:
            
            fees_collected = []
            total_fees_amount = 0
            total_late_fees = 0
            total_discount_amount = 0
            total_paid_fees_amount = 0
            total_partial_paid =0
            partial_paid = 0
            check_due =0
            student_profile = StudentProfile.objects.get(id=student_id)
            name = student_profile.student_field.name
            class_id = student_profile.class_id.class_group_id.class_id.name
            
            group_id = (student_profile.class_id.class_group_id.group_id.name if student_profile.class_id and student_profile.class_id.class_group_id.group_id
                else " ")
            
            section_id = (
                student_profile.class_id.section_id.name
                if student_profile.class_id and student_profile.class_id.section_id
                else " "
            )
            
            phone_number = student_profile.parent_id.phone_number
            formatted_number = '880' + phone_number.lstrip('0')
            total_fee_paid=None
            
            for fee_id, discount in zip(selected_fee_ids, discounts):
                
                fee = Fees.objects.get(id=fee_id)
                fee_month = fee.month_id.name
                latefee = fee.calculate_late_fee() or fee.late_amount
                
                if fees_status == 'collectfees' and input_data.get(str(fee_id)):
                    payment_amount = Decimal(input_data.get(str(fee_id), '0.00'))
                    fee.late_amount = latefee
                    fee.discount_amount =discount
                    fee.created_by = current_user
                    fee.save(update_fields=['late_amount',
                     'discount_amount', 'updated_by'])
                    fee.record_payment(payment_amount, current_user)
                    partial_instance = PartialPayment.objects.filter(fee=fee.id).first()
                    create_receive_and_journal(fee, partial_instance.amount, current_user)
                elif fees_status == 'collectfees':
                    staus_cheking = fee.status
                    if staus_cheking == 'unpaid':
                        status = 'paid'
                        fee.status = status
                        fee.late_amount = latefee
                        fee.discount_amount = discount
                        fee.created_by = current_user
                        fee.save()
                        create_receive_and_journal(fee, fee.total_netTotal(), current_user)

                    elif staus_cheking == 'partial':
                        status = 'paid'
                        fee.status = status
                        fee.late_amount = latefee
                        fee.discount_amount = discount or fee.discount_amount
                        previous_amount=fee.total_paid_amount()
                        payment_amount = fee.total_fee() -previous_amount - fee.discount_amount
                        fee.created_by = current_user
                        fee.save(update_fields=['late_amount',
                        'discount_amount', 'updated_by'])
                        fee.record_payment(payment_amount, current_user)
                        # partial_instance = PartialPayment.objects.filter(fee=fee.id).first()
                        create_receive_and_journal(fee, payment_amount, current_user)

                total_fee_amount = fee.amount  
                total_late_fee=fee.late_amount
                
                if fee.status == 'paid':
                    total_fee_paid = fee.total_netTotal()
                    
                elif fee.status == 'partial':
                    partial_instance = PartialPayment.objects.filter(fee=fee.id).first()
                    total_fee_paid = partial_instance.amount
                    if partial_instance:
                        total_fee_paid = partial_instance.amount
                    else:
                        total_fee_paid = Decimal('0.00')

                partial_payments = PartialPayment.objects.filter(fee=fee.id).order_by('payment_date')
                if partial_payments.exists():
                    previous_partial_payments = partial_payments.exclude(id=partial_payments.last().id)
                    partial_paid = previous_partial_payments.aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
                    total_fee_paid = PartialPayment.objects.filter(fee=fee.id).last().amount
                else:
                    partial_paid = Decimal('0.00')

                check_due += fee.total_netTotal()    
                total_discount_amount += fee.discount_amount
                
                fees_collected.append({ 
                    'fee_id': fee_id,
                    'fees_name': fee.feetype_id.fees_title,
                    'fees_amount': fee.amount,
                    'discount_amount': fee.discount_amount,
                    'late_amount': fee.late_amount,
                    'total_fee': total_fee_paid,
                    'partial_paid':partial_paid,
                    'status': fee.status
                })
               
                if total_fee_paid is not None:
                    total_paid_fees_amount += total_fee_paid
                if total_fee_amount is not None:
                    total_fees_amount += total_fee_amount   
                if total_late_fee is not None:
                    total_late_fees += total_late_fee   
                
                
            due_total =(total_fees_amount+total_late_fees)-(total_discount_amount+check_due)

            full_and_final = total_fee_paid

            # template = SMSTemplateNotification.objects.filter(
            #         notification_type='Pay Slip Info',
            #         notification_status='Active'
            #         ).first()
            # if fees_status == 'collectfees' and template:
                
            #     sms_body = template.body.format(
            #                 name=name,
            #                 fee_month=fee_month,
            #                 total_amount=total_paid_fees_amount
            #                 )
            #     sms_count = count_sms(sms_body)
            #     sms_limit_obj = SMSUsage.objects.filter(
            #         Msg_type='NONMASKING').first()
            #     if sms_limit_obj.total_sms < 1:
            #         print('SMS LIMIT OVER')
            #     else:
            #         receiver = f'{formatted_number}'
            #         send_sms(receiver, sms_body)
            #         sms_limit_obj.total_sms -= sms_count
            #         sms_limit_obj.save()
            #         SMS.objects.create(
            #             mobile=phone_number, title='Paid Msg', msg=sms_body, created_by=current_user)


            student_info = {
                'id': student_profile.id,
                'name': name,
                'class': class_id,
                'group_id':group_id,
                'section_id': section_id,
                'phone_number': phone_number
            }

            return JsonResponse({
                'fees': fees_collected,
                'total_fees_amount': total_fees_amount,
                'total_late_fees': total_late_fees,
                'total_paid_fees_amount': total_paid_fees_amount,
                'total_discount_amount': total_discount_amount,
                'due_total':due_total,
                'full_and_final': full_and_final,
                'student_info': student_info,
            }, status=200)

        except StudentProfile.DoesNotExist:
            return JsonResponse({'error': 'Student profile not found'}, status=404)

        except Fees.DoesNotExist:
            return JsonResponse({'error': 'Fee not found'}, status=404)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'msg': 'Invalid request'}, status=400)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
@csrf_exempt
# def fee_put_back(request):
#     current_user = request.user
#     fee_id = request.POST.get('id')
#     btn_status = request.POST.get('status')
    
#     if btn_status == "partial_payment":
#         partial_instance = get_object_or_404(PartialPayment, pk=fee_id)
#         pid = partial_instance.id
#         fee_instance = partial_instance.fee
#         p_count = PartialPayment.objects.filter(fee=fee_instance).count()
#         transaction_number = str(fee_instance.transaction_no) + '-' + str(pid)
        
#         if p_count > 1:
#             partial_instance.delete()
#             IncomeitemList.objects.filter(invoice_number=transaction_number).delete()
#         else:
#             late_amount = 0
#             discount_amount = 0
#             partial_instance.delete()
#             transaction_no = fee_instance.transaction_no
#             fee_instance.status = 'unpaid'
#             fee_instance.late_amount = late_amount
#             fee_instance.discount_amount = discount_amount
#             fee_instance.updated_by = current_user
#             fee_instance.save()
#             deleted_count, _ = IncomeitemList.objects.filter(invoice_number__startswith=transaction_no).delete()
        
#         return JsonResponse({'msg': 'Partial payment deleted'}, safe=False)
    
#     elif btn_status == "fee":
#         fees_instance = get_object_or_404(Fees, pk=fee_id)
#         PartialPayment.objects.filter(fee=fees_instance).delete()
#         transaction_no = fees_instance.transaction_no
#         late_amount = 0
#         discount_amount = 0
#         fees_instance.status = 'unpaid'
#         fees_instance.late_amount = late_amount
#         fees_instance.discount_amount = discount_amount
#         fees_instance.updated_by = current_user
#         fees_instance.save()
#         deleted_count, _ = IncomeitemList.objects.filter(invoice_number__startswith=transaction_no).delete()
        
#         return JsonResponse({'msg': 'Fee and related partial payments deleted'}, safe=False)
    
#     return JsonResponse({'msg': 'Successfully' }, safe=False)

@login_required(login_url='login')
@user_passes_test(lambda user: user.is_staff or user.groups.filter(name__in=['Manager', 'Accountant']).exists())
@csrf_exempt
def fee_put_back(request):
    current_user = request.user
    fee_id = request.POST.get('id')
    btn_status = request.POST.get('status')

    try:
        with transaction.atomic(): 

            if btn_status == "partial_payment":
                partial_instance = get_object_or_404(PartialPayment, pk=fee_id)
                pid = partial_instance.id
                fee_instance = partial_instance.fee

                transaction_number = f"FEE-{fee_instance.id}-{pid}"

                logger.info(f"Deleting Receive, Ledger Entries, and JournalTransaction for partial payment (voucher_no: {transaction_number})")

                Receive.objects.filter(voucher_no=transaction_number).delete()
                Journal.objects.filter(voucher_no=transaction_number).delete()
                JournalEntry.objects.filter(journal__voucher_no=transaction_number).delete()
                LedgerEntry.objects.filter(description__icontains=transaction_number).delete()

                logger.info("Deleted related Receive, Ledger, and Journal entries")

                if PartialPayment.objects.filter(fee=fee_instance).count() > 1:
                    partial_instance.delete()
                else:
                    partial_instance.delete()
                    fee_instance.status = 'unpaid'
                    fee_instance.late_amount = 0
                    fee_instance.discount_amount = 0
                    fee_instance.updated_by = current_user
                    fee_instance.save()

                return JsonResponse({'msg': 'Partial payment and related records deleted successfully'}, safe=False)

            elif btn_status == "fee":
                fees_instance = get_object_or_404(Fees, pk=fee_id)
                transaction_no_prefix = f"FEE-{fees_instance.id}-"

                logger.info(f"Deleting all Receive, Ledger Entries, and JournalTransaction for fee (voucher_no starts with: {transaction_no_prefix})")

                Receive.objects.filter(voucher_no__startswith=transaction_no_prefix).delete()
                Journal.objects.filter(voucher_no__startswith=transaction_no_prefix).delete()
                JournalEntry.objects.filter(journal__voucher_no__startswith=transaction_no_prefix).delete()
                LedgerEntry.objects.filter(description__icontains=transaction_no_prefix).delete()

                logger.info("Deleted related Receive, Ledger, and Journal entries for full fee reversal")

                fees_instance.status = 'unpaid'
                fees_instance.late_amount = 0
                fees_instance.discount_amount = 0
                fees_instance.updated_by = current_user
                fees_instance.save()

                return JsonResponse({'msg': 'Fee and all related records deleted successfully'}, safe=False)

    except Exception as e:
        logger.error(f"Error in fee_put_back: {str(e)}", exc_info=True)
        return JsonResponse({'msg': f"Error: {str(e)}"}, safe=False)

    return JsonResponse({'msg': 'Invalid action or parameters'}, safe=False)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def put_back_fee(request):
    academic_year = Admission_Year.objects.latest('updated_at')
    stulist = StudentProfile.objects.filter(
        Q(admission_year_id=academic_year) & Q(student_field__status="Active"))
    
    fee_instance = None
    student_instance = None

    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        student_instance = get_object_or_404(StudentProfile, pk=student_id)
        
        fee_instance = Fees.objects.filter(student_id=student_instance).filter(Q(status='paid') | Q(status='partial'))
        

    context = {
        'fee_instance': fee_instance,
        'stulist': stulist,
        'student_instance': student_instance,
        'heading': 'Fees',
        'subheading': 'Put Back Fees',
    }

    return render(request, 'crucial/finance/put_back_fee.html', context)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
@csrf_exempt
def fee_del(request):
    fee_id = request.POST.get('id')
    fees_instance = Fees.objects.filter(pk=fee_id).last()
    fees_instance.delete()
   
    return JsonResponse({'msg': 'Suceesfully'}, safe=False)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def del_fee(request):
    fee_instance,student_instance = None,None
    academic_year = Admission_Year.objects.latest('updated_at')
    stulist = StudentProfile.objects.filter(
        Q(admission_year_id=academic_year))
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        student_instance = get_object_or_404(StudentProfile, pk=student_id)
        try:
            fee_instance = Fees.objects.filter(student_id=student_instance)
            

        except Fees.DoesNotExist:
            fee_instance = None
    context = {
        'stulist':stulist,
        'fee_instance':fee_instance,
        'student_instance':student_instance,
        'heading': 'Fees',
        'subheading': 'Delete Fees',
    }
    return render(request, 'crucial/finance/fee_delete.html', context) 
    

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def student_fee(request):
    current_year = str(datetime.now().year)
    admission_year = Admission_Year.objects.get(name=current_year)

    studentlist = StudentProfile.objects.filter(
        Q(admission_year_id=admission_year) & Q(student_field__status="Active"))

    context = {
        'heading': 'Fees',
        'subheading': 'Student Fees',
        'studentlist': studentlist
    }
    return render(request, 'crucial/finance/student_fee.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user, "student", "parent"))
def showfee(request):
    feelist = []
    if request.user.is_authenticated and is_staff_or_in_group(request.user, "parent", "student"):
        requested_user = request.user
        if requested_user.groups.filter(name='student').exists():
            student_profile = StudentProfile.objects.get(
                student_field=requested_user)
            feelist = Fees.objects.filter(student_id=student_profile)

        if requested_user.groups.filter(name='parent').exists():
            children_profiles = StudentProfile.objects.filter(
                parent_id=requested_user)
            children = [
                child_profile.student_field for child_profile in children_profiles]
            student_profiles = StudentProfile.objects.filter(
                student_field__in=children)
            feelist = Fees.objects.filter(student_id__in=student_profiles)

        total_late_fee = Decimal('0.00')
        for fee in feelist:
            if fee.status in ['paid', 'partial']:
                total_late_fee += fee.late_amount
            elif fee.status == 'unpaid':
                total_late_fee += fee.calculate_late_fee()

        total_late_fee = round(total_late_fee, 0)

        total_amount = feelist.aggregate(total_amount=Sum('amount'))['total_amount'] or 0
        total_amount = total_amount+total_late_fee
        total_discount_amount = feelist.aggregate(total_discount=Sum('discount_amount'))['total_discount'] or 0
        total_amount_paid_fees = feelist.filter(status='paid').aggregate(
            total_amount=Sum(
                ExpressionWrapper(F('amount') + F('late_amount') - F('discount_amount'), output_field=DecimalField())
            )
        )['total_amount'] or 0
        partial_feelist = feelist.filter(status='partial')

        total_partial_payments = PartialPayment.objects.filter(fee__in=partial_feelist).aggregate(
            total_paid=Sum('amount')
        )['total_paid'] or 0

        total_amount_paid = total_amount_paid_fees + total_partial_payments

        total_amount_unpaid = total_amount - total_amount_paid  -total_discount_amount

    context = {
        'total_amount': total_amount,
        'total_discount_amount': total_discount_amount,
        'total_amount_paid': total_amount_paid,
        'total_amount_unpaid': total_amount_unpaid,
        'feelist': feelist,
        'heading': 'Fees',
        'subheading': 'Student Fees',
    }
    return render(request, 'crucial/finance/show_fee_list.html', context) 


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def student_date_fee_report(request):
    datewise_paid_counts = None
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

        # Filter for both 'paid' and 'partial' statuses
        filtered_students = Fees.objects.filter(
            Q(status='paid') | Q(status='partial'),
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        )

        # Annotations to compute the total fee after discount
        total_fee_after_discount = (
            F('amount') +
            F('late_amount') -
            F('discount_amount')
        )

        
        partial_payment_sum = PartialPayment.objects.filter(
            fee=OuterRef('id')
        ).values('fee').annotate(
            total_paid=Sum('amount')
        ).values('total_paid')[:1]

        datewise_paid_counts = filtered_students.values('created_at__date').annotate(
            total_students=Count('student_id', distinct=True),
            paid_fees_count=Count(
                Case(When(status="paid", then=1), output_field=IntegerField())
            ),
            partial_fees_count=Count(
                Case(When(status="partial", then=1), output_field=IntegerField())
            ),
            paid_fees_amount=Sum(
                Case(When(status="paid", then=total_fee_after_discount),
                     default=Decimal('0.0'), output_field=DecimalField())
            ),
            partial_fees_amount=Sum(
                Case(When(status="partial", then=Subquery(partial_payment_sum)),
                     default=Decimal('0.0'), output_field=DecimalField())
            ),
            total_amount=Sum(
                Case(When(Q(status="paid") | Q(status="partial"), then=total_fee_after_discount),
                     default=Decimal('0.0'), output_field=DecimalField())
            )
        )

    context = {
        'datewise_paid_counts': datewise_paid_counts,
        'heading': 'Fees',
        'subheading': 'Date Wise Paid Fee',
    }
    return render(request, 'crucial/report/student_date_fee_report.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def student_time_fee_report(request):
    paid_students = None

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

        # Filter for both 'paid' and 'partial' statuses
        filtered_students = Fees.objects.filter(
            Q(status='paid') | Q(status='partial'),
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        )

        # Subquery for summing partial payments
        partial_payment_sum = PartialPayment.objects.filter(
            fee=OuterRef('id')
        ).values('fee').annotate(
            total_paid=Sum('amount')
        ).values('total_paid')[:1]

        paid_students = filtered_students.values(
            'student_id__student_field__name',
            'student_id__class_id__class_group_id__class_id__name',
            'student_id__class_id__class_group_id__group_id__name',
            'student_id__class_id__section_id__name',
            'student_id__roll_no',
            'created_at__date',
            'status'  # Include status in the query
        ).annotate(
            total_paid=Sum(
                Case(
                    When(status='paid', then=F('amount')),
                    When(status='partial', then=Subquery(partial_payment_sum)),
                    default=0,
                    output_field=DecimalField()
                )
            )
        )

    context = {
        'paid_students': paid_students,
        'heading': 'Fees',
        'subheading': 'Student Wise Paid Fees',
    }
    return render(request, 'crucial/report/student_time_fee_report.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def duefee(request):
    academic_year_list = Admission_Year.objects.all()
    session_year_list = AcademicSession.objects.all()
    classList = ClassConfig.objects.all()
    version_list = StudentProfile.Version.choices

    total_unpaid_amount = None
    total_student_unpaid_amount = None 
    class_instance = None
    selected_version = None

    if request.method == 'POST':
        class_id = request.POST.get('class_name_id')
        class_instance = get_object_or_404(ClassConfig, pk=class_id)
        academic_year_id = request.POST.get('academic_year_id')
        session_year_id = request.POST.get('session_year_id')
        selected_version = request.POST.get('version')
        
        # Base query
        query = Q(
            student_id__class_id=class_instance,
            status__in=['unpaid', 'partial']
        )

        if academic_year_id and academic_year_id.isdigit():
            query &= Q(academic_year=academic_year_id)
        else:
            query

        # Add session year condition
        if session_year_id and session_year_id.isdigit():
            query &= Q(student_id__academic_session_year=session_year_id)
        else:
            query 

        # Add version filter
        if selected_version:
            query &= Q(student_id__version=selected_version)

        unpaid_feeslist = Fees.objects.filter(query).order_by('-student_id__version')

        # Calculate totals
        total_unpaid_amount = sum(fee.total_fee_after_partial_payments() for fee in unpaid_feeslist)
        
        try:
            total_student_unpaid_amount = {}
            for fee in unpaid_feeslist:
                student_id = fee.student_id.id
                fees_amount = fee.total_fee_after_discount()
                discount = fee.discount_amount
                partial_pay = fee.total_paid_amount()
                dues = fee.total_fee_after_partial_payments()

                if student_id not in total_student_unpaid_amount:
                    total_student_unpaid_amount[student_id] = {
                        'name': fee.student_id.student_field.name,
                        'roll': fee.student_id.roll_no,
                        'class': fee.student_id.class_id,
                        'version': fee.student_id.version,  # Added version here
                        'total_fees_amount': fees_amount,
                        'total_discount': discount,
                        'total_partial_pay': partial_pay,
                        'total_dues': dues,
                        'feetypes': {}
                    }
                else:
                    total_student_unpaid_amount[student_id]['total_fees_amount'] += fees_amount
                    total_student_unpaid_amount[student_id]['total_discount'] += discount
                    total_student_unpaid_amount[student_id]['total_partial_pay'] += partial_pay
                    total_student_unpaid_amount[student_id]['total_dues'] += dues

                # Update fee types details
                feetype = fee.feetype_id.fees_title
                if feetype not in total_student_unpaid_amount[student_id]['feetypes']:
                    total_student_unpaid_amount[student_id]['feetypes'][feetype] = {
                        'fees_amount': fees_amount,
                        'discount': discount,
                        'partial_pay': partial_pay,
                        'dues': dues,
                        'status': fee.status
                    }
                else:
                    total_student_unpaid_amount[student_id]['feetypes'][feetype]['fees_amount'] += fees_amount
                    total_student_unpaid_amount[student_id]['feetypes'][feetype]['discount'] += discount
                    total_student_unpaid_amount[student_id]['feetypes'][feetype]['partial_pay'] += partial_pay
                    total_student_unpaid_amount[student_id]['feetypes'][feetype]['dues'] += dues

        except Exception as e:
            print(f"Error processing fees: {str(e)}")

    context = {
        'classList': classList,
        'academic_year_list': academic_year_list,
        'session_year_list': session_year_list,
        'version_list': version_list,  
        'total_unpaid_amount': total_unpaid_amount,
        'total_student_unpaid_amount': total_student_unpaid_amount,
        'class_instance': class_instance,
        'selected_version': selected_version,
        'heading': 'Fees',
        'subheading': 'Due Fees',
    }
    return render(request, 'crucial/report/due_fee.html', context)


import xlwt
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from .models import Fees, ClassConfig
from django.db.models import Q
from django.shortcuts import get_object_or_404


# def export_due_fees(request):
#     export_type = request.GET.get('export_type', 'pdf')  # Default to 'pdf' if not specified

#     class_id = request.GET.get('class_id')
#     class_instance = get_object_or_404(ClassConfig, pk=class_id)
#     academic_year_id = request.GET.get('academic_year_id')
#     session_year_id = request.GET.get('session_year_id')
#     selected_version = request.GET.get('version')

#     query = Q(student_id__class_id=class_instance, status__in=['unpaid', 'partial'])

#     if academic_year_id and academic_year_id.isdigit():
#         query &= Q(academic_year=academic_year_id)
#     if session_year_id and session_year_id.isdigit():
#         query &= Q(student_id__academic_session_year=session_year_id)
#     if selected_version:
#         query &= Q(student_id__version=selected_version)

#     unpaid_feeslist = Fees.objects.filter(query).order_by('-student_id__version')

#     # Summarize fees by student
#     student_fees_summary = {}
#     for fee in unpaid_feeslist:
#         student_id = fee.student_id.id
#         fees_amount = fee.total_fee_after_discount()
#         discount = fee.discount_amount
#         partial_pay = fee.total_paid_amount()
#         dues = fee.total_fee_after_partial_payments()

#         if student_id not in student_fees_summary:
#             student_fees_summary[student_id] = {
#                 'name': fee.student_id.student_field.name,
#                 'roll': fee.student_id.roll_no,
#                 'class': str(fee.student_id.class_id),
#                 'version': fee.student_id.version,
#                 'total_fees_amount': fees_amount,
#                 'total_discount': discount,
#                 'total_partial_pay': partial_pay,
#                 'total_dues': dues,
#             }
#         else:
#             student_fees_summary[student_id]['total_fees_amount'] += fees_amount
#             student_fees_summary[student_id]['total_discount'] += discount
#             student_fees_summary[student_id]['total_partial_pay'] += partial_pay
#             student_fees_summary[student_id]['total_dues'] += dues

#     if export_type == 'excel':
#         response = HttpResponse(content_type='application/ms-excel')
#         response['Content-Disposition'] = 'attachment; filename="due_fees_report.xls"'

#         wb = xlwt.Workbook(encoding='utf-8')
#         ws = wb.add_sheet('Due Fees')

#         # Header
#         columns = ['Student Name', 'Roll No', 'Class', 'Version', 'Total Fees', 'Total Discount', 'Total Partial Pay', 'Total Dues']
#         for col_num, column_title in enumerate(columns):
#             ws.write(0, col_num, column_title, xlwt.easyxf('font: bold 1'))

#         row_num = 1
#         total_fees = total_discount = total_partial_pay = total_dues = 0

#         for student in student_fees_summary.values():
#             ws.write(row_num, 0, student['name'])
#             ws.write(row_num, 1, student['roll'])
#             ws.write(row_num, 2, student['class'])
#             ws.write(row_num, 3, student['version'])
#             ws.write(row_num, 4, student['total_fees_amount'])
#             ws.write(row_num, 5, student['total_discount'])
#             ws.write(row_num, 6, student['total_partial_pay'])
#             ws.write(row_num, 7, student['total_dues'])

#             total_fees += student['total_fees_amount']
#             total_discount += student['total_discount']
#             total_partial_pay += student['total_partial_pay']
#             total_dues += student['total_dues']
#             row_num += 1

#         # Add a total row
#         ws.write(row_num, 0, 'Total')
#         ws.write(row_num, 4, total_fees)
#         ws.write(row_num, 5, total_discount)
#         ws.write(row_num, 6, total_partial_pay)
#         ws.write(row_num, 7, total_dues)

#         wb.save(response)
#         return response

#     elif export_type == 'pdf':
#         response = HttpResponse(content_type='application/pdf')
#         response['Content-Disposition'] = 'attachment; filename="due_fees_report.pdf"'

#         c = canvas.Canvas(response, pagesize=letter)
#         c.setFont("Helvetica-Bold", 14)
#         c.drawString(30, 750, "Due Fees Report")

#         y_position = 720
#         total_fees = total_discount = total_partial_pay = total_dues = 0

#         c.setFont("Helvetica", 12)
#         for student in student_fees_summary.values():
#             c.drawString(30, y_position, f"Student: {student['name']} | Roll No: {student['roll']} | Class: {student['class']} | Version: {student['version']}")
#             c.drawString(30, y_position - 20, f"Total Fees: {student['total_fees_amount']} | Discount: {student['total_discount']} | Partial Paid: {student['total_partial_pay']} | Dues: {student['total_dues']}")
#             y_position -= 50

#             total_fees += student['total_fees_amount']
#             total_discount += student['total_discount']
#             total_partial_pay += student['total_partial_pay']
#             total_dues += student['total_dues']

#             if y_position < 50:
#                 c.showPage()
#                 c.setFont("Helvetica", 12)
#                 y_position = 750

#         # Add total at the end of the PDF
#         c.setFont("Helvetica-Bold", 12)
#         c.drawString(30, y_position, f"Grand Total: Total Fees: {total_fees}, Total Discount: {total_discount}, Total Partial Pay: {total_partial_pay}, Total Dues: {total_dues}")
#         c.save()
#         return response

#     else:
#         return HttpResponse("Invalid export type", status=400)

def export_due_fees(request):
    export_type = request.GET.get('export_type', 'pdf')
    class_id = request.GET.get('class_id')
    class_instance = get_object_or_404(ClassConfig, pk=class_id)
    academic_year_id = request.GET.get('academic_year_id')
    session_year_id = request.GET.get('session_year_id')
    selected_version = request.GET.get('version')  # This should be the key (e.g., 'bn', 'en')

    # Build the query with all filters including version
    query = Q(
        student_id__class_id=class_instance,
        status__in=['unpaid', 'partial']
    )

    if academic_year_id and academic_year_id.isdigit():
        query &= Q(academic_year=academic_year_id)
    
    if session_year_id and session_year_id.isdigit():
        query &= Q(student_id__academic_session_year=session_year_id)
    
    # Apply version filter if selected_version is provided
    if selected_version:
        query &= Q(student_id__version=selected_version)

    unpaid_feeslist = Fees.objects.filter(query).order_by('-student_id__version')

    # Summarize fees by student
    student_fees_summary = {}
    for fee in unpaid_feeslist:
        student_id = fee.student_id.id
        fees_amount = fee.total_fee_after_discount()
        discount = fee.discount_amount
        partial_pay = fee.total_paid_amount()
        dues = fee.total_fee_after_partial_payments()

        if student_id not in student_fees_summary:
            student_fees_summary[student_id] = {
                'name': fee.student_id.student_field.name,
                'roll': fee.student_id.roll_no,
                'class': str(fee.student_id.class_id),
                'version': fee.student_id.get_version_display(),  
                'total_fees_amount': fees_amount,
                'total_discount': discount,
                'total_partial_pay': partial_pay,
                'total_dues': dues,
            }
        else:
            student_fees_summary[student_id]['total_fees_amount'] += fees_amount
            student_fees_summary[student_id]['total_discount'] += discount
            student_fees_summary[student_id]['total_partial_pay'] += partial_pay
            student_fees_summary[student_id]['total_dues'] += dues


    # Excel Export Logic
    if export_type == 'excel':
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="due_fees_report.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Due Fees')

        # Header
        columns = ['Student Name', 'Roll No', 'Class', 'Version', 'Total Fees', 'Total Discount', 'Total Partial Pay', 'Total Dues']
        for col_num, column_title in enumerate(columns):
            ws.write(0, col_num, column_title, xlwt.easyxf('font: bold 1'))

        row_num = 1
        total_fees = total_discount = total_partial_pay = total_dues = 0

        for student in student_fees_summary.values():
            # Since the query already filters by version, no need to check version here
            ws.write(row_num, 0, student['name'])
            ws.write(row_num, 1, student['roll'])
            ws.write(row_num, 2, student['class'])
            ws.write(row_num, 3, student['version'])
            ws.write(row_num, 4, student['total_fees_amount'])
            ws.write(row_num, 5, student['total_discount'])
            ws.write(row_num, 6, student['total_partial_pay'])
            ws.write(row_num, 7, student['total_dues'])

            total_fees += student['total_fees_amount']
            total_discount += student['total_discount']
            total_partial_pay += student['total_partial_pay']
            total_dues += student['total_dues']
            row_num += 1

        # Add a total row at the end
        ws.write(row_num, 0, 'Total')
        ws.write(row_num, 4, total_fees)
        ws.write(row_num, 5, total_discount)
        ws.write(row_num, 6, total_partial_pay)
        ws.write(row_num, 7, total_dues)

        wb.save(response)
        return response

    elif export_type == 'pdf':
        # PDF generation logic
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="due_fees_report.pdf"'

        c = canvas.Canvas(response, pagesize=letter)
        c.setFont("Helvetica-Bold", 14)
        c.drawString(30, 750, "Due Fees Report")

        y_position = 720
        total_fees = total_discount = total_partial_pay = total_dues = 0

        c.setFont("Helvetica", 12)
        for student in student_fees_summary.values():
            c.drawString(30, y_position, f"Student: {student['name']} | Roll No: {student['roll']} | Class: {student['class']} | Version: {student['version']}")
            c.drawString(30, y_position - 20, f"Total Fees: {student['total_fees_amount']} | Discount: {student['total_discount']} | Partial Paid: {student['total_partial_pay']} | Dues: {student['total_dues']}")
            y_position -= 50

            total_fees += student['total_fees_amount']
            total_discount += student['total_discount']
            total_partial_pay += student['total_partial_pay']
            total_dues += student['total_dues']

            if y_position < 50:
                c.showPage()
                c.setFont("Helvetica", 12)
                y_position = 750

        c.setFont("Helvetica-Bold", 12)
        c.drawString(30, y_position, f"Grand Total: Total Fees: {total_fees}, Total Discount: {total_discount}, Total Partial Pay: {total_partial_pay}, Total Dues: {total_dues}")
        c.save()
        return response

    else:
        return HttpResponse("Invalid export type", status=400)



@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))

def fee_type_wise_class_report(request):
    academic_year_list = Admission_Year.objects.all()
    session_year_list=AcademicSession.objects.all()
    classList = ClassConfig.objects.all()
    monthList = Fee_month.objects.all()
    class_instance,month_instance = None,None
    report_data = None
    report_summary = None
    
    if request.method == 'POST':
        class_id = request.POST.get('class_name_id')
        month_id = request.POST.get('fee_month_id')
        academic_year_id = request.POST.get('academic_year_id')
        session_year_id = request.POST.get('session_year_id')
        
        class_instance = get_object_or_404(ClassConfig, pk=class_id)
        month_instance = get_object_or_404(Fee_month, pk=month_id)
        
        students_in_class = StudentProfile.objects.filter(
            Q(class_id=class_instance) & 
            Q(student_field__status="Active") & 
            Q(admission_year_id=academic_year_id) & 
            (Q(academic_session_year=session_year_id) | Q(academic_session_year__isnull=True))
        )
        
        report_data = Fees.objects.filter(
            academic_year=academic_year_id,
            student_id__in=students_in_class,
            month_id=month_instance
        ).select_related('student_id', 'feetype_id')
        
        fee_statuses = ['paid', 'unpaid', 'partial']
        report_summary = {}
        
        for status in fee_statuses:
            status_fees = report_data.filter(status=status)
            if status=='unpaid':
                total_amount = sum(fee.total_fee() for fee in status_fees)
            else:
                total_amount = sum(fee.total_netTotal() for fee in status_fees)

            students = []
            for fee in status_fees:
                students.append({
                    'student_id': fee.student_id.id,
                    'student_name': str(fee.student_id),
                    'fee_type':fee.feetype_id.fees_type.fee_head, 
                    'student_roll': fee.student_id.roll_no or "N/A",
                    'paid_amount': round(fee.total_netTotal(), 2),
                    'unpaid_amount': round(fee.total_fee(), 2),
                    'partial_amount': round(fee.total_paid_amount(), 2),
                })

            sorted_students = sorted(students, key=lambda x: x['student_id'])

            report_summary[status] = {
                'students': sorted_students,
                'total_amount': round(total_amount, 2)
            }
            
             
    context = {
        'heading': 'Fees',
        'subheading': 'Class wise Total Month Fee',
        'classList': classList,
        'monthList': monthList,
        'session_year_list':session_year_list,
        'academic_year_list':academic_year_list,
        'class_instance':class_instance,
        'month_instance':month_instance,
        'report_data': report_data,
        'report_summary': report_summary,
        'selected_class': class_id if request.method == 'POST' else None,
        'selected_month': month_id if request.method == 'POST' else None,
    }
    
    return render(request, 'crucial/report/month_wise_fee_report_cls.html', context)



@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))

def fee_type_wise_report(request):
    academic_year_list = Admission_Year.objects.all()
    monthList = Fee_month.objects.all()
    month_instance,admission_year = None,None
    report_data = None
    report_summary = None
    
    if request.method == 'POST':

        month_id = request.POST.get('fee_month_id')
        academic_year_id = request.POST.get('academic_year_id')
        admission_year= Admission_Year.objects.get(id=academic_year_id)
        month_instance = get_object_or_404(Fee_month, pk=month_id)
        
        # Filter fees based on the selected class, month, and session year
        report_data = Fees.objects.filter(
            academic_year=academic_year_id,
            month_id=month_instance
        ).select_related('student_id', 'feetype_id')
        print(report_data)
        # Separate students by fee status and aggregate the amounts
        fee_statuses = ['paid', 'unpaid', 'partial']
        report_summary = {}
        
        for status in fee_statuses:
            status_fees = report_data.filter(status=status)
            if status=='unpaid':
                total_amount = sum(fee.total_fee() for fee in status_fees)
            else:
                total_amount = sum(fee.total_netTotal() for fee in status_fees)

            # Prepare student data with each student's fee details
            students = []
            for fee in status_fees:
                students.append({
                    'student_id': fee.student_id.id,
                    'student_name': str(fee.student_id),
                    'student_class': str(fee.student_id.class_id),
                    'fee_type':fee.feetype_id.fees_type.fee_head,
                    'student_roll': fee.student_id.roll_no or "N/A",
                    'paid_amount': round(fee.total_netTotal(), 2),
                    'unpaid_amount': round(fee.total_fee(), 2),
                    'partial_amount': round(fee.total_paid_amount(), 2),
                })

            sorted_students = sorted(students, key=lambda x: x['student_id'])

            report_summary[status] = {
                'students': sorted_students,
                'total_amount': round(total_amount, 2)
            }
            
            
    context = {
        'heading': 'Fees',
        'subheading': 'Monthly wise Total Fee',
        'monthList': monthList,
        'academic_year_list':academic_year_list,
        'month_instance':month_instance,
        'admission_year':admission_year,
        'report_data': report_data,
        'report_summary': report_summary,
        'selected_month': month_id if request.method == 'POST' else None,
    }
    
    return render(request, 'crucial/report/month_wise_fee_report.html', context) 
    

# ------------------------------Hostel Package--------------------------


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def packagetype(request):
    context = {
        'heading': 'Hostel',
        'subheading': 'Hostel Package Type',
    }
    return render(request, 'crucial/hostel/package_type.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def list_package_type(request):
    admission_year = Admission_Year.objects.latest('updated_at')
    package_type_list = Hostel_package.objects.filter(
        academic_year=admission_year).order_by("-id")
    context = {
        'package_type_list': package_type_list
    }
    return render(request, 'crucial/hostel/list_package_type.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def add_package_type(request):
    admission_year = Admission_Year.objects.latest('updated_at')
    form = Hosteltypeform(initial={'academic_year': admission_year})

    if request.method == 'POST':
        form = Hosteltypeform(request.POST)
        if form.is_valid():
            # Replace with the relevant field(s)
            package_name = form.cleaned_data['package_name']
            # Replace with the relevant field(s)
            academic_year = form.cleaned_data['academic_year']
            if Hostel_package.objects.filter(package_name=package_name, academic_year=academic_year).exists():
                messages.success(
                    request, 'This package type for the selected session year already exists ! ! !')
                return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                    "messages": 'Success',
                    "packageChanged": "packageChanged"
                })})
            else:
                form.save()
                messages.success(
                    request, 'Hostel Package has been Saved ! ! !')
                return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                    "messages": 'Success',
                    "packageChanged": "packageChanged"
                })})

    context = {
        'form': form
    }

    return render(request, 'crucial/hostel/add_package.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def edit_package_type(request, pk):
    package_type = get_object_or_404(Hostel_package, pk=pk)
    if request.method == 'POST':
        form = Hosteltypeform(request.POST, instance=package_type)
        if form.is_valid():
            form.save()
            messages.success(
                request, 'Hostel Package type has been Updated ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "packageChanged": "packageChanged"
            })})
    else:
        form = Hosteltypeform(instance=package_type)

    context = {
        'form': form
    }

    return render(request, 'crucial/hostel/add_package.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def del_package_type(request, pk):
    package_type = get_object_or_404(Hostel_package, pk=pk)
    package_type.delete()
    return HttpResponse(status=204, headers={'HX-Trigger': 'packageChanged'})
# ------------------------------Hostel view --------------------------


@login_required(login_url='auth_login')
def hostel_package_allocate(request):
    admission_year = Admission_Year.objects.latest('updated_at')
    classList = ClassConfig.objects.all()
    package_list = Hostel_package.objects.filter(academic_year=admission_year)
    hostel_student_instance = Hostel.objects.filter(
        academic_year=admission_year
    ).exclude(student_id__isnull=True).exclude(hostel_package__isnull=True)  # Exclude invalid entries

    student_ids = hostel_student_instance.values_list('student_id', flat=True)
    class_instance, id_profiles = None, None

    if request.method == "POST":
        class_id = request.POST.get('class_id')
        class_instance = get_object_or_404(ClassConfig, pk=class_id)
        id_profiles = StudentProfile.objects.filter(
            Q(class_id=class_instance) & Q(student_field__status="Active")).order_by("roll_no")
    else:
        id_profiles = StudentProfile.objects.filter(id__in=student_ids)

    # Create map while handling None values
    student_package_map = {
        hostel.student_id.id: hostel.hostel_package.id
        for hostel in hostel_student_instance
    }

    context = {
        "id_profiles": id_profiles,
        "class_instance": class_instance,
        "classList": classList,
        "package_list": package_list,
        'student_package_map': json.dumps(student_package_map),
        'heading': 'Student',
        'subheading': 'Hostel Package Allocation'
    }
    return render(request, 'crucial/hostel/student_hostel.html', context)


@login_required(login_url='auth_login')
@csrf_exempt
def update_hostel_package(request):
    if request.method == "POST":
        student_id = request.POST.get('student_id')
        new_package_id = request.POST.get('package')
        try:
            student = StudentProfile.objects.get(id=student_id)
            package = Hostel_package.objects.get(id=new_package_id)
            admission_year = Admission_Year.objects.latest('updated_at')

            # Update or create the Hostel record for the student
            updated_values = {'hostel_package': package,
                              'academic_year': admission_year}
            obj, created = Hostel.objects.update_or_create(
                student_id=student, defaults=updated_values
            )
            return JsonResponse({'status': 'success'})
        except StudentProfile.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Student not found'})
        except Hostel_package.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Package not found'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})


# ------------------------------Tution Package--------------------------

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def tution_packagetype(request):
    context = {
        'heading': 'Coaching',
        'subheading': 'Coaching Package Type',
    }
    return render(request, 'crucial/tution/package_type.html', context)


@login_required(login_url='login')
# @user_passes_test(lambda user: is_staff_or_in_group(user, role='Accountant'))
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def list_tution_package_type(request):
    admission_year = Admission_Year.objects.latest('updated_at')
    package_type_list = Tution_package.objects.filter(
        academic_year=admission_year).order_by("-id")
    context = {
        'package_type_list': package_type_list
    }
    return render(request, 'crucial/tution/list_package_type.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def add_tution_package_type(request):
    admission_year = Admission_Year.objects.latest('updated_at')
    form = Tutiontypeform(initial={'academic_year': admission_year})

    if request.method == 'POST':
        form = Tutiontypeform(request.POST)
        if form.is_valid():
            # Replace with the relevant field(s)
            package_name = form.cleaned_data['package_name']
            # Replace with the relevant field(s)
            academic_year = form.cleaned_data['academic_year']
            if Tution_package.objects.filter(package_name=package_name, academic_year=academic_year).exists():
                messages.success(
                    request, 'This package type for the selected session year already exists ! ! !')
                return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                    "messages": 'Success',
                    "packageChanged": "tpackageChanged"
                })})
            else:
                form.save()
                messages.success(
                    request, 'Tution Package has been Saved ! ! !')
                return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                    "messages": 'Success',
                    "packageChanged": "tpackageChanged"
                })})

    context = {
        'form': form
    }

    return render(request, 'crucial/tution/add_package.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def edit_tution_package_type(request, pk):
    package_type = get_object_or_404(Tution_package, pk=pk)
    if request.method == 'POST':
        form = Tutiontypeform(request.POST, instance=package_type)
        if form.is_valid():
            form.save()
            messages.success(
                request, 'Tution Package type has been Updated ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "packageChanged": "tpackageChanged"
            })})
    else:
        form = Tutiontypeform(instance=package_type)

    context = {
        'form': form
    }

    return render(request, 'crucial/tution/add_package.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def del_tution_package_type(request, pk):
    package_type = get_object_or_404(Tution_package, pk=pk)
    package_type.delete()
    return HttpResponse(status=204, headers={'HX-Trigger': 'tpackageChanged'})
# ------------------------------ Tution --------------------------


@login_required(login_url='auth_login')
def tution_package_allocate(request):
    admission_year = Admission_Year.objects.latest('updated_at')
    classList = ClassConfig.objects.all()
    package_list = Tution_package.objects.filter(academic_year=admission_year)
    tution_student_instance = Tution.objects.filter(
        academic_year=admission_year)
    student_ids = tution_student_instance.values_list('student_id', flat=True)
    class_instance, id_profiles = None, None
    if request.method == "POST":
        class_id = request.POST.get('class_id')
        class_instance = get_object_or_404(ClassConfig, pk=class_id)
        id_profiles = StudentProfile.objects.filter(
            Q(class_id=class_instance) & Q(student_field__status="Active")).order_by("roll_no")
    else:
        id_profiles = StudentProfile.objects.filter(id__in=student_ids)

    student_package_map = {
        tution.student_id.id: tution.tution_package.id for tution in tution_student_instance}
    context = {
        "id_profiles": id_profiles,
        "class_instance": class_instance,
        "classList": classList,
        "package_list": package_list,
        'student_package_map': json.dumps(student_package_map),
        'heading': 'Coaching',
        'subheading': 'Coaching Package Allocation'
    }
    return render(request, 'crucial/tution/student_tution.html', context)


@login_required(login_url='auth_login')
@csrf_exempt
def update_tution_package(request):
    if request.method == "POST":
        student_id = request.POST.get('student_id')
        new_package_id = request.POST.get('package')
        try:
            student = StudentProfile.objects.get(id=student_id)
            package = Tution_package.objects.get(id=new_package_id)
            admission_year = Admission_Year.objects.latest('updated_at')

            # Update or create the Hostel record for the student
            updated_values = {'tution_package': package,
                              'academic_year': admission_year}
            obj, created = Tution.objects.update_or_create(
                student_id=student, defaults=updated_values
            )
            return JsonResponse({'status': 'success'})
        except StudentProfile.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Student not found'})
        except Hostel_package.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Package not found'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

# ------------------------------Transport Package--------------------------


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def transport_packagetype(request):
    context = {
        'heading': 'Transport',
        'subheading': 'Transport Package Type',
    }
    return render(request, 'crucial/transport/package_type.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def list_transport_package_type(request):
    admission_year = Admission_Year.objects.latest('updated_at')
    package_type_list = Transport_package.objects.filter(
        academic_year=admission_year).order_by("-id")
    context = {
        'package_type_list': package_type_list
    }
    return render(request, 'crucial/transport/list_package_type.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def add_transport_package_type(request):
    admission_year = Admission_Year.objects.latest('updated_at')
    form = Transporttypeform(initial={'academic_year': admission_year})

    if request.method == 'POST':
        form = Transporttypeform(request.POST)
        if form.is_valid():
            # Replace with the relevant field(s)
            package_name = form.cleaned_data['package_name']
            # Replace with the relevant field(s)
            academic_year = form.cleaned_data['academic_year']
            if Transport_package.objects.filter(package_name=package_name, academic_year=academic_year).exists():
                messages.success(
                    request, 'This package type for the selected session year already exists ! ! !')
                return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                    "messages": 'Success',
                    "packageChanged": "ttpackageChanged"
                })})
            else:
                form.save()
                messages.success(
                    request, 'Transport Package has been Saved ! ! !')
                return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                    "messages": 'Success',
                    "packageChanged": "ttpackageChanged"
                })})

    context = {
        'form': form
    }

    return render(request, 'crucial/transport/add_package.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def edit_transport_package_type(request, pk):
    package_type = get_object_or_404(Transport_package, pk=pk)
    if request.method == 'POST':
        form = Transporttypeform(request.POST, instance=package_type)
        if form.is_valid():
            form.save()
            messages.success(
                request, 'Transport Package type has been Updated ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "packageChanged": "ttpackageChanged"
            })})
    else:
        form = Transporttypeform(instance=package_type)

    context = {
        'form': form
    }

    return render(request, 'crucial/transport/add_package.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def del_transport_package_type(request, pk):
    package_type = get_object_or_404(Transport_package, pk=pk)
    package_type.delete()
    return HttpResponse(status=204, headers={'HX-Trigger': 'ttpackageChanged'})
# ------------------------------expense head view --------------------------


@login_required(login_url='auth_login')
def transport_package_allocate(request):
    admission_year = Admission_Year.objects.latest('updated_at')
    classList = ClassConfig.objects.all()
    package_list = Transport_package.objects.filter(
        academic_year=admission_year)
    transport_student_instance = Transport.objects.filter(
        academic_year=admission_year)
    student_ids = transport_student_instance.values_list(
        'student_id', flat=True)
    class_instance, id_profiles = None, None
    if request.method == "POST":
        class_id = request.POST.get('class_id')
        class_instance = get_object_or_404(ClassConfig, pk=class_id)
        id_profiles = StudentProfile.objects.filter(
            Q(class_id=class_instance) & Q(student_field__status="Active")).order_by("roll_no")
    else:
        id_profiles = StudentProfile.objects.filter(id__in=student_ids)

    student_package_map = {
        transport.student_id.id: transport.transport_package.id for transport in transport_student_instance}
    context = {
        "id_profiles": id_profiles,
        "class_instance": class_instance,
        "classList": classList,
        "package_list": package_list,
        'student_package_map': json.dumps(student_package_map),
        'heading': 'Tution',
        'subheading': 'Tution Package Allocation'
    }
    return render(request, 'crucial/transport/student_transport.html', context)


@login_required(login_url='auth_login')
@csrf_exempt
def update_transport_package(request):
    if request.method == "POST":
        student_id = request.POST.get('student_id')
        new_package_id = request.POST.get('package')
        try:
            student = StudentProfile.objects.get(id=student_id)
            package = Transport_package.objects.get(id=new_package_id)
            admission_year = Admission_Year.objects.latest('updated_at')

            # Update or create the Hostel record for the student
            updated_values = {'transport_package': package,
                              'academic_year': admission_year}
            obj, created = Transport.objects.update_or_create(
                student_id=student, defaults=updated_values
            )
            return JsonResponse({'status': 'success'})
        except StudentProfile.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Student not found'})
        except Hostel_package.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Package not found'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})
# ------------------------------expense head view --------------------------


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def expensetype(request):
    context = {
        'heading': 'Expense',
        'subheading': 'Expense Head',
    }
    return render(request, 'crucial/expense/expense_type.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def list_expensetype(request):
    expenselist = ExpenseHead.objects.all()
    context = {
        'expenselist': expenselist
    }
    return render(request, 'crucial/expense/list_expense_type.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def add_expensetype(request):
    if request.method == 'POST':
        form = Expensetypeform(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Expensetype has been Saved ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "expenseListChanged": "expenseListChanged"
            })})

    else:
        form = Expensetypeform()

    context = {
        'form': form
    }

    return render(request, 'crucial/expense/add_expense_type.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def edit_expensetype(request, pk):
    expensetype = get_object_or_404(ExpenseHead, pk=pk)
    if request.method == 'POST':
        form = Expensetypeform(request.POST, instance=expensetype)
        if form.is_valid():
            form.save()
            messages.success(request, 'Expensetype has been Updated ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "expenseListChanged": "expenseListChanged"
            })})
    else:
        form = Expensetypeform(instance=expensetype)

    context = {
        'form': form
    }

    return render(request, 'crucial/expense/add_expense_type.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def del_expensetype(request, pk):
    expensetype = get_object_or_404(ExpenseHead, pk=pk)
    expensetype.delete()
    return HttpResponse(status=204, headers={'HX-Trigger': 'expenseListChanged'})


# ------------------------------expense item view --------------------------
@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant', 'HR']))
def expenseitem(request):

    context = {
        'heading': 'Expense',
        'subheading': 'Expense List',
    }

    return render(request, 'crucial/expense/expense.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant', 'HR']))
def list_expenseitem(request):
    expenseitem = Expenseitemlist.objects.all().order_by("-id")
    totalexpense = Expenseitemlist.objects.all().aggregate(Sum('amount'))
    totalexpense = totalexpense['amount__sum']
    context = {
        'totalexpense': totalexpense,
        'expenseitem': expenseitem,
    }
    return render(request, 'crucial/expense/list_expense.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant', 'HR']))
def add_expenseitem(request):
    if request.method == 'POST':
        form = Expenselistform(request.POST, request.FILES)
        if form.is_valid():
            current_user = request.user
            expense_instance = form.save(commit=False)
            expense_instance.created_by = current_user
            expense_instance.save()
            messages.success(request, 'Expenseitem has been Saved ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "expenseitemChanged": "expenseitemChanged"
            })})
    else:
        form = Expenselistform()

    context = {
        'form': form
    }

    return render(request, 'crucial/expense/add_expense.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant', 'HR']))
def edit_expenseitem(request, pk):
    expenseitem = get_object_or_404(Expenseitemlist, pk=pk)

    if request.method == 'POST':
        form = Expenselistform(
            request.POST, request.FILES, instance=expenseitem)
        if form.is_valid():
            current_user = request.user
            expense_instance = form.save(commit=False)
            expense_instance.updated_by = current_user
            expense_instance.save()
            messages.success(request, 'Expenseitem has been Updated ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "expenseitemChanged": "expenseitemChanged"
            })})
    else:
        form = Expenselistform(instance=expenseitem)

    context = {
        'form': form

    }

    return render(request, 'crucial/expense/add_expense.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def del_expenseitem(request, pk):
    expenseitem = get_object_or_404(Expenseitemlist, pk=pk)
    expenseitem.delete()
    return HttpResponse(status=204, headers={'HX-Trigger': 'expenseitemChanged'})


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def search_date_range_expense(request):
    total_expense = None
    expenses = None
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        
        print(start_date,end_date)

        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        expensesP=Expenseitemlist.objects.values_list()
        print("expensesP",expensesP)
        expenses = Expenseitemlist.objects.filter(
            Q(expense_date__gte=start_date) & Q(expense_date__lte=end_date)
        ).order_by('-created_at')
        print(expenses.query)  # To see the generated SQL query
        print(expenses.count())
        total_expense = expenses.aggregate(
            total_amount=Sum('amount'))['total_amount']

    context = {
        'expenses': expenses,
        'total_expense': total_expense,
        'heading': 'Accounting Report',
        'subheading': 'Date Wise Expense',
    }
    return render(request, 'crucial/expense/search_date_range_expense.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def search_head_expense(request):
    expenseheadlist=ExpenseHead.objects.all()
    total_expense = None
    expenses = None
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        expense_head_id = request.POST.get('expense_head_id')
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        expenses = Expenseitemlist.objects.filter(
            Q(expense_date__gte=start_date) & Q(expense_date__lte=end_date) & Q(expensetype_id=expense_head_id)
        ).order_by('-created_at')
        total_expense = expenses.aggregate(
            total_amount=Sum('amount'))['total_amount']

    context = {
        'expenseheadlist':expenseheadlist,
        'expenses': expenses,
        'total_expense': total_expense,
        'heading': 'Accounting Report',
        'subheading': 'Head Wise Expense',
    }
    return render(request, 'crucial/expense/search_head_expense.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def search_entry_expense(request):
    total_expense = None
    expenses = None
    search_date = None
    if request.method == 'POST':
        date = request.POST.get('date')
        search_date = datetime.strptime(date, '%Y-%m-%d').date()
        expenses = Expenseitemlist.objects.filter(created_at=search_date)
        total_expense = sum(expense.amount for expense in expenses)

    context = {
        'expenses': expenses,
        'total_expense': total_expense,
        'search_date': search_date,
        'heading': 'Expense Report',
        'subheading': 'Date Wise Expense',
    }
    return render(request, 'crucial/expense/expense_entry_search.html', context)


# ------------------------------Income Head view --------------------------
@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def incomehead(request):
    context = {
        'heading': 'Income',
        'subheading': 'Income Head',
    }
    return render(request, 'crucial/income/income_head.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def list_incomehead(request):
    incomelist = IncomeHead.objects.all()
    context = {
        'incomelist': incomelist
    }
    return render(request, 'crucial/income/list_income_type.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def add_incomehead(request):
    if request.method == 'POST':
        form = IncomeHeadform(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Income Head has been Saved ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "incomeHeadChanged": "incomeHeadChanged"
            })})

    else:
        form = IncomeHeadform()

    context = {
        'form': form
    }

    return render(request, 'crucial/income/add_income_head.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def edit_incomehead(request, pk):
    income_head = get_object_or_404(IncomeHead, pk=pk)
    if request.method == 'POST':
        form = IncomeHeadform(request.POST, instance=income_head)
        if form.is_valid():
            form.save()
            messages.success(request, 'Income Head has been Updated ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "incomeHeadChanged": "incomeHeadChanged"
            })})
    else:
        form = IncomeHeadform(instance=income_head)

    context = {
        'form': form
    }

    return render(request, 'crucial/income/add_income_head.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def del_incomehead(request, pk):
    income_head = get_object_or_404(IncomeHead, pk=pk)
    income_head.delete()
    return HttpResponse(status=204, headers={'HX-Trigger': 'incomeHeadChanged'})


# ------------------------------Income item view --------------------------
@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def incomeitem(request):

    context = {
        'heading': 'Income',
        'subheading': 'Income List',
    }

    return render(request, 'crucial/income/income.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def list_incomeitem(request):

    incomeitem = IncomeitemList.objects.all().order_by("-id")
    totalincome = IncomeitemList.objects.all().aggregate(Sum('amount'))
    totalincome = totalincome['amount__sum']
    context = {
        'totalincome': totalincome,
        'incomeitem': incomeitem,
    }
    return render(request, 'crucial/income/list_income.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def add_incomeitem(request):

    if request.method == 'POST':
        form = Incomelistform(request.POST, request.FILES)
        if form.is_valid():
            current_user = request.user
            income_instance = form.save(commit=False)
            income_instance.created_by = current_user
            income_instance.save()
            messages.success(request, 'Incomelist has been Saved ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "incomeitemChanged": "incomeitemChanged"
            })})
    else:
        form = Incomelistform()

    context = {
        'form': form
    }

    return render(request, 'crucial/income/add_income.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def edit_incomeitem(request, pk):
    incomeitem = get_object_or_404(IncomeitemList, pk=pk)

    if request.method == 'POST':
        form = Incomelistform(request.POST, request.FILES, instance=incomeitem)
        if form.is_valid():
            current_user = request.user
            income_instance = form.save(commit=False)
            income_instance.updated_by = current_user
            income_instance.save()
            messages.success(request, 'Incomelist has been Updated ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "incomeitemChanged": "incomeitemChanged"
            })})
    else:
        form = Incomelistform(instance=incomeitem)

    context = {
        'form': form

    }

    return render(request, 'crucial/income/add_income.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def del_incomeitem(request, pk):
    incomeitem = get_object_or_404(IncomeitemList, pk=pk)
    incomeitem.delete()
    return HttpResponse(status=204, headers={'HX-Trigger': 'incomeitemChanged'})


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def search_date_range_income(request):
    total_income = None
    incomes = None
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

        incomes = IncomeitemList.objects.filter(
            Q(income_date__gte=start_date) & Q(income_date__lte=end_date)
        ).order_by("-created_at")
        total_income = incomes.aggregate(total_amount=Sum('amount'))[
            'total_amount']

    context = {
        'incomes': incomes,
        'total_income': total_income,
        'heading': 'Accounting Report',
        'subheading': 'Date Wise Income',
    }
    return render(request, 'crucial/income/search_date_range_income.html', context)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def search_head_income(request):
    incomeheadlist=IncomeHead.objects.all()
    total_income = None
    incomes = None
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        income_head_id = request.POST.get('income_head_id')

        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

        incomes = IncomeitemList.objects.filter(
            Q(income_date__gte=start_date) & Q(income_date__lte=end_date) & Q(incometype_id=income_head_id)
        ).order_by("-created_at")
        total_income = incomes.aggregate(total_amount=Sum('amount'))[
            'total_amount']

    context = {
        'incomeheadlist':incomeheadlist,
        'incomes': incomes,
        'total_income': total_income,
        'heading': 'Accounting Report',
        'subheading': 'Head Wise Income',
    }
    return render(request, 'crucial/income/search_head_income.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def search_entry_income(request):
    total_income = None
    incomes = None
    search_date = None
    if request.method == 'POST':
        date = request.POST.get('date')
        search_date = datetime.strptime(date, '%Y-%m-%d').date()
        incomes = IncomeitemList.objects.filter(created_at=search_date)
        total_income = sum(income.amount for income in incomes)

    context = {
        'incomes': incomes,
        'total_income': total_income,
        'search_date': search_date,
        'heading': 'Income Report',
        'subheading': 'Date Wise Expense',
    }
    return render(request, 'crucial/income/income_entry_search.html', context)

# ------------------------------Withdraw --------------------------


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def withdraw(request):
    context = {
        'heading': 'Withdraw',
        'subheading': 'Withdraw',
    }
    return render(request, 'crucial/withdraw/withdraw.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def list_withdraw(request):
    withdraw = Withdraw.objects.all().aggregate(Sum('amount'))
    totalwithdraw = withdraw['amount__sum']
    withdrawlist = Withdraw.objects.all().order_by("-id")
    context = {
        'totalwithdraw': totalwithdraw,
        'withdrawlist': withdrawlist
    }
    return render(request, 'crucial/withdraw/list_withdraw.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def add_withdraw(request):
    if request.method == 'POST':
        form = WithdrawForm(request.POST)
        if form.is_valid():
            current_user = request.user
            withdraw_instance = form.save(commit=False)
            withdraw_instance.created_by = current_user
            withdraw_instance.save()

            messages.success(request, 'Withdraw has been Saved ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "withdrawChanged": "withdrawChanged"
            })})
    else:
        form = WithdrawForm()

    context = {
        'form': form
    }

    return render(request, 'crucial/withdraw/add_withdraw.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def edit_withdraw(request, pk):
    withdraw = get_object_or_404(Withdraw, pk=pk)
    if request.method == 'POST':
        form = WithdrawForm(request.POST, instance=withdraw)
        if form.is_valid():
            current_user = request.user
            withdraw_instance = form.save(commit=False)
            withdraw_instance.updated_by = current_user
            withdraw_instance.save()
            messages.success(request, 'Withdraw has been Updated ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "withdrawChanged": "withdrawChanged"
            })})
    else:
        form = WithdrawForm(instance=withdraw)

    context = {
        'form': form
    }

    return render(request, 'crucial/withdraw/add_withdraw.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def del_withdraw(request, pk):
    withdraw = get_object_or_404(Withdraw, pk=pk)
    withdraw.delete()
    return HttpResponse(status=204, headers={'HX-Trigger': 'withdrawChanged'})


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def search_date_range_withdraw(request):
    total_withdraw = None
    withdraws = None
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

        withdraws = Withdraw.objects.filter(
            Q(created_at__gte=start_date) & Q(created_at__lte=end_date)
        ).order_by("-created_at")
        total_withdraw = withdraws.aggregate(
            total_amount=Sum('amount'))['total_amount']

    context = {
        'withdraws': withdraws,
        'total_withdraw': total_withdraw,
        'heading': 'Accounting Report',
        'subheading': 'Date Wise Withdraws',
    }
    return render(request, 'crucial/withdraw/search_date_range_withdraw.html', context)
# ------------------------------Accounting --------------------------


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def account_statement(request):
    expenses_by_date = defaultdict(Decimal)
    incomes_by_date = defaultdict(Decimal)
    withdraws_by_date = defaultdict(Decimal)
    start_date = None
    end_date = None
    total_withdraw = None
    total_income = None
    total_expense = None
    status = None

    final_withdraw = Withdraw.objects.aggregate(total_amount=Sum('amount')).get('total_amount', Decimal('0')) or Decimal('0')
    final_income = IncomeitemList.objects.aggregate(total_amount=Sum('amount')).get('total_amount', Decimal('0')) or Decimal('0')
    final_expense = Expenseitemlist.objects.aggregate(total_amount=Sum('amount')).get('total_amount', Decimal('0')) or Decimal('0')


    # final_withdraw = Withdraw.objects.aggregate(total_amount=Sum('amount'))[
    #     'total_amount'] or Decimal('0')
    # final_income = IncomeitemList.objects.aggregate(total_amount=Sum('amount'))[
    #     'total_amount'] or Decimal('0')
    # final_expense = Expenseitemlist.objects.aggregate(total_amount=Sum('amount'))[
    #     'total_amount'] or Decimal('0')
    
    if final_income:
        check = final_income-final_expense
        if check > 0:
            status = "Profit"
        elif check == 0:
            status = "Break-even"
        else:
            status = "Loss"

    try:
        main_balance_instance = MainBalance.objects.first()
        if main_balance_instance:
            main_balance = main_balance_instance.balance
        else:
            main_balance = Decimal('0.00') 
    except Exception as e:
        main_balance = Decimal('0.00')
        print(f"Error retrieving MainBalance: {e}")


    if request.method == 'POST':
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')

        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        withdraws = Withdraw.objects.filter(
            created_at__range=(start_date, end_date)
        )
        for withdraw in withdraws:
            withdraws_by_date[withdraw.created_at] += withdraw.amount
        total_withdraw = withdraws.aggregate(
            total_amount=Sum('amount'))['total_amount']

        expenses = Expenseitemlist.objects.filter(
            expense_date__range=(start_date, end_date)
        )
        for expense in expenses:
            expenses_by_date[expense.expense_date] += expense.amount
        total_expense = expenses.aggregate(
            total_amount=Sum('amount'))['total_amount']

        incomes = IncomeitemList.objects.filter(
            income_date__range=(start_date, end_date)
        )
        for income in incomes:
            incomes_by_date[income.income_date] += income.amount
        total_income = incomes.aggregate(total_amount=Sum('amount'))[
            'total_amount']

    # Create a list of all unique dates from the three dictionaries
    all_dates = sorted(list(set(incomes_by_date.keys()) | set(
        expenses_by_date.keys()) | set(withdraws_by_date.keys())))
    table_rows = []
    for date in all_dates:
        income_amount = Decimal(str(incomes_by_date.get(date, Decimal('0'))))
        expense_amount = Decimal(str(expenses_by_date.get(date, Decimal('0'))))
        withdraw_amount = Decimal(
            str(withdraws_by_date.get(date, Decimal('0'))))

        table_row = {
            'date': date,
            'income': income_amount,
            'expense': expense_amount,
            'withdraw': withdraw_amount
        }
        table_row['result'] = income_amount - expense_amount - withdraw_amount
        table_rows.append(table_row)

    context = {
        'status': status,
        'table_rows': table_rows,
        'start_date_str': start_date,
        'end_date_str': end_date,
        'final_withdraw': final_withdraw,
        'final_income': final_income,
        'final_expense': final_expense,
        'main_balance': main_balance,
        'total_withdraw': total_withdraw,
        'total_income': total_income,
        'total_expense': total_expense,
        'heading': 'Accounting Report',
        'subheading': 'Account Statement',
    }
    return render(request, 'crucial/accounting/account_statement.html', context)


# ------------------------------Payroll --------------------------

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))

def salary_config(request):
    all_addition_types = Addition_type.objects.all()
    all_deduction_types = Deduction_type.objects.all()

    addition_types = Addition_type.objects.values_list('addition_type', flat=True)
    deduction_types = Deduction_type.objects.values_list('deduction_type', flat=True)

    non_monthly_addition_types = all_addition_types.filter(is_every_month=False).values_list('addition_type', flat=True)
    non_monthly_deduction_types = all_deduction_types.filter(is_every_month=False).values_list('deduction_type', flat=True)


    staff_list = StaffProfile.objects.select_related('staff_field', 'role').all()

    employee_data = []
    for employee in staff_list:
        user_data = {
            'id': employee.staff_field.user_id,  
            'name': employee.staff_field.name,  
            'designation': employee.designation if employee.designation else 'N/A',
            'basic_salary': 0, 
            'additions': {},
            'deductions': {},
        }

        salary_conf = SalaryConfiguration.objects.filter(employee=employee).first()

        if salary_conf:
            user_data['basic_salary'] = salary_conf.basic_salary

        for addition_type in addition_types:
            total_addition = Addition.objects.filter(
                employee_id=employee, 
                addition_type_id__addition_type=addition_type
            ).aggregate(Sum('amount'))['amount__sum'] or 0
            user_data['additions'][addition_type] = total_addition

        for deduction_type in deduction_types:
            total_deduction = Deduction.objects.filter(
                employee_id=employee,  
                deduction_type_id__deduction_type=deduction_type
            ).aggregate(Sum('amount'))['amount__sum'] or 0
            user_data['deductions'][deduction_type] = total_deduction

        employee_data.append(user_data)

    context = {
        'addition_types': addition_types,
        'deduction_types': deduction_types,
        'non_monthly_addition_types': non_monthly_addition_types,
        'non_monthly_deduction_types': non_monthly_deduction_types,
        'employees': employee_data,
        'heading': 'Payroll',
        'subheading': 'Salary Config',
    }
    return render(request, 'crucial/payroll/salary_conf/salary_conf.html', context)


@csrf_exempt
def update_salary(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        user_id = data.get('id') 
        column = data.get('column')  
        value = data.get('value')  

        try:
            employee = StaffProfile.objects.get(staff_field__user_id=user_id)
            if column == 'basic_salary':
                salary_conf, created = SalaryConfiguration.objects.get_or_create(
                    employee=employee,
                    defaults={'basic_salary': Decimal(value)}
                )
                if not created:
                    salary_conf.basic_salary = Decimal(value)
                    salary_conf.save()

            elif column.startswith('addition_'):
                addition_type_name = column.replace('addition_', '')
                addition_type = Addition_type.objects.get(addition_type=addition_type_name)
                addition, created = Addition.objects.get_or_create(
                    addition_type_id=addition_type,
                    employee_id=employee,
                    defaults={'amount': Decimal(value), 'date': timezone.now().date()}
                )
                if not created:
                    addition.amount = Decimal(value)
                    addition.save()

            elif column.startswith('deduction_'):
                deduction_type_name = column.replace('deduction_', '')
                deduction_type = Deduction_type.objects.get(deduction_type=deduction_type_name)
                deduction, created = Deduction.objects.get_or_create(
                    deduction_type_id=deduction_type,
                    employee_id=employee,
                    defaults={'amount': Decimal(value), 'date': timezone.now().date()}
                )
                if not created:
                    deduction.amount = Decimal(value)
                    deduction.save()

            return JsonResponse({'success': True})
        except CustomUser.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'User not found.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})




@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def salary_increment(request):
    context = { 
        'heading': 'Payroll',
        'subheading': 'Increment',
    }
    return render(request, 'crucial/payroll/increment/increment.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def list_salary_increment(request):
    incrementlist = SalaryIncrement.objects.all()
    context = {
        'incrementlist': incrementlist
    }
    return render(request, 'crucial/payroll/increment/list_increment.html', context)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))

def add_salary_increment(request):
    employee_list = StaffProfile.objects.select_related('staff_field', 'role').all()

    if request.method == 'POST':
        form = SalaryIncrementform(request.POST)
        if form.is_valid():
            salary_instance = form.save(commit=False)

            if isinstance(salary_instance.employee, StaffProfile):
                salary_instance.employee = salary_instance.employee  
            
            salary_instance.save()
            messages.success(request, 'Salary Increment has been Saved!')

            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                "incrementListChanged": "incrementListChanged"
            })})
    else:
        form = SalaryIncrementform()

    context = {
        'employee_list': employee_list,
        'form': form
    }

    return render(request, 'crucial/payroll/increment/add_increment.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def edit_salary_increment(request, pk):
    increment = get_object_or_404(SalaryIncrement, pk=pk)
    if request.method == 'POST':
        form = SalaryIncrementform(request.POST, instance=increment)
        if form.is_valid():
            form.save()
            messages.success(request, 'Salary Increment has been Updated ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "incrementListChanged": "incrementListChanged"
            })})
    else:
        form = SalaryIncrementform(instance=increment)

    context = {
        'form': form
    }

    return render(request, 'crucial/payroll/increment/edit_increment.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def del_salary_increment(request, pk):
    increment = get_object_or_404(SalaryIncrement, pk=pk)
    increment.delete()
    return HttpResponse(status=204, headers={'HX-Trigger': 'incrementListChanged'})


from django.utils.timezone import now
@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))

def salary_allocation(request):
    staff_list = StaffProfile.objects.select_related('staff_field', 'role').all()
    special_additions = Addition_type.objects.filter(is_every_month=False)
    special_deductions = Deduction_type.objects.filter(is_every_month=False)

    if request.method == 'POST':
        selected_employee_ids = request.POST.getlist("selected_employees[]") 

        if not selected_employee_ids:
            messages.error(request, "No employees selected.")
            return redirect("salary_allocation")

        # Fetch selected employees
        selected_employees = StaffProfile.objects.filter(id__in=selected_employee_ids)
        salary_month = request.POST.get("salary_month_id", None)

        # Fetch selected special additions and deductions
        special_addition_list = request.POST.getlist("addition_id[]")
        special_deduction_list = request.POST.getlist("deduction_id[]")

        for employee in selected_employees:
            salary_config = SalaryConfiguration.objects.filter(employee=employee).first()
            if not salary_config:
                continue  # Skip employees without a salary configuration

            # Use the `total_salary()` method from the model
            if not special_addition_list and not special_deduction_list:
                total_salary = salary_config.total_salary()
            else:
                
                total_salary = salary_config.total_salary()  
                special_addition_amount = Addition.objects.filter(
                    addition_type_id__in=special_addition_list,
                    employee_id=employee
                ).aggregate(Sum('amount'))['amount__sum'] or 0

                special_deduction_amount = Deduction.objects.filter(
                    deduction_type_id__in=special_deduction_list,
                    employee_id=employee
                ).aggregate(Sum('amount'))['amount__sum'] or 0

                total_salary += special_addition_amount
                total_salary -= special_deduction_amount
            # Create or update salary process using `total_salary()` method
            salary_process, created = SalaryProcess.objects.get_or_create(
                employee_salary=salary_config,
                salary_month=salary_month,
                defaults={
                    'total_salary': total_salary,  # Calls method dynamically
                    'salary_status': 'unpaid',
                }
            )

            

    context = {
        'employee_list': staff_list,
        'special_additions': special_additions,
        'special_deductions': special_deductions,
        'heading': 'Payroll',
        'subheading': 'Salary Allocation',
    }
    return render(request, 'crucial/payroll/salary_process/addsalary.html', context)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def addition(request):
    context = { 
        'heading': 'Payroll',
        'subheading': 'Addition',
    }
    return render(request, 'crucial/payroll/addition/addition.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def list_addition(request):
    additionTypelist = Addition_type.objects.all()
    context = {
        'additionTypelist': additionTypelist
    }
    return render(request, 'crucial/payroll/addition/list_addition.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def add_addition(request):
    if request.method == 'POST':
        form = Addition_type_form(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Salary Addition Type has been Saved ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "additionListChanged": "additionListChanged"
            })})

    else:
        form = Addition_type_form()

    context = {
        'form': form
    }

    return render(request, 'crucial/payroll/addition/add_addition.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def edit_addition(request, pk):
    addition = get_object_or_404(Addition_type, pk=pk)
    if request.method == 'POST':
        form = Addition_type_form(request.POST, instance=addition)
        if form.is_valid():
            form.save()
            messages.success(request, 'Salary Addition Type has been Updated ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "additionListChanged": "additionListChanged"
            })})
    else:
        form = Addition_type_form(instance=addition)

    context = {
        'form': form
    }

    return render(request, 'crucial/payroll/addition/add_addition.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def del_addition(request, pk):
    addition = get_object_or_404(Addition_type, pk=pk)
    addition.delete()
    return HttpResponse(status=204, headers={'HX-Trigger': 'additionListChanged'})


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def deduction(request):
    context = {
        'heading': 'Payroll',
        'subheading': 'Deduction Type',
    }
    return render(request, 'crucial/payroll/deduction/deduction.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def list_deduction(request):
    deductionTypelist = Deduction_type.objects.all()
    context = {
        'deductionTypelist': deductionTypelist
    }
    return render(request, 'crucial/payroll/deduction/list_deduction.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def add_deduction(request):
    if request.method == 'POST':
        form = Deduction_type_form(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Salary Deduction Type has been Saved ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "deductionListChanged": "deductionListChanged"
            })})
    else:
        form = Deduction_type_form()

    context = {
        'form': form
    }

    return render(request, 'crucial/payroll/deduction/add_deduction.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def edit_deduction(request, pk):
    deduction = get_object_or_404(Deduction_type, pk=pk)
    if request.method == 'POST':
        form = Deduction_type_form(request.POST, instance=deduction)
        if form.is_valid():
            form.save()
            messages.success(request, 'Salary Deduction Type has been Updated ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "deductionListChanged": "deductionListChanged"
            })})
    else:
        form = Deduction_type_form(instance=deduction)

    context = {
        'form': form
    }

    return render(request, 'crucial/payroll/deduction/add_deduction.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def del_deduction(request, pk):
    deduction = get_object_or_404(Deduction_type, pk=pk)
    deduction.delete()
    return HttpResponse(status=204, headers={'HX-Trigger': 'deductionListChanged'})

from django.template.loader import render_to_string
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
@login_required(login_url='login')
@csrf_exempt
def salary_process(request):
    """
    View to handle salary processing: bulk, selected, and individual payments.
    """
    salary_process_list = SalaryProcess.objects.select_related('employee_salary__employee__staff_field').order_by("-id")

    # **Filtering based on user input**
    search_id = request.GET.get('search_id', '') or ''
    search_name = request.GET.get('search_name', '') or ''
    search_month = request.GET.get('search_month', '') or ''

    if search_id:
        salary_process_list = salary_process_list.filter(employee_salary__employee__staff_field__user_id__icontains=search_id)

    if search_name:
        salary_process_list = salary_process_list.filter(employee_salary__employee__staff_field__name__icontains=search_name)

    if search_month:
        salary_process_list = salary_process_list.filter(salary_month=search_month)

    # **Fixing Pending Amount Calculation**
    for salary in salary_process_list:
        salary.paid_amount = salary.payment_amount or 0 
        if salary.salary_status == "paid":
            salary.pending_amount = 0
        else:
            salary.pending_amount = max(Decimal(salary.total_salary) - Decimal(salary.paid_amount), 0)

    # **Pagination**
    page = request.GET.get('page', 1)
    paginator = Paginator(salary_process_list, 10)

    try:
        salary_process_list = paginator.page(page)
    except PageNotAnInteger:
        salary_process_list = paginator.page(1)
    except EmptyPage:
        salary_process_list = paginator.page(paginator.num_pages)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            invoice_type = data.get('invoice_type')

            if invoice_type == "bulk":
                selected_month = data.get('selected_month')
                if not selected_month:
                    return JsonResponse({'error': 'No month selected for bulk payment'}, status=400)

                pending_salaries = SalaryProcess.objects.filter(salary_month=selected_month, salary_status='unpaid')

                for salary in pending_salaries:
                    payment_amount = Decimal(salary.total_salary) - Decimal(salary.payment_amount or 0)
                    if payment_amount > 0:
                        process_individual_payment(salary, payment_amount, "cash", "full", request.user)

                return JsonResponse({'message': f"Bulk payments for {selected_month} processed successfully."})

            elif invoice_type == "selected":
                selected_salaries = data.get('selected_salaries', {})

                if not selected_salaries:
                    return JsonResponse({'error': 'No employees selected for payment'}, status=400)

                for salary_id, payment_amount in selected_salaries.items():
                    salary = get_object_or_404(SalaryProcess, id=salary_id)
                    payment_amount = Decimal(payment_amount)  
                    payment_method = data.get(f'payment_method_{salary_id}', "cash")
                    payment_status = data.get(f'payment_status_{salary_id}', "partial")

                    if payment_status == "full":
                        payment_amount = Decimal(salary.total_salary) - Decimal(salary.payment_amount or 0)

                    if payment_amount > 0:
                        process_individual_payment(salary, payment_amount, payment_method, payment_status, request.user)

                return JsonResponse({'message': "Selected payments processed successfully."})

            elif invoice_type.startswith("individual_"):  
                salary_id = invoice_type.split("_")[1]  
                salary = get_object_or_404(SalaryProcess, id=salary_id)

                payment_amount = Decimal(data.get('payment_amount', 0))
                payment_method = data.get('payment_method', "cash")
                payment_status = data.get('payment_status', "partial")

                if payment_status == "full":
                    payment_amount = Decimal(salary.total_salary) - Decimal(salary.payment_amount or 0)

                if payment_amount > 0:
                    process_individual_payment(salary, payment_amount, payment_method, payment_status, request.user)

                return JsonResponse({'message': "Payment processed successfully."})

            else:
                return JsonResponse({'error': 'Invalid payment request'}, status=400)

        except Exception as e:
            return JsonResponse({'error': f"Error processing payment: {str(e)}"}, status=500)

    context = {
        'salary_process_list': salary_process_list,
        'months': [
            "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December"
        ],
        'heading': 'Payroll',
        'subheading': 'Salary Process',
    }
    return render(request, 'crucial/payroll/salary_process/list_salary.html', context)


def process_individual_payment(salary, payment_amount, payment_method, payment_status, user):
    """
    Processes individual salary payments and updates the salary record.
    """
    pending_amount = Decimal(salary.total_salary) - Decimal(salary.payment_amount or 0)

    if payment_amount <= 0 or payment_amount > pending_amount:
        raise ValueError("Invalid payment amount.")

    total_paid = Decimal(salary.payment_amount or 0)
    total_paid += payment_amount

    if payment_status == 'full' or total_paid >= Decimal(salary.total_salary):
        salary.salary_status = 'paid'
        salary.payment_amount = Decimal(salary.total_salary)
        salary.pending_amount = Decimal(0)  # Ensure pending amount is correctly set to zero
    else:
        salary.salary_status = 'partial'
        salary.payment_amount = total_paid
        salary.pending_amount = max(0, Decimal(salary.total_salary) - total_paid) 
    salary.payment_method = payment_method
    salary.updated_by = user
    salary.save()

    trigger_payment_and_journal(salary, payment_amount)

def trigger_payment_and_journal(salary, payment_amount):
    """
    Handles the accounting entries for salary payments by recording them in the Payment table only.
    """
    try:
        payment_amount = Decimal(payment_amount)

        if payment_amount <= 0:
            raise ValueError("Payment amount must be greater than zero.")

        debit_ledger = Ledger.objects.get(name="Salaries & Wages")  
        if salary.payment_method == 'cash':
            credit_ledger = Ledger.objects.get(name="Cash in Hand")  
        else:
            credit_ledger = Ledger.objects.get(name="Bank Account (Main)")  

        current_time = timezone.now().strftime("%Y%m%d%H%M%S")
        voucher_number = f"PAY{current_time}"

        with transaction.atomic(): 

            staff_member = salary.employee_salary.employee  

            Payment.objects.create(
                voucher_no=voucher_number,
                date=timezone.now().date(),
                amount=payment_amount,
                staff=staff_member, 
                expense_ledger=debit_ledger,
                cash_ledger=credit_ledger,
                description=f"Salary payment for {salary.salary_month} - {voucher_number}",
                created_by=salary.created_by 
            )

    except Ledger.DoesNotExist as e:
        raise ValueError(f"Ledger not found: {str(e)}")
    except Exception as e:
        raise ValueError(f"Error processing payment: {str(e)}")  
    

def update_tax_profile_on_payment(salary_process):
    try:
        tax_profile, created = TaxProfile.objects.get_or_create(
            employee=salary_process.employee_salary.employee,
            defaults={'tax_type': 'Income Tax'}
        )
        tax_profile.salary_process = salary_process
        tax_profile.calculate_tax()
        tax_profile.save()
    except Exception as e:
        print(f"Error updating TaxProfile: {e}")



def send_salary_sms(template, salary_process, amount):
    """
    Send SMS notifications to employees about their salary.
    """
    employee = salary_process.employee_salary.employee
    employee_number = employee.phone_number
    formatted_number = '880' + employee_number.lstrip('0')

    sms_body = template.body.format(
        employee_name=employee.name,
        month=salary_process.salary_month,
        salary_amount=amount,
        payment_date=salary_process.payment_date.date()
    )

    sms_count = count_sms(sms_body)
    sms_limit_obj = SMSUsage.objects.filter(Msg_type='NONMASKING').first()

    if sms_limit_obj and sms_limit_obj.total_sms >= sms_count:
        send_sms(formatted_number, sms_body)
        sms_limit_obj.total_sms -= sms_count
        sms_limit_obj.save()
        SMS.objects.create(
            mobile=employee_number, title='Salary MSG', msg=sms_body, created_by=salary_process.created_by
        )
    else:
        print('SMS LIMIT OVER')


def tax_summary(request):
    query = request.GET.get('q')  # Search by employee name or ID
    tax_profiles = TaxProfile.objects.all()

    if query:
        tax_profiles = tax_profiles.filter(
            models.Q(employee__name__icontains=query) | models.Q(employee__id__icontains=query)
        )

    for profile in tax_profiles:
        profile.calculate_tax()

    context = {
        'tax_profiles': tax_profiles,
        'heading': 'Tax Summary',
        'subheading': 'Employee CPF and Income Tax Details',
    }
    return render(request, 'crucial/finance/tax_summary.html', context)

def tax_detail(request, profile_id):
    profile = get_object_or_404(TaxProfile, id=profile_id)
    if request.method == 'POST':
        profile.salary_process = get_object_or_404(SalaryProcess, id=request.POST.get('salary_process_id'))
        profile.gender = request.POST.get('gender', profile.gender)
        profile.is_senior_citizen = request.POST.get('is_senior_citizen', 'off') == 'on'
        profile.is_disabled = request.POST.get('is_disabled', 'off') == 'on'
        profile.is_freedom_fighter = request.POST.get('is_freedom_fighter', 'off') == 'on'
        profile.calculate_tax()
        messages.success(request, 'Tax details updated successfully.')
        return redirect('tax_detail', profile_id=profile.id)

    context = {
        'profile': profile,
        'heading': 'Tax Detail',
        'subheading': f"{profile.employee.name}'s Tax Information",
    }
    return render(request, 'crucial/finance/tax_detail.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, "staff", "teacher", roles=['Manager', 'Accountant','HR']))
def employee_salary(request):
    if is_staff_or_has_role(request.user, "teacher", "staff", roles=["Manager", 'Accountant', 'HR']):
        salary_process_list = []
        requested_user = request.user
        salary_config = get_object_or_404(
            SalaryConfiguration, employee=requested_user)
        salary_process_list = SalaryProcess.objects.filter(
            employee_salary=salary_config)
        
        context = {
        'salary_process_list': salary_process_list,
        'heading': 'Payroll',
        'subheading': 'Salary List'}
    return render(request, 'crucial/payroll/salary_process/employee_salary.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def generate_pay_report(request, salary_id):
    institute = Institute.objects.latest('id')
    salary = SalaryProcess.objects.get(id=salary_id)
    context = {
        'institute': institute,
        'salary': salary
    }
    return render(request, 'report/payroll.html', context)


# -------------------------------HomeWork--------------------------


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def homework(request):
    return render(request, 'crucial/homework/homework.html')


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def list_homework(request):
    homeworklist = Homework.objects.all()
    context = {
        'homeworklist': homeworklist
    }
    return render(request, 'crucial/homework/list_homework.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def add_homework(request):
    if request.method == 'POST':
        form = Homeworkform(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return HttpResponse(status=204, headers={'HX-Trigger': 'homeworkListChanged'})
    else:
        form = Homeworkform()

    context = {
        'form': form
    }

    return render(request, 'crucial/homework/add_homework.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def edit_homework(request, pk):
    homework = get_object_or_404(Homework, pk=pk)
    if request.method == 'POST':
        form = Homeworkform(request.POST, instance=homework)
        if form.is_valid():
            form.save()
            return HttpResponse(status=204, headers={'HX-Trigger': 'homeworkListChanged'})
    else:
        form = Homeworkform(instance=homework)

    context = {
        'form': form
    }

    return render(request, 'crucial/homework/add_homework.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user))
def del_homework(request, pk):
    homework = get_object_or_404(Homework, pk=pk)
    homework.delete()
    return HttpResponse(status=204, headers={'HX-Trigger': 'homeworkListChanged'})


# -------------------------------HomeWork--------------------------

@login_required(login_url='login')
# @user_passes_test(lambda user: is_staff_or_in_group(user, "teacher"))
@user_passes_test(lambda user: is_staff_or_has_role(user, "teacher", roles=['Manager', 'HR']))
def add_download(request):
    stuclass = ClassConfig.objects.all()
    downloadlist = Download.objects.all()
    admission_year_list=Admission_Year.objects.all()
    academic_session_list=AcademicSession.objects.all()
    if request.method == 'POST':
        title = request.POST.get('title')
        download_type = request.POST.get('download_type')
        files = request.FILES.get('files')
        class_id = request.POST.get('class_id')
        class_instance = get_object_or_404(ClassConfig, pk=class_id)
        admission_year_id = request.POST.get('admission_year')
        admission_year = get_object_or_404(Admission_Year, pk=admission_year_id)
        academic_session_id = request.POST.get('academic_session_year')
        academic_session = get_object_or_404(AcademicSession, pk=academic_session_id)
        Download.objects.create(
            title=title,
            download_type=download_type,
            files=files,
            class_id=class_instance,
            academic_year=admission_year,
            academic_session_year=academic_session
        )

        messages.success(request, f" {download_type} has been Saved !")
        return redirect('study_material')

    context = {
        'downloadList': downloadlist,
        'stuclass': stuclass,
        'admission_year_list':admission_year_list,
        'academic_session_list':academic_session_list,
        'heading': 'Study Material',
        'subheading': 'Study Material'
    }
    return render(request, 'crucial/download/download_list.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, "teacher", roles=['Manager', 'HR']))
def del_download(request, pk):
    download = get_object_or_404(Download, pk=pk)
    download.delete()
    return redirect('download')


# ------------------------------Fees Head--------------------------

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def fees_head(request):
    context = {
        'heading': 'Fees',
        'subheading': 'Fees Head',
    }
    return render(request, 'crucial/fees_head/fees_head.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def list_fees_head(request):
    fees_head_list = FeeHead.objects.all()
    context = {
        'fees_head_list': fees_head_list
    }
    return render(request, 'crucial/fees_head/list_fees_head.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def add_fees_head(request):
    form = FeeHeadForm()

    if request.method == 'POST':
        form = FeeHeadForm(request.POST)
        if form.is_valid():
            # Replace with the relevant field(s)
            fees_head_name = form.cleaned_data['name']
            if FeeHead.objects.filter(name=fees_head_name).exists():
                messages.success(
                    request, 'This Fee Head already exists ! ! !')
                return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                    "messages": 'Success',
                    "feesheadChanged": "feesheadChanged"
                })})
            else:
                form.save()
                messages.success(
                    request, 'FeeHead has been Saved ! ! !')
                return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                    "messages": 'Success',
                    "feesheadChanged": "feesheadChanged"
                })})

    context = {
        'form': form
    }

    return render(request, 'crucial/fees_head/add_fees_head.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def edit_fees_head(request, pk):
    fees_head = get_object_or_404(FeeHead, pk=pk)
    if request.method == 'POST':
        form = FeeHeadForm(request.POST, instance= fees_head)
        if form.is_valid():
            form.save()
            messages.success(
                request, 'This Fee Head has been Updated ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "feesheadChanged": "feesheadChanged"
            })})
    else:
        form = FeeHeadForm(instance=fees_head)

    context = {
        'form': form
    }

    return render(request, 'crucial/fees_head/add_fees_head.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def del_fees_head(request, pk):
    fees_head = get_object_or_404(FeeHead, pk=pk)
    fees_head.delete()
    return HttpResponse(status=204, headers={'HX-Trigger': 'feesheadChanged'})




# ------------------------------Fees Head--------------------------

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def fees_head_ledger_config(request):
    context = {
        'heading': 'Fees',
        'subheading': 'Fees Head Ledger Config',
    }
    return render(request, 'crucial/fees_head_ledger_config/fees_head_ledger_config.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def list_fees_head_ledger_config(request):
    fees_head_ledger_config_list = FeeHeadLedgerConfig.objects.all()
    context = {
        'fees_head_ledger_config_list': fees_head_ledger_config_list
    }
    return render(request, 'crucial/fees_head_ledger_config/list_fees_head_ledger_config.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def add_fees_head_ledger_config(request):
    form = FeeHeadLedgerConfigForm()

    if request.method == 'POST':
        form = FeeHeadLedgerConfigForm(request.POST)
        if form.is_valid():
            # Replace with the relevant field(s)
            fees_config_head = form.cleaned_data['head']
            fees_config_ledger = form.cleaned_data['ledger']
            if FeeHeadLedgerConfig.objects.filter(head=fees_config_head, ledger=fees_config_ledger).exists():
                messages.success(
                    request, 'This Fee Head Ledger Config already exists ! ! !')
                return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                    "messages": 'Success',
                    "feesheadLconfChanged": "feesheadLconfChanged"
                })})
            else:
                form.save()
                messages.success(
                    request, 'Fee Head Ledger Config has been Saved ! ! !')
                return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                    "messages": 'Success',
                    "feesheadLconfChanged": "feesheadLconfChanged"
                })})
            
        else:
            print("Form errors:", form.errors)

    context = {
        'form': form
    }

    return render(request, 'crucial/fees_head_ledger_config/add_fees_head_ledger_config.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def edit_fees_head_ledger_config(request, pk):
    fees_head_ledger_config = get_object_or_404(FeeHeadLedgerConfig, pk=pk)
    if request.method == 'POST':
        form = FeeHeadLedgerConfigForm(request.POST, instance= fees_head_ledger_config)
        if form.is_valid():
            form.save()
            messages.success(
                request, 'This Fee Head Ledger Config has been Updated ! ! !')
            return HttpResponse(status=204, headers={'HX-Trigger': json.dumps({
                "messages": 'Success',
                            "feesheadLconfChanged": "feesheadLconfChanged"
            })})
    else:
        form = FeeHeadLedgerConfigForm(instance=fees_head_ledger_config)

    context = {
        'form': form
    }

    return render(request, 'crucial/fees_head_ledger_config/add_fees_head_ledger_config.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def del_fees_head_ledger_config(request, pk):
    fees_head_ledger_config = get_object_or_404(FeeHeadLedgerConfig, pk=pk)
    fees_head_ledger_config.delete()
    return HttpResponse(status=204, headers={'HX-Trigger': 'feesheadLconfChanged'})



from datetime import datetime, date
from decimal import Decimal, InvalidOperation
import calendar
import pandas as pd
from django.db import transaction
from django.contrib import messages
from django.shortcuts import render, redirect
from django.db import models


from decimal import Decimal, InvalidOperation
from datetime import date, datetime
import calendar

import pandas as pd
from django.contrib import messages
from django.db import transaction, models
from django.shortcuts import redirect, render


def bulk_upload_unpaid_fees(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        month_id = request.POST.get('month_id')
        academic_year_id = request.POST.get('academic_year_id')

        if not all([excel_file, month_id, academic_year_id]):
            messages.error(request, "All fields are required")
            return redirect('bulk_upload_unpaid')

        try:
            month = Fee_month.objects.get(id=month_id)
            academic_year = Admission_Year.objects.get(id=academic_year_id)
        except (Fee_month.DoesNotExist, Admission_Year.DoesNotExist):
            messages.error(request, "Invalid month or academic year selected")
            return redirect('bulk_upload_unpaid')

        # -----------------------------
        # Read Excel + normalize headers
        # -----------------------------
        try:
            df = pd.read_excel(excel_file)

            # header clean
            df.columns = df.columns.astype(str).str.strip()

            required_columns = {
                'FullName', 'RollNo', 'FeeHeadName', 'FeeAmount',
                'ClassName', 'SectionName', 'VersionName'
            }
            if not required_columns.issubset(df.columns):
                missing = required_columns - set(df.columns)
                messages.error(request, f"Missing columns: {', '.join(sorted(missing))}")
                return redirect('bulk_upload_unpaid')

            # IMPORTANT: RollNo is string (User_id)
            df['RollNo'] = df['RollNo'].astype(str).str.strip()
            df['FullName'] = df['FullName'].astype(str).str.strip().str.upper()

            df['FeeHeadName'] = df['FeeHeadName'].astype(str).str.strip()
            df['FeeAmount'] = df['FeeAmount']  # amount parse later per-row for better error msg

            df['ClassName'] = df['ClassName'].astype(str).str.strip()
            df['SectionName'] = df['SectionName'].astype(str).str.strip()
            df['VersionName'] = df['VersionName'].astype(str).str.strip()

        except Exception as e:
            messages.error(request, f"Error reading Excel file: {str(e)}")
            return redirect('bulk_upload_unpaid')

        errors = []
        success_count = 0

        # -----------------------------
        # Pre-compute dates once
        # -----------------------------
        year = int(academic_year.name) if str(academic_year.name).isdigit() else date.today().year
        try:
            month_number = datetime.strptime(month.name, "%B").month
        except ValueError:
            month_number = date.today().month

        start_date = date(year, month_number, 1)
        end_day = calendar.monthrange(year, month_number)[1]
        end_date = date(year, month_number, end_day)

        # -----------------------------
        # Process rows
        # -----------------------------
        with transaction.atomic():
            for idx, row in df.iterrows():
                try:
                    with transaction.atomic():
                        rollno = str(row.get('RollNo', '')).strip()
                        fullname = str(row.get('FullName', '')).strip().upper()

                        if not rollno or rollno.lower() in ('nan', 'none'):
                            raise ValueError("Missing RollNo (User_id)")
                        if not fullname or fullname.lower() in ('nan', 'none'):
                            raise ValueError("Missing FullName")

                        # -----------------------------
                        # Student lookup (RollNo = User_id)
                        # -----------------------------
                        try:
                            # ✅ এখানে user_id ফিল্ড ধরে করা হলো
                            student = StudentProfile.objects.get(user_id=rollno)
                        except StudentProfile.DoesNotExist:
                            # fallback: name + class + section + version
                            student = StudentProfile.objects.annotate(
                                upper_name=models.functions.Upper(models.F('student_field__name'))
                            ).get(
                                upper_name=fullname,
                                class_id__class_id__name__iexact=str(row['ClassName']).strip(),
                                class_id__section__name__iexact=str(row['SectionName']).strip(),
                                version__iexact=str(row['VersionName']).strip()
                            )

                        # -----------------------------
                        # FeeHead & Feetype
                        # -----------------------------
                        fee_head_name = str(row['FeeHeadName']).strip()[:50]
                        if not fee_head_name:
                            raise ValueError("Missing FeeHeadName")

                        fee_head, _ = FeeHead.objects.get_or_create(name=fee_head_name)

                        feetype, _ = Feetype.objects.get_or_create(
                            fee_head=fee_head,
                            fee_Schedule='Annually',
                            defaults={'status': 'Active', 'created_by': request.user}
                        )

                        # -----------------------------
                        # Fees_name (create for each row as your old logic)
                        # NOTE: if you want 1 Fees_name per FeeHead per month/year,
                        #       you should use get_or_create instead of create.
                        # -----------------------------
                        fees_name = Fees_name.objects.create(
                            fees_type=feetype,
                            month=month,
                            startdate=start_date,
                            enddate=end_date,
                            academic_year=academic_year,
                            created_by=request.user,
                            updated_by=request.user,
                            fees_title=f"{fee_head_name[:45]}..."
                        )

                        # -----------------------------
                        # Validate amount
                        # -----------------------------
                        try:
                            fee_amount = Decimal(str(row['FeeAmount']).strip())
                        except (InvalidOperation, ValueError, AttributeError):
                            raise ValueError(f"Invalid FeeAmount: {row['FeeAmount']}")

                        # -----------------------------
                        # Save fee record
                        # -----------------------------
                        Fees.objects.create(
                            student_id=student,
                            feetype_id=fees_name,
                            amount=fee_amount,
                            status='unpaid',
                            description=f"{fee_head_name[:47]}...",
                            month_id=month,
                            academic_year=academic_year,
                            created_by=request.user
                        )

                        success_count += 1

                except StudentProfile.DoesNotExist:
                    errors.append(f"Row {idx+2}: Student not found - {fullname} (User_id: {rollno})")
                except Exception as e:
                    errors.append(f"Row {idx+2}: {str(e)}")

        if errors:
            messages.warning(request, f"Uploaded with {len(errors)} errors")
        else:
            messages.success(request, f"Successfully created {success_count} fee records")

        return render(request, 'fees/bulk_upload.html', {'errors': errors})

    # GET request
    months = Fee_month.objects.all()
    academic_years = Admission_Year.objects.all()

    return render(request, 'fees/bulk_upload.html', {
        'months': months,
        'academic_years': academic_years
    })


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager']))
def search_and_edit_fee(request):
    query = request.GET.get('q', '')
    fees = Fees.objects.none()
    class_id = None
    version = None
    name = None
    if query:
        fees = Fees.objects.filter(student_id__student_field__user_id__icontains=query)
        if fees.exists():
            name = fees.first().student_id.student_field.name
            class_id = fees.first().student_id.class_id
            version = fees.first().student_id.version

    if request.method == 'POST':
        fee_id = request.POST.get('fee_id')
        new_amount = Decimal(request.POST.get('amount')) 
        waiver_checked = request.POST.get('waiver') 
        less_than_checked = request.POST.get('less_than')  

        if fee_id and new_amount:
            fee = get_object_or_404(Fees, id=fee_id)
            original_amount = fee.amount 

            fee.amount = new_amount
            fee.save()

            if waiver_checked or less_than_checked:
                waiver_amount = original_amount - new_amount
                if waiver_amount > 0:
                    Waiver.objects.create(
                        student_id=fee.student_id,
                        waiver_amount=waiver_amount,
                        description=f"Waiver applied for {fee.feetype_id}",
                        academic_year=fee.academic_year,
                        created_by=request.user,
                        updated_by=request.user
                    )

            messages.success(request, "Amount updated successfully!")
            return redirect('search_and_edit_fee')

    context = {
        'fees': fees,
        'query': query,
        'class_id': class_id,
        'version': version,
        'name': name,
        'heading': 'Fees',
        'subheading': 'Edit Fees',
    }
    return render(request, 'crucial/finance/search_edit_fee.html', context)



def fees_collection_report_excel(request):
    context = generate_fees_collection_context(request)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="fees_collection_report.xlsx"'
    
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet('Fees Report')
    
    bold = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
    normal = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})

    headers = ["Student Name", "Student ID", "Version", "Class", "Section", "Shift", "Group"]
    fee_head_titles = [fee_head.fees_title for fee_head in context['fee_heads']]
    headers.extend(fee_head_titles)
    headers.append("Total Paid")

    for col_num, header in enumerate(headers):
        worksheet.write(0, col_num, header, bold)

    for row_num, student_data in enumerate(context['students_data'], start=1):
        worksheet.write(row_num, 0, student_data['student'].student_field.name if student_data['student'].student_field else '', normal)
        worksheet.write(row_num, 1, student_data['student'].student_field.user_id if student_data['student'].student_field else '', normal)
        worksheet.write(row_num, 2, student_data['student'].version or '', normal)
        
        worksheet.write(row_num, 3, getattr(student_data['student'].class_id.class_group_id.class_id, 'name', ''), normal)
        worksheet.write(row_num, 4, getattr(student_data['student'].class_id.section_id, 'name', ''), normal)
        worksheet.write(row_num, 5, getattr(student_data['student'].class_id.shift_id, 'name', ''), normal)
        worksheet.write(row_num, 6, getattr(student_data['student'].class_id.class_group_id.group_id, 'name', ''), normal)
        
        for col_num, payment in enumerate(student_data['payments'], start=7):
            worksheet.write(row_num, col_num, payment, normal)
        
        worksheet.write(row_num, len(headers) - 1, student_data['total'], normal)

    total_row = len(context['students_data']) + 1
    worksheet.write(total_row, 6, "Total", bold)
    for col_num, fee_head in enumerate(context['fee_heads'], start=7):
        worksheet.write(total_row, col_num, context['column_totals'].get(fee_head.id, 0), bold)
    worksheet.write(total_row, len(headers) - 1, context['grand_total'], bold)

    workbook.close()
    output.seek(0)
    response.write(output.read())
    return response


# views.py (updated)
from django.db.models import Q, F, Sum
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.shortcuts import render
from weasyprint import HTML, CSS
import io

# def generate_fees_collection_context(request):
#     # Get filter parameters
#     selected_class = request.GET.get('class')
#     selected_section = request.GET.get('section')
#     selected_group = request.GET.get('group')
#     selected_shift = request.GET.get('shift')
#     selected_version = request.GET.get('version')
#     from_date = request.GET.get('from_date')
#     to_date = request.GET.get('to_date')

#     context = {}
    
#     if from_date and to_date:
#         # Base queryset with all relations
#         students = StudentProfile.objects.select_related(
#             'student_field',
#             'class_id__class_group_id__class_id',
#             'class_id__class_group_id__group_id',
#             'class_id__section_id',
#             'class_id__shift_id'
#         )

#         # Apply filters
#         filters = Q()
#         if selected_class: 
#             filters &= Q(class_id__class_group_id__class_id=selected_class)
#         if selected_section:
#             filters &= Q(class_id__section_id=selected_section)
#         if selected_group:
#             filters &= Q(class_id__class_group_id__group_id=selected_group)
#         if selected_shift:
#             filters &= Q(class_id__shift_id=selected_shift)
#         if selected_version:
#             filters &= Q(version=selected_version)

#         students = students.filter(filters)
#         student_ids = students.values_list('id', flat=True)

#         # Get fees data with date range filter
#         fees = Fees.objects.filter(
#             student_id__in=student_ids,
#             created_at__date__range=[from_date, to_date]
#         )

#         # Separate regular fees and previous dues
#         regular_fees = fees.filter(feetype_id__fee_amount_id__isnull=False)
#         previous_due_fees = fees.filter(feetype_id__fee_amount_id__isnull=True)

#         # Get fee heads for regular fees
#         regular_fee_heads = Fees_name.objects.filter(
#             id__in=regular_fees.values_list('feetype_id', flat=True)
#         ).distinct().order_by('fees_title')

#         # Prepare data structure
#         students_data = []
#         column_totals = {fh.id: Decimal('0') for fh in regular_fee_heads}
#         previous_due_total = Decimal('0')
#         grand_total = Decimal('0')

#         # Student-level calculations
#         for student in students:
#             student_payments = {}  # Track payments by feetype_id
#             student_regular_total = Decimal('0')
            
#             # Regular fees
#             for fh in regular_fee_heads:
#                 fee_total = regular_fees.filter(
#                     student_id=student,
#                     feetype_id=fh
#                 ).aggregate(
#                     total=Sum('amount')
#                 )['total'] or Decimal('0')
                
#                 partial_payments = PartialPayment.objects.filter(
#                     fee__student_id=student,
#                     fee__feetype_id=fh
#                 ).aggregate(
#                     total=Sum('amount')
#                 )['total'] or Decimal('0')
                
#                 total_paid = fee_total + partial_payments
#                 student_payments[fh.id] = total_paid
#                 student_regular_total += total_paid


#             # Previous dues
#             previous_due = previous_due_fees.filter(
#                 student_id=student
#             ).aggregate(
#                 total=Sum('amount')
#             )['total'] or Decimal('0')
            
#             previous_partial = PartialPayment.objects.filter(
#                 fee__student_id=student,
#                 fee__feetype_id__fee_amount_id__isnull=True
#             ).aggregate(
#                 total=Sum('amount')
#             )['total'] or Decimal('0')
            
#             total_previous = previous_due + previous_partial
#             student_total = student_regular_total + total_previous

#             # Only include students with payments > 0
#             if student_total > 0:
#                 # Update totals
#                 for fh_id, total in student_payments.items():
#                     column_totals[fh_id] += total
#                 previous_due_total += total_previous
#                 grand_total += student_total

#                 students_data.append({
#                     'student': student,
#                     'payments_dict': student_payments,  # Store as dict
#                     'previous_due': total_previous,
#                     'total': student_total
#                 })
                
#         non_zero_head_ids = [fh_id for fh_id, total in column_totals.items() if total > 0]
#         regular_fee_heads = regular_fee_heads.filter(id__in=non_zero_head_ids).order_by('fees_title')

#         # Rebuild column totals and adjust student payments
#         column_totals = {fh.id: column_totals[fh.id] for fh in regular_fee_heads}
#         for student_data in students_data:
#             payments_dict = student_data.pop('payments_dict')
#             student_data['payments'] = [payments_dict.get(fh.id, Decimal('0')) for fh in regular_fee_heads]


#         context = {
#             'students_data': students_data,
#             'fee_heads': regular_fee_heads,
#             'previous_due_total': previous_due_total,
#             'column_totals': column_totals,
#             'grand_total': grand_total,
#             'from_date': from_date,
#             'to_date': to_date,
#             'selected_class': selected_class,
#             'selected_section': selected_section,
#             'selected_group': selected_group,
#             'selected_shift': selected_shift,
#             'selected_version': selected_version,
#         }

#     return context

def generate_fees_collection_context(request):
    selected_class = request.GET.get('class')
    selected_section = request.GET.get('section')
    selected_group = request.GET.get('group')
    selected_shift = request.GET.get('shift')
    selected_version = request.GET.get('version')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    context = {}
    
    if from_date and to_date:
        students = StudentProfile.objects.select_related(
            'student_field',
            'class_id__class_group_id__class_id',
            'class_id__class_group_id__group_id',
            'class_id__section_id',
            'class_id__shift_id'
        )

        # Apply filters
        filters = Q()
        if selected_class: 
            filters &= Q(class_id__class_group_id__class_id=selected_class)
        if selected_section:
            filters &= Q(class_id__section_id=selected_section)
        if selected_group:
            filters &= Q(class_id__class_group_id__group_id=selected_group)
        if selected_shift:
            filters &= Q(class_id__shift_id=selected_shift)
        if selected_version:
            filters &= Q(version=selected_version)

        students = students.filter(filters)
        student_ids = students.values_list('id', flat=True)

        # Get Partial Payments and Full Payments (without partials) for Regular Fees
        partial_payments_regular = PartialPayment.objects.filter(
            payment_date__date__range=[from_date, to_date],
            fee__student_id__in=student_ids,
            fee__feetype_id__fee_amount_id__isnull=False
        ).values('fee__student_id', 'fee__feetype_id').annotate(partial_total=Sum('amount'))

        full_payments_regular = Fees.objects.filter(
            status='paid',
            created_at__date__range=[from_date, to_date],
            student_id__in=student_ids,
            feetype_id__fee_amount_id__isnull=False
        ).values('student_id', 'feetype_id').annotate(full_total=Sum('amount'))

        # Get Partial Payments and Full Payments for Previous Dues
        partial_payments_previous = PartialPayment.objects.filter(
            payment_date__date__range=[from_date, to_date],
            fee__student_id__in=student_ids,
            fee__feetype_id__fee_amount_id__isnull=True
        ).values('fee__student_id').annotate(partial_total=Sum('amount'))

        full_payments_previous = Fees.objects.filter(
            status='paid',
            created_at__date__range=[from_date, to_date],
            student_id__in=student_ids,
            feetype_id__fee_amount_id__isnull=True
        ).values('student_id').annotate(full_total=Sum('amount'))

        # Prepare Regular Fee Heads
        regular_fee_heads = Fees_name.objects.filter(
            fee_amount_id__isnull=False,
            id__in=Subquery(Fees.objects.filter(
                student_id__in=student_ids,
                feetype_id__fee_amount_id__isnull=False
            ).values('feetype_id'))
        ).distinct().order_by('fees_title')

        # Prepare data structure
        students_data = []
        column_totals = {fh.id: Decimal('0.00') for fh in regular_fee_heads}
        previous_due_total = Decimal('0.00')
        grand_total = Decimal('0.00')

        # Build student payment data
        for student in students:
            student_id = student.id
            # Regular Fees
            regular_dict = defaultdict(Decimal)
            
            # Add partial payments
            for pp in partial_payments_regular.filter(fee__student_id=student_id):
                feetype_id = pp['fee__feetype_id']
                regular_dict[feetype_id] += pp['partial_total']
            
            # Add full payments (no partials)
            for fp in full_payments_regular.filter(student_id=student_id):
                feetype_id = fp['feetype_id']
                regular_dict[feetype_id] += fp['full_total']
            
            # Previous Dues
            previous_total = Decimal('0.00')
            
            # Add partial payments
            prev_partial = partial_payments_previous.filter(fee__student_id=student_id).order_by('fee__student_id').first()
            if prev_partial:
                previous_total += prev_partial['partial_total']
            
            # Add full payments
            prev_full = full_payments_previous.filter(student_id=student_id).order_by('student_id').first()
            if prev_full:
                previous_total += prev_full['full_total']
            
            # Prepare student row with payments dictionary
            student_payments_dict = {fh.id: regular_dict.get(fh.id, Decimal('0.00')) for fh in regular_fee_heads}
            student_total = sum(student_payments_dict.values()) + previous_total
            
            if student_total > 0:
                students_data.append({
                    'student': student,
                    'payments_dict': student_payments_dict,  # Store as dictionary
                    'previous_due': previous_total,
                    'total': student_total
                })
                
                # Update totals using the dictionary
                for fh in regular_fee_heads:
                    column_totals[fh.id] += student_payments_dict[fh.id]
                previous_due_total += previous_total
                grand_total += student_total

        # Filter fee_heads to only those with non-zero totals
        filtered_fee_heads = [fh for fh in regular_fee_heads if column_totals[fh.id] > 0]

        # Update students_data to have payments lists based on filtered fee_heads
        for student_data in students_data:
            payments_dict = student_data.pop('payments_dict')
            student_data['payments'] = [payments_dict[fh.id] for fh in filtered_fee_heads]

        # Update column_totals to only include filtered fee_heads
        filtered_column_totals = {fh.id: column_totals[fh.id] for fh in filtered_fee_heads}

        context = {
            'students_data': students_data,
            'fee_heads': filtered_fee_heads,
            'previous_due_total': previous_due_total,
            'column_totals': filtered_column_totals,
            'grand_total': grand_total,
            'from_date': from_date,
            'to_date': to_date,
            'selected_class': selected_class,
            'selected_section': selected_section,
            'selected_group': selected_group,
            'selected_shift': selected_shift,
            'selected_version': selected_version,
        }

    return context

def fees_collection_report(request):
    context = generate_fees_collection_context(request)
    return render(request, 'crucial/report/fees_report.html', context)

def fees_collection_report_pdf(request):
    context = generate_fees_collection_context(request)
    html_string = render_to_string('crucial/report/fees_report_pdf.html', context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="fees_collection_report.pdf"'
    
    HTML(string=html_string).write_pdf(response, stylesheets=[
        CSS(string='''
            @page { size: A4 landscape; margin: 10mm; }
            th { writing-mode: vertical-lr; text-orientation: upright; padding: 4px 2px; }
            table { border-collapse: collapse; width: 100%; font-size: 10px; }
            th, td { border: 1px solid #000; padding: 4px; text-align: center; }
        ''')
    ])
    
    return response




# def generate_previous_dues_context(request):
#     # Get filter parameters
#     selected_class = request.GET.get('class')
#     selected_section = request.GET.get('section')
#     selected_group = request.GET.get('group')
#     selected_shift = request.GET.get('shift')
#     selected_version = request.GET.get('version')
#     from_date = request.GET.get('from_date')
#     to_date = request.GET.get('to_date')

#     context = {}
    
#     if from_date and to_date:
#         # Get students with filters
#         students = StudentProfile.objects.select_related(
#             'student_field',
#             'class_id__class_group_id__class_id',
#             'class_id__class_group_id__group_id',
#             'class_id__section_id',
#             'class_id__shift_id'
#         )

#         filters = Q()
#         if selected_class: 
#             filters &= Q(class_id__class_group_id__class_id=selected_class)
#         if selected_section:
#             filters &= Q(class_id__section_id=selected_section)
#         if selected_group:
#             filters &= Q(class_id__class_group_id__group_id=selected_group)
#         if selected_shift:
#             filters &= Q(class_id__shift_id=selected_shift)
#         if selected_version:
#             filters &= Q(version=selected_version)

#         students = students.filter(filters)
#         student_ids = students.values_list('id', flat=True)

#         # Get previous dues fees
#         previous_dues = Fees.objects.filter(
#             feetype_id__fee_amount_id__isnull=True,
#             status='paid', 
#             student_id__in=student_ids,
#             created_at__date__range=[from_date, to_date]
#         )

#         # Get unique previous due fee heads
#         previous_fee_heads = FeeHead.objects.filter(
#             fee_types__fees_name_type__fee_amount_id__isnull=True
#         ).distinct().order_by('name')

#         # Prepare data structure
#         students_data = []
#         column_totals = {fh.id: Decimal('0') for fh in previous_fee_heads}
#         grand_total = Decimal('0')

#         for student in students:
#             student_dues = []
#             student_total = Decimal('0')
            
#             for fh in previous_fee_heads:
#                 fee_total = previous_dues.filter(
#                     student_id=student,
#                     feetype_id__fees_type__fee_head=fh
#                 ).aggregate(
#                     total=Sum('amount')
#                 )['total'] or Decimal('0')
                
#                 partial_payments = PartialPayment.objects.filter(
#                     fee__student_id=student,
#                     fee__feetype_id__fees_type__fee_head=fh,
#                     fee__feetype_id__fee_amount_id__isnull=True
                    
#                 ).aggregate(
#                     total=Sum('amount')
#                 )['total'] or Decimal('0')
                
#                 total_due = fee_total + partial_payments
#                 student_dues.append(total_due)
#                 column_totals[fh.id] += total_due
#                 student_total += total_due

#             if student_total > 0:
#                 students_data.append({
#                     'student': student,
#                     'dues': student_dues,
#                     'total': student_total
#                 })
#                 grand_total += student_total

#         context = {
#             'students_data': students_data,
#             'fee_heads': previous_fee_heads,
#             'column_totals': column_totals,
#             'grand_total': grand_total,
#             'from_date': from_date,
#             'to_date': to_date,
#             'selected_class': selected_class,
#             'selected_section': selected_section,
#             'selected_group': selected_group,
#             'selected_shift': selected_shift,
#             'selected_version': selected_version,
#         }

#     return context

def generate_previous_dues_context(request):
    selected_class = request.GET.get('class')
    selected_section = request.GET.get('section')
    selected_group = request.GET.get('group')
    selected_shift = request.GET.get('shift')
    selected_version = request.GET.get('version')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    context = {}
    
    if from_date and to_date:
        students = StudentProfile.objects.select_related(
            'student_field',
            'class_id__class_group_id__class_id',
            'class_id__class_group_id__group_id',
            'class_id__section_id',
            'class_id__shift_id'
        )

        # Apply filters
        filters = Q()
        if selected_class: 
            filters &= Q(class_id__class_group_id__class_id=selected_class)
        if selected_section:
            filters &= Q(class_id__section_id=selected_section)
        if selected_group:
            filters &= Q(class_id__class_group_id__group_id=selected_group)
        if selected_shift:
            filters &= Q(class_id__shift_id=selected_shift)
        if selected_version:
            filters &= Q(version=selected_version)

        students = students.filter(filters)
        student_ids = students.values_list('id', flat=True)

        # Get full payments (paid fees created in the date range)
        full_payments = Fees.objects.filter(
            status='paid',
            created_at__date__range=[from_date, to_date],
            feetype_id__fee_amount_id__isnull=True,
            student_id__in=student_ids
        ).values(
            'student_id', 'feetype_id__fees_type__fee_head'
        ).annotate(full_total=Sum('amount'))

        # Get partial payments (partial payments within date range)
        partial_payments = PartialPayment.objects.filter(
            payment_date__date__range=[from_date, to_date],
            fee__feetype_id__fee_amount_id__isnull=True,
            fee__student_id__in=student_ids
        ).values(
            'fee__student_id', 'fee__feetype_id__fees_type__fee_head'
        ).annotate(partial_total=Sum('amount'))

        # Collect all unique fee heads from both full and partial payments
        fee_head_ids = set()
        for fp in full_payments:
            fee_head_ids.add(fp['feetype_id__fees_type__fee_head'])
        for pp in partial_payments:
            fee_head_ids.add(pp['fee__feetype_id__fees_type__fee_head'])
        previous_fee_heads = FeeHead.objects.filter(id__in=fee_head_ids).order_by('name')

        # Prepare data structure
        students_data = []
        column_totals = defaultdict(Decimal)  
        grand_total = Decimal('0.00')

        for student in students:
            student_id = student.id
            dues_dict = defaultdict(Decimal)

            # Process full payments for the student
            for fp in full_payments.filter(student_id=student_id):
                head_id = fp['feetype_id__fees_type__fee_head']
                amount = fp['full_total']
                dues_dict[head_id] += amount

            # Process partial payments for the student
            for pp in partial_payments.filter(fee__student_id=student_id):
                head_id = pp['fee__feetype_id__fees_type__fee_head']
                amount = pp['partial_total']
                dues_dict[head_id] += amount

            total = sum(dues_dict.values())
            if total > 0:
                students_data.append({
                    'student': student,
                    'dues_dict': dues_dict,  # Store as dictionary
                    'total': total
                })
                # Update column totals
                for head_id, amount in dues_dict.items():
                    column_totals[head_id] += amount
                grand_total += total

        # Filter fee heads to only those with non-zero totals
        filtered_fee_heads = [fh for fh in previous_fee_heads if column_totals[fh.id] > 0]
        
        # Convert student dues_dict to list format based on filtered fee heads
        for student_data in students_data:
            dues_dict = student_data.pop('dues_dict')
            student_data['dues'] = [dues_dict.get(fh.id, Decimal('0.00')) for fh in filtered_fee_heads]

        # Create final column totals dict for template
        filtered_column_totals = {fh.id: column_totals[fh.id] for fh in filtered_fee_heads}
        context = {
            'students_data': students_data,
            'fee_heads': filtered_fee_heads,
            'column_totals': filtered_column_totals,
            'grand_total': grand_total,
            'from_date': from_date,
            'to_date': to_date,
            'selected_class': selected_class,
            'selected_section': selected_section,
            'selected_group': selected_group,
            'selected_shift': selected_shift,
            'selected_version': selected_version,
        }

    return context

def previous_dues_report(request):
    context = generate_previous_dues_context(request)
    return render(request, 'crucial/report/previous_dues_report.html', context)

def download_previous_dues_pdf(request):
    context = generate_previous_dues_context(request)
    html_string = render_to_string('crucial/report/previous_dues_pdf.html', context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="previous_dues_report.pdf"'
    
    HTML(string=html_string).write_pdf(response, stylesheets=[
        CSS(string='''
            @page { 
                size: A4 landscape; 
                margin: 10mm;
            }
            table { 
                border-collapse: collapse; 
                width: 100%; 
                font-size: 8px;
            }
            th { 
                writing-mode: vertical-lr; 
                text-orientation: upright; 
                padding: 4px 1px !important; 
            }
            td { 
                padding: 3px !important; 
                text-align: right; 
            }
            td:nth-child(-n+7) { 
                text-align: center !important; 
            }
        ''')
    ])
    
    return response



import io
import pandas as pd
from django.http import HttpResponse
from django.shortcuts import render
from django.template.loader import get_template
from xhtml2pdf import pisa  # For PDF generation
from .models import Fees, Fees_name
from django.db.models import Sum, DecimalField
from decimal import Decimal

# def generate_master_head_wise_context(request):
#     from_date = request.GET.get('from_date')
#     to_date = request.GET.get('to_date')

#     data = []
#     all_fee_heads_dict = {}
#     overall_totals = {}
#     overall_grand_total = 0

#     if from_date and to_date:
#         fees = Fees.objects.filter(status='paid', created_at__date__range=[from_date, to_date])

#         shifts = ['Morning', 'Day']
#         versions = ['Bangla', 'English']

#         all_fee_heads = FeeHead.objects.filter(
#             fee_types__fees_name_type__fee_amount_id__isnull=False
#         ).distinct().order_by('name')

#         all_fee_heads_dict = {fee_head.id: fee_head.name for fee_head in all_fee_heads}

#         for version in versions:
#             for shift in shifts:
#                 shift_fees = fees.filter(student_id__version=version, student_id__class_id__shift_id__name=shift)

#                 head_totals = shift_fees.annotate(
#                     fee_head_id=F('feetype_id__fees_type__fee_head__id')
#                 ).values('fee_head_id').annotate(total=Sum('amount'))
#                 head_totals_dict = {ht['fee_head_id']: ht['total'] for ht in head_totals}

#                 fee_head_totals = {
#                     fee_head.id: head_totals_dict.get(fee_head.id, Decimal('0.00'))
#                     for fee_head in all_fee_heads
#                 }

#                 previous_due_total = (
#                     shift_fees.filter(feetype_id__fee_amount_id__isnull=True)
#                     .aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
#                 )

#                 row_grand_total = sum(fee_head_totals.values()) + previous_due_total

#                 data.append({
#                     'version': version,
#                     'shift': shift,
#                     'fee_head_totals': fee_head_totals,
#                     'previous_due_total': previous_due_total,
#                     'grand_total': row_grand_total
#                 })

#         overall_totals = {fee_head.id: Decimal('0.00') for fee_head in all_fee_heads}
#         overall_totals['previous_due_total'] = Decimal('0.00')

#         for item in data:
#             for fee_head in all_fee_heads:
#                 overall_totals[fee_head.id] += item['fee_head_totals'].get(fee_head.id, Decimal('0.00'))
#             overall_totals['previous_due_total'] += item['previous_due_total']

#         overall_grand_total = sum(overall_totals.values())

#     context = {
#         'data': data,
#         'all_fee_heads': all_fee_heads_dict,
#         'overall_totals': overall_totals,
#         'overall_grand_total': overall_grand_total
#     }
#     return context

from django.db.models import Sum, F, Q, DecimalField, Value
from django.db.models.functions import Coalesce
def generate_master_head_wise_context(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    context = {  # Ensure context is always defined
        'data': [],
        'all_fee_heads': {},
        'overall_totals': {},
        'overall_grand_total': 0
    }

    if from_date and to_date:
        # Get only paid or partial fees within date range
        fees = Fees.objects.filter(
            Q(
                Q(status='paid', created_at__date__range=[from_date, to_date]) |
                Q(partial_payments__payment_date__date__range=[from_date, to_date])
            )
        ).prefetch_related(
            Prefetch('partial_payments', 
                    queryset=PartialPayment.objects.filter(payment_date__date__range=[from_date, to_date]),
                    to_attr='relevant_partials')
        ).distinct()

        shifts = ['Morning', 'Day']
        versions = ['Bangla', 'English']

        all_fee_heads = FeeHead.objects.filter(
            fee_types__fees_name_type__fee_amount_id__isnull=False
        ).distinct().order_by('name')

        all_fee_heads_dict = {fee_head.id: fee_head.name for fee_head in all_fee_heads}
        data = []
        overall_totals = {fee_head.id: Decimal('0.00') for fee_head in all_fee_heads}
        overall_totals['previous_due_total'] = Decimal('0.00')

        for version in versions:
            for shift in shifts:
                shift_fees = fees.filter(
                    student_id__version=version,
                    student_id__class_id__shift_id__name=shift
                )

                regular_payments = shift_fees.filter(
                    feetype_id__fee_amount_id__isnull=False
                ).annotate(
                    fee_head_id=F('feetype_id__fees_type__fee_head__id')
                ).values('fee_head_id').annotate(
                    total=Coalesce(
                        Sum(
                            Case(
                                When(status='paid', then='amount'),
                                default=0,
                                output_field=DecimalField()
                            )
                        ), 
                        Value(0, output_field=DecimalField())
                    )
                )

                head_totals_dict = {item['fee_head_id']: item['total'] for item in regular_payments}
                previous_due_total = sum(
                    fee.amount if fee.status == 'paid' else 0
                    for fee in shift_fees.filter(feetype_id__fee_amount_id__isnull=True)
                )

                for fee in shift_fees.filter(feetype_id__fee_amount_id__isnull=True):
                    previous_due_total += sum(partial.amount for partial in getattr(fee, 'relevant_partials', []))

                fee_head_totals = {fee_head.id: head_totals_dict.get(fee_head.id, Decimal('0.00')) for fee_head in all_fee_heads}
                row_grand_total = sum(fee_head_totals.values()) + previous_due_total

                data.append({
                    'version': version,
                    'shift': shift,
                    'fee_head_totals': fee_head_totals,
                    'previous_due_total': previous_due_total,
                    'grand_total': row_grand_total
                })

                for fee_head in all_fee_heads:
                    overall_totals[fee_head.id] += fee_head_totals.get(fee_head.id, Decimal('0.00'))
                overall_totals['previous_due_total'] += previous_due_total

        overall_grand_total = sum(overall_totals.values())

        active_fee_head_ids = [fh_id for fh_id, total in overall_totals.items() if fh_id != 'previous_due_total' and total > 0]

        active_fee_heads_dict = {fh_id: name for fh_id, name in all_fee_heads_dict.items() if fh_id in active_fee_head_ids}

        for entry in data:
            entry['fee_head_totals'] = {fh_id: amount for fh_id, amount in entry['fee_head_totals'].items() if fh_id in active_fee_head_ids}

        filtered_overall_totals = {fh_id: overall_totals[fh_id] for fh_id in active_fee_head_ids}
        filtered_overall_totals['previous_due_total'] = overall_totals['previous_due_total']

        context = {
            'data': data,
            'all_fee_heads': active_fee_heads_dict,
            'overall_totals': filtered_overall_totals,
            'overall_grand_total': overall_grand_total
        }

    return context


def master_head_wise_report(request):
    context = generate_master_head_wise_context(request)
    return render(request, 'crucial/report/master_head_wise_report.html', context)


def download_pdf(request):
    context = generate_master_head_wise_context(request)
    template = get_template('crucial/report/master_head_wise_report_pdf.html')
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="master_head_wise_report.pdf"'
    pisa_status = pisa.CreatePDF(io.BytesIO(html.encode("UTF-8")), dest=response, encoding='UTF-8')
    
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response


def download_excel(request):
    context = generate_master_head_wise_context(request)
    data = []

    for item in context['data']:
        row = {
            'Version': item['version'],
            'Shift': item['shift'],
            'Previous Software\'s Due': item['previous_due_total'],
            'Total': item['grand_total']
        }
        for fee_head_id, fee_title in context['all_fee_heads'].items():
            row[fee_title] = item['fee_head_totals'].get(fee_head_id, Decimal('0.00'))
        data.append(row)

    df = pd.DataFrame(data)
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Master Report')

    workbook = writer.book
    worksheet = writer.sheets['Master Report']
    format_header = workbook.add_format({'bold': True, 'bg_color': '#D7E4BC', 'border': 1})
    for col_num, value in enumerate(df.columns.values):
        worksheet.write(0, col_num, value, format_header)

    writer.close()
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="master_head_wise_report.xlsx"'
    return response


def generate_previous_due_context(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    context = {}

    if from_date and to_date:
        # Get Fee Heads that have previous due entries
        previous_heads = FeeHead.objects.filter(
            fee_types__fees_name_type__fee_amount_id__isnull=True
        ).distinct().order_by('name')

        # Get relevant payments
        fees = Fees.objects.filter(
            Q(
                Q(status='paid', created_at__date__range=[from_date, to_date]) |
                Q(status='partial', partial_payments__payment_date__date__range=[from_date, to_date])
            ),
            feetype_id__fee_amount_id__isnull=True
        ).prefetch_related('partial_payments').distinct()

        versions = ['Bangla', 'English']
        shifts = ['Morning', 'Day']
        data = []
        head_totals = {head.id: Decimal('0.00') for head in previous_heads}
        grand_total = Decimal('0.00')

        for version in versions:
            for shift in shifts:
                shift_fees = fees.filter(
                    student_id__version=version,
                    student_id__class_id__shift_id__name=shift
                )

                # Calculate totals with Case/When
                fee_breakdown = shift_fees.annotate(
                    head_id=F('feetype_id__fees_type__fee_head__id')
                ).values('head_id').annotate(
                    total=Coalesce(
                        Sum(
                            Case(
                                When(status='paid', then='amount'),
                                When(status='partial', then=0),
                                output_field=DecimalField()
                            )
                        ), 
                        Value(0, output_field=DecimalField())
                    ) +
                    Coalesce(
                        Sum(
                            Case(
                                When(status='partial', then='partial_payments__amount'),
                                default=0,
                                output_field=DecimalField()
                            )
                        ),
                        Value(0, output_field=DecimalField())
                    )
                )

                row_totals = {head.id: Decimal('0.00') for head in previous_heads}
                for item in fee_breakdown:
                    head_id = item['head_id']
                    if head_id in row_totals:
                        amount = item['total']
                        row_totals[head_id] = amount
                        head_totals[head_id] += amount

                row_total = sum(row_totals.values())
                grand_total += row_total

                data.append({
                    'version': version,
                    'shift': shift,
                    'head_totals': row_totals,
                    'row_total': row_total
                })

        # Filter out columns with zero totals
        active_head_ids = [head_id for head_id, total in head_totals.items() if total > 0]
        active_previous_heads = [head for head in previous_heads if head.id in active_head_ids]

        # Filter data entries to only show active columns
        for entry in data:
            entry['head_totals'] = {
                head_id: amount 
                for head_id, amount in entry['head_totals'].items()
                if head_id in active_head_ids
            }

        # Create filtered totals dictionary
        filtered_overall_totals = {
            head_id: total 
            for head_id, total in head_totals.items()
            if head_id in active_head_ids
        }

        context = {
            'data': data,
            'previous_heads': {head.id: head.name for head in active_previous_heads},
            'overall_totals': filtered_overall_totals,
            'overall_grand_total': grand_total,
            'from_date': from_date,
            'to_date': to_date
        }
    
    return context


def previous_software_due_report(request):
    context = generate_previous_due_context(request)
    return render(request, 'crucial/report/previous_software_due.html', context)

def download_previous_due_pdf(request):
    context = generate_previous_due_context(request)
    template = get_template('crucial/report/previous_software_due_pdf.html')
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    filename = f"Previous_Dues_{timezone.now().strftime('%Y-%m-%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    pisa.CreatePDF(
        io.BytesIO(html.encode("UTF-8")),
        dest=response,
        encoding='UTF-8'
    )
    return response

# from django.shortcuts import render
# from django.db.models import Prefetch, DecimalField, Sum
# from django.db.models.functions import Coalesce
# from django.http import HttpResponse
# from django.template.loader import render_to_string
# from weasyprint import HTML
# from decimal import Decimal
# from core.models import StudentSection

# def fee_reports(request):
#     months = Fee_month.objects.all()
#     classes = StudentClass.objects.all()
#     sections = StudentSection.objects.all()
#     groups = StuGroup.objects.all()
#     versions = [v[0] for v in StudentProfile.Version.choices]
    
#     paid_students = []
#     unpaid_students = []
#     context = {
#         'months': months,
#         'classes': classes,
#         'sections': sections,
#         'groups': groups,
#         'versions': versions,
#     }

#     if request.method == 'GET' and 'class' in request.GET:
#         class_id = request.GET.get('class')
#         section_id = request.GET.get('section')
#         version = request.GET.get('version')
#         group_id = request.GET.get('group')
#         month_id = request.GET.get('month')
#         status = request.GET.get('status')

#         filters = {
#             'class_id__class_group_id__class_id': class_id,
#         }
#         if section_id:
#             filters['class_id__section_id'] = section_id
#         if version:
#             filters['version'] = version
#         if group_id:
#             filters['class_id__class_group_id__group_id'] = group_id

#         students = StudentProfile.objects.filter(**filters).prefetch_related(
#             Prefetch(
#                 'fees_set',
#                 queryset=Fees.objects.filter(month_id=month_id),
#                 to_attr='month_fees'
#             )
#         )

#         if status == 'Paid':
#             for student in students:
#                 if student.month_fees and all(fee.status == 'paid' for fee in student.month_fees):
#                     paid_students.append(student)
#         elif status == 'Unpaid':
#             for student in students:
#                 total_unpaid = Decimal('0.00')
#                 has_unpaid = False
#                 for fee in student.month_fees:
#                     if fee.status in ['unpaid', 'partial']:
#                         total_unpaid += fee.total_fee_after_partial_payments()
#                         has_unpaid = True
#                 if has_unpaid:
#                     unpaid_students.append({
#                         'student': student,
#                         'total_unpaid': total_unpaid
#                     })

#         if 'download' in request.GET:
#             template = 'crucial/report/pdf_report.html' if status == 'Paid' else 'crucial/report/pdf_unpaid_report.html'
#             html_string = render_to_string(template, {
#                 'students': paid_students if status == 'Paid' else unpaid_students,
#                 'status': status,
#                 'month': Fee_month.objects.get(id=month_id) if month_id else '',
#                 'class': StudentClass.objects.get(id=class_id) if class_id else '',
#                 'section': StudentSection.objects.get(id=section_id) if section_id else '',
#                 'group': StuGroup.objects.get(id=group_id) if group_id else '',
#                 'version': version
#             })
#             html = HTML(string=html_string)
#             pdf = html.write_pdf()
#             response = HttpResponse(pdf, content_type='application/pdf')
#             filename = f"{status}_report.pdf"
#             response['Content-Disposition'] = f'attachment; filename="{filename}"'
#             return response

#         context.update({
#             'paid_students': paid_students,
#             'unpaid_students': unpaid_students,
#             'selected_class': class_id,
#             'selected_section': section_id,
#             'selected_version': version,
#             'selected_group': group_id,
#             'selected_month': month_id,
#             'selected_status': status,
#         })

#     return render(request, 'crucial/report/fee_reports.html', context)

from core.models import StudentSection
from django.shortcuts import render
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
from decimal import Decimal
from django.db.models import Prefetch
from .tasks import generate_pdf_report_task
from django.core.cache import cache
import time

def fee_reports(request):
    months = Fee_month.objects.all()
    classes = StudentClass.objects.all()
    sections = StudentSection.objects.all()
    groups = StuGroup.objects.all()
    versions = [v[0] for v in StudentProfile.Version.choices]
    
    paid_students = []
    unpaid_students = []
    context = {
        'months': months,
        'classes': classes,
        'sections': sections,
        'groups': groups,
        'versions': versions,
    }

    if request.method == 'GET' and 'class' in request.GET:
        class_id = request.GET.get('class')
        section_id = request.GET.get('section')
        version = request.GET.get('version')
        group_id = request.GET.get('group')
        month_id = request.GET.get('month')
        status = request.GET.get('status')

        filters = {
            'class_id__class_group_id__class_id': class_id,
        }
        if section_id:
            filters['class_id__section_id'] = section_id
        if version:
            filters['version'] = version
        if group_id:
            filters['class_id__class_group_id__group_id'] = group_id

        students = StudentProfile.objects.filter(**filters).prefetch_related(
            Prefetch(
                'fees_set',
                queryset=Fees.objects.filter(month_id=month_id),
                to_attr='month_fees'
            )
        )

        if status == 'Paid':
            for student in students:
                if student.month_fees and all(fee.status == 'paid' for fee in student.month_fees):
                    paid_students.append(student)
        elif status == 'Unpaid':
            for student in students:
                total_unpaid = Decimal('0.00')
                has_unpaid = False
                for fee in student.month_fees:
                    if fee.status in ['unpaid', 'partial']:
                        total_unpaid += fee.total_fee_after_partial_payments()
                        has_unpaid = True
                if has_unpaid:
                    unpaid_students.append({
                        'student': student,
                        'total_unpaid': total_unpaid
                    })

        if 'download' in request.GET:
            serialized_students = []
            for student in (paid_students if status == 'Paid' else unpaid_students):
                if status == 'Paid':
                    serialized_students.append(student.id)
                else:  
                    serialized_students.append({
                        'student_id': student['student'].id,
                        'total_unpaid': str(student['total_unpaid']) 
                    })

            report_params = {
                'template': 'crucial/report/pdf_report.html' if status == 'Paid' else 'crucial/report/pdf_unpaid_report.html',
                'context_data': {  
                    'student_ids': serialized_students,
                    'status': status,
                    'month_id': month_id,
                    'class_id': class_id,
                    'section_id': section_id,
                    'group_id': group_id,
                    'version': version
                },
                'filename': f"{status}_report_{int(time.time())}.pdf"
            }

            task = generate_pdf_report_task.delay(report_params)
            
            cache.set(task.id, {
                'status': 'PENDING',
                'filename': report_params['filename']
            }, 300)
            
            return HttpResponse(task.id)
            # return HttpResponse(
            #     f"Report generation started! Task ID: {task.id}. "
            #     f"Check /check-report/?task_id={task.id} for status."
            # )

        context.update({
            'paid_students': paid_students,
            'unpaid_students': unpaid_students,
            'selected_class': class_id,
            'selected_section': section_id,
            'selected_version': version,
            'selected_group': group_id,
            'selected_month': month_id,
            'selected_status': status,
        })

    return render(request, 'crucial/report/fee_reports.html', context)

from core.models import StudentShift
import datetime
import os

def feedate_reports(request):
    classes = StudentClass.objects.all()
    sections = StudentSection.objects.all()
    groups = StuGroup.objects.all()
    versions = [v[0] for v in StudentProfile.Version.choices]
    shifts = StudentShift.objects.all()
    academic_sessions = AcademicSession.objects.all()
    
    paid_students = []
    unpaid_students = []
    context = {
        'classes': classes,
        'sections': sections,
        'groups': groups,
        'versions': versions,
        'shifts': shifts,
        'academic_sessions': academic_sessions,
    }

    if request.method == 'GET' and 'class' in request.GET:
        class_id = request.GET.get('class')
        section_id = request.GET.get('section')
        version = request.GET.get('version')
        group_id = request.GET.get('group')
        from_date_str = request.GET.get('from_date')
        to_date_str = request.GET.get('to_date')
        status = request.GET.get('status')
        shift_id = request.GET.get('shift')
        academic_session_id = request.GET.get('academic_session')

        # Parse dates
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
        # from_date = datetime.datetime.strptime(from_date_str, '%Y-%m-%d').date()
        # to_date = datetime.datetime.strptime(to_date_str, '%Y-%m-%d').date()


        # Create datetime objects for the query
        start_datetime = timezone.make_aware(datetime.combine(from_date, datetime.min.time()))
        end_datetime = timezone.make_aware(datetime.combine(to_date, datetime.max.time()))
        
        # start_datetime = timezone.make_aware(datetime.datetime.combine(from_date, datetime.datetime.min.time()))
        # end_datetime = timezone.make_aware(datetime.datetime.combine(to_date, datetime.datetime.max.time()))


        filters = {
            'class_id__class_group_id__class_id': class_id,
            'student_field__status': 'Active',
        }
        if section_id:
            filters['class_id__section_id'] = section_id
        if version:
            filters['version'] = version
        if group_id:
            filters['class_id__class_group_id__group_id'] = group_id
        if shift_id:  
            filters['class_id__shift_id'] = shift_id
        if academic_session_id: 
            filters['academic_session_year_id'] = academic_session_id
            

        students = StudentProfile.objects.filter(**filters).prefetch_related(
            Prefetch(
                'fees_set',
                queryset=Fees.objects.filter(created_at__range=(start_datetime, end_datetime)),
                to_attr='date_range_fees'
            )
        )

        if status == 'Paid':
            paid_students = [
                student for student in students 
                if student.date_range_fees and 
                all(fee.status == 'paid' for fee in student.date_range_fees)
            ]
            
        elif status == 'Unpaid':
            for student in students:
                total_unpaid = Decimal('0.00')
                has_unpaid = any(
                    fee.status in ['unpaid', 'partial'] 
                    for fee in student.date_range_fees
                ) or not student.date_range_fees

                if has_unpaid:
                    for fee in student.date_range_fees:
                        if fee.status in ['unpaid', 'partial']:
                            total_unpaid += fee.total_fee_after_partial_payments()
                    if not student.date_range_fees:
                        total_unpaid = 'No fee records'  
                        
                    unpaid_students.append({
                        'student': student,
                        'total_unpaid': total_unpaid
                    })
                    
        if 'download' in request.GET:
            # Serialize student data
            serialized_students = []
            if status == 'Paid':
                for student in paid_students:
                    serialized_students.append({
                        'name': student.student_field.name,
                        'phone_number': student.parent_id.phone_number,
                        'roll_no': student.roll_no,
                        'class_name': student.class_id.class_group_id.class_id.name,
                        'section': student.class_id.section_id.name if student.class_id.section_id else '-',
                        'version': student.version if student.version else '-',
                        'shift': student.class_id.shift_id.name if student.class_id.shift_id else '-',
                        'group': student.class_id.class_group_id.group_id.name if student.class_id.class_group_id.group_id else '-',
                    })
            else:
                for entry in unpaid_students:
                    serialized_students.append({
                        'name': entry['student'].student_field.name,
                        'phone_number': entry['student'].parent_id.phone_number,
                        'roll_no': entry['student'].roll_no,
                        'class_name': entry['student'].class_id.class_group_id.class_id.name,
                        'section': entry['student'].class_id.section_id.name if entry['student'].class_id.section_id else '-',
                        'version': entry['student'].version if entry['student'].version else '-',
                        'shift': entry['student'].class_id.shift_id.name if entry['student'].class_id.shift_id else '-',
                        'group': entry['student'].class_id.class_group_id.group_id.name if entry['student'].class_id.class_group_id.group_id else '-',
                        'total_unpaid': f"{Decimal(entry['total_unpaid']):.2f}" if isinstance(entry['total_unpaid'], (Decimal, int, float)) else 'No fee records'
                    })

            # Create PDF context
            pdf_context = {
                        'students': serialized_students,
                        'status': status,
                        'from_date': from_date_str,
                        'to_date': to_date_str,
                        'class_name': StudentClass.objects.get(id=class_id).name,
                        'section': StudentSection.objects.get(id=section_id).name if section_id else 'All',
                        'group': StuGroup.objects.get(id=group_id).name if group_id else 'All',
                        'version': version if version else 'All',
                        'shift': StudentShift.objects.get(id=shift_id).name if shift_id else 'All',
                        'academic_session': AcademicSession.objects.get(id=academic_session_id).__str__() if academic_session_id else 'All',
                        'generated_date': timezone.now(),
                        'total_unpaid_sum': sum(
                            Decimal(s['total_unpaid']) for s in serialized_students if s['total_unpaid'].replace('.', '', 1).isdigit()
                        ) if status == 'Unpaid' else Decimal('0.00')
                    }


            # Render PDF template
            template_path = 'crucial/report/pdf_paid.html' if status == 'Paid' else 'crucial/report/pdf_unpaid.html'            
            html_string = render_to_string(template_path, pdf_context)
            
            # Create PDF response
            response = HttpResponse(content_type='application/pdf')
            filename = f"{status}_report_{from_date_str}_to_{to_date_str}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            # Generate PDF
            pisa_status = pisa.CreatePDF(
                html_string,
                dest=response,
                encoding='UTF-8',
                link_callback=lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, '')))
            
            if pisa_status.err:
                return HttpResponse('PDF generation error', status=500)
            return response
            
        context.update({
            'paid_students': paid_students,
            'unpaid_students': unpaid_students,
            'selected_class': class_id,
            'selected_section': section_id,
            'selected_version': version,
            'selected_group': group_id,
            'selected_from_date': from_date_str,
            'selected_to_date': to_date_str,
            'selected_status': status,
            'selected_shift': shift_id,
            'selected_academic_session': academic_session_id,
        })

    return render(request, 'crucial/report/fee_date_reports.html', context)


def export_fees_to_excel(request):
    class_id = request.GET.get('class')
    section_id = request.GET.get('section')
    version = request.GET.get('version')
    group_id = request.GET.get('group')
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    status = request.GET.get('status')
    shift_id = request.GET.get('shift')

    from_date = datetime.datetime.strptime(from_date_str, '%Y-%m-%d').date()
    to_date = datetime.datetime.strptime(to_date_str, '%Y-%m-%d').date()
    start_datetime = timezone.make_aware(datetime.datetime.combine(from_date, datetime.datetime.min.time()))
    end_datetime = timezone.make_aware(datetime.datetime.combine(to_date, datetime.datetime.max.time()))

    filters = {'class_id__class_group_id__class_id': class_id, 'student_field__status': 'Active',}
    if section_id:
        filters['class_id__section_id'] = section_id
    if version:
        filters['version'] = version
    if group_id:
        filters['class_id__class_group_id__group_id'] = group_id
    if shift_id:
        filters['class_id__shift_id'] = shift_id

    students = StudentProfile.objects.filter(**filters).prefetch_related(
        Prefetch(
            'fees_set',
            queryset=Fees.objects.filter(created_at__range=(start_datetime, end_datetime)),
            to_attr='date_range_fees'
        )
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Fee Reports"
    
    headers = ["Name", "Roll No", "Class", "Section", "Group", "Version", "Shift", "Total Unpaid" if status == "Unpaid" else "Status"]

    sheet.append(headers)

    for student in students:
        if status == "Paid":
            if student.date_range_fees and all(fee.status == 'paid' for fee in student.date_range_fees):
                sheet.append([
                    student.student_field.name,
                    student.roll_no,
                    student.class_id.class_group_id.class_id.name,
                    student.class_id.section_id.name if student.class_id.section_id else '-',
                    student.class_id.class_group_id.group_id.name if student.class_id.class_group_id.group_id else '-',
                    student.version if student.version else '-',
                    student.class_id.shift_id.name if student.class_id.shift_id else '-',
                    "Paid"
                ])
        elif status == "Unpaid":
            total_unpaid = Decimal('0.00')
            has_unpaid = any(fee.status in ['unpaid', 'partial'] for fee in student.date_range_fees) or not student.date_range_fees
            if has_unpaid:
                for fee in student.date_range_fees:
                    if fee.status in ['unpaid', 'partial']:
                        total_unpaid += fee.total_fee_after_partial_payments()
                if not student.date_range_fees:
                    total_unpaid = 'No fee records'
                sheet.append([
                    student.student_field.name,
                    student.roll_no,
                    student.class_id.class_group_id.class_id.name,
                    student.class_id.section_id.name if student.class_id.section_id else '-',
                    student.class_id.class_group_id.group_id.name if student.class_id.class_group_id.group_id else '-',
                    student.version if student.version else '-',
                    student.class_id.shift_id.name if student.class_id.shift_id else '-',
                    f"{Decimal(total_unpaid):.2f}" if isinstance(total_unpaid, (Decimal, int, float)) else 'No fee records'
                ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="fee_report_{from_date_str}_to_{to_date_str}.xlsx"'
    workbook.save(response)
    return response



from django.http import JsonResponse
def check_report_status(request):
    task_id = request.GET.get('task_id')
    if not task_id:
        return JsonResponse({'status': 'ERROR', 'error': 'Missing task ID'}, status=400)
    
    task_result = cache.get(task_id, {})
    
    if not task_result:
        return JsonResponse({'status': 'ERROR', 'error': 'Invalid task ID'}, status=404)
    
    response_data = {
        'status': task_result.get('status', 'UNKNOWN'),
        'error': task_result.get('error')
    }
    
    if task_result.get('status') == 'SUCCESS':
        response_data['download_url'] = f'/serve-pdf/?file={task_result["file_path"]}'
    
    return JsonResponse(response_data)
    
    
from django.core.files.storage import default_storage
from django.core.exceptions import SuspiciousFileOperation
from django.http import HttpResponse, FileResponse

def serve_generated_pdf(request):
    file_path = request.GET.get('file')
    if not file_path:
        return HttpResponse("File parameter missing", status=400)

    try:
        if '..' in file_path or not file_path.startswith('reports/'):
            raise SuspiciousFileOperation("Invalid file path")

        if not default_storage.exists(file_path):
            return HttpResponse("File not found", status=404)

        file = default_storage.open(file_path)
        response = FileResponse(file)
        response['Content-Type'] = 'application/pdf'
        response['Content-Disposition'] = f'attachment; filename="{file_path.split("/")[-1]}"'
        return response
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)
    
    
    
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Fees, Fees_name
def update_fee_type(request):
    if not request.user.is_staff:
        messages.error(request, "Permission denied")
        return redirect('admin:index')

    fee_types = Fees_name.objects.all()
    
    if request.method == 'POST':
        old_id = request.POST.get('old_feetype')
        new_id = request.POST.get('new_feetype')
        
        if not old_id or not new_id:
            messages.error(request, "Both fee types must be selected")
            return render(request, 'fees/update_feetype.html', {'fee_types': fee_types})
        
        try:
            # Get the fee type objects
            old_feetype = Fees_name.objects.get(id=old_id)
            new_feetype = Fees_name.objects.get(id=new_id)

            # Perform the update
            updated_count = Fees.objects.filter(feetype_id=old_feetype).update(feetype_id=new_feetype)
            
            messages.success(request, f"Successfully updated {updated_count} fees from {old_feetype.fees_title} to {new_feetype.fees_title}")
            
            return redirect('update_fee_type')

        except Fees_name.DoesNotExist:
            messages.error(request, "One of the selected fee types does not exist")
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")

    return render(request, 'fees/update_feetype.html', {'fee_types': fee_types})


def fees_list(request):
    fees = Fees.objects.all() 
    return render(request, 'fees/fees_list.html', {'fees': fees})


def transfer_fees(request):
    if not request.user.is_staff:
        messages.error(request, "Permission denied")
        return redirect('admin:index')

    fee_types = Fees_name.objects.all()
    
    if request.method == 'POST':
        old_id = request.POST.get('old_feetype')
        new_id = request.POST.get('new_feetype')
        
        if not old_id or not new_id:
            messages.error(request, "Both fee types must be selected")
            return render(request, 'fees/transfer_fees.html', {'fee_types': fee_types})
        
        if old_id == new_id:
            messages.error(request, "Source and destination fee types cannot be the same")
            return render(request, 'fees/transfer_fees.html', {'fee_types': fee_types})
        
        try:
            old_feetype = Fees_name.objects.get(id=old_id)
            new_feetype = Fees_name.objects.get(id=new_id)

            # Update only unpaid fees
            updated_count = Fees.objects.filter(
                feetype_id=old_feetype,
                status='unpaid'
            ).update(feetype_id=new_feetype)
            
            messages.success(
                request, 
                f"Transferred {updated_count} unpaid fees from {old_feetype.fees_title} to {new_feetype.fees_title}"
            )
            
            return redirect('transfer_fees')

        except Fees_name.DoesNotExist:
            messages.error(request, "One of the selected fee types does not exist")
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")

    return render(request, 'fees/transfer_fees.html', {'fee_types': fee_types})

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Count, Exists, OuterRef
from .models import Fees, Fees_name

def duplicate_fees_list(request):
    if not request.user.is_staff:
        messages.error(request, "Permission denied")
        return redirect('admin:index')

    # Subquery to identify duplicate student_id and feetype_id pairs
    duplicate_pairs = Fees.objects.filter(
        student_id=OuterRef('student_id'),
        feetype_id=OuterRef('feetype_id')
    ).values('student_id', 'feetype_id').annotate(count=Count('id')).filter(count__gt=1)

    # Get all fees that are part of duplicate pairs
    duplicate_fees = Fees.objects.annotate(
        is_duplicate=Exists(duplicate_pairs)
    ).filter(is_duplicate=True).select_related(
        'student_id__student_field',
        'student_id__class_id__class_group_id__class_id',
        'student_id__class_id__class_group_id__group_id',
        'student_id__class_id__section_id',
        'student_id__class_id__shift_id',
        'feetype_id',
        'month_id',
    ).order_by('student_id__student_field__name')

    fees_data = []
    for fee in duplicate_fees:
        student = fee.student_id
        class_config = student.class_id
        class_group_config = class_config.class_group_id if class_config else None

        # Initialize values with defaults
        class_name = '-'
        group = '-'
        section = '-'
        shift = '-'
        version = student.version if student.version else '-'
        roll = student.roll_no if student.roll_no is not None else '-'

        if class_group_config:
            if class_group_config.class_id:
                class_name = class_group_config.class_id.name
            if class_group_config.group_id:
                group = class_group_config.group_id.name
        
        if class_config:
            if class_config.section_id:
                section = class_config.section_id.name
            if class_config.shift_id:
                shift = class_config.shift_id.name

        fees_data.append({
            'id': fee.id,
            'student_name': student.student_field.name,
            'roll': roll,
            'class': class_name,
            'group': group,
            'section': section,
            'version': version,
            'shift': shift,
            'feetype_id': fee.feetype_id.fees_title if fee.feetype_id else '-',
            'amount': fee.amount,
            'status': fee.status,
            'month': fee.month_id.name if fee.month_id else '-',
        })

    return render(request, 'fees/duplicate_fees_list.html', {'fees': fees_data})

def delete_duplicate_fee(request, fee_id):
    if not request.user.is_staff:
        messages.error(request, "Permission denied")
        return redirect('admin:index')

    fee = get_object_or_404(Fees, id=fee_id)
    if request.method == 'POST':
        fee.delete()
        messages.success(request, 'Fee entry deleted successfully.')
    return redirect('duplicate_fees_list')


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Count, Q
from .models import Fees, Fees_name, StudentProfile

def find_duplicate_fees(request):
    if not request.user.is_staff:
        messages.error(request, "Permission denied")
        return redirect('admin:index')

    fee_types = Fees_name.objects.all()
    students = []
    
    if request.method == 'POST':
        feetype1_id = request.POST.get('feetype1')
        feetype2_id = request.POST.get('feetype2')
        
        if not feetype1_id or not feetype2_id:
            messages.error(request, "Please select both fee types")
            return render(request, 'fees/find_duplicates.html', {'fee_types': fee_types})
        
        # Get students with BOTH fee types using intersection
        fee1_students = Fees.objects.filter(feetype_id=feetype1_id).values_list('student_id', flat=True)
        fee2_students = Fees.objects.filter(feetype_id=feetype2_id).values_list('student_id', flat=True)
        common_student_ids = set(fee1_students).intersection(set(fee2_students))
        
        # Get full student objects and their fees
        students = []
        for student in StudentProfile.objects.filter(id__in=common_student_ids):
            fees = Fees.objects.filter(
                student_id=student,
                feetype_id__in=[feetype1_id, feetype2_id]
            ).select_related(
                'feetype_id',
                'student_id__class_id__class_group_id__class_id',
                'student_id__class_id__class_group_id__group_id',
                'student_id__class_id__section_id',
                'student_id__class_id__shift_id',
            ).order_by('feetype_id')
            
            if fees.count() >= 2:
                students.append({
                    'student': student,
                    'fees': fees,
                    'student_info': get_student_info(student)
                })

    return render(request, 'fees/find_duplicates.html', {
        'fee_types': fee_types,
        'students': students
    })

def get_student_info(student):
    class_config = student.class_id
    class_group_config = class_config.class_group_id if class_config else None

    return {
        'roll': student.roll_no or '-',
        'class': class_group_config.class_id.name if class_group_config and class_group_config.class_id else '-',
        'group': class_group_config.group_id.name if class_group_config and class_group_config.group_id else '-',
        'section': class_config.section_id.name if class_config and class_config.section_id else '-',
        'version': student.version or '-',
        'shift': class_config.shift_id.name if class_config and class_config.shift_id else '-'
    }

def delete_fee_entry(request, fee_id):
    if not request.user.is_staff:
        messages.error(request, "Permission denied")
        return redirect('admin:index')

    fee = get_object_or_404(Fees, id=fee_id)
    if request.method == 'POST':
        fee.delete()
        messages.success(request, 'Fee entry deleted successfully')
    return redirect('find_duplicate_fees')


from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Fees, Fees_name

def manage_fees(request):
    if not request.user.is_staff:
        messages.error(request, "Permission denied")
        return redirect('admin:index')

    fee_types = Fees_name.objects.all()
    selected_feetype = None
    unpaid_students = []
    paid_students = []

    if request.method == 'POST':
        # Handle fee type selection
        if 'feetype_submit' in request.POST:
            selected_feetype_id = request.POST.get('feetype')
            if selected_feetype_id:
                try:
                    selected_feetype = Fees_name.objects.get(id=selected_feetype_id)
                    # Include 'partial' in unpaid status
                    unpaid_students = Fees.objects.filter(feetype_id=selected_feetype, status__in=['unpaid', 'partial'])
                    paid_students = Fees.objects.filter(feetype_id=selected_feetype, status='paid')
                except Fees_name.DoesNotExist:
                    messages.error(request, "Invalid fee type selected")
        
        # Handle deletion requests
        elif 'delete_unpaid' in request.POST or 'delete_paid' in request.POST:
            selected_feetype_id = request.POST.get('selected_feetype')
            selected_student_ids = request.POST.getlist('selected_students')
            
            if selected_feetype_id and selected_student_ids:
                try:
                    selected_feetype = Fees_name.objects.get(id=selected_feetype_id)
                    # Delete selected fees entries
                    deleted_count, _ = Fees.objects.filter(id__in=selected_student_ids).delete()
                    messages.success(request, f"Deleted {deleted_count} records successfully!")
                    
                    # Refresh student lists after deletion
                    unpaid_students = Fees.objects.filter(feetype_id=selected_feetype, status__in=['unpaid', 'partial'])
                    paid_students = Fees.objects.filter(feetype_id=selected_feetype, status='paid')
                
                except Fees_name.DoesNotExist:
                    messages.error(request, "Invalid fee type during deletion")
                except Exception as e:
                    messages.error(request, f"Error occurred: {str(e)}")

    return render(request, 'fees/manage_fees.html', {
        'fee_types': fee_types,
        'selected_feetype': selected_feetype,
        'unpaid_students': unpaid_students,
        'paid_students': paid_students,
    })
    


# def generate_master_head_wise_context_two(request):
#     from_date = request.GET.get('from_date')
#     to_date = request.GET.get('to_date')

#     context = {
#         'data': [],
#         'all_fee_heads': {},
#         'overall_totals': {},
#         'overall_grand_total': Decimal('0.00'),
#         'from_date': from_date,
#         'to_date': to_date,
#         'other_fee_total': Decimal('0.00')
#     }

#     if from_date and to_date:
#         try:
#             from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
#             to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()
#         except ValueError:
#             return context        
        
#         receives = Receive.objects.filter(
#             date__gte=from_date_obj,
#             date__lte=to_date_obj,
#         ).exclude(voucher_no__startswith='FEE').select_related('student').order_by('student__student_field__name')

#         # Apply additional filters
#         filters = {
#             'student__class_id__class_group_id__class_id': request.GET.get('class'),
#             'student__class_id__section_id': request.GET.get('section'),
#             'student__class_id__class_group_id__group_id': request.GET.get('group'),
#             'student__class_id__shift_id': request.GET.get('shift'),
#             'student__version': request.GET.get('version'),
#         }
#         filters = {k: v for k, v in filters.items() if v}
#         receives = receives.filter(**filters)

#         # Separate student and non-student receives
#         student_receives = []
#         other_fee_total = Decimal('0.00')
#         for receive in receives:
#             # ========== MODIFIED: Handle REC2 as Other Fee automatically ========== #
#             if receive.voucher_no.startswith('REC2') or not receive.student:
#                 other_fee_total += receive.amount
#             else:
#                 student_receives.append(receive)

#         # ========== REST OF THE CODE REMAINS THE SAME ========== #
#         student_ids = [r.student_id for r in student_receives]
#         receive_dates = {r.created_at.date() for r in student_receives}

#         # Fetch Fees entries ordered by id for deterministic processing
#         fees_entries = Fees.objects.filter(
#             student_id__in=student_ids,
#             updated_at__date__in=receive_dates
#         ).select_related('feetype_id__fees_type__fee_head').order_by('id')

#         # Group fees entries by (student_id, date, amount)
#         fees_grouped = defaultdict(list)
#         for fee in fees_entries:
#             key = (fee.student_id_id, fee.updated_at.date(), fee.amount)
#             if fee.feetype_id and fee.feetype_id.fees_type:
#                 fees_grouped[key].append({
#                     'fee_head': fee.feetype_id.fees_type.fee_head,
#                     'fee_id': fee.id
#                 })

#         # Track used fee entries
#         used_fees = set()
#         aggregated_data = defaultdict(lambda: defaultdict(Decimal))
#         grand_total = Decimal('0.00')

#         for receive in student_receives:
#             student = receive.student
#             version = student.version if student else "Unknown"
#             shift = student.class_id.shift_id.name if student and student.class_id and student.class_id.shift_id else "Unknown"
            
#             key = (student.id, receive.created_at.date(), receive.amount)
#             fee_head = None

#             # Find first unused fee entry for this key
#             possible_fees = fees_grouped.get(key, [])
#             for fee_entry in possible_fees:
#                 if fee_entry['fee_id'] not in used_fees:
#                     fee_head = fee_entry['fee_head']
#                     used_fees.add(fee_entry['fee_id'])
#                     break

#             if fee_head:
#                 aggregated_data[(version, shift)][fee_head.id] += receive.amount
#                 aggregated_data[(version, shift)]['total'] += receive.amount
#                 grand_total += receive.amount

#         # Add Other Fee to grand total (now includes REC2 vouchers)
#         grand_total += other_fee_total

#         # Get only fee heads that have actual values
#         used_fee_heads = FeeHead.objects.filter(
#             id__in=set(fh_id for _, fh_totals in aggregated_data.items() 
#                       for fh_id in fh_totals.keys() if fh_id != 'total')
#         ).order_by('name')

#         # Format data for template
#         data = []
#         for (version, shift), fee_totals in aggregated_data.items():
#             data.append({
#                 'version': version,
#                 'shift': shift,
#                 'fee_head_totals': {fh.id: fee_totals[fh.id] for fh in used_fee_heads},
#                 'grand_total': fee_totals['total']
#             })

#         # Calculate Overall Totals
#         overall_totals = defaultdict(Decimal)
#         for item in data:
#             for fh_id, amount in item['fee_head_totals'].items():
#                 overall_totals[fh_id] += amount

#         context.update({
#             'data': data,
#             'all_fee_heads': {fh.id: fh.name for fh in used_fee_heads},
#             'overall_totals': overall_totals,
#             'overall_grand_total': grand_total,
#             'other_fee_total': other_fee_total
#         })

#     return context


# def master_head_wise_report_two(request):
#     context = generate_master_head_wise_context_two(request)
#     return render(request, 'crucial/report/master_head_wise_report_two.html', context)


# # Update views.py for Master Head Report PDF
# def download_pdf_two(request):
#     context = generate_master_head_wise_context_two(request)
#     html = render_to_string('crucial/report/master_head_wise_report_pdf_two.html', context)
    
#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="master_head_wise_report.pdf"'
    
#     # Use WeasyPrint like in Student Report with explicit CSS
#     HTML(string=html).write_pdf(response, stylesheets=[
#         CSS(string='''
#             th { 
#                 height: 150px !important;
#                 width: 35px !important;
#                 padding: 0 !important;
#                 position: relative !important;
#             }
#             .vertical-header {
#                 position: absolute !important;
#                 bottom: 15px !important;  /* Adjusted positioning */
#                 left: 50% !important;
#                 transform: translateX(-50%) rotate(-90deg) !important;
#                 transform-origin: center !important;
#                 width: 150px !important;  /* Match th height */
#                 font-size: 7px !important;
#                 text-align: left !important;
#             }
#             table {
#                 border-collapse: collapse !important;
#                 margin-top: 10px !important;
#             }
#         ''')
#     ])
#     return response


def generate_master_head_wise_context_two(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    institute = Institute.objects.order_by('-id').first() 

    context = {
        'data': [],
        'chunks_with_info': [],
        'overall_grand_total': Decimal('0.00'),
        'from_date': from_date,
        'to_date': to_date,
        'other_fee_total': Decimal('0.00'),
        'institute_logo': request.build_absolute_uri(institute.institute_logo.url) if institute and institute.institute_logo else None
        
    }

    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
            to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()
        except ValueError:
            return context        
        
        receives = Receive.objects.filter(
            date__gte=from_date_obj,
            date__lte=to_date_obj,
        ).exclude(voucher_no__startswith='FEE').select_related('student').order_by('student__student_field__name')

        # Apply additional filters
        filters = {
            'student__class_id__class_group_id__class_id': request.GET.get('class'),
            'student__class_id__section_id': request.GET.get('section'),
            'student__class_id__class_group_id__group_id': request.GET.get('group'),
            'student__class_id__shift_id': request.GET.get('shift'),
            'student__version': request.GET.get('version'),
        }
        filters = {k: v for k, v in filters.items() if v}
        receives = receives.filter(**filters)

        # Separate student and non-student receives
        student_receives = []
        other_fee_total = Decimal('0.00')
        for receive in receives:
            if receive.voucher_no.startswith('REC2') or not receive.student:
                other_fee_total += receive.amount
            else:
                student_receives.append(receive)

        student_ids = [r.student_id for r in student_receives]
        receive_dates = {r.created_at.date() for r in student_receives}

        # Fetch Fees entries
        fees_entries = Fees.objects.filter(
            student_id__in=student_ids,
            updated_at__date__in=receive_dates
        ).select_related('feetype_id__fees_type__fee_head').order_by('id')

        # Group fees entries
        fees_grouped = defaultdict(list)
        for fee in fees_entries:
            key = (fee.student_id_id, fee.updated_at.date(), fee.amount)
            if fee.feetype_id and fee.feetype_id.fees_type:
                fees_grouped[key].append({
                    'fee_head': fee.feetype_id.fees_type.fee_head,
                    'fee_id': fee.id
                })

        # Process receives
        used_fees = set()
        aggregated_data = defaultdict(lambda: defaultdict(Decimal))
        grand_total = Decimal('0.00')

        for receive in student_receives:
            student = receive.student
            version = student.version if student else "Unknown"
            shift = student.class_id.shift_id.name if student and student.class_id and student.class_id.shift_id else "Unknown"
            
            key = (student.id, receive.created_at.date(), receive.amount)
            fee_head = None

            possible_fees = fees_grouped.get(key, [])
            for fee_entry in possible_fees:
                if fee_entry['fee_id'] not in used_fees:
                    fee_head = fee_entry['fee_head']
                    used_fees.add(fee_entry['fee_id'])
                    break

            if fee_head:
                aggregated_data[(version, shift)][fee_head.id] += receive.amount
                aggregated_data[(version, shift)]['total'] += receive.amount
                grand_total += receive.amount

        # Add Other Fee to grand total
        grand_total += other_fee_total

        # Get used fee heads
        used_fee_heads = FeeHead.objects.filter(
            id__in=set(fh_id for _, fh_totals in aggregated_data.items() 
                      for fh_id in fh_totals.keys() if fh_id != 'total')
        ).order_by('name')

        # Format data
        data = []
        for (version, shift), fee_totals in aggregated_data.items():
            data.append({
                'version': version,
                'shift': shift,
                'fee_head_totals': {fh.id: fee_totals[fh.id] for fh in used_fee_heads},
                'grand_total': fee_totals['total']
            })

        # Calculate Overall Totals
        overall_totals = defaultdict(Decimal)
        for item in data:
            for fh_id, amount in item['fee_head_totals'].items():
                overall_totals[fh_id] += amount

        # Split into chunks (10 columns per page)
        max_columns_per_page = 11
        fee_head_chunks = [list(used_fee_heads)[i:i+max_columns_per_page] 
                          for i in range(0, len(used_fee_heads), max_columns_per_page)]

        # Prepare chunked data
        chunks_with_info = []
        for chunk in fee_head_chunks:
            chunk_fee_heads = {fh.id: fh.name for fh in chunk}
            chunk_totals = {fh.id: overall_totals[fh.id] for fh in chunk}
            
            chunks_with_info.append({
                'fee_heads': chunk_fee_heads,
                'totals': chunk_totals,
                'fee_head_objects': chunk
            })

        context.update({
            'data': data,
            'chunks_with_info': chunks_with_info,
            'all_fee_heads': {fh.id: fh.name for fh in used_fee_heads},  # Add back
            'overall_totals': overall_totals,  # Crucial for PDF
            'overall_grand_total': grand_total,
            'other_fee_total': other_fee_total
        })

    return context

def master_head_wise_report_two(request):
    context = generate_master_head_wise_context_two(request)
    return render(request, 'crucial/report/master_head_wise_report_two.html', context)

def download_pdf_two(request):
    context = generate_master_head_wise_context_two(request)
    html = render_to_string('crucial/report/master_head_wise_report_pdf_two.html', context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="master_head_wise_report.pdf"'
    
    HTML(string=html).write_pdf(response, stylesheets=[
        CSS(string='''
            th { 
                height: 150px !important;
                width: 25px !important;  /* Reduced width */
                padding: 0 !important;
                position: relative !important;
            }
            .vertical-header {
                position: absolute !important;
                bottom: 10px !important;  /* Adjusted position */
                left: 50% !important;
                transform: translateX(-50%) rotate(-90deg) !important;
                transform-origin: center !important;
                width: 100px !important;  /* Reduced width */
                font-size: 6px !important;  /* Smaller font */
                text-align: left !important;
            }
            table {
                border-collapse: collapse !important;
                margin-top: 10px !important;
                table-layout: fixed !important;  /* Ensure fixed layout */
            }
            .page-break {
                page-break-inside: avoid;
            }
            @media print {
                .page-break {
                    display: block;
                    page-break-before: always;
                }
            }
        ''')
    ])
    return response



from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side


def download_excel_two(request):
    context = generate_master_head_wise_context_two(request)
    
    # Create the Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Head Report"

    # Prepare fee heads sorted by name
    fee_heads = sorted(context['all_fee_heads'].items(), key=lambda x: x[1])
    
    # ========== NEW: Add Report Header and Date Range ========== #
    # Add title row
    ws.append(['Master Head Wise Report'])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(fee_heads)+3)

    # Add date range row
    date_range_str = f"Date Range: {context['from_date']} to {context['to_date']}"
    ws.append([date_range_str])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(fee_heads)+3)

    # Add empty row before headers
    ws.append([])

    # Create headers (now starting at row 4)
    headers = ['Version', 'Shift'] + [name for _, name in fee_heads] + ['Grand Total']
    ws.append(headers)

    # Add data rows
    for item in context['data']:
        row = [
            item['version'],
            item['shift'],
            *[item['fee_head_totals'].get(fh_id, Decimal('0.00')) for fh_id, _ in fee_heads],
            item['grand_total']
        ]
        ws.append(row)

    # Add Overall Totals row (excludes Other Fee)
    totals_row = [
        'Total', 
        '', 
        *[context['overall_totals'].get(fh_id, Decimal('0.00')) for fh_id, _ in fee_heads],
        context['overall_grand_total'] - context['other_fee_total']
    ]
    ws.append(totals_row)

    # ========== Apply Enhanced Styling ========== #
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
    title_font = Font(bold=True, size=14)
    date_font = Font(italic=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    currency_format = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # Style title row
    title_cell = ws['A1']
    title_cell.font = title_font
    title_cell.alignment = center_alignment

    # Style date range row
    date_cell = ws['A2']
    date_cell.font = date_font
    date_cell.alignment = center_alignment

    # Style headers (now at row 4)
    for cell in ws[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Set column widths
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            cell_value = str(cell.value) if cell.value else ""
            if isinstance(cell.value, Decimal):
                formatted = "{:,.2f}".format(cell.value)
                max_length = max(max_length, len(formatted))
            else:
                max_length = max(max_length, len(cell_value))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Apply number formatting
    for row in ws.iter_rows(min_row=5):  # Start formatting from data rows
        for cell in row:
            if cell.column >= 3:  # Format currency columns
                cell.number_format = currency_format

    # Style totals row
    totals_row_num = ws.max_row
    for cell in ws[totals_row_num]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")

    # Apply borders to data cells only
    for row in ws.iter_rows(min_row=4):  # Start from headers
        for cell in row:
            cell.border = thin_border

    # Freeze header row (now at row 4)
    ws.freeze_panes = 'A5'

    # Prepare response
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="master_head_wise_report.xlsx"'
    wb.save(response)
    
    return response 
  
  
    
from django.db.models import Q, Sum, Exists, OuterRef, Subquery, DecimalField
from django.db.models.functions import Coalesce
from django.shortcuts import render
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML, CSS
from decimal import Decimal

def generate_new_fees_context(request):
    # Get filter parameters
    selected_class = request.GET.get('class')
    selected_section = request.GET.get('section')
    selected_group = request.GET.get('group')
    selected_shift = request.GET.get('shift')
    selected_version = request.GET.get('version')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    context = {}
    
    if from_date and to_date:
        # Base student query
        students = StudentProfile.objects.select_related(
            'student_field',
            'class_id__class_group_id__class_id',
            'class_id__class_group_id__group_id',
            'class_id__section_id',
            'class_id__shift_id'
        )

        # Apply filters
        filters = Q()
        if selected_class: 
            filters &= Q(class_id__class_group_id__class_id=selected_class)
        if selected_section:
            filters &= Q(class_id__section_id=selected_section)
        if selected_group:
            filters &= Q(class_id__class_group_id__group_id=selected_group)
        if selected_shift:
            filters &= Q(class_id__shift_id=selected_shift)
        if selected_version:
            filters &= Q(version=selected_version)

        students = students.filter(filters)
        student_ids = students.values_list('id', flat=True)

        # Subquery for successful transactions
        success_tx = PaymentTransaction.objects.filter(
            student_profile=OuterRef('fee__student_profile_id'),  # Updated
            status='SUCCESS',
            tran_date__date=OuterRef('payment_date')
        )

        # Get valid partial payments with success status
        partial_payments = PartialPayment.objects.filter(
            payment_date__date__range=[from_date, to_date],
            fee__student_profile_id__in=student_ids  # Updated
        ).annotate(
            has_success=Exists(success_tx)
        ).filter(has_success=True).values(
            'fee__student_profile_id', 'fee__feetype_id'  # Updated
        ).annotate(partial_total=Sum('amount'))

        # Get valid full payments with success status
        full_payments = Fees.objects.filter(
            status='paid',
            created_at__date__range=[from_date, to_date],
            student_profile_id__in=student_ids  # Updated
        ).annotate(
            has_success=Exists(
                PaymentTransaction.objects.filter(
                    student_profile=OuterRef('student_profile_id'),  # Updated
                    status='SUCCESS',
                    tran_date__date=OuterRef('created_at__date')
                )
            )
        ).filter(has_success=True).values(
            'student_profile_id', 'feetype_id'  # Updated
        ).annotate(full_total=Sum('amount'))

        # Get all unique fee heads with payments
        fee_heads = Fees_name.objects.filter(
            Q(id__in=Subquery(full_payments.values('feetype_id')) |
            Q(id__in=Subquery(partial_payments.values('fee__feetype_id'))
        ).distinct().order_by('fees_title')))

        # Prepare data structure
        students_data = []
        column_totals = {fh.id: Decimal('0.00') for fh in fee_heads}
        grand_total = Decimal('0.00')

        for student in students:
            student_id = student.id
            payments_dict = defaultdict(Decimal)
            
            # Process partial payments
            for pp in partial_payments.filter(fee__student_profile_id=student_id):  # Updated
                feetype_id = pp['fee__feetype_id']
                payments_dict[feetype_id] += pp['partial_total']
            
            # Process full payments
            for fp in full_payments.filter(student_profile_id=student_id):  # Updated
                feetype_id = fp['feetype_id']
                payments_dict[feetype_id] += fp['full_total']
            
            student_total = sum(payments_dict.values())
            
            if student_total > 0:
                students_data.append({
                    'student': student,
                    'payments_dict': payments_dict,
                    'total': student_total
                })
                
                # Update totals
                for fh in fee_heads:
                    amount = payments_dict.get(fh.id, Decimal('0.00'))
                    column_totals[fh.id] += amount
                    grand_total += amount

        # Convert payments to ordered list
        for student_data in students_data:
            payments_dict = student_data.pop('payments_dict')
            student_data['payments'] = [payments_dict.get(fh.id, Decimal('0.00')) for fh in fee_heads]

        context = {
            'students_data': students_data,
            'fee_heads': fee_heads,
            'column_totals': [column_totals[fh.id] for fh in fee_heads],
            'grand_total': grand_total,
            'from_date': from_date,
            'to_date': to_date,
            'selected_class': selected_class,
            'selected_section': selected_section,
            'selected_group': selected_group,
            'selected_shift': selected_shift,
            'selected_version': selected_version,
        }

    return context


def new_fees_collection_report(request):
    context = generate_new_fees_context(request)
    return render(request, 'crucial/report/new_fees_report.html', context)

def new_fees_collection_report_pdf(request):
    context = generate_new_fees_context(request)
    html_string = render_to_string('crucial/report/new_fees_report_pdf.html', context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="new_fees_collection_report.pdf"'
    
    HTML(string=html_string).write_pdf(response, stylesheets=[
        CSS(string='''
            @page { size: A4 landscape; margin: 10mm; }
            th { writing-mode: vertical-lr; text-orientation: upright; padding: 4px 2px; }
            table { border-collapse: collapse; width: 100%; font-size: 10px; }
            th, td { border: 1px solid #000; padding: 4px; text-align: center; }
        ''')
    ])
    
    return response
  
  
    
from collections import defaultdict
from decimal import Decimal
from datetime import datetime
from django.utils import timezone
from django.shortcuts import render
from .models import Receive, Fees, Fees_name
import pandas as pd
from django.http import HttpResponse
from io import BytesIO




# def generate_student_fees_context(request):
#     context = {
#         'students_data': [],
#         'fee_heads': [],
#         'column_totals': defaultdict(Decimal),
#         'grand_total': Decimal('0.00'),
#         'from_date': None,
#         'to_date': None,
#         'chunks_with_info': [],
#         'other_fee_total': Decimal('0.00'),
#         'other_fee_total_formatted': '0',
#         'grand_total_formatted': '0'
#     }

#     from_date = request.GET.get('from_date')
#     to_date = request.GET.get('to_date')

#     if from_date and to_date:
#         try:
#             from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
#             to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()
#         except ValueError:
#             return context

#         receives = Receive.objects.filter(
#             date__gte=from_date_obj,
#             date__lte=to_date_obj,
#         ).exclude(voucher_no__startswith='FEE').select_related('student').order_by('student__student_field__name')

#         filters = {
#             'student__class_id__class_group_id__class_id': request.GET.get('class'),
#             'student__class_id__section_id': request.GET.get('section'),
#             'student__class_id__class_group_id__group_id': request.GET.get('group'),
#             'student__class_id__shift_id': request.GET.get('shift'),
#             'student__version': request.GET.get('version'),
#         }
#         filters = {k: v for k, v in filters.items() if v}
#         receives = receives.filter(**filters)

#         other_fee_total = Decimal('0.00')
#         student_receives = []

#         for receive in receives:
#             if receive.voucher_no.startswith('REC2') or not receive.student:
#                 other_fee_total += receive.amount
#             else:
#                 student_receives.append(receive)

#         student_ids = [receive.student_id for receive in student_receives if receive.student_id]
#         receive_dates = {receive.created_at.date() for receive in student_receives}

#         fees_entries = Fees.objects.filter(
#             student_id__in=student_ids,
#             updated_at__date__in=receive_dates
#         ).select_related('feetype_id__fees_type__fee_head').order_by('id')

#         fees_grouped = defaultdict(list)
#         for fee in fees_entries:
#             key = (fee.student_id_id, fee.updated_at.date(), fee.amount)
#             if fee.feetype_id:
#                 fees_grouped[key].append({
#                     'feetype': fee.feetype_id,
#                     'fee_id': fee.id
#                 })

#         students_data = defaultdict(lambda: {
#             'student': None,
#             'payments': defaultdict(Decimal),
#             'total': Decimal('0.00'),
#             'class': '',
#             'group': '',
#             'section': '',
#             'shift': '',
#             'version': ''
#         })

#         column_totals = defaultdict(Decimal)
#         grand_total = Decimal('0.00')
#         used_fees = set()

#         for receive in student_receives:
#             student = receive.student
#             key = (student.id, receive.created_at.date(), receive.amount)
#             fee_type = None

#             possible_fees = fees_grouped.get(key, [])
#             for fee_entry in possible_fees:
#                 if fee_entry['fee_id'] not in used_fees:
#                     fee_type = fee_entry['feetype']
#                     used_fees.add(fee_entry['fee_id'])
#                     break

#             if fee_type:
#                 fee_head = fee_type.fees_type.fee_head
#                 students_data[student.id]['student'] = student
#                 students_data[student.id]['payments'][fee_head.id] += receive.amount
#                 students_data[student.id]['total'] += receive.amount

#                 class_config = student.class_id
#                 if class_config:
#                     class_group = class_config.class_group_id
#                     students_data[student.id]['class'] = class_group.class_id.name if class_group else ''
#                     group_name = class_group.group_id.name if class_group and class_group.group_id else ''
#                     students_data[student.id]['group'] = group_name.split()[0] if group_name and len(group_name.split()) > 1 else group_name
#                     students_data[student.id]['section'] = class_config.section_id.name if class_config.section_id else ''
#                     students_data[student.id]['shift'] = class_config.shift_id.name if class_config.shift_id else ''
#                 students_data[student.id]['version'] = student.version

#                 column_totals[fee_head.id] += receive.amount
#                 grand_total += receive.amount

#         grand_total += other_fee_total
#         used_fee_heads = FeeHead.objects.filter(id__in=column_totals.keys()).order_by('name')

#         students_list = []
#         for student_id, data in students_data.items():
#             student = data['student']
#             payments = [data['payments'].get(fh.id, Decimal('0.00')) for fh in used_fee_heads]
            
#             # Process student name
#             full_name = student.student_field.name if (student and student.student_field) else "Undefined Student"
#             name_parts = full_name.split()
#             last_name = name_parts[-1] if len(name_parts) > 1 else full_name

#             # Process roll number
#             roll_no = str(student.roll_no) if (student and student.roll_no) else ""
#             roll_last_four = roll_no[-4:] if len(roll_no) >= 4 else roll_no

#             # Format amounts
#             payments_formatted = [
#                 f"{x:.0f}" if x == x.to_integral() else f"{x:.2f}"
#                 for x in payments
#             ]
#             total_formatted = f"{data['total']:.0f}" if data['total'] == data['total'].to_integral() else f"{data['total']:.2f}"

#             students_list.append({
#                 'student': student,
#                 'payments': payments,
#                 'payments_formatted': payments_formatted,
#                 'total': data['total'],
#                 'total_formatted': total_formatted,
#                 'class': data['class'],
#                 'group': data['group'] or '-',
#                 'section': data['section'],
#                 'shift': data['shift'],
#                 'version': data['version'],
#                 'last_name': last_name,
#                 'roll_last_four': roll_last_four
#             })

#         # Process totals formatting
#         other_fee_total_formatted = f"{other_fee_total:.0f}" if other_fee_total == other_fee_total.to_integral() else f"{other_fee_total:.2f}"
#         grand_total_formatted = f"{grand_total:.0f}" if grand_total == grand_total.to_integral() else f"{grand_total:.2f}"

#         # Chunking logic
#         max_headers_per_page = 20
#         fee_head_chunks = [used_fee_heads[i:i+max_headers_per_page] 
#                          for i in range(0, len(used_fee_heads), max_headers_per_page)]

#         chunks_with_info = []
#         start = 0
#         for chunk in fee_head_chunks:
#             end = start + len(chunk)
#             chunks_with_info.append({
#                 'chunk': chunk,
#                 'start': start,
#                 'end': end,
#                 'totals': [column_totals[fh.id] for fh in chunk],
#             })
#             start = end

#         # Attach formatted payment chunks
#         for chunk_index, chunk in enumerate(chunks_with_info):
#             start, end = chunk['start'], chunk['end']
#             for student in students_list:
#                 student[f'payments_formatted_chunk_{chunk_index}'] = student['payments_formatted'][start:end]

#         context.update({
#             'students_data': students_list,
#             'fee_heads': used_fee_heads,
#             'column_totals': column_totals,
#             'grand_total': grand_total,
#             'from_date': from_date,
#             'to_date': to_date,
#             'chunks_with_info': chunks_with_info,
#             'other_fee_total': other_fee_total,
#             'other_fee_total_formatted': other_fee_total_formatted,
#             'grand_total_formatted': grand_total_formatted
#         })

#     return context


# ABBREVIATIONS = {
#     # Groups
#     'Science': 'Sc',
#     'Humanities': 'Hum',
#     'Business Studies': 'Bst',
#     # Shifts
#     'Morning': 'M',
#     'Day': 'D',
#     # Versions
#     'English': 'Eng',
#     'Bangla': 'Ban'
# }

# def generate_student_fees_context(request):
#     context = {
#         'students_data': [],
#         'chunks_with_info': [],
#         'grand_total': Decimal('0.00'),
#         'from_date': None,
#         'to_date': None,
#         'other_fee_total': Decimal('0.00'),
#         'classes': StudentClass.objects.all(),
#         'sections': StudentSection.objects.all(),
#         'groups': StuGroup.objects.all(),
#         'shifts': StudentShift.objects.all(),
#         'selected_class': 'All',
#         'selected_section': 'All',
#         'selected_group': 'All',
#         'selected_shift': 'All',
#         'selected_version': 'All',
#     }
    
#     # Get selected filters from request
#     selected_class_id = request.GET.get('class')
#     selected_section_id = request.GET.get('section')
#     selected_group_id = request.GET.get('group')
#     selected_shift_id = request.GET.get('shift')
#     selected_version = request.GET.get('version')

#     # Get display names for selected filters
#     if selected_class_id:
#         try:
#             context['selected_class'] = StudentClass.objects.get(id=selected_class_id).name
#         except StudentClass.DoesNotExist:
#             pass

#     if selected_section_id:
#         try:
#             context['selected_section'] = StudentSection.objects.get(id=selected_section_id).name
#         except StudentSection.DoesNotExist:
#             pass

#     if selected_group_id:
#         try:
#             context['selected_group'] = StuGroup.objects.get(id=selected_group_id).name
#         except StuGroup.DoesNotExist:
#             pass

#     if selected_shift_id:
#         try:
#             context['selected_shift'] = StudentShift.objects.get(id=selected_shift_id).name
#         except StudentShift.DoesNotExist:
#             pass

#     if selected_version:
#         context['selected_version'] = selected_version

#     from_date = request.GET.get('from_date')
#     to_date = request.GET.get('to_date')

#     if from_date and to_date:
#         try:
#             from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
#             to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()
#         except ValueError:
#             return context

        
#         receives = Receive.objects.filter(
#             date__gte=from_date_obj,
#             date__lte=to_date_obj,
#         ).exclude(voucher_no__startswith='FEE').select_related('student').order_by('student__student_field__name')

#         # Apply additional filters from the request
#         filters = {
#             'student__class_id__class_group_id__class_id': request.GET.get('class'),
#             'student__class_id__section_id': request.GET.get('section'),
#             'student__class_id__class_group_id__group_id': request.GET.get('group'),
#             'student__class_id__shift_id': request.GET.get('shift'),
#             'student__version': request.GET.get('version'),
#         }
#         filters = {k: v for k, v in filters.items() if v}
#         receives = receives.filter(**filters)

#         # Calculate Other Fee total for entries without students and REC2 vouchers
#         other_fee_total = Decimal('0.00')
#         student_receives = []

#         for receive in receives:
#             if receive.voucher_no.startswith('REC2') or not receive.student:
#                 other_fee_total += receive.amount
#             else:
#                 student_receives.append(receive)

#         student_ids = [receive.student_id for receive in student_receives if receive.student_id]
#         receive_dates = {receive.created_at.date() for receive in student_receives}

#         fees_entries = Fees.objects.filter(
#             student_id__in=student_ids,
#             updated_at__date__in=receive_dates
#         ).select_related('feetype_id__fees_type__fee_head').order_by('id')

#         fees_grouped = defaultdict(list)
#         for fee in fees_entries:
#             key = (fee.student_id_id, fee.updated_at.date(), fee.amount)
#             if fee.feetype_id:
#                 fees_grouped[key].append({
#                     'feetype': fee.feetype_id,
#                     'fee_id': fee.id
#                 })

#         students_data = defaultdict(lambda: {
#             'student': None,
#             'payments': defaultdict(Decimal),
#             'total': Decimal('0.00'),
#             'class': '',
#             'group': '',
#             'section': '',
#             'shift': '',
#             'version': ''
#         })

#         column_totals = defaultdict(Decimal)
#         grand_total = Decimal('0.00')
#         used_fees = set()

#         for receive in student_receives:
#             student = receive.student
#             key = (student.id, receive.created_at.date(), receive.amount)
#             fee_type = None

#             possible_fees = fees_grouped.get(key, [])
#             for fee_entry in possible_fees:
#                 if fee_entry['fee_id'] not in used_fees:
#                     fee_type = fee_entry['feetype']
#                     used_fees.add(fee_entry['fee_id'])
#                     break

#             if fee_type:
#                 fee_head = fee_type.fees_type.fee_head
#                 students_data[student.id]['student'] = student
#                 students_data[student.id]['payments'][fee_head.id] += receive.amount
#                 students_data[student.id]['total'] += receive.amount

#                 # Populate class details
#                 class_config = student.class_id
#                 if class_config:
#                     class_group = class_config.class_group_id
#                     students_data[student.id]['class'] = class_group.class_id.name if class_group else ''
#                     students_data[student.id]['group'] = class_group.group_id.name if class_group and class_group.group_id else ''
#                     students_data[student.id]['section'] = class_config.section_id.name if class_config.section_id else ''
#                     students_data[student.id]['shift'] = class_config.shift_id.name if class_config.shift_id else ''
#                 students_data[student.id]['version'] = student.version

#                 # Update totals
#                 column_totals[fee_head.id] += receive.amount
#                 grand_total += receive.amount

#         grand_total += other_fee_total

#         used_fee_heads = FeeHead.objects.filter(id__in=column_totals.keys()).order_by('name')

#         students_list = []
#         for student_id, data in students_data.items():
#             student = data['student']
#             payments = [data['payments'].get(fh.id, Decimal('0.00')) for fh in used_fee_heads]
#             students_list.append({
#                 'student': student,
#                 'payments': payments,
#                 'total': data['total'],
#                 'class': data['class'],
#                 'group': data['group'],
#                 'section': data['section'],
#                 'shift': data['shift'],
#                 'version': data['version']
#             })

#         max_columns_per_page = 1
#         fee_head_chunks = [list(used_fee_heads)[i:i+max_columns_per_page] 
#                           for i in range(0, len(used_fee_heads), max_columns_per_page)]

#         # Prepare chunked data
#         # chunks_with_info = []
#         # for chunk in fee_head_chunks:
#         #     chunk_totals = {fh.id: column_totals[fh.id] for fh in chunk}
            
#         #     chunks_with_info.append({
#         #         'fee_head_objects': chunk,
#         #         'totals': chunk_totals,
#         #         'chunk_fee_heads': {fh.id: fh.name for fh in chunk}
#         #     })
        
#         for student in students_list:
#             student['payment_chunks'] = []
#             for chunk in fee_head_chunks:
#                 chunk_payments = [student['payments'][i] for i, fh in enumerate(used_fee_heads) if fh in chunk]
#                 student['payment_chunks'].append(chunk_payments)
        
#         # Prepare chunked data
#         chunks_with_info = []
#         for idx, chunk in enumerate(fee_head_chunks):
#             chunk_totals = {fh.id: column_totals[fh.id] for fh in chunk}
#             # Collect student payments for this chunk
#             student_payments_for_chunk = []
#             for student in students_list:
#                 payment_chunk = student['payment_chunks'][idx]
#                 student_payments_for_chunk.append({
#                     'student': student['student'],
#                     'payments': payment_chunk,
#                     'total': student['total'],
#                     'class': student['class'],
#                     'group': student['group'],
#                     'section': student['section'],
#                     'shift': student['shift'],
#                     'version': student['version'],
#                 })
#             chunks_with_info.append({
#                 'fee_head_objects': chunk,
#                 'totals': chunk_totals,
#                 'chunk_fee_heads': {fh.id: fh.name for fh in chunk},
#                 'student_payments': student_payments_for_chunk,
#             })
        

#         # Attach sliced payments to students
#         # for student in students_list:
#         #     student['payment_chunks'] = []
#         #     for chunk in fee_head_chunks:
#         #         chunk_payments = [student['payments'][i] for i, fh in enumerate(used_fee_heads) if fh in chunk]
#         #         student['payment_chunks'].append(chunk_payments)


#         context.update({
#             'students_data': students_list,
#             'chunks_with_info': chunks_with_info,
#             'all_fee_heads': {fh.id: fh.name for fh in used_fee_heads},
#             'used_fee_heads': used_fee_heads, 
#             'column_totals': column_totals,
#             'grand_total': grand_total,
#             'other_fee_total': other_fee_total,
#             'from_date': from_date,
#             'to_date': to_date
#         })

#     return context

# def student_fees_report_pdf(request):
#     context = generate_student_fees_context(request)
#     html = render_to_string('crucial/report/student_fees_report_pdf.html', context)
    
#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="student_fees_report.pdf"'
    
#     HTML(string=html).write_pdf(response, stylesheets=[
#         CSS(string='''
#             th { 
#                 height: 150px !important;
#                 width: 7px !important;  /* Reduced by 80% from 35px */
#                 padding: 0 !important;
#                 position: relative !important;
#             }
#             /* Specific narrow columns */
#             th:nth-child(1),  /* SL NO */
#             th:nth-child(5),  /* Class */
#             th:nth-child(6),  /* Group */
#             th:nth-child(7),  /* Section */
#             th:nth-child(8),  /* Shift */
#             th:nth-child(9) { /* Version */
#                 width: 7px !important;
#             }
#             .vertical-header {
#                 position: absolute !important;
#                 bottom: 10px !important;
#                 left: 50% !important;
#                 transform: translateX(-50%) rotate(-90deg) !important;
#                 transform-origin: center !important;
#                 width: 50px !important;
#                 font-size: 6px !important;
#                 text-align: left !important;
#             }
#             .page-break {
#                 page-break-inside: avoid;
#             }
#             @media print {
#                 .page-break {
#                     display: block;
#                     page-break-before: always;
#                 }
#             }
#         ''')
#     ])
#     return response


from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from celery.result import AsyncResult
from .tasks import generate_pdf_report_task, generate_excel_report_task
from .utils import generate_student_fees_context
from django.core.paginator import Paginator


ABBREVIATIONS = {
    # Groups
    'Science': 'Sc',
    'Humanities': 'Hum',
    'Business Studies': 'Bst',
    # Shifts
    'Morning': 'M',
    'Day': 'D',
    # Versions
    'English': 'Eng',
    'Bangla': 'Ban'
}

def generate_student_fees_context(request):
    institute = Institute.objects.order_by('-id').first()
    context = {
        'students_data': [],
        'chunks_with_info': [],
        'grand_total': Decimal('0.00'),
        'from_date': request.GET.get('from_date', ''),  
        'to_date': request.GET.get('to_date', ''),
        'other_fee_total': Decimal('0.00'),
        'classes': StudentClass.objects.all(),
        'sections': StudentSection.objects.all(),
        'groups': StuGroup.objects.all(),
        'shifts': StudentShift.objects.all(),
        'selected_class': 'All',
        'selected_section': 'All',
        'selected_group': 'All',
        'selected_shift': 'All',
        'selected_version': 'All',
        'institute_logo': request.build_absolute_uri(institute.institute_logo.url) if institute and institute.institute_logo else None
    }
    
    # Get selected filters from request
    selected_class_id = request.GET.get('class')
    selected_section_id = request.GET.get('section')
    selected_group_id = request.GET.get('group')
    selected_shift_id = request.GET.get('shift')
    selected_version = request.GET.get('version')

    # Get display names for selected filters
    if selected_class_id:
        try:
            context['selected_class'] = StudentClass.objects.get(id=selected_class_id).name
        except StudentClass.DoesNotExist:
            pass

    if selected_section_id:
        try:
            context['selected_section'] = StudentSection.objects.get(id=selected_section_id).name
        except StudentSection.DoesNotExist:
            pass

    if selected_group_id:
        try:
            context['selected_group'] = ABBREVIATIONS.get(
                StuGroup.objects.get(id=selected_group_id).name.strip(),
                StuGroup.objects.get(id=selected_group_id).name[:3]
            )
        except StuGroup.DoesNotExist:
            pass

    if selected_shift_id:
        try:
            context['selected_shift'] = ABBREVIATIONS.get(
                StudentShift.objects.get(id=selected_shift_id).name.strip(),
                StudentShift.objects.get(id=selected_shift_id).name[:1]
            )
        except StudentShift.DoesNotExist:
            pass

    if selected_version:
        context['selected_version'] = ABBREVIATIONS.get(
            selected_version.strip(),
            selected_version[:3]
        )

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
            to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()
        except ValueError:
            return context

        receives = Receive.objects.filter(
            date__gte=from_date_obj,
            date__lte=to_date_obj,
        ).exclude(voucher_no__startswith='FEE').select_related('student').order_by('student__student_field__name')

        # Apply additional filters from the request
        filters = {
            'student__class_id__class_group_id__class_id': request.GET.get('class'),
            'student__class_id__section_id': request.GET.get('section'),
            'student__class_id__class_group_id__group_id': request.GET.get('group'),
            'student__class_id__shift_id': request.GET.get('shift'),
            'student__version': request.GET.get('version'),
        }
        filters = {k: v for k, v in filters.items() if v}
        receives = receives.filter(**filters)

        # Calculate Other Fee total for entries without students and REC2 vouchers
        other_fee_total = Decimal('0.00')
        student_receives = []

        for receive in receives:
            if receive.voucher_no.startswith('REC2') or not receive.student:
                other_fee_total += receive.amount
            else:
                student_receives.append(receive)

        student_ids = [receive.student_id for receive in student_receives if receive.student_id]
        receive_dates = {receive.created_at.date() for receive in student_receives}

        fees_entries = Fees.objects.filter(
            student_id__in=student_ids,
            updated_at__date__in=receive_dates
        ).select_related('feetype_id__fees_type__fee_head').order_by('id')

        fees_grouped = defaultdict(list)
        for fee in fees_entries:
            key = (fee.student_id_id, fee.updated_at.date(), fee.amount)
            if fee.feetype_id:
                fees_grouped[key].append({
                    'feetype': fee.feetype_id,
                    'fee_id': fee.id
                })

        students_data = defaultdict(lambda: {
            'student': None,
            'payments': defaultdict(Decimal),
            'total': Decimal('0.00'),
            'class': '',
            'group': '',
            'section': '',
            'shift': '',
            'version': ''
        })

        column_totals = defaultdict(Decimal)
        grand_total = Decimal('0.00')
        used_fees = set()

        for receive in student_receives:
            student = receive.student
            key = (student.id, receive.created_at.date(), receive.amount)
            fee_type = None

            possible_fees = fees_grouped.get(key, [])
            for fee_entry in possible_fees:
                if fee_entry['fee_id'] not in used_fees:
                    fee_type = fee_entry['feetype']
                    used_fees.add(fee_entry['fee_id'])
                    break

            if fee_type:
                fee_head = fee_type.fees_type.fee_head
                students_data[student.id]['student'] = student
                students_data[student.id]['payments'][fee_head.id] += receive.amount
                students_data[student.id]['total'] += receive.amount

                # Populate class details with abbreviations
                class_config = student.class_id
                if class_config:
                    class_group = class_config.class_group_id
                    students_data[student.id]['class'] = class_group.class_id.name if class_group else ''
                    
                    # Group abbreviation
                    group_name = class_group.group_id.name if class_group and class_group.group_id else ''
                    students_data[student.id]['group'] = ABBREVIATIONS.get(
                        group_name.strip(), 
                        group_name[:3] if group_name else ''
                    )
                    
                    # Section
                    students_data[student.id]['section'] = class_config.section_id.name if class_config.section_id else ''
                    
                    # Shift abbreviation
                    shift_name = class_config.shift_id.name if class_config.shift_id else ''
                    students_data[student.id]['shift'] = ABBREVIATIONS.get(
                        shift_name.strip(),
                        shift_name[:1] if shift_name else ''
                    )
                
                # Version abbreviation
                version = student.version or ''
                students_data[student.id]['version'] = ABBREVIATIONS.get(
                    version.strip(),
                    version[:3]
                )

                # Update totals
                column_totals[fee_head.id] += receive.amount
                grand_total += receive.amount

        grand_total += other_fee_total

        used_fee_heads = FeeHead.objects.filter(id__in=column_totals.keys()).order_by('name')

        students_list = []
        for student_id, data in students_data.items():
            student = data['student']
            payments = [data['payments'].get(fh.id, Decimal('0.00')) for fh in used_fee_heads]
            students_list.append({
                'student': student,
                'payments': payments,
                'total': data['total'],
                'class': data['class'],
                'group': data['group'],
                'section': data['section'],
                'shift': data['shift'],
                'version': data['version']
            })

        max_columns_per_page = 10
        fee_head_chunks = [list(used_fee_heads)[i:i+max_columns_per_page] 
                          for i in range(0, len(used_fee_heads), max_columns_per_page)]

        for student in students_list:
            student['payment_chunks'] = []
            for chunk in fee_head_chunks:
                chunk_payments = [student['payments'][i] for i, fh in enumerate(used_fee_heads) if fh in chunk]
                student['payment_chunks'].append(chunk_payments)
        
        chunks_with_info = []
        for idx, chunk in enumerate(fee_head_chunks):
            chunk_totals = {fh.id: column_totals[fh.id] for fh in chunk}
            student_payments_for_chunk = []
            for student in students_list:
                payment_chunk = student['payment_chunks'][idx]
                student_payments_for_chunk.append({
                    'student': student['student'],
                    'payments': payment_chunk,
                    'total': student['total'],
                    'class': student['class'],
                    'group': student['group'],
                    'section': student['section'],
                    'shift': student['shift'],
                    'version': student['version'],
                })
            chunks_with_info.append({
                'fee_head_objects': chunk,
                'totals': chunk_totals,
                'chunk_fee_heads': {fh.id: fh.name for fh in chunk},
                'student_payments': student_payments_for_chunk,
            })
            
        page = request.GET.get('page', 1)
        paginator = Paginator(students_list, 50)  # 50 items per page
        
        try:
            students_list = paginator.page(page)
        except PageNotAnInteger:
            students_list = paginator.page(1)
        except EmptyPage:
            students_list = paginator.page(paginator.num_pages)
        

        context.update({
            'students_data': students_list,
            'chunks_with_info': chunks_with_info,
            'all_fee_heads': {fh.id: fh.name for fh in used_fee_heads},
            'used_fee_heads': used_fee_heads, 
            'column_totals': column_totals,
            'grand_total': grand_total,
            'other_fee_total': other_fee_total,
            'from_date': from_date,
            'to_date': to_date
        })

    return context




# Update the student_fees_report view
def student_fees_report(request):
    context = generate_student_fees_context(request)
    context.update({
        'classes': StudentClass.objects.all(),
        'sections': StudentSection.objects.all(),
        'groups': StuGroup.objects.all(),
        'shifts': StudentShift.objects.all(),
    })
    return render(request, 'crucial/report/student_fees_report.html', context)

from django.http import JsonResponse
from celery.result import AsyncResult
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def check_task_status(request):
    task_id = request.GET.get("task_id")
    task_result = AsyncResult(task_id)

    response_data = {
        'task_id': task_id,
        'status': task_result.status,
    }

    if task_result.successful():
        response_data['result'] = task_result.result
    elif task_result.failed():
        # Don't serialize the actual exception
        response_data['error'] = str(task_result.result)  # Just string, not the object

    return JsonResponse(response_data)

#using this
# def student_fees_report_pdf(request):
#     context = generate_student_fees_context(request)
#     html = render_to_string('crucial/report/student_fees_report_pdf.html', context)
    
#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="student_fees_report.pdf"'
    
#     HTML(string=html).write_pdf(response, stylesheets=[
#         CSS(string='''
#             /* Table structure enforcement */
#             .styled-table {
#                 table-layout: fixed !important;
#                 border-collapse: collapse !important;
#                 width: 100% !important;
#             }

#             /* Default column sizing */
#             .styled-table th,
#             .styled-table td {
#                 width: auto !important;
#                 min-width: 40px !important;
#                 padding: 2px !important;
#                 overflow: hidden !important;
#                 text-overflow: ellipsis !important;
#             }

#             /* Narrow columns: 1(SL), 4(Roll), 5(Class), 6(Group), 7(Section), 8(Shift) */
#             .styled-table th:nth-child(1),
#             .styled-table th:nth-child(4),
#             .styled-table th:nth-child(5),
#             .styled-table th:nth-child(6),
#             .styled-table th:nth-child(7),
#             .styled-table th:nth-child(8),
#             .styled-table td:nth-child(1),
#             .styled-table td:nth-child(4),
#             .styled-table td:nth-child(5),
#             .styled-table td:nth-child(6),
#             .styled-table td:nth-child(7),
#             .styled-table td:nth-child(8) {
#                 width: 20px !important;  /* Increased from 7px for better visibility */
#                 min-width: 20px !important;
#                 max-width: 20px !important;
#             }

#             /* Wide columns: Student(2-3), Version(9) */
#             .styled-table th:nth-child(2),
#             .styled-table th:nth-child(3),
#             .styled-table th:nth-child(9),
#             .styled-table td:nth-child(2),
#             .styled-table td:nth-child(3),
#             .styled-table td:nth-child(9) {
#                 min-width: 60px !important;
#                 max-width: 120px !important;
#             }

#             /* Vertical header adjustments */
#             .vertical-header {
#                 position: absolute !important;
#                 bottom: 8px !important;
#                 left: 50% !important;
#                 transform: translateX(-50%) rotate(-90deg) !important;
#                 transform-origin: center !important;
#                 width: 80px !important;
#                 font-size: 8px !important;
#                 line-height: 1.2 !important;
#                 text-align: center !important;
#                 padding: 0 !important;
#                 margin: 0 !important;
#             }

#             /* Student name column specific */
#             td:nth-child(2) {
#                 text-align: left !important;
#                 padding-left: 4px !important;
#             }

#             .styled-table th {
#                 height: 155px !important;
#                 padding: 0 !important;
#                 position: relative !important;
#             }

#             /* Page break handling */
#             .page-break {
#                 page-break-inside: avoid !important;
#             }

#             @media print {
#                 .page-break {
#                     display: block !important;
#                     page-break-before: always !important;
#                 }
#             }
#         ''')
#     ])
#     return response


from django_tenants.utils import get_tenant

def student_fees_report_pdf(request):
    if request.method == 'POST':
        tenant = get_tenant(request)

        task_params = {
            'from_date': request.GET.get('from_date'),
            'to_date': request.GET.get('to_date'),
            'class': request.GET.get('class'),
            'section': request.GET.get('section'),
            'group': request.GET.get('group'),
            'shift': request.GET.get('shift'),
            'version': request.GET.get('version'),
            'template': 'crucial/report/student_fees_report_pdf.html',
            'filename': f"student_fees_report_{int(time.time())}.pdf",
            'schema_name': tenant.schema_name
        }

        if not task_params['from_date'] or not task_params['to_date']:
            return JsonResponse({'error': 'Date parameters are required'}, status=400)

        task = generate_pdf_report_task.delay(task_params)
        return JsonResponse({'task_id': task.id})

    return HttpResponseBadRequest()

#using this
# import xlsxwriter
# def student_fees_report_excel(request):
#     context = generate_student_fees_context(request)
#     students_data = context.get('students_data', [])
#     used_fee_heads = context.get('used_fee_heads', [])  # Changed from fee_heads to used_fee_heads
#     column_totals = context.get('column_totals', defaultdict(Decimal))
#     grand_total = context.get('grand_total', Decimal('0.00'))
#     other_fee_total = context.get('other_fee_total', Decimal('0.00'))

#     # Prepare column headers with used fee heads
#     columns = [
#         'SL NO', 'Student Name', 
#         'Class', 'Group', 'Section', 
#         'Shift', 'Version', 'Roll No'
#     ] + [fh.name for fh in used_fee_heads] + ['Other Fee', 'Total'] 

#     # Prepare data rows
#     data_rows = []
#     serial_no = 1
#     for student in students_data:
#         row = [
#             serial_no,
#             student['student'].student_field.name,
#             student['class'],
#             student['group'],
#             student['section'],
#             student['shift'],
#             student['version'],
#             student['student'].roll_no or '',
#         ] + student['payments'] + [
#             Decimal('0.00'),  
#             student['total']
#         ]
#         data_rows.append(row)
#         serial_no += 1

#     # Add grand total row (now using used_fee_heads)
#     total_row = [
#         'Grand Total', '', '', '', '', '', '', ''
#     ] + [column_totals.get(fh.id, Decimal('0.00')) for fh in used_fee_heads] + [
#         other_fee_total, 
#         grand_total
#     ]
#     data_rows.append(total_row)


#     # Convert to Pandas DataFrame
#     df = pd.DataFrame(data_rows, columns=columns)

#     # Create Excel file with enhanced borders
#     output = BytesIO()
#     with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
#         df.to_excel(writer, sheet_name='Student Fees Report', index=False, startrow=3)
#         workbook = writer.book
#         worksheet = writer.sheets['Student Fees Report']

#         # Define border styles
#         thin_border = {
#             'border': 1,
#             'border_color': '#000000'
#         }
#         thick_top_border = {
#             'top': 2,
#             'border_color': '#000000'
#         }

#         # Create formats
#         header_format = workbook.add_format({
#             **thin_border,
#             'bold': True,
#             'text_wrap': True,
#             'valign': 'top',
#             'fg_color': '#4F81BD',
#             'font_color': 'white'
#         })

#         title_format = workbook.add_format({
#             'bold': True,
#             'font_size': 16,
#             'align': 'center',
#             'valign': 'vcenter'
#         })

#         currency_format = workbook.add_format({
#             **thin_border,
#             'num_format': '#,##0.00'
#         })

#         total_format = workbook.add_format({
#             **thin_border,
#             **thick_top_border,
#             'num_format': '#,##0.00',
#             'bold': True,
#             'fg_color': '#D9E1F2'
#         })

#         center_format = workbook.add_format({
#             **thin_border,
#             'align': 'center'
#         })

#         # Add titles
#         worksheet.merge_range('A1:Z1', 'STUDENT FEES COLLECTION REPORT', title_format)
#         worksheet.merge_range('A2:Z2', 
#                             f"Date Range: {context.get('from_date')} to {context.get('to_date')}", 
#                             workbook.add_format({'align': 'center', 'italic': True}))

#         # Format headers
#         for col_num, value in enumerate(columns):
#             worksheet.write(3, col_num, value, header_format)

#         # Format data cells
#         for row_num in range(4, len(data_rows) + 4):
#             for col_num in range(len(columns)):
#                 cell_value = data_rows[row_num-4][col_num]
                
#                 # Apply different formats based on cell type
#                 if row_num == len(data_rows) + 3:  # Grand total row
#                     if col_num >= 8:
#                         worksheet.write(row_num, col_num, cell_value, total_format)
#                     else:
#                         worksheet.write(row_num, col_num, cell_value, header_format)
#                 else:
#                     if col_num in [0, 2, 3, 4, 5, 6, 7]:
#                         worksheet.write(row_num, col_num, cell_value, center_format)
#                     elif col_num >= 8:
#                         worksheet.write(row_num, col_num, cell_value, currency_format)
#                     else:
#                         worksheet.write(row_num, col_num, cell_value, workbook.add_format(thin_border))

#         # Set column widths
#         for i, col in enumerate(columns):
#             max_len = max(
#                 df[col].astype(str).apply(len).max(),
#                 len(str(col)) + 2
#             )
#             worksheet.set_column(i, i, max_len + 2)

#         # Freeze panes and set zoom
#         worksheet.freeze_panes(4, 2)
#         worksheet.set_zoom(90)

#         # Add autofilter and print settings
#         worksheet.autofilter(3, 0, 3, len(columns)-1)
#         worksheet.set_landscape()
#         worksheet.set_margins(left=0.5, right=0.5, top=0.75, bottom=0.75)
#         worksheet.repeat_rows(0, 3)
#         worksheet.print_area(0, 0, len(data_rows)+3, len(columns)-1)

#     response = HttpResponse(
#         output.getvalue(),
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="student_fees_report.xlsx"'
#     return response

from django.core.paginator import Paginator, Page, PageNotAnInteger, EmptyPage
def student_fees_report_excel(request):
    if request.method == 'POST':
        tenant = get_tenant(request)
        params = request.GET.dict()
        params.pop('page', None)
        params.update({
            'filename': f"student_fees_report_{int(time.time())}.xlsx",
            'schema_name': tenant.schema_name
        })
        task = generate_excel_report_task.delay(params)
        return JsonResponse({'task_id': task.id})
    
from django.conf import settings

def download_report(request):
    file_path = request.GET.get('file')
    full_path = os.path.join(settings.MEDIA_ROOT, file_path)
    
    if os.path.exists(full_path):
        with open(full_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/octet-stream')
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
            return response
    return HttpResponseNotFound('File not found')

from django.shortcuts import render
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
import tempfile
import csv
from datetime import datetime
from django.db.models import Subquery, OuterRef
from accounting.models import *

# def hostel_fees_report(request):
#     start_date = request.GET.get('start_date')
#     end_date = request.GET.get('end_date')
#     status_filter = request.GET.get('status', 'all')

#     paid_students = []
#     unpaid_students = []
#     total_amount = 0
#     total_paid = 0
#     total_unpaid = 0
#     fee_months = []

#     if start_date and end_date:
#         start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
#         end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()

#         # Get all months in date range
#         current_month = start_date_obj.replace(day=1)
#         while current_month <= end_date_obj:
#             month_name = current_month.strftime('%B')
#             fee_months.extend(
#                 Fee_month.objects.filter(name=month_name)
#             )
#             if current_month.month == 12:
#                 current_month = current_month.replace(year=current_month.year+1, month=1)
#             else:
#                 current_month = current_month.replace(month=current_month.month+1)

#         # Get fees for all months in range
#         hostel_feetypes = Feetype.objects.filter(is_hostel_fee=True)
#         fees_names = Fees_name.objects.filter(
#             fees_type__in=hostel_feetypes,
#             month_id__in=[m.id for m in fee_months]
#         )
#         base_fees = Fees.objects.filter(
#             feetype_id__in=fees_names,
#             month_id__in=[m.id for m in fee_months]
#         )

#         # Paid filter logic
#         if status_filter == 'paid':
#             fees = base_fees.filter(status='paid').annotate(
#                 latest_payment_date=Subquery(
#                     PartialPayment.objects.filter(fee=OuterRef('pk'))
#                     .order_by('-payment_date')
#                     .values('payment_date')[:1]
#                 )
#             ).filter(
#                 latest_payment_date__date__range=[start_date_obj, end_date_obj]
#             )
#         # Unpaid filter logic
#         elif status_filter == 'unpaid':
#             fees = base_fees.filter(status='unpaid')
#         # All statuses
#         else:  
#             paid_fees = base_fees.filter(status='paid').annotate(
#                 latest_payment_date=Subquery(
#                     PartialPayment.objects.filter(fee=OuterRef('pk'))
#                     .order_by('-payment_date')
#                     .values('payment_date')[:1]
#                 )
#             ).filter(
#                 latest_payment_date__date__range=[start_date_obj, end_date_obj]
#             )
#             unpaid_fees = base_fees.filter(status='unpaid')
#             fees = paid_fees | unpaid_fees

#         serial_paid = 1
#         serial_unpaid = 1
#         for fee in fees:
#             student_profile = fee.student_id
#             class_config = student_profile.class_id
#             class_group = class_config.class_group_id if class_config else None
#             class_name = class_group.class_id.name if class_group and class_group.class_id else ""
#             section = class_config.section_id.name if class_config and class_config.section_id else ""
#             shift = class_config.shift_id.name if class_config and class_config.shift_id else ""
#             group = class_group.group_id.name if class_group and class_group.group_id else ""
#             version = student_profile.version
#             roll = student_profile.roll_no

#             student_data = {
#                 'serial': serial_paid if fee.status == 'paid' else serial_unpaid,
#                 'name': student_profile.student_field.name,
#                 'class': class_name,
#                 'section': section,
#                 'shift': shift,
#                 'group': group,
#                 'version': version,
#                 'roll': roll,
#                 'selected_month': fee.month_id.name,
#             }

#             if fee.status == 'paid':
#                 latest_payment = fee.partial_payments.order_by('-payment_date').first()
#                 if latest_payment:
#                     student_data['paid_date'] = latest_payment.payment_date.strftime('%Y-%m-%d')
#                 else:
#                     student_data['paid_date'] = "N/A"
#                 paid_students.append(student_data)
#                 serial_paid += 1
#             else:
#                 student_data['amount'] = fee.total_fee_after_partial_payments()
#                 unpaid_students.append(student_data)
#                 total_amount += student_data['amount']
#                 serial_unpaid += 1

#         total_paid = len(paid_students)
#         total_unpaid = len(unpaid_students)

#     # Handle CSV download
#     if 'download' in request.GET:
#         response = HttpResponse(content_type='text/csv')
#         response['Content-Disposition'] = 'attachment; filename="hostel_fees_report.csv"'

#         writer = csv.writer(response)
#         writer.writerow([
#             'SL No', 'Name', 'Class', 'Section', 'Shift', 'Group', 
#             'Version', 'Roll', 'Month', 'Amount/Paid Date'
#         ])
        
#         for student in paid_students:
#             writer.writerow([
#                 student['serial'],
#                 student['name'],
#                 student['class'],
#                 student['section'],
#                 student['shift'],
#                 student['group'],
#                 student['version'],
#                 student['roll'],
#                 student['selected_month'],
#                 student.get('paid_date', '')
#             ])
        
#         for student in unpaid_students:
#             writer.writerow([
#                 student['serial'],
#                 student['name'],
#                 student['class'],
#                 student['section'],
#                 student['shift'],
#                 student['group'],
#                 student['version'],
#                 student['roll'],
#                 student['selected_month'],
#                 student.get('amount', '')
#             ])
        
#         writer.writerow([])
#         writer.writerow(['Total Amount', total_amount])
#         writer.writerow(['Total Paid Students', total_paid])
#         writer.writerow(['Total Unpaid Students', total_unpaid])

#         return response
    
#     elif 'download_pdf' in request.GET:
#         html_string = render_to_string('crucial/report/pdf_hostel.html', {
#             'paid_students': paid_students,
#             'unpaid_students': unpaid_students,
#             'total_amount': total_amount,
#             'total_paid': total_paid,
#             'total_unpaid': total_unpaid,
#             'selected_month_name': fee_month.name if fee_month else '',
#             'selected_date': selected_date,
#             'status_filter': status_filter,
#         })

#         html = HTML(string=html_string)
#         result = html.write_pdf()

#         response = HttpResponse(content_type='application/pdf')
#         response['Content-Disposition'] = f'attachment; filename="hostel_fees_report_{selected_date}.pdf"'
#         response['Content-Transfer-Encoding'] = 'binary'
        
#         with tempfile.NamedTemporaryFile(delete=True) as output:
#             output.write(result)
#             output.flush()
#             output.seek(0)
#             response.write(output.read())

#         return response

#     context = {
#         'start_date': start_date,
#         'end_date': end_date,
#         'fee_months': [m.name for m in fee_months],
#         'status_filter': status_filter,
#         'paid_students': paid_students,
#         'unpaid_students': unpaid_students,
#         'total_amount': total_amount,
#         'total_paid': total_paid,
#         'total_unpaid': total_unpaid,
#     }
#     return render(request, 'crucial/report/hostel_fees_report.html', context)

from django.shortcuts import render
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
import tempfile
import csv
from datetime import datetime
from django.db.models import Subquery, OuterRef
from accounting.models import *

def hostel_fees_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status_filter = request.GET.get('status', 'all')

    paid_students = []
    unpaid_students = []
    total_amount = 0
    total_paid = 0
    total_unpaid = 0
    fee_months = []

    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            return render(request, 'crucial/report/hostel_fees_report.html', {'error': 'Invalid date format'})

        # Get all unique months in date range
        current_month = start_date_obj.replace(day=1)
        seen_months = set()
        
        while current_month <= end_date_obj:
            month_name = current_month.strftime('%B')
            try:
                month_obj = Fee_month.objects.get(name=month_name)
                if month_obj.id not in seen_months:
                    fee_months.append(month_obj)
                    seen_months.add(month_obj.id)
            except Fee_month.DoesNotExist:
                pass
            
            # Move to next month
            if current_month.month == 12:
                current_month = current_month.replace(year=current_month.year+1, month=1)
            else:
                current_month = current_month.replace(month=current_month.month+1)

        if fee_months:
            # Get hostel fee types
            hostel_feetypes = Feetype.objects.filter(is_hostel_fee=True)
            fees_names = Fees_name.objects.filter(
                fees_type__in=hostel_feetypes,
                month_id__in=[m.id for m in fee_months]
            ).distinct()

            base_fees = Fees.objects.filter(
                feetype_id__in=fees_names,
                month_id__in=[m.id for m in fee_months]
            ).select_related('student_id', 'month_id')

            # Paid filter logic
            if status_filter == 'paid':
                fees = base_fees.filter(status='paid').annotate(
                    latest_payment_date=Subquery(
                        PartialPayment.objects.filter(fee=OuterRef('pk'))
                        .order_by('-payment_date')
                        .values('payment_date')[:1]
                    )
                ).filter(
                    latest_payment_date__date__range=(start_date_obj, end_date_obj)
                ).distinct()

            # Unpaid filter logic
            elif status_filter == 'unpaid':
                fees = base_fees.filter(status='unpaid')

            # All statuses
            else:
                paid_fees = base_fees.filter(status='paid').annotate(
                    latest_payment_date=Subquery(
                        PartialPayment.objects.filter(fee=OuterRef('pk'))
                        .order_by('-payment_date')
                        .values('payment_date')[:1]
                    )
                ).filter(
                    latest_payment_date__date__range=(start_date_obj, end_date_obj)
                )
                unpaid_fees = base_fees.filter(status='unpaid')
                fees = paid_fees.union(unpaid_fees)

            # Process results
            serial_paid = 1
            serial_unpaid = 1
            
            for fee in fees:
                student_profile = fee.student_id
                class_config = student_profile.class_id
                class_group = class_config.class_group_id if class_config else None
                
                student_data = {
                    'serial': serial_paid if fee.status == 'paid' else serial_unpaid,
                    'name': student_profile.student_field.name,
                    'class': class_group.class_id.name if class_group else "N/A",
                    'section': class_config.section_id.name if class_config else "N/A",
                    'shift': class_config.shift_id.name if class_config else "N/A",
                    'group': class_group.group_id.name if class_group else "N/A",
                    'version': student_profile.version,
                    'roll': student_profile.roll_no,
                    'selected_month': fee.month_id.name,
                }

                if fee.status == 'paid':
                    # Get actual payment date
                    latest_payment = fee.partial_payments.order_by('-payment_date').first()
                    if latest_payment:
                        student_data['paid_date'] = latest_payment.payment_date.strftime('%Y-%m-%d')
                    else:
                        student_data['paid_date'] = fee.updated_at.strftime('%Y-%m-%d')
                    paid_students.append(student_data)
                    serial_paid += 1
                else:
                    student_data['amount'] = fee.total_fee_after_partial_payments()
                    unpaid_students.append(student_data)
                    total_amount += student_data['amount']
                    serial_unpaid += 1

            total_paid = len(paid_students)
            total_unpaid = len(unpaid_students)

    # Handle CSV download
    if 'download' in request.GET and start_date and end_date:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="hostel_fees_{start_date}_to_{end_date}.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'SL No', 'Name', 'Class', 'Section', 'Shift', 'Group', 
            'Version', 'Roll', 'Month', 'Amount/Paid Date'
        ])
        
        for student in paid_students:
            writer.writerow([
                student['serial'],
                student['name'],
                student['class'],
                student['section'],
                student['shift'],
                student['group'],
                student['version'],
                student['roll'],
                student['selected_month'],
                student.get('paid_date', '')
            ])
        
        for student in unpaid_students:
            writer.writerow([
                student['serial'],
                student['name'],
                student['class'],
                student['section'],
                student['shift'],
                student['group'],
                student['version'],
                student['roll'],
                student['selected_month'],
                student.get('amount', '')
            ])
        
        writer.writerow([])
        writer.writerow(['Total Amount', total_amount])
        writer.writerow(['Total Paid Students', total_paid])
        writer.writerow(['Total Unpaid Students', total_unpaid])

        return response
    
    # Handle PDF download
    elif 'download_pdf' in request.GET and start_date and end_date:
        html_string = render_to_string('crucial/report/pdf_hostel.html', {
            'paid_students': paid_students,
            'unpaid_students': unpaid_students,
            'total_amount': total_amount,
            'total_paid': total_paid,
            'total_unpaid': total_unpaid,
            'start_date': start_date,
            'end_date': end_date,
            'fee_months': [m.name for m in fee_months],
            'status_filter': status_filter,
        })

        html = HTML(string=html_string)
        result = html.write_pdf()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="hostel_fees_{start_date}_to_{end_date}.pdf"'
        
        with tempfile.NamedTemporaryFile(delete=True) as output:
            output.write(result)
            output.flush()
            output.seek(0)
            response.write(output.read())

        return response

    context = {
        'start_date': start_date,
        'end_date': end_date,
        'fee_months': [m.name for m in fee_months],
        'status_filter': status_filter,
        'paid_students': paid_students,
        'unpaid_students': unpaid_students,
        'total_amount': total_amount,
        'total_paid': total_paid,
        'total_unpaid': total_unpaid,
        'error': None,
    }
    return render(request, 'crucial/report/hostel_fees_report.html', context)



from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from io import BytesIO
from ssl_commerz.models import PaymentTransaction

def download_success_transactions_excel(request):
    transactions = PaymentTransaction.objects.filter(status='SUCCESS')

    wb = Workbook()
    ws = wb.active
    ws.title = "Successful Transactions"

    headers = ['Transaction ID', 'Amount', 'Date', 'Time']
    ws.append(headers)

    for txn in transactions:
        tran_date = txn.tran_date
        date = tran_date.date()
        time = tran_date.strftime("%I:%M:%S %p") 
        row = [
            txn.tran_id,
            float(txn.amount),  
            date,
            time, 
        ]
        ws.append(row)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="successful_transactions.xlsx"'
    return response


# import re
# from django.http import HttpResponse
# from openpyxl import Workbook
# from openpyxl.styles import Font
# from openpyxl.utils import get_column_letter

# def download_receive_excel(request):
#     # Filter receives and exclude voucher numbers starting with 'FEE'
#     receives = Receive.objects.exclude(voucher_no__startswith='FEE')
    
#     # Process data and group entries
#     grouped_data = {}
#     for receive in receives:
#         # Extract transaction ID from description
#         tran_id = None
#         if receive.description:
#             match = re.search(r'\(([^)]+)\)$', receive.description)
#             if match:
#                 tran_id = match.group(1)
        
#         # Skip entries without student or transaction ID
#         if not receive.student or not tran_id:
#             continue
        
#         # Create grouping key
#         created_at = receive.created_at
#         key = (receive.student.id, created_at, tran_id)
        
#         # Group entries and sum amounts
#         if key in grouped_data:
#             grouped_data[key]['amount'] += receive.amount
#         else:
#             grouped_data[key] = {
#                 'student': receive.student,
#                 'date': created_at.date(),
#                 'time': created_at.time().strftime('%H:%M:%S'),
#                 'tran_id': tran_id,
#                 'amount': receive.amount,
#             }
    
#     # Prepare Excel response
#     response = HttpResponse(content_type='application/ms-excel')
#     response['Content-Disposition'] = 'attachment; filename="student_payments.xlsx"'
    
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Student Payments"
    
#     # Write headers
#     headers = ['Student Name', 'Roll Number', 'Amount', 'Date', 'Time', 'Transaction ID']
#     for col_num, header in enumerate(headers, 1):
#         col_letter = get_column_letter(col_num)
#         ws[f'{col_letter}1'] = header
#         ws[f'{col_letter}1'].font = Font(bold=True)
    
#     # Write data rows
#     row_num = 2
#     for entry in grouped_data.values():
#         student = entry['student']
#         ws.cell(row=row_num, column=1, value=student.student_field.name)
#         ws.cell(row=row_num, column=2, value=student.roll_no)
#         ws.cell(row=row_num, column=3, value=float(entry['amount']))
#         ws.cell(row=row_num, column=4, value=entry['date'])
#         ws.cell(row=row_num, column=5, value=entry['time'])
#         ws.cell(row=row_num, column=6, value=entry['tran_id'])
#         row_num += 1
    
#     # Adjust column widths
#     for column in ws.columns:
#         max_length = 0
#         column_letter = get_column_letter(column[0].column)
#         for cell in column:
#             try:
#                 if len(str(cell.value)) > max_length:
#                     max_length = len(str(cell.value))
#             except:
#                 pass
#         adjusted_width = (max_length + 2)
#         ws.column_dimensions[column_letter].width = adjusted_width
    
#     wb.save(response)
#     return response

import re
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def download_receive_excel(request):
    # Get filter parameters from request
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    # Base query
    receives = Receive.objects.exclude(voucher_no__startswith='FEE')
    
    # Apply date filter if dates are provided
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            receives = receives.filter(created_at__date__range=(start_date, end_date))
        except ValueError:
            # Handle invalid date format
            pass
    
    # Rest of the processing remains the same
    grouped_data = {}
    for receive in receives:
        tran_id = None
        if receive.description:
            match = re.search(r'\(([^)]+)\)$', receive.description)
            if match:
                tran_id = match.group(1)
        
        if not receive.student or not tran_id:
            continue
        
        created_at = receive.created_at
        key = (receive.student.id, created_at, tran_id)
        
        if key in grouped_data:
            grouped_data[key]['amount'] += receive.amount
        else:
            grouped_data[key] = {
                'student': receive.student,
                'date': created_at.date(),
                'time': created_at.time().strftime('%H:%M:%S'),
                'tran_id': tran_id,
                'amount': receive.amount,
            }

    # Excel generation remains the same
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="student_payments.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Payments"
    
    headers = ['Student Name', 'Roll Number', 'Amount', 'Date', 'Time', 'Transaction ID']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
        ws[f'{col_letter}1'].font = Font(bold=True)
    
    row_num = 2
    for entry in grouped_data.values():
        student = entry['student']
        ws.cell(row=row_num, column=1, value=student.student_field.name)
        ws.cell(row=row_num, column=2, value=student.roll_no)
        ws.cell(row=row_num, column=3, value=float(entry['amount']))
        ws.cell(row=row_num, column=4, value=entry['date'])
        ws.cell(row=row_num, column=5, value=entry['time'])
        ws.cell(row=row_num, column=6, value=entry['tran_id'])
        row_num += 1
    
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(response)
    return response


from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from dateutil.relativedelta import relativedelta
from datetime import datetime
from .models import (
    FeeHead, Feetype, Fees_name, Fee_month, Fees,
    Fee_package, StudentProfile
)

from datetime import date
from dateutil.relativedelta import relativedelta
from django.utils import timezone
from django.contrib import messages
from django.shortcuts import render, redirect
import logging

logger = logging.getLogger(__name__)

def get_allocation_month(date_obj):
    """Determine the allocation month based on creation date"""
    # Always return current month only (simplified logic)
    return date_obj.replace(day=1)

# def create_fee_structure():
#     """Ensure required fee structure exists in the system"""
#     heads = {
#         'tuition': ('Tuition Fee', {'version': 'English'}),
#         'late': ('Fine for Delayed Tuition Fee Payment', {'version': 'English'}),
#         'readmission': ('Re-admission Fee (Late Pay)', {'version': 'English'}),
#         'hostel': ('Fine for Delayed Hostel Charge Payment', {'is_hostel_fee': True, 'version': 'English'})
#     }
    
#     feetypes = {}
#     for key, (name, extra) in heads.items():
#         head, _ = FeeHead.objects.get_or_create(name=name)
        
#         defaults = {
#             'fee_Schedule': 'Monthly',
#             'status': 'Active',
#             'late_fee_percentage': 0,
#             **extra
#         }
        
#         existing_feetype = Feetype.objects.filter(
#             fee_head=head,
#             fee_Schedule=defaults['fee_Schedule'],
#             status=defaults['status'],
#             version=defaults.get('version', 'English'),
#             **{k: v for k, v in extra.items() if k in ['is_hostel_fee']}
#         ).first()
        
#         if existing_feetype:
#             feetypes[key] = existing_feetype
#         else:
#             feetypes[key], _ = Feetype.objects.get_or_create(
#                 fee_head=head,
#                 defaults=defaults
#             )
#             for field, value in defaults.items():
#                 setattr(feetypes[key], field, value)
#             feetypes[key].save()
    
#     return tuple(feetypes.values())

def create_fee_structure():
    """Ensure required fee structure exists in the system for both versions"""
    heads = {
        'tuition': ('Tuition Fee', {'version': None}),  # Version will be set per student
        'late': ('Fine for Delayed Tuition Fee Payment', {'version': None}),
        'readmission': ('Re-admission Fee (Late Pay)', {'version': None}),
        'hostel': ('Fine for Delayed Hostel Charge Payment', {'is_hostel_fee': True, 'version': None})
    }
    
    # We'll create two sets of fee types - one for each version
    feetypes = {'English': {}, 'Bangla': {}}
    
    for version in ['English', 'Bangla']:
        for key, (name, extra) in heads.items():
            head, _ = FeeHead.objects.get_or_create(name=name)
            
            defaults = {
                'fee_Schedule': 'Monthly',
                'status': 'Active',
                'late_fee_percentage': 0,
                'version': version,
                **extra
            }
            
            existing_feetype = Feetype.objects.filter(
                fee_head=head,
                fee_Schedule=defaults['fee_Schedule'],
                status=defaults['status'],
                version=version,
                **{k: v for k, v in extra.items() if k in ['is_hostel_fee']}
            ).first()
            
            if existing_feetype:
                feetypes[version][key] = existing_feetype
            else:
                feetypes[version][key], _ = Feetype.objects.get_or_create(
                    fee_head=head,
                    version=version,
                    defaults=defaults
                )
                for field, value in defaults.items():
                    setattr(feetypes[version][key], field, value)
                feetypes[version][key].save()
    
    # Return both versions' fee types
    return (
        feetypes['English']['tuition'], 
        feetypes['English']['late'], 
        feetypes['English']['readmission'], 
        feetypes['English']['hostel'],
        feetypes['Bangla']['tuition'], 
        feetypes['Bangla']['late'], 
        feetypes['Bangla']['readmission'], 
        feetypes['Bangla']['hostel']
    )

def get_or_create_fee_month(month_date):
    """Get or create Fee_month for a given date"""
    month_name = month_date.strftime('%B')
    month_obj, _ = Fee_month.objects.get_or_create(name=month_name)
    return month_obj

def create_fee_record(student, fee_type, amount, month_date, academic_year, user):
    """Create a fee record if it doesn't exist"""
    month_obj = get_or_create_fee_month(month_date)
    
    fee_name = Fees_name.objects.filter(
        fees_type=fee_type,
        month=month_obj,
        academic_year=academic_year,
        fees_title=f"{fee_type.fee_head.name} - {month_obj.name}"
    ).first()
    
    if not fee_name:
        fee_name = Fees_name.objects.create(
            fees_type=fee_type,
            month=month_obj,
            academic_year=academic_year,
            fees_title=f"{fee_type.fee_head.name} - {month_obj.name}",
            startdate=month_date.replace(day=1),
            enddate=(month_date + relativedelta(months=1)).replace(day=1) - relativedelta(days=1),
            created_by=user
        )
    
    if not Fees.objects.filter(
        student_id=student,
        feetype_id=fee_name,
        month_id=month_obj,
        academic_year=academic_year
    ).exists():
        Fees.objects.create(
            student_id=student,
            feetype_id=fee_name,
            amount=amount,
            status='unpaid',
            month_id=month_obj,
            academic_year=academic_year,
            created_by=user
        )
        return True
    return False

def delete_existing_late_fees(student, late_feetype, readmission_feetype, academic_year):
    """Delete existing late fees and readmission fees for a student"""
    # Get all fee names for these types
    fee_names = Fees_name.objects.filter(
        fees_type__in=[late_feetype, readmission_feetype],
        academic_year=academic_year
    )
    
    # Delete all fees records
    Fees.objects.filter(
        student_id=student,
        feetype_id__in=fee_names
    ).delete()

# def process_student_fees(student, tuition_feetype, late_feetype, readmission_feetype, 
#                         hostel_feetype, current_date, request):
#     """Process all fees for a single student"""
#     current_month = current_date.replace(day=1)
    
#     # Process tuition fees
#     tuition_fees = Fees.objects.filter(
#         student_id=student,
#         feetype_id__fees_type=tuition_feetype,
#         status__in=['unpaid', 'partial']
#     ).order_by('month_id')
    
#     if tuition_fees.exists():
#         # Get the oldest unpaid tuition fee
#         oldest_unpaid = tuition_fees.first()
#         allocation_month = get_allocation_month(oldest_unpaid.created_at.date())
        
#         # Calculate months overdue (only up to current month)
#         months_overdue = 0
#         while allocation_month < current_month:
#             allocation_month += relativedelta(months=1)
#             months_overdue += 1
        
#         # Delete any existing late fees before creating new ones
#         delete_existing_late_fees(student, late_feetype, readmission_feetype, oldest_unpaid.academic_year)
        
#         if months_overdue > 0:
#             try:
#                 tuition_package = Fee_package.objects.get(
#                     student_class=student.class_id.class_group_id,
#                     fees_type=tuition_feetype,
#                     academic_year=oldest_unpaid.academic_year
#                 )
                
#                 if months_overdue == 1:
#                     # Only 1 month late - apply 50 Taka fine
#                     create_fee_record(
#                         student=student,
#                         fee_type=late_feetype,
#                         amount=50,
#                         month_date=current_month,
#                         academic_year=oldest_unpaid.academic_year,
#                         user=request.user
#                     )
#                 elif months_overdue > 1:
#                     # More than 1 month late - apply re-admission fee
#                     create_fee_record(
#                         student=student,
#                         fee_type=readmission_feetype,
#                         amount=tuition_package.amount,
#                         month_date=current_month,
#                         academic_year=oldest_unpaid.academic_year,
#                         user=request.user
#                     )
#             except Fee_package.DoesNotExist:
#                 logger.error(f"No tuition package found for student {student.id}")
    
#     # Process hostel fees
#     hostel_fees = Fees.objects.filter(
#         student_id=student,
#         feetype_id__fees_type=hostel_feetype,
#         status__in=['unpaid', 'partial']
#     ).order_by('month_id')
    
#     if hostel_fees.exists():
#         oldest_unpaid = hostel_fees.first()
#         allocation_month = get_allocation_month(oldest_unpaid.created_at.date())
        
#         # Calculate months overdue (only up to current month)
#         months_overdue = 0
#         while allocation_month < current_month:
#             allocation_month += relativedelta(months=1)
#             months_overdue += 1
        
#         # Apply 300 Taka fine for each overdue month
#         if months_overdue > 0:
#             create_fee_record(
#                 student=student,
#                 fee_type=hostel_feetype,
#                 amount=300,
#                 month_date=current_month,
#                 academic_year=oldest_unpaid.academic_year,
#                 user=request.user
#             )


def process_student_fees(student, tuition_feetype, late_feetype, readmission_feetype, 
                        hostel_feetype, current_date, request):
    """Process all fees for a single student"""
    # Get the appropriate fee types based on student version
    if student.version == 'Bangla':
        tuition_feetype = Feetype.objects.get(fee_head=tuition_feetype.fee_head, version='Bangla')
        late_feetype = Feetype.objects.get(fee_head=late_feetype.fee_head, version='Bangla')
        readmission_feetype = Feetype.objects.get(fee_head=readmission_feetype.fee_head, version='Bangla')
        hostel_feetype = Feetype.objects.get(fee_head=hostel_feetype.fee_head, version='Bangla', is_hostel_fee=True)
    
    current_month = current_date.replace(day=1)
    
    # Process tuition fees - CRITICAL FIX: Filter by student's version-specific tuition feetype
    tuition_fees = Fees.objects.filter(
        student_id=student,
        feetype_id__fees_type=tuition_feetype,  # This now uses the version-correct feetype
        status__in=['unpaid', 'partial']
    ).order_by('month_id')
    
    if tuition_fees.exists():
        # Rest of the function remains the same...
        oldest_unpaid = tuition_fees.first()
        allocation_month = get_allocation_month(oldest_unpaid.created_at.date())
        
        months_overdue = 0
        while allocation_month < current_month:
            allocation_month += relativedelta(months=1)
            months_overdue += 1
        
        delete_existing_late_fees(student, late_feetype, readmission_feetype, oldest_unpaid.academic_year)
        
        if months_overdue > 0:
            try:
                tuition_package = Fee_package.objects.get(
                    student_class=student.class_id.class_group_id,
                    fees_type=tuition_feetype,  # Using version-correct feetype here too
                    academic_year=oldest_unpaid.academic_year
                )
                
                if months_overdue == 1:
                    create_fee_record(
                        student=student,
                        fee_type=late_feetype,
                        amount=50,
                        month_date=current_month,
                        academic_year=oldest_unpaid.academic_year,
                        user=request.user
                    )
                elif months_overdue > 1:
                    create_fee_record(
                        student=student,
                        fee_type=readmission_feetype,
                        amount=tuition_package.amount,
                        month_date=current_month,
                        academic_year=oldest_unpaid.academic_year,
                        user=request.user
                    )
            except Fee_package.DoesNotExist:
                logger.error(f"No tuition package found for student {student.id}")
    
    # Similar fix for hostel fees
    hostel_fees = Fees.objects.filter(
        student_id=student,
        feetype_id__fees_type=hostel_feetype,  # Version-correct hostel feetype
        status__in=['unpaid', 'partial']
    ).order_by('month_id')
    
    if hostel_fees.exists():
        oldest_unpaid = hostel_fees.first()
        allocation_month = get_allocation_month(oldest_unpaid.created_at.date())
        
        months_overdue = 0
        while allocation_month < current_month:
            allocation_month += relativedelta(months=1)
            months_overdue += 1
        
        if months_overdue > 0:
            create_fee_record(
                student=student,
                fee_type=hostel_feetype,
                amount=300,
                month_date=current_month,
                academic_year=oldest_unpaid.academic_year,
                user=request.user
            )


# def add_late_fees(request):
#     if request.method == 'POST':
#         try:
#             # Initialize fee structure for both versions
#             (en_tuition, en_late, en_readmission, en_hostel, 
#              bn_tuition, bn_late, bn_readmission, bn_hostel) = create_fee_structure()
            
#             current_date = timezone.now().date()
            
#             # Process students in batches
#             batch_size = 50
#             student_count = StudentProfile.objects.count()
            
#             for offset in range(0, student_count, batch_size):
#                 students = StudentProfile.objects.all()[offset:offset+batch_size]
#                 for student in students:
#                     try:
#                         # Select appropriate fee types based on student version
#                         if student.version == 'Bangla':
#                             tuition = bn_tuition
#                             late = bn_late
#                             readmission = bn_readmission
#                             hostel = bn_hostel
#                         else:
#                             tuition = en_tuition
#                             late = en_late
#                             readmission = en_readmission
#                             hostel = en_hostel
                            
#                         process_student_fees(
#                             student=student,
#                             tuition_feetype=tuition,
#                             late_feetype=late,
#                             readmission_feetype=readmission,
#                             hostel_feetype=hostel,
#                             current_date=current_date,
#                             request=request
#                         )
#                     except Exception as e:
#                         logger.error(f"Error processing student {student.id}: {str(e)}")
#                         continue
            
#             messages.success(request, "Late fees processed successfully for all students")
#         except Exception as e:
#             logger.error(f"Error in add_late_fees: {str(e)}")
#             messages.error(request, f"Error processing fees: {str(e)}")
        
#         return redirect('add_late_fees')
    
#     return render(request, 'crucial/finance/add_late_fees.html')

def add_late_fees(request):
    # Initialize filter options
    classes = StudentClass.objects.all()
    sections = StudentSection.objects.all()
    groups = StuGroup.objects.all()
    versions = [v[0] for v in StudentProfile.Version.choices]
    shifts = StudentShift.objects.all()
    academic_sessions = AcademicSession.objects.all()

    context = {
        'classes': classes,
        'sections': sections,
        'groups': groups,
        'versions': versions,
        'shifts': shifts,
        'academic_sessions': academic_sessions,
    }

    if request.method == 'POST':
        try:
            # Get filter parameters from POST data
            class_id = request.POST.get('class')
            section_id = request.POST.get('section')
            version = request.POST.get('version')
            group_id = request.POST.get('group')
            shift_id = request.POST.get('shift')
            academic_session_id = request.POST.get('academic_session')

            # Initialize fee structure
            en_tuition, en_late, en_readmission, en_hostel, bn_tuition, bn_late, bn_readmission, bn_hostel = create_fee_structure()
            current_date = timezone.now().date()
            
            # Build student filter
            filters = {
                'class_id__class_group_id__class_id': class_id,
                'student_field__status': 'Active',
            }
            if section_id:
                filters['class_id__section_id'] = section_id
            if version:
                filters['version'] = version
            if group_id:
                filters['class_id__class_group_id__group_id'] = group_id
            if shift_id:  
                filters['class_id__shift_id'] = shift_id
            if academic_session_id: 
                filters['academic_session_year_id'] = academic_session_id

            students = StudentProfile.objects.filter(**filters)
            
            # Process filtered students
            for student in students:
                try:
                    if student.version == 'Bangla':
                        tuition = bn_tuition
                        late = bn_late
                        readmission = bn_readmission
                        hostel = bn_hostel
                    else:
                        tuition = en_tuition
                        late = en_late
                        readmission = en_readmission
                        hostel = en_hostel
                        
                    process_student_fees(
                        student=student,
                        tuition_feetype=tuition,
                        late_feetype=late,
                        readmission_feetype=readmission,
                        hostel_feetype=hostel,
                        current_date=current_date,
                        request=request
                    )
                except Exception as e:
                    logger.error(f"Error processing student {student.id}: {str(e)}")
                    continue
            
            messages.success(request, "Late fees processed successfully for filtered students")
        except Exception as e:
            logger.error(f"Error in add_late_fees: {str(e)}")
            messages.error(request, f"Error processing fees: {str(e)}")
        
        # Return to same page with context
        return render(request, 'crucial/finance/add_late_fees.html', context)
    
    # GET request - show filter form
    return render(request, 'crucial/finance/add_late_fees.html', context)


def generate_head_report_context(request):
    institute = Institute.objects.order_by('-id').first()
    context = {
        'students_data': [],
        'selected_head': None,
        'grand_total': Decimal('0.00'),
        'from_date': None,
        'to_date': None,
        'classes': StudentClass.objects.all(),
        'sections': StudentSection.objects.all(),
        'groups': StuGroup.objects.all(),
        'shifts': StudentShift.objects.all(),
        'selected_class': 'All',
        'selected_section': 'All',
        'selected_group': 'All',
        'selected_shift': 'All',
        'selected_version': 'All',
        'heads': FeeHead.objects.all(),
        'search_term': '',
        'institute_logo': request.build_absolute_uri(institute.institute_logo.url) if institute and institute.institute_logo else None
    }
    
    selected_head_id = request.GET.get('head')
    search_term = request.GET.get('search', '')
    
    if selected_head_id:
        try:
            context['selected_head'] = FeeHead.objects.get(id=selected_head_id)
        except FeeHead.DoesNotExist:
            pass

    # Get selected filters from request
    selected_class_id = request.GET.get('class')
    selected_section_id = request.GET.get('section')
    selected_group_id = request.GET.get('group')
    selected_shift_id = request.GET.get('shift')
    selected_version = request.GET.get('version')
    context['search_term'] = search_term

    # Get display names for selected filters
    if selected_class_id:
        try:
            context['selected_class'] = StudentClass.objects.get(id=selected_class_id).name
        except StudentClass.DoesNotExist:
            pass

    if selected_section_id:
        try:
            context['selected_section'] = StudentSection.objects.get(id=selected_section_id).name
        except StudentSection.DoesNotExist:
            pass

    if selected_group_id:
        try:
            context['selected_group'] = ABBREVIATIONS.get(
                StuGroup.objects.get(id=selected_group_id).name.strip(),
                StuGroup.objects.get(id=selected_group_id).name[:3]
            )
        except StuGroup.DoesNotExist:
            pass

    if selected_shift_id:
        try:
            context['selected_shift'] = ABBREVIATIONS.get(
                StudentShift.objects.get(id=selected_shift_id).name.strip(),
                StudentShift.objects.get(id=selected_shift_id).name[:1]
            )
        except StudentShift.DoesNotExist:
            pass

    if selected_version:
        context['selected_version'] = ABBREVIATIONS.get(
            selected_version.strip(),
            selected_version[:3]
        )

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if from_date and to_date and context['selected_head']:
        try:
            from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
            to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()
        except ValueError:
            return context

        receives = Receive.objects.filter(
            date__gte=from_date_obj,
            date__lte=to_date_obj,
        ).exclude(voucher_no__startswith='FEE').exclude(voucher_no__startswith='REC2').exclude(student=None).select_related('student').order_by('student__student_field__name')

        # Apply additional filters from the request
        filters = {
            'student__class_id__class_group_id__class_id': selected_class_id,
            'student__class_id__section_id': selected_section_id,
            'student__class_id__class_group_id__group_id': selected_group_id,
            'student__class_id__shift_id': selected_shift_id,
            'student__version': selected_version,
        }
        filters = {k: v for k, v in filters.items() if v}
        receives = receives.filter(**filters)

        # Apply search filter
        if search_term:
            receives = receives.filter(
                Q(student__student_field__name__icontains=search_term) |
                Q(student__roll_no__icontains=search_term)
            )

        student_ids = [receive.student_id for receive in receives if receive.student_id]
        receive_dates = {receive.created_at.date() for receive in receives}

        # Get fees entries for the selected head
        fees_entries = Fees.objects.filter(
            student_id__in=student_ids,
            updated_at__date__in=receive_dates,
            feetype_id__fees_type__fee_head=context['selected_head']
        ).select_related('feetype_id__fees_type__fee_head').order_by('id')

        fees_grouped = defaultdict(list)
        for fee in fees_entries:
            key = (fee.student_id_id, fee.updated_at.date(), fee.amount)
            if fee.feetype_id:
                fees_grouped[key].append({
                    'feetype': fee.feetype_id,
                    'fee_id': fee.id
                })

        students_data = []
        grand_total = Decimal('0.00')
        used_fees = set()

        for receive in receives:
            student = receive.student
            key = (student.id, receive.created_at.date(), receive.amount)
            fee_type = None

            possible_fees = fees_grouped.get(key, [])
            for fee_entry in possible_fees:
                if fee_entry['fee_id'] not in used_fees:
                    fee_type = fee_entry['feetype']
                    used_fees.add(fee_entry['fee_id'])
                    break

            if fee_type and fee_type.fees_type.fee_head == context['selected_head']:
                # Get student details
                # Get student details with null checks
                class_config = student.class_id
                class_name = class_config.class_group_id.class_id.name if (
                    class_config and 
                    class_config.class_group_id and 
                    class_config.class_group_id.class_id
                ) else ''
                group_name = class_config.class_group_id.group_id.name if (
                    class_config and 
                    class_config.class_group_id and 
                    class_config.class_group_id.group_id
                ) else ''
                group = ABBREVIATIONS.get(group_name.strip(), group_name[:3]) if group_name else ''
                section = class_config.section_id.name if (
                    class_config and 
                    class_config.section_id
                ) else ''
                shift_name = class_config.shift_id.name if (
                    class_config and 
                    class_config.shift_id
                ) else ''
                shift = ABBREVIATIONS.get(shift_name.strip(), shift_name[:1] if shift_name else '')
                version = student.version or ''
                version_abbr = ABBREVIATIONS.get(version.strip(), version[:3])

                students_data.append({
                    'student': student,
                    'roll_no': student.roll_no,
                    'class': class_name,
                    'group': group,
                    'section': section,
                    'shift': shift,
                    'version': version_abbr,
                    'amount': receive.amount
                })
                grand_total += receive.amount

        context.update({
            'students_data': students_data,
            'grand_total': grand_total,
            'from_date': from_date,
            'to_date': to_date
        })

    return context

def student_head_report(request):
    context = generate_head_report_context(request)
    # Add pagination
    paginator = Paginator(context['students_data'], 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context['page_obj'] = page_obj
    return render(request, 'crucial/report/student_head_report.html', context)

def student_head_report_pdf(request):
    context = generate_head_report_context(request)
    html_string = render_to_string('crucial/report/student_head_report_pdf.html', context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="student_head_report_{int(time.time())}.pdf"'
    
    # Create PDF
    pisa_status = pisa.CreatePDF(
        html_string, 
        dest=response,
        encoding='UTF-8'
    )
    
    if pisa_status.err:
        return HttpResponse('PDF generation error', status=500)
    return response

def student_head_report_excel(request):
    context = generate_head_report_context(request)
    students_data = context['students_data']
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="student_head_report_{int(time.time())}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Head Report"
    
    # Add headers
    headers = [
        'SL NO', 'Student', 'Roll No', 'Class', 
        'Group', 'Section', 'Shift', 'Version', 'Amount'
    ]
    ws.append(headers)
    
    # Add data rows
    for idx, student in enumerate(students_data, start=1):
        row = [
            idx,
            student['student'].student_field.name,
            student['roll_no'],
            student['class'],
            student['group'],
            student['section'],
            student['shift'],
            student['version'],
            student['amount']
        ]
        ws.append(row)
    
    # Add grand total
    ws.append([''] * 8 + [context['grand_total']])
    
    wb.save(response)
    return response



def delete_late_fees(request):
    if request.method == 'POST':
        try:
            # Get the three fee types we want to delete
            late_fee_heads = [
                'Fine for Delayed Tuition Fee Payment',
                'Re-admission Fee (Late Pay)', 
                'Fine for Delayed Hostel Charge Payment'
            ]
            
            # Get the FeeHead objects
            fee_heads = FeeHead.objects.filter(name__in=late_fee_heads)
            
            if not fee_heads.exists():
                messages.warning(request, "No late fee types found to delete")
                return redirect('delete_late_fees')
            
            # Get all Feetype objects under these heads
            fee_types = Feetype.objects.filter(fee_head__in=fee_heads)
            
            if not fee_types.exists():
                messages.warning(request, "No late fee allocations found to delete")
                return redirect('delete_late_fees')
            
            # Get all Fees_name records for these fee types
            fee_names = Fees_name.objects.filter(fees_type__in=fee_types)
            
            # Delete in batches to avoid memory issues
            batch_size = 100
            total_deleted = 0
            
            # First delete the Fees records
            for i in range(0, fee_names.count(), batch_size):
                batch = fee_names[i:i+batch_size]
                deleted, _ = Fees.objects.filter(feetype_id__in=batch).delete()
                total_deleted += deleted
            
            # Then delete the Fees_name records
            fee_names.delete()
            
            # Optionally delete the Feetype records (comment out if you want to keep them)
            # fee_types.delete()
            
            messages.success(request, f"Successfully deleted {total_deleted} late fee allocations")
            
        except Exception as e:
            messages.error(request, f"Error deleting late fees: {str(e)}")
        
        return redirect('delete_late_fees')
    
    # For GET request, just render the template
    return render(request, 'crucial/finance/delete_late_fees.html')

import zipfile
import logging

def admit_card_download_view(request):
    # Handle download requests first
    if 'download_type' in request.GET:
        return handle_admit_card_download(request)
    
    # Show filter form
    classes = StudentClass.objects.all()
    sections = StudentSection.objects.all()
    groups = StuGroup.objects.all()
    shifts = StudentShift.objects.all()
    exams = Examname.objects.all() 
    academic_sessions = AcademicSession.objects.all() 
    
    context = {
        'classes': classes,
        'sections': sections,
        'groups': groups,
        'shifts': shifts,
        'exams': exams,
        'academic_sessions': academic_sessions,
    }
    return render(request, 'admit_card/download_form.html', context)


# def handle_admit_card_download(request):
#     download_type = request.GET.get('download_type')
#     exam_id = request.GET.get('exam_id')
    
#     exam = None
#     if exam_id:
#         try:
#             exam = Examname.objects.get(id=exam_id)
#         except Examname.DoesNotExist:
#             pass
    
#     if download_type == 'search_by_roll':
#         roll_no = request.GET.get('roll_no')
#         class_id = request.GET.get('class_id')

#         if not roll_no or not class_id:
#             return JsonResponse({'error': 'Roll number and class are required for search'}, status=400)

#         try:
#             student = StudentProfile.objects.get(
#                 roll_no=roll_no,
#                 class_id__class_group_id__class_id=class_id
#             )

#             unpaid_fees = Fees.objects.filter(
#                 student_id=student,
#                 status='unpaid'
#             ).exists()

#             if unpaid_fees:
#                 return JsonResponse({'error': 'Student has unpaid dues'}, status=400)

#             return generate_admit_card_pdf(student)

#         except StudentProfile.DoesNotExist:
#             return JsonResponse({'error': 'Student not found'}, status=404)

#     elif download_type == 'list':
#         class_id = request.GET.get('class_id')
#         if not class_id:
#             return JsonResponse({'error': 'Class is required'}, status=400)

#         filters = {
#             'class_id__class_group_id__class_id': class_id
#         }

#         section_id = request.GET.get('section_id')
#         if section_id:
#             filters['class_id__section_id'] = section_id

#         group_id = request.GET.get('group_id')
#         if group_id:
#             filters['class_id__class_group_id__group_id'] = group_id

#         shift_id = request.GET.get('shift_id')
#         if shift_id:
#             filters['class_id__shift_id'] = shift_id

#         version = request.GET.get('version')
#         if version:
#             filters['version'] = version

#         students = StudentProfile.objects.filter(
#             **filters
#         ).select_related(
#             'student_field',
#             'class_id',
#             'class_id__class_group_id',
#             'class_id__section_id',
#             'class_id__shift_id'
#         )

#         eligible_students = []
#         for student in students:
#             unpaid_fees = Fees.objects.filter(
#                 student_id=student,
#                 status='unpaid'
#             ).exists()

#             if not unpaid_fees:
#                 eligible_students.append({
#                     'id': student.id,
#                     'name': student.student_field.name,
#                     'roll_no': student.roll_no,
#                     'class_name': student.class_id.class_group_id.class_id.name,
#                     'section_name': student.class_id.section_id.name if student.class_id.section_id else 'N/A',
#                 })

#         return JsonResponse({'students': eligible_students})

#     elif download_type == 'single':
#         student_id = request.GET.get('student_id')
#         student = get_object_or_404(StudentProfile, id=student_id)

#         unpaid_fees = Fees.objects.filter(
#             student_id=student,
#             status='unpaid'
#         ).exists()

#         if unpaid_fees:
#             return JsonResponse({'error': 'Student has unpaid dues'}, status=400)

#         return generate_admit_card_pdf(student)

#     elif download_type == 'bulk':
#         class_id = request.GET.get('class_id')
#         if not class_id:
#             return JsonResponse({'error': 'Class is required'}, status=400)

#         filters = {
#             'class_id__class_group_id__class_id': class_id
#         }

#         section_id = request.GET.get('section_id')
#         if section_id:
#             filters['class_id__section_id'] = section_id

#         group_id = request.GET.get('group_id')
#         if group_id:
#             filters['class_id__class_group_id__group_id'] = group_id

#         shift_id = request.GET.get('shift_id')
#         if shift_id:
#             filters['class_id__shift_id'] = shift_id

#         version = request.GET.get('version')
#         if version:
#             filters['version'] = version

#         students = StudentProfile.objects.filter(
#             **filters
#         ).select_related('student_field')

#         eligible_students = [
#             student for student in students
#             if not Fees.objects.filter(student_id=student, status='unpaid').exists()
#         ]

#         if not eligible_students:
#             return JsonResponse({'error': 'No eligible students found'}, status=400)

#         return generate_admit_cards_zip(eligible_students)

#     return JsonResponse({'error': 'Invalid download type'}, status=400)


def handle_admit_card_download(request):
    download_type = request.GET.get('download_type')
    exam_id = request.GET.get('exam_id')  # Get selected exam ID
    
    # Get exam object if provided
    exam = None
    if exam_id:
        try:
            exam = Examname.objects.get(id=exam_id)
        except Examname.DoesNotExist:
            pass
    
    if download_type == 'search_by_roll':
        roll_no = request.GET.get('roll_no')
        class_id = request.GET.get('class_id')

        if not roll_no or not class_id:
            return JsonResponse({'error': 'Roll number and class are required for search'}, status=400)

        try:
            student = StudentProfile.objects.get(
                roll_no=roll_no,
                class_id__class_group_id__class_id=class_id
            )

            unpaid_fees = Fees.objects.filter(
                student_id=student,
                status='unpaid'
            ).exists()

            if unpaid_fees:
                return JsonResponse({'error': 'Student has unpaid dues'}, status=400)

            return generate_admit_card_pdf(student, exam)

        except StudentProfile.DoesNotExist:
            return JsonResponse({'error': 'Student not found'}, status=404)

    elif download_type == 'list':
        class_id = request.GET.get('class_id')
        if not class_id:
            return JsonResponse({'error': 'Class is required'}, status=400)

        filters = {
            'class_id__class_group_id__class_id': class_id
        }

        section_id = request.GET.get('section_id')
        if section_id:
            filters['class_id__section_id'] = section_id

        group_id = request.GET.get('group_id')
        if group_id:
            filters['class_id__class_group_id__group_id'] = group_id

        shift_id = request.GET.get('shift_id')
        if shift_id:
            filters['class_id__shift_id'] = shift_id

        version = request.GET.get('version')
        if version:
            filters['version'] = version

        students = StudentProfile.objects.filter(
            **filters
        ).select_related(
            'student_field',
            'class_id',
            'class_id__class_group_id',
            'class_id__section_id',
            'class_id__shift_id'
        )

        eligible_students = []
        for student in students:
            unpaid_fees = Fees.objects.filter(
                student_id=student,
                status='unpaid'
            ).exists()

            if not unpaid_fees:
                eligible_students.append({
                    'id': student.id,
                    'name': student.student_field.name,
                    'roll_no': student.roll_no,
                    'class_name': student.class_id.class_group_id.class_id.name,
                    'section_name': student.class_id.section_id.name if student.class_id.section_id else 'N/A',
                })

        return JsonResponse({'students': eligible_students})

    elif download_type == 'single':
        student_id = request.GET.get('student_id')
        student = get_object_or_404(StudentProfile, id=student_id)

        unpaid_fees = Fees.objects.filter(
            student_id=student,
            status='unpaid'
        ).exists()

        if unpaid_fees:
            return JsonResponse({'error': 'Student has unpaid dues'}, status=400)

        return generate_admit_card_pdf(student, exam)

    elif download_type == 'bulk':
        class_id = request.GET.get('class_id')
        if not class_id:
            return JsonResponse({'error': 'Class is required'}, status=400)

        filters = {
            'class_id__class_group_id__class_id': class_id
        }

        section_id = request.GET.get('section_id')
        if section_id:
            filters['class_id__section_id'] = section_id

        group_id = request.GET.get('group_id')
        if group_id:
            filters['class_id__class_group_id__group_id'] = group_id

        shift_id = request.GET.get('shift_id')
        if shift_id:
            filters['class_id__shift_id'] = shift_id

        version = request.GET.get('version')
        if version:
            filters['version'] = version

        students = StudentProfile.objects.filter(
            **filters
        ).select_related('student_field')

        eligible_students = [
            student for student in students
            if not Fees.objects.filter(student_id=student, status='unpaid').exists()
        ]

        if not eligible_students:
            return JsonResponse({'error': 'No eligible students found'}, status=400)

        return generate_admit_cards_zip(eligible_students, exam)

    return JsonResponse({'error': 'Invalid download type'}, status=400)


# def generate_admit_card_pdf(student):
#     template_path = 'admit_card/admit_card_template.html'
#     academic_session = student.academic_session_year
#     context = {
#         'student': student,
#         'exam_name': exam.name if exam else "PreTest",  # Use selected exam or default
#         'academic_session': academic_session,
#         'college_name': 'Rajuk Uttara Model College',
#         'college_address': 'Sector #6, Uttara Model Town, Dhaka-1230',
#         # 'static_path': os.path.join(settings.BASE_DIR, 'static')
#         'MEDIA_ROOT': settings.MEDIA_ROOT,  # Add this
#         'STATIC_ROOT': settings.STATIC_ROOT,
#     }
    
#     try:
#         template = get_template(template_path)
#         html = template.render(context)
        
#         response = HttpResponse(content_type='application/pdf')
#         filename = f"admit_card_{student.student_field.name.replace(' ', '_')}.pdf"
#         response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
#         # Generate PDF directly to the response
#         pisa_status = pisa.CreatePDF(
#             html, 
#             dest=response,
#             encoding='UTF-8'
#         )
        
#         if pisa_status.err:
#             logger.error(f"Error generating PDF: {pisa_status.err}")
#             return HttpResponse('Error generating PDF', status=500)
        
#         return response
    
#     except Exception as e:
#         logger.error(f"Error in generate_admit_card_pdf: {str(e)}")
#         return HttpResponse(f'Error generating PDF: {str(e)}', status=500)

def generate_admit_card_pdf(student, exam=None):
    template_path = 'admit_card/admit_card_template.html'
    
    # Get academic session from student's database record
    academic_session = student.academic_session_year
    
    context = {
        'student': student,
        'exam_name': exam.name if exam else "Test",  # Use selected exam or default
        'academic_session': academic_session,  # Use academic session from database
        'college_name': 'Rajuk Uttara Model College',
        'college_address': 'Sector #6, Uttara Model Town, Dhaka-1230',
        'MEDIA_ROOT': settings.MEDIA_ROOT,
        'STATIC_ROOT': settings.STATIC_ROOT,
    }
    
    try:
        template = get_template(template_path)
        html = template.render(context)
        
        response = HttpResponse(content_type='application/pdf')
        filename = f"admit_card_{student.student_field.name.replace(' ', '_')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Generate PDF directly to the response
        pisa_status = pisa.CreatePDF(
            html, 
            dest=response,
            encoding='UTF-8'
        )
        
        if pisa_status.err:
            logger.error(f"Error generating PDF: {pisa_status.err}")
            return HttpResponse('Error generating PDF', status=500)
        
        return response
    
    except Exception as e:
        logger.error(f"Error in generate_admit_card_pdf: {str(e)}")
        return HttpResponse(f'Error generating PDF: {str(e)}', status=500)


# def generate_admit_cards_zip(students):
#     import zipfile
#     from io import BytesIO
    
#     try:
#         # Create in-memory zip file
#         zip_buffer = BytesIO()
        
#         with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
#             for student in students:
#                 # Generate PDF content directly
#                 template_path = 'admit_card/admit_card_template.html'
#                 context = {
#                     'student': student,
#                     'college_name': 'Rajuk Uttara Model College',
#                     'college_address': 'Sector #6, Uttara Model Town, Dhaka-1230'
#                 }
                
#                 template = get_template(template_path)
#                 html = template.render(context)
                
#                 pdf_buffer = BytesIO()
#                 pisa_status = pisa.CreatePDF(
#                     BytesIO(html.encode("UTF-8")), 
#                     dest=pdf_buffer
#                 )
                
#                 if not pisa_status.err:
#                     filename = f"admit_card_{student.student_field.name.replace(' ', '_')}.pdf"
#                     zip_file.writestr(filename, pdf_buffer.getvalue())
        
#         zip_buffer.seek(0)
#         response = HttpResponse(zip_buffer, content_type='application/zip')
#         response['Content-Disposition'] = 'attachment; filename="admit_cards.zip"'
#         return response
        
#     except Exception as e:
#         logger.error(f"Error in generate_admit_cards_zip: {str(e)}")
#         return HttpResponse(f'Error generating ZIP file: {str(e)}', status=500)
    
def generate_admit_cards_zip(students, exam=None):
    import zipfile
    from io import BytesIO
    
    try:
        # Create in-memory zip file
        zip_buffer = BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for student in students:
                # Get academic session from student's database record
                academic_session = student.academic_session_year
                
                # Generate PDF content directly
                template_path = 'admit_card/admit_card_template.html'
                context = {
                    'student': student,
                    'exam_name': exam.name if exam else "Test",  # Use selected exam or default
                    'academic_session': academic_session,  # Use academic session from database
                    'college_name': 'Rajuk Uttara Model College',
                    'college_address': 'Sector #6, Uttara Model Town, Dhaka-1230'
                }
                
                template = get_template(template_path)
                html = template.render(context)
                
                pdf_buffer = BytesIO()
                pisa_status = pisa.CreatePDF(
                    BytesIO(html.encode("UTF-8")), 
                    dest=pdf_buffer
                )
                
                if not pisa_status.err:
                    filename = f"admit_card_{student.student_field.name.replace(' ', '_')}.pdf"
                    zip_file.writestr(filename, pdf_buffer.getvalue())
        
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="admit_cards.zip"'
        return response
        
    except Exception as e:
        logger.error(f"Error in generate_admit_cards_zip: {str(e)}")
        return HttpResponse(f'Error generating ZIP file: {str(e)}', status=500)
 
    
# from django.shortcuts import render, redirect, get_object_or_404
# from django.http import HttpResponse, HttpResponseNotFound
# from django.db.models import Q
# from io import BytesIO
# import zipfile
# from weasyprint import HTML
# from django.template.loader import render_to_string

# def check_fee_cleared(student, academic_session=None):
#     """Check if student has cleared fees for the given academic session"""
#     try:
#         # If academic session is provided, get the admission year
#         if academic_session:
#             admission_year = Admission_Year.objects.get(name=academic_session.start_year)
#             fee_records = Fees.objects.filter(
#                 student_id=student,
#                 academic_year=admission_year
#             )
#         else:
#             # Check all fee records if no specific session is provided
#             fee_records = Fees.objects.filter(student_id=student)
        
#         # Check if all fees are paid
#         for fee in fee_records:
#             if fee.status != 'paid':
#                 return False
#         return True
    
#     except (Admission_Year.DoesNotExist, Fees.DoesNotExist):
#         return False

# def generate_admit_card_pdf(student, academic_session=None):
#     institute = Institute.objects.order_by('-id').first()
#     institute_logo = institute.institute_logo.url if institute and institute.institute_logo else None
    
#     context = {
#         'student': student,
#         'academic_session': academic_session,
#         'institute_logo': institute_logo 
#     }
#     html_string = render_to_string('admit_card/admit_card_template.html', context)
#     html = HTML(string=html_string)
#     return html.write_pdf()

# def bulk_admit_card_download(request):
#     if request.method == 'POST':
#         # Get filter parameters
#         # class_id = request.POST.get('class')
#         class_group_id = request.POST.get('class_group')
#         section_id = request.POST.get('section')
#         group_id = request.POST.get('group')
#         version = request.POST.get('version')
#         shift_id = request.POST.get('shift')
#         academic_session_id = request.POST.get('academic_session')
        
#         # Start with base queryset
#         # students = StudentProfile.objects.filter(class_id=class_id)
#         students = StudentProfile.objects.filter(
#             class_id__class_group_id=class_group_id
#         )
        
#         # Apply filters
#         if section_id and section_id != 'all':
#             students = students.filter(class_id__section_id=section_id)
#         if group_id and group_id != 'all':
#             students = students.filter(class_id__class_group_id__group_id=group_id)
#         if version and version != 'all':
#             students = students.filter(version=version)
#         if shift_id and shift_id != 'all':
#             students = students.filter(class_id__shift_id=shift_id)
        
#         # Handle academic session
#         academic_session = None
#         if academic_session_id and academic_session_id != 'all':
#             academic_session = AcademicSession.objects.get(id=academic_session_id)

#         # Create zip file with admit cards
#         zip_buffer = BytesIO()
#         with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
#             count = 0
#             for student in students:
#                 # Check fee status
#                 if check_fee_cleared(student, academic_session):
#                     try:
#                         pdf_bytes = generate_admit_card_pdf(student, academic_session)
#                         filename = f"admit_card_{student.student_field.name}_{student.roll_no}.pdf"
#                         zip_file.writestr(filename, pdf_bytes)
#                         count += 1
#                     except Exception as e:
#                         # Log error but continue processing other students
#                         print(f"Error generating admit card for {student}: {str(e)}")
        
#         if count == 0:
#             return HttpResponse("No students with cleared fees found for the selected criteria", status=400)
        
#         zip_buffer.seek(0)
#         response = HttpResponse(zip_buffer, content_type='application/zip')
#         response['Content-Disposition'] = 'attachment; filename="admit_cards.zip"'
#         return response

#     # GET request - show filter form
#     # classes = ClassConfig.objects.all()
#     class_groups = ClassGroupConfig.objects.distinct().select_related('class_id', 'group_id')
#     sections = StudentSection.objects.all()
#     groups = StuGroup.objects.all()
#     shifts = StudentShift.objects.all()
#     academic_sessions = AcademicSession.objects.all()
    
#     context = {
#         'class_groups': class_groups,
#         'sections': sections,
#         'groups': groups,
#         'shifts': shifts,
#         'academic_sessions': academic_sessions,
#     }
#     return render(request, 'admit_card/bulk_admit_card_form.html', context)

# def single_admit_card_download(request):
#     if request.method == 'POST':
#         search_term = request.POST.get('search_term')
#         academic_session_id = request.POST.get('academic_session')
        
#         # Handle academic session
#         academic_session = None
#         if academic_session_id and academic_session_id != 'all':
#             academic_session = AcademicSession.objects.get(id=academic_session_id)
        
#         # Get matching students
#         students = StudentProfile.objects.filter(
#             Q(roll_no__icontains=search_term) | 
#             Q(student_field__name__icontains=search_term)
#         ).distinct()
        
#         # Handle cases
#         if not students.exists():
#             return HttpResponse("No students found", status=404)
            
#         if students.count() == 1:
#             student = students.first()
#             if check_fee_cleared(student, academic_session):
#                 try:
#                     pdf_bytes = generate_admit_card_pdf(student, academic_session)
#                     response = HttpResponse(pdf_bytes, content_type='application/pdf')
#                     filename = f"admit_card_{student.student_field.name}.pdf"
#                     response['Content-Disposition'] = f'attachment; filename="{filename}"'
#                     return response
#                 except Exception as e:
#                     return HttpResponse(f"Error generating PDF: {str(e)}", status=500)
#             else:
#                 return HttpResponse("Fee not cleared for this student", status=400)
#         else:
#             # Show list of matching students
#             context = {
#                 'students': students,
#                 'academic_session': academic_session,
#                 'search_term': search_term
#             }
#             return render(request, 'admit_card/student_list.html', context)
    
#     # GET request - show search form
#     academic_sessions = AcademicSession.objects.all()
#     return render(request, 'admit_card/single_admit_card_form.html', {'academic_sessions': academic_sessions})

# def download_student_admit_card(request, student_id):
#     academic_session_id = request.GET.get('academic_session')
#     student = get_object_or_404(StudentProfile, id=student_id)
    
#     # Handle academic session
#     academic_session = None
#     if academic_session_id and academic_session_id != 'all':
#         academic_session = AcademicSession.objects.get(id=academic_session_id)
    
#     if check_fee_cleared(student, academic_session):
#         try:
#             pdf_bytes = generate_admit_card_pdf(student, academic_session)
#             response = HttpResponse(pdf_bytes, content_type='application/pdf')
#             filename = f"admit_card_{student.student_field.name}.pdf"
#             response['Content-Disposition'] = f'attachment; filename="{filename}"'
#             return response
#         except Exception as e:
#             return HttpResponse(f"Error generating PDF: {str(e)}", status=500)
#     else:
#         return HttpResponse("Fee not cleared for this student", status=400)
    
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseNotFound
from django.db.models import Q
from io import BytesIO
import zipfile
from weasyprint import HTML
from django.template.loader import render_to_string
from django.contrib import messages

def check_fee_cleared(student, academic_session=None):
    """Check if student has cleared fees for the given academic session"""
    try:
        # Use the student's academic_session_year from database
        if student.academic_session_year:
            admission_year = Admission_Year.objects.get(name=student.academic_session_year.start_year)
            fee_records = Fees.objects.filter(
                student_id=student,
                academic_year=admission_year
            )
        else:
            # Check all fee records if no specific session is provided
            fee_records = Fees.objects.filter(student_id=student)
        
        # Check if all fees are paid
        for fee in fee_records:
            if fee.status != 'paid':
                return False
        return True
    
    except (Admission_Year.DoesNotExist, Fees.DoesNotExist):
        return False

def generate_admit_card_pdf(student, exam=None):
    # Use the student's academic_session_year from database
    academic_session = student.academic_session_year
    institute = Institute.objects.order_by('-id').first()
    institute_logo = institute.institute_logo.path if institute and institute.institute_logo else None
    
    context = {
        'student': student,
        'academic_session': academic_session,
        'exam': exam,  # Pass the exam to template
        'institute_logo': institute_logo,
        'institute_name': institute.institute_name if institute else '',
        'institute_address': institute.institute_address if institute else '',
    }
    html_string = render_to_string('admit_card/admit_card_template.html', context)
    html = HTML(string=html_string)
    return html.write_pdf()

def bulk_admit_card_download(request):
    if request.method == 'POST':
        # Get filter parameters
        class_group_id = request.POST.get('class_group')
        section_id = request.POST.get('section')
        group_id = request.POST.get('group')
        version = request.POST.get('version')
        shift_id = request.POST.get('shift')
        exam_id = request.POST.get('exam')
        academic_session_id = request.POST.get('academic_session')
        
        # Validate required fields
        if not class_group_id:
            messages.error(request, "Class selection is required.")
            return redirect('bulk_admit_card_download')
        
        if not exam_id:
            messages.error(request, "Exam selection is required.")
            return redirect('bulk_admit_card_download')
        
        # Start with base queryset
        students = StudentProfile.objects.filter(
            class_id__class_group_id=class_group_id
        )
        
        # Apply filters
        if section_id and section_id != 'all':
            students = students.filter(class_id__section_id=section_id)
        if group_id and group_id != 'all':
            students = students.filter(class_id__class_group_id__group_id=group_id)
        if version and version != 'all':
            students = students.filter(version=version)
        if shift_id and shift_id != 'all':
            students = students.filter(class_id__shift_id=shift_id)
        
        # Apply academic session filter
        if academic_session_id and academic_session_id != 'all':
            students = students.filter(academic_session_year_id=academic_session_id)
        
        # Get exam object
        exam = None
        if exam_id and exam_id != 'all':
            try:
                exam = Examname.objects.get(id=exam_id)
            except Examname.DoesNotExist:
                messages.error(request, "Selected exam does not exist.")
                return redirect('bulk_admit_card_download')

        # Check if any students match the criteria
        if not students.exists():
            messages.error(request, "No students found matching the selected criteria.")
            return redirect('bulk_admit_card_download')

        # Create zip file with admit cards
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            count = 0
            for student in students:
                # Check fee status
                if check_fee_cleared(student):
                    try:
                        pdf_bytes = generate_admit_card_pdf(student, exam)
                        filename = f"admit_card_{student.student_field.name}_{student.roll_no}.pdf"
                        zip_file.writestr(filename, pdf_bytes)
                        count += 1
                    except Exception as e:
                        # Log error but continue processing other students
                        print(f"Error generating admit card for {student}: {str(e)}")
        
        if count == 0:
            messages.error(request, "No students with cleared fees found for the selected criteria.")
            return redirect('bulk_admit_card_download')
        
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="admit_cards.zip"'
        return response

    # GET request - show filter form
    class_groups = ClassGroupConfig.objects.distinct().select_related('class_id', 'group_id')
    sections = StudentSection.objects.all()
    groups = StuGroup.objects.all()
    shifts = StudentShift.objects.all()
    exams = Examname.objects.all()
    academic_sessions = AcademicSession.objects.all()
    institute = Institute.objects.order_by('-id').first()
    
    context = {
        'class_groups': class_groups,
        'sections': sections,
        'groups': groups,
        'shifts': shifts,
        'exams': exams,
        'academic_sessions': academic_sessions,
        'institute_name': institute.institute_name,
        'institute_address': institute.institute_address,
    }
    return render(request, 'admit_card/bulk_admit_card_form.html', context)

def single_admit_card_download(request):
    if request.method == 'POST':
        search_term = request.POST.get('search_term')
        exam_id = request.POST.get('exam')
        academic_session_id = request.POST.get('academic_session')
        
        # Validate required fields
        if not search_term:
            messages.error(request, "Search term is required.")
            return redirect('single_admit_card_download')
        
        if not exam_id:
            messages.error(request, "Exam selection is required.")
            return redirect('single_admit_card_download')
        
        # Get exam object
        exam = None
        if exam_id and exam_id != 'all':
            try:
                exam = Examname.objects.get(id=exam_id)
            except Examname.DoesNotExist:
                messages.error(request, "Selected exam does not exist.")
                return redirect('single_admit_card_download')
        
        # Get matching students
        students = StudentProfile.objects.filter(
            Q(roll_no__icontains=search_term) | 
            Q(student_field__name__icontains=search_term)
        )
        
        # Apply academic session filter
        if academic_session_id and academic_session_id != 'all':
            students = students.filter(academic_session_year_id=academic_session_id)
        
        # Handle cases
        if not students.exists():
            messages.error(request, "No students found matching your search.")
            return redirect('single_admit_card_download')
            
        if students.count() == 1:
            student = students.first()
            if check_fee_cleared(student):
                try:
                    pdf_bytes = generate_admit_card_pdf(student, exam)
                    response = HttpResponse(pdf_bytes, content_type='application/pdf')
                    filename = f"admit_card_{student.student_field.name}.pdf"
                    response['Content-Disposition'] = f'attachment; filename="{filename}"'
                    return response
                except Exception as e:
                    messages.error(request, f"Error generating PDF: {str(e)}")
                    return redirect('single_admit_card_download')
            else:
                messages.error(request, "Fee not cleared for this student.")
                return redirect('single_admit_card_download')
        else:
            # Show list of matching students
            context = {
                'students': students,
                'exam': exam,
                'search_term': search_term
            }
            return render(request, 'admit_card/student_list.html', context)
    
    # GET request - show search form
    exams = Examname.objects.all()
    academic_sessions = AcademicSession.objects.all()
    
    return render(request, 'admit_card/single_admit_card_form.html', {
        'exams': exams,
        'academic_sessions': academic_sessions
    })

def download_student_admit_card(request, student_id):
    exam_id = request.GET.get('exam')
    
    # Get exam object
    exam = None
    if exam_id and exam_id != 'all':
        try:
            exam = Examname.objects.get(id=exam_id)
        except Examname.DoesNotExist:
            return HttpResponse("Selected exam does not exist.", status=400)
    
    student = get_object_or_404(StudentProfile, id=student_id)
    
    if check_fee_cleared(student):
        try:
            pdf_bytes = generate_admit_card_pdf(student, exam)
            response = HttpResponse(pdf_bytes, content_type='application/pdf')
            filename = f"admit_card_{student.student_field.name}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except Exception as e:
            return HttpResponse(f"Error generating PDF: {str(e)}", status=500)
    else:
        return HttpResponse("Fee not cleared for this student.", status=400)
    
     
# crucial/finance/views.py
import io
from django.http import HttpResponse
from django.shortcuts import render
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from .models import Fees, FeeHead
from .forms import FeeReportForm
from django.db.models import F, Sum
from collections import defaultdict

def fee_report(request):
    if request.method == 'POST':
        form = FeeReportForm(request.POST)
        if form.is_valid():
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            format = form.cleaned_data['format']
            
            # Get all paid fees in date range
            fees = Fees.objects.filter(
                status='paid',
                updated_at__date__range=(start_date, end_date)
            ).select_related(
                'student_id__student_field',
                'feetype_id__fees_type__fee_head'
            ).annotate(
                student_name=F('student_id__student_field__name'),
                roll_no=F('student_id__roll_no')
            )
            
            # Create a dictionary to organize data
            report_data = defaultdict(lambda: {
                'student_name': '',
                'roll_no': '',
                'fee_heads': defaultdict(int)
            })
            
            # Collect all unique fee heads
            fee_heads = set()
            
            # Process each fee entry
            for fee in fees:
                student_key = fee.student_id_id
                fee_head_name = fee.feetype_id.fees_type.fee_head.name
                
                # Update student data
                report_data[student_key]['student_name'] = fee.student_name
                report_data[student_key]['roll_no'] = fee.roll_no
                
                # Add amount to fee head
                report_data[student_key]['fee_heads'][fee_head_name] += fee.amount
                
                # Track unique fee heads
                fee_heads.add(fee_head_name)
            
            # Convert to list and sort fee heads
            fee_heads = sorted(fee_heads)
            data_list = list(report_data.values())
            
            if format == 'pdf':
                return generate_pdf(data_list, fee_heads, start_date, end_date)
            else:
                return generate_excel(data_list, fee_heads, start_date, end_date)
    else:
        form = FeeReportForm()
    
    return render(request, 'crucial/finance/fee_report.html', {'form': form})

def generate_pdf(data, fee_heads, start_date, end_date):
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer)
    
    # Add header
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 800, "Fee Report Summary")
    p.setFont("Helvetica", 12)
    p.drawString(100, 780, f"Date Range: {start_date} to {end_date}")
    
    # Table setup
    col_width = 100
    x_start = 50
    y_start = 750
    row_height = 20
    
    # Table headers
    p.setFont("Helvetica-Bold", 10)
    headers = ["Student Name", "Roll No"] + fee_heads
    x = x_start
    
    # Draw headers
    for header in headers:
        p.drawString(x, y_start, header[:15])  # Limit header length
        x += col_width
    
    # Table data
    p.setFont("Helvetica", 9)
    y = y_start - row_height
    
    for row in data:
        x = x_start
        p.drawString(x, y, row['student_name'][:15])  # Limit name length
        x += col_width
        
        p.drawString(x, y, str(row['roll_no']))
        x += col_width
        
        for head in fee_heads:
            amount = row['fee_heads'].get(head, 0)
            p.drawString(x, y, str(amount))
            x += col_width
        
        y -= row_height
        # New page if needed
        if y < 50:
            p.showPage()
            y = 750
            p.setFont("Helvetica", 9)
    
    p.showPage()
    p.save()
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="fee_report_{start_date}_to_{end_date}.pdf"'
    return response

def generate_excel(data, fee_heads, start_date, end_date):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fee Report"
    
    # Add header
    ws.append(["Fee Report Summary"])
    ws.append(["Date Range:", f"{start_date} to {end_date}"])
    ws.append([])  # Empty row
    
    # Table headers
    headers = ["Student Name", "Roll Number"] + fee_heads
    ws.append(headers)
    
    # Table data
    for row in data:
        data_row = [
            row['student_name'],
            row['roll_no']
        ]
        for head in fee_heads:
            data_row.append(row['fee_heads'].get(head, 0))
        ws.append(data_row)
    
    # Add totals row
    ws.append([])
    totals_row = ["TOTAL", ""] + [f"=SUM({chr(67+i)}3:{chr(67+i)}{len(data)+3})" 
                                 for i in range(len(fee_heads))]
    ws.append(totals_row)
    
    # Auto-size columns
    for i, column in enumerate(ws.columns):
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="fee_report_{start_date}_to_{end_date}.xlsx"'
    return response


# import openpyxl
# from decimal import Decimal
# from django.shortcuts import render, redirect
# from django.contrib import messages
# from django.utils import timezone
# from datetime import datetime, time, date
# from crucial.models import Fees, StudentProfile, FeeHead, Feetype, Fee_month, Fees_name

# def fee_bulk_upload(request):
#     error_details = []  # Store detailed error messages
#     success_count = 0
    
#     if request.method == 'POST' and request.FILES.get('excel_file'):
#         excel_file = request.FILES['excel_file']
        
#         try:
#             wb = openpyxl.load_workbook(excel_file, data_only=True)
#             sheet = wb.active
            
#             for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
#                 try:
#                     # Skip empty rows
#                     if not any(row):
#                         continue
                    
#                     # Validate minimum columns
#                     if len(row) < 6:
#                         raise ValueError("Insufficient data (need 6 columns)")
                    
#                     # Unpack row data
#                     student_name = row[0] or ""
#                     roll_no = row[1]
#                     transaction_id = row[2] or ""
#                     amount = row[3]
#                     date_value = row[4]
#                     time_value = row[5] or "00:00:00"
                    
#                     # Validate required fields
#                     if not all([roll_no, amount, date_value]):
#                         raise ValueError("Missing required data (Roll, Amount, or Date)")
                    
#                     # Convert roll_no to integer
#                     try:
#                         roll_no = int(roll_no)
#                     except (TypeError, ValueError):
#                         raise ValueError(f"Invalid roll number: {roll_no}")
                    
#                     # Convert amount to Decimal
#                     try:
#                         amount = Decimal(str(amount))
#                     except:
#                         raise ValueError(f"Invalid amount: {amount}")
                    
#                     # Handle different date/time types from Excel
#                     if isinstance(date_value, datetime):
#                         # Already a datetime object from Excel
#                         excel_dt = date_value
#                     elif isinstance(date_value, date):
#                         # Date object - combine with time
#                         if isinstance(time_value, time):
#                             excel_dt = datetime.combine(date_value, time_value)
#                         elif isinstance(time_value, str):
#                             # Try to parse time string
#                             try:
#                                 time_obj = datetime.strptime(time_value, "%H:%M:%S").time()
#                                 excel_dt = datetime.combine(date_value, time_obj)
#                             except:
#                                 excel_dt = datetime.combine(date_value, time(0, 0))
#                         else:
#                             excel_dt = datetime.combine(date_value, time(0, 0))
#                     elif isinstance(date_value, str):
#                         # String date - parse with time
#                         try:
#                             excel_dt = datetime.strptime(f"{date_value} {time_value}", "%m/%d/%y %H:%M:%S")
#                         except:
#                             try:
#                                 excel_dt = datetime.strptime(f"{date_value} {time_value}", "%Y-%m-%d %H:%M:%S")
#                             except:
#                                 try:
#                                     excel_dt = datetime.strptime(date_value, "%m/%d/%y")
#                                 except:
#                                     raise ValueError(f"Unsupported date format: {date_value}")
#                     else:
#                         raise ValueError(f"Unsupported date type: {type(date_value)}")
                    
#                     aware_dt = timezone.make_aware(excel_dt)
                    
#                     # Determine fee head
#                     if amount == Decimal('50.00'):
#                         head_str = "Fine for Delayed Tuition Fee Payment"
#                     else:
#                         head_str = "Re-admission Fee (Late Pay)"
                    
#                     # Get related objects
#                     try:
#                         student = StudentProfile.objects.get(roll_no=roll_no)
#                     except StudentProfile.DoesNotExist:
#                         raise ValueError(f"Student with roll {roll_no} not found")
                    
#                     try:
#                         fee_head = FeeHead.objects.get(name=head_str)
#                     except FeeHead.DoesNotExist:
#                         raise ValueError(f"Fee head '{head_str}' not found. Please create it first.")
                    
#                     # Get active Feetype
#                     feetype = Feetype.objects.filter(
#                         fee_head=fee_head,
#                         status='Active'
#                     ).first()
                    
#                     if not feetype:
#                         raise ValueError(f"No active Feetype found for '{head_str}'")
                    
#                     month_name = aware_dt.strftime("%B")
#                     try:
#                         fee_month = Fee_month.objects.get(name=month_name)
#                     except Fee_month.DoesNotExist:
#                         raise ValueError(f"Month '{month_name}' not configured")
                    
#                     # Create Fees_name if needed
#                     fees_title = f"{month_name} {head_str}"
#                     fees_name, created = Fees_name.objects.get_or_create(
#                         fees_title=fees_title,
#                         defaults={
#                             'fees_type': feetype,
#                             'month': fee_month,
#                             'academic_year': student.admission_year_id,
#                             'created_by': request.user,
#                             'updated_by': request.user
#                         }
#                     )
                    
#                     # Step 1: Create the fee record (auto fields will be set to current time)
#                     fee = Fees.objects.create(
#                         student_id=student,
#                         amount=amount,
#                         status='paid',
#                         month_id=fee_month,
#                         feetype_id=fees_name,
#                         description=f"Bulk import: {student_name}",
#                         academic_year=student.admission_year_id,
#                         created_by=request.user,
#                         updated_by=request.user
#                     )
                    
#                     # Step 2: Update datetime fields directly in database (bypasses auto_now)
#                     Fees.objects.filter(pk=fee.pk).update(
#                         created_at=aware_dt,
#                         updated_at=aware_dt
#                     )
                    
#                     success_count += 1
                    
#                 except Exception as e:
#                     error_details.append(f"Row {row_idx}: {str(e)}")
            
#             # Show result messages
#             if success_count:
#                 messages.success(request, f"Successfully imported {success_count} records")
            
#             if error_details:
#                 # Combine all errors into one message
#                 error_message = f"Completed with {len(error_details)} errors:<br>" + "<br>".join(error_details[:10])  # Show first 10 errors
#                 if len(error_details) > 10:
#                     error_message += f"<br>...and {len(error_details) - 10} more errors"
#                 messages.error(request, error_message)
        
#         except Exception as e:
#             messages.error(request, f"File processing error: {str(e)}")
    
#     return render(request, 'crucial/finance/fee_bulk_upload.html', {'error_details': error_details})

import openpyxl
from decimal import Decimal
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from datetime import datetime, time, date
from crucial.models import Fees, StudentProfile, FeeHead, Feetype, Fee_month, Fees_name

def fee_bulk_upload(request):
    error_details = []  # Store detailed error messages
    success_count = 0
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Skip empty rows
                    if not any(row):
                        continue
                    
                    # Validate minimum columns
                    if len(row) < 6:
                        raise ValueError("Insufficient data (need 6 columns)")
                    
                    # Unpack row data
                    student_name = row[0] or ""
                    roll_no = row[1]
                    transaction_id = row[2] or ""
                    amount = row[3]
                    date_value = row[4]
                    time_value = row[5] or "00:00:00"
                    
                    # Validate required fields
                    if not all([roll_no, amount, date_value]):
                        raise ValueError("Missing required data (Roll, Amount, or Date)")
                    
                    # Convert roll_no to integer
                    try:
                        roll_no = int(roll_no)
                    except (TypeError, ValueError):
                        raise ValueError(f"Invalid roll number: {roll_no}")
                    
                    # Convert amount to Decimal
                    try:
                        amount = Decimal(str(amount))
                    except:
                        raise ValueError(f"Invalid amount: {amount}")
                    
                    # Handle different date/time types from Excel
                    if isinstance(date_value, datetime):
                        # Already a datetime object from Excel - use as is
                        excel_dt = date_value
                        
                        # If we have a separate time value, combine it
                        if time_value and not isinstance(time_value, time):
                            try:
                                # Parse time string if needed
                                if isinstance(time_value, str):
                                    time_obj = datetime.strptime(time_value, "%H:%M:%S").time()
                                    excel_dt = datetime.combine(date_value.date(), time_obj)
                            except:
                                # If time parsing fails, use the time from the datetime object
                                pass
                    
                    elif isinstance(date_value, date):
                        # Date object - combine with time
                        if isinstance(time_value, time):
                            excel_dt = datetime.combine(date_value, time_value)
                        elif isinstance(time_value, str):
                            # Try to parse time string
                            try:
                                time_obj = datetime.strptime(time_value, "%H:%M:%S").time()
                                excel_dt = datetime.combine(date_value, time_obj)
                            except:
                                # Fallback to midnight if time parsing fails
                                excel_dt = datetime.combine(date_value, time(0, 0))
                        else:
                            excel_dt = datetime.combine(date_value, time(0, 0))
                    
                    elif isinstance(date_value, str):
                        # String date - parse with time
                        try:
                            # First try parsing with both date and time
                            excel_dt = datetime.strptime(f"{date_value} {time_value}", "%m/%d/%y %H:%M:%S")
                        except:
                            try:
                                # Try alternative datetime format
                                excel_dt = datetime.strptime(f"{date_value} {time_value}", "%Y-%m-%d %H:%M:%S")
                            except:
                                try:
                                    # Try just the date if time parsing fails
                                    excel_dt = datetime.strptime(date_value, "%m/%d/%y")
                                except:
                                    try:
                                        # Try different date format
                                        excel_dt = datetime.strptime(date_value, "%d/%m/%y")
                                    except:
                                        raise ValueError(f"Unsupported date format: {date_value}")
                    
                    else:
                        raise ValueError(f"Unsupported date type: {type(date_value)}")
                    
                    # Make datetime timezone-aware
                    aware_dt = timezone.make_aware(excel_dt)
                    
                    # Determine fee head
                    if amount == Decimal('50.00'):
                        head_str = "Fine for Delayed Tuition Fee Payment"
                    else:
                        head_str = "Re-admission Fee (Late Pay)"
                    
                    # Get related objects
                    try:
                        student = StudentProfile.objects.get(roll_no=roll_no)
                    except StudentProfile.DoesNotExist:
                        raise ValueError(f"Student with roll {roll_no} not found")
                    
                    try:
                        fee_head = FeeHead.objects.get(name=head_str)
                    except FeeHead.DoesNotExist:
                        raise ValueError(f"Fee head '{head_str}' not found. Please create it first.")
                    
                    # Get active Feetype
                    feetype = Feetype.objects.filter(
                        fee_head=fee_head,
                        status='Active'
                    ).first()
                    
                    if not feetype:
                        raise ValueError(f"No active Feetype found for '{head_str}'")
                    
                    month_name = aware_dt.strftime("%B")
                    try:
                        fee_month = Fee_month.objects.get(name=month_name)
                    except Fee_month.DoesNotExist:
                        raise ValueError(f"Month '{month_name}' not configured")
                    
                    # Create Fees_name if needed
                    fees_title = f"{month_name} {head_str}"
                    fees_name, created = Fees_name.objects.get_or_create(
                        fees_title=fees_title,
                        defaults={
                            'fees_type': feetype,
                            'month': fee_month,
                            'academic_year': student.admission_year_id,
                            'created_by': request.user,
                            'updated_by': request.user
                        }
                    )
                    
                    # Create fee record (auto fields will be set to current time)
                    fee = Fees.objects.create(
                        student_id=student,
                        amount=amount,
                        status='paid',
                        month_id=fee_month,
                        feetype_id=fees_name,
                        description=f"Bulk import: {student_name}",
                        academic_year=student.admission_year_id,
                        created_by=request.user,
                        updated_by=request.user
                    )
                    
                    # Update datetime fields directly in database (bypasses auto_now)
                    Fees.objects.filter(pk=fee.pk).update(
                        created_at=aware_dt,
                        updated_at=aware_dt
                    )
                    
                    success_count += 1
                    
                except Exception as e:
                    error_details.append(f"Row {row_idx}: {str(e)}")
            
            # Show result messages
            if success_count:
                messages.success(request, f"Successfully imported {success_count} records")
            
            if error_details:
                # Combine all errors into one message
                error_message = f"Completed with {len(error_details)} errors:<br>" + "<br>".join(error_details[:10])  # Show first 10 errors
                if len(error_details) > 10:
                    error_message += f"<br>...and {len(error_details) - 10} more errors"
                messages.error(request, error_message)
        
        except Exception as e:
            messages.error(request, f"File processing error: {str(e)}")
    
    return render(request, 'crucial/finance/fee_bulk_upload.html', {'error_details': error_details})

from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.conf import settings
from weasyprint import HTML
from datetime import datetime
import tempfile
from django.db.models import Q

def deposit_form_ui(request):
    student = None
    due_fees = []
    total_due = 0
    error_message = None
    
    if request.method == 'GET' and 'search_term' in request.GET:
        search_term = request.GET.get('search_term')
        try:
            # Search by either student name or student ID
            student = StudentProfile.objects.filter(
                Q(student_field__name__icontains=search_term) | 
                Q(student_field__user_id=search_term)
            ).first()
            
            if student:
                # Get all unpaid or partially paid fees
                due_fees = Fees.objects.filter(
                    student_id=student, 
                    status__in=['unpaid', 'partial']
                )
                
                # Calculate total due amount
                for fee in due_fees:
                    total_due += fee.total_fee_after_partial_payments()
            else:
                error_message = "Student not found with the given search term."
                
        except ValueError:
            error_message = "Invalid search term. Please enter a valid name or student ID."
    
    # Get institute details (assuming there's only one institute)
    institute = Institute.objects.first()
    
    context = {
        'student': student,
        'due_fees': due_fees,
        'total_due': total_due,
        'institute': institute,
        'current_date': datetime.now().date(),
        'error_message': error_message
    }
    
    return render(request, 'crucial/report/deposit_form_ui.html', context)

def generate_deposit_pdf(request, student_id):
    # Get student by ID (primary key)
    student = get_object_or_404(StudentProfile, id=student_id)
    due_fees = Fees.objects.filter(
        student_id=student, 
        status__in=['unpaid', 'partial']
    )
    
    total_due = 0
    for fee in due_fees:
        total_due += fee.total_fee_after_partial_payments()
    
    # Get institute details
    institute = Institute.objects.first()
    
    # Get number of copies from request
    copies = int(request.GET.get('copies', 1))
    
    context = {
        'student': student,
        'due_fees': due_fees,
        'total_due': total_due,
        'institute': institute,
        'current_date': datetime.now().date(),
        'copies': range(copies)  # For generating multiple copies in template
    }
    
    # Render HTML template
    html_string = render_to_string('crucial/report/deposit_form_pdf.html', context)
    
    # Generate PDF
    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    result = html.write_pdf()
    
    # Create HTTP response with PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="deposit_form_{student.student_field.name}.pdf"'
    response.write(result)
    
    return response



from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from core.models import StudentClass, ClassGroupConfig, StuGroup, AcademicSession
from crucial.models import Fees, Fees_name, Fee_package, Fee_month
from user.models import StudentProfile
import json

def download_fee_data(request):
    # Get filter parameters
    class_id = request.GET.get('class')
    status = request.GET.get('status', 'paid')  # Default to paid
    session_id = request.GET.get('session')
    
    # Get all classes, groups, and sessions for the filter form
    classes = StudentClass.objects.all().distinct()
    sessions = AcademicSession.objects.all().order_by('-start_year')
    
    if request.method == 'GET' and class_id:
        # Filter fees based on class, status, and session
        fees = Fees.objects.filter(status=status).select_related(
            'student_id',
            'student_id__student_field',
            'student_id__class_id',
            'student_id__class_id__class_group_id',
            'student_id__class_id__class_group_id__class_id',
            'student_id__class_id__class_group_id__group_id',
            'student_id__academic_session_year',
            'feetype_id',
            'feetype_id__fees_type',
            'feetype_id__fees_type__fee_head',
            'feetype_id__month',
            'month_id'
        )
        
        # Filter by class
        fees = fees.filter(student_id__class_id__class_group_id__class_id__id=class_id)
        
        # Filter by session if provided
        if session_id:
            fees = fees.filter(student_id__academic_session_year_id=session_id)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Fee Data - {status.capitalize()}"
        
        # Create headers
        headers = ['Student Name', 'Student ID', 'Session']
        
        # Get all unique fee combinations for column headers
        fee_combinations = set()
        fee_data_map = {}
        
        for fee in fees:
            student_id = fee.student_id.id
            student_name = fee.student_id.student_field.name
            student_user_id = fee.student_id.student_field.user_id or fee.student_id.student_field.id
            session_name = f"{fee.student_id.academic_session_year.start_year}-{fee.student_id.academic_session_year.end_year}" if fee.student_id.academic_session_year else "No Session"
            
            # Build the column header: Month - Fee Head - Class - Group
            month_name = fee.feetype_id.month.name if fee.feetype_id and fee.feetype_id.month else fee.month_id.name
            fee_head_name = fee.feetype_id.fees_type.fee_head.name if fee.feetype_id and fee.feetype_id.fees_type else "Unknown Fee Head"
            
            # Get class and group info from student
            class_name = fee.student_id.class_id.class_group_id.class_id.name if fee.student_id.class_id else "Unknown Class"
            group_name = fee.student_id.class_id.class_group_id.group_id.name if fee.student_id.class_id and fee.student_id.class_id.class_group_id.group_id else ""
            
            column_header = f"{month_name} - {fee_head_name} - {class_name}"
            if group_name:
                column_header += f" - {group_name}"
            
            fee_combinations.add(column_header)
            
            # Initialize student data if not exists
            if student_id not in fee_data_map:
                fee_data_map[student_id] = {
                    'student_name': student_name,
                    'student_id': student_user_id,
                    'session': session_name,
                    'fees': {}
                }
            
            # Store fee amount for this column
            fee_data_map[student_id]['fees'][column_header] = fee.amount
        
        # Convert to sorted list for consistent column order
        fee_combinations = sorted(list(fee_combinations))
        
        # Add fee combination columns to headers
        headers.extend(fee_combinations)
        
        # Add total column
        headers.append('Total')
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Add student data
        for row, (student_id, student_data) in enumerate(fee_data_map.items(), 2):
            # Basic student info
            ws.cell(row=row, column=1, value=student_data['student_name'])
            ws.cell(row=row, column=2, value=student_data['student_id'])
            ws.cell(row=row, column=3, value=student_data['session'])
            
            total_amount = 0
            col_offset = 3  # After basic info columns (increased due to session column)
            
            # Fill in fee amounts for each column
            for fee_combination in fee_combinations:
                amount = student_data['fees'].get(fee_combination, 0)
                if amount:
                    ws.cell(row=row, column=col_offset + 1, value=float(amount))
                    total_amount += float(amount)
                else:
                    ws.cell(row=row, column=col_offset + 1, value="")
                col_offset += 1
            
            # Add total amount
            ws.cell(row=row, column=col_offset + 1, value=float(total_amount))
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)  # Cap at 50 to avoid too wide columns
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response with Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Create filename with filters
        filename_parts = [f"fee_data_{status}"]
        if class_id:
            class_obj = StudentClass.objects.get(id=class_id)
            filename_parts.append(class_obj.name)
        if session_id:
            session_obj = AcademicSession.objects.get(id=session_id)
            filename_parts.append(f"{session_obj.start_year}-{session_obj.end_year}")
        
        filename = "_".join(filename_parts) + ".xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    context = {
        'classes': classes,
        'sessions': sessions,
        'status_choices': [
            ('paid', 'Paid'),
            ('unpaid', 'Unpaid'),
            ('partial', 'Partial')
        ]
    }
    return render(request, 'crucial/backup/download_fee_data.html', context)

import pandas as pd
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db.models import Q
from .models import Fees, Fees_name, Fee_month, FeeHead

def download_exam_half_november_fees(request):
    """
    Download Excel file for Fee Head = Exam (Half) and Fee Month = November
    """
    try:
        # Filter fees for Exam (Half) fee head, November month, and unpaid status
        fees = Fees.objects.filter(
            feetype_id__fees_type__fee_head__name='Exam (Half)',
            month_id__name='November',
            status='unpaid',
            is_enable=True
        ).select_related('student_id', 'feetype_id', 'month_id')
        
        # Check if there are any unpaid fees to download
        if not fees.exists():
            messages.warning(request, "No unpaid fees found for Exam (Half) - November.")
            return redirect('fees_management')
        
        # Create DataFrame with Student ID, Roll No, and Amount
        data = []
        for fee in fees:
            data.append({
                'Student ID': fee.student_id.id,
                'Roll No': fee.student_id.roll_no or '',
                'Amount': float(fee.amount)
            })
        
        df = pd.DataFrame(data)
        
        # Create HTTP response with Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exam_half_november_unpaid_fees.xlsx"'
        
        # Write DataFrame to Excel
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Exam_Half_November_Unpaid', index=False)
        
        # Set session flag to indicate download was completed
        request.session['download_completed'] = True
        request.session.save()  # Ensure session is saved
        
        # Add success message
        messages.success(request, "File downloaded successfully! You can now delete the fees records.")
        
        return response
        
    except Exception as e:
        messages.error(request, f"Error downloading file: {str(e)}")
        return redirect('fees_management')

def delete_exam_half_november_fees(request):
    """
    Delete fees for Fee Head = Exam (Half) and Fee Month = November with unpaid status
    """
    # Check if download was completed
    if not request.session.get('download_completed'):
        messages.error(request, "Please download the file first before deleting fees.")
        return redirect('fees_management')
    
    try:
        # Get unpaid fees to delete
        fees_to_delete = Fees.objects.filter(
            feetype_id__fees_type__fee_head__name='Exam (Half)',
            month_id__name='November',
            status='unpaid',
            is_enable=True
        )
        
        # Count before deletion
        count_before = fees_to_delete.count()
        
        if count_before == 0:
            messages.warning(request, "No unpaid fees found to delete for Exam (Half) - November.")
            # Clear the session flag
            request.session['download_completed'] = False
            request.session.save()
            return redirect('fees_management')
        
        # Soft delete by setting is_enable=False
        fees_to_delete.update(is_enable=False)
        
        # Clear the session flag
        request.session['download_completed'] = False
        request.session.save()
        
        messages.success(request, f"Successfully deleted {count_before} unpaid fees records for Exam (Half) - November.")
        
    except Exception as e:
        messages.error(request, f"Error deleting fees: {str(e)}")
    
    return redirect('fees_management')

def fees_management_view(request):
    """
    Main view for fees management with download and delete functionality
    """
    # Check if download was completed to enable delete button
    download_completed = request.session.get('download_completed', False)
    
    # Get counts for display - only unpaid fees
    exam_half_november_count = Fees.objects.filter(
        feetype_id__fees_type__fee_head__name='Exam (Half)',
        month_id__name='November',
        status='unpaid',
        is_enable=True
    ).count()
    
    context = {
        'download_completed': download_completed,
        'exam_half_november_count': exam_half_november_count,
    }
    
    return render(request, 'crucial/backup/fees_management.html', context)


import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from datetime import datetime
from decimal import Decimal
import traceback


@login_required
def upload_fee_payments(request):
    context = {'page_title': 'Upload Fee Payments'}
    
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        
        if not excel_file:
            messages.error(request, 'Please select an Excel file to upload')
            return render(request, 'crucial/finance/upload_fee_payments.html', context)
        
        try:
            # Read the Excel file
            df = pd.read_excel(excel_file)
            
            # Validate required columns
            required_columns = ['roll_number', 'date']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                messages.error(request, f'Missing columns in Excel: {", ".join(missing_columns)}')
                return render(request, 'crucial/finance/upload_fee_payments.html', context)
            
            success_count = 0
            error_count = 0
            error_details = []
            
            for index, row in df.iterrows():
                try:
                    roll_number = str(row['roll_number']).strip()
                    date_str = str(row['date']).strip()
                    
                    # Convert date string to datetime
                    try:
                        # Try multiple date formats
                        date_formats = ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y', '%Y/%m/%d']
                        payment_date = None
                        
                        for fmt in date_formats:
                            try:
                                payment_date = datetime.strptime(date_str, fmt)
                                break
                            except ValueError:
                                continue
                        
                        if not payment_date:
                            error_details.append(f"Row {index+2}: Invalid date format for roll {roll_number}")
                            error_count += 1
                            continue
                            
                        # Set time to 09:52 PM
                        payment_date = payment_date.replace(hour=21, minute=52, second=0)
                        
                    except Exception as e:
                        error_details.append(f"Row {index+2}: Date parsing error for roll {roll_number}: {str(e)}")
                        error_count += 1
                        continue
                    
                    # Find student by roll number
                    try:
                        student_profile = StudentProfile.objects.get(roll_no=roll_number)
                    except StudentProfile.DoesNotExist:
                        error_details.append(f"Row {index+2}: Student with roll number {roll_number} not found")
                        error_count += 1
                        continue
                    
                    # Get all unpaid fees for this student
                    unpaid_fees = Fees.objects.filter(
                        student_id=student_profile,
                        status__in=['unpaid', 'partial']
                    )
                    
                    if not unpaid_fees.exists():
                        error_details.append(f"Row {index+2}: No unpaid fees found for roll {roll_number}")
                        error_count += 1
                        continue
                    
                    # Process each unpaid fee
                    for fee in unpaid_fees:
                        try:
                            # Calculate total amount to pay (fee amount + late fee - discount)
                            total_due = fee.total_fee_after_discount()
                            
                            # Record full payment
                            fee.record_payment(total_due, request.user)
                            
                            # Update the fee's updated_at with Excel date (bypass auto_now)
                            Fees.objects.filter(id=fee.id).update(updated_at=payment_date)
                            
                            # Update partial payment's created_at if exists
                            if fee.status == 'paid':
                                PartialPayment.objects.filter(fee=fee).update(
                                    payment_date=payment_date,
                                    created_by=request.user
                                )
                            
                            # Update corresponding Receive entry's date
                            Receive.objects.filter(
                                voucher_no=f"FEE-{fee.transaction_no}"
                            ).update(date=payment_date.date())
                            
                            success_count += 1
                            
                        except Exception as e:
                            error_details.append(f"Row {index+2}, Fee {fee.id}: {str(e)}")
                            error_count += 1
                            continue
                    
                except Exception as e:
                    error_details.append(f"Row {index+2}: Unexpected error: {str(e)}")
                    error_count += 1
                    continue
            
            # Display results
            if success_count > 0:
                messages.success(request, f'Successfully processed {success_count} fee payments')
            
            if error_count > 0:
                messages.warning(request, f'{error_count} errors occurred during processing')
                context['error_details'] = error_details[:10]  # Show first 10 errors
            
            context['success_count'] = success_count
            context['error_count'] = error_count
            
        except Exception as e:
            messages.error(request, f'Error processing Excel file: {str(e)}')
            print(traceback.format_exc())
    
    return render(request, 'crucial/finance/upload_fee_payments.html', context)


@login_required
def download_sample_excel(request):
    """Download a sample Excel template"""
    import io
    from django.http import HttpResponse
    
    # Create sample data
    data = {
        'roll_number': [1001, 1002, 1003],
        'date': ['2024-01-15', '2024-01-16', '2024-01-17'],
        'note': ['January fee payment', 'January fee payment', 'January fee payment']
    }
    
    df = pd.DataFrame(data)
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Fee Payments')
    
    output.seek(0)
    
    # Create HTTP response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="fee_payment_template.xlsx"'
    
    return response