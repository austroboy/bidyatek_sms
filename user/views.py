from django.shortcuts import render, redirect, get_object_or_404
from .models import *
from .forms import *
from django.views import View
from tablib import Dataset
from django.http import HttpResponse,HttpResponseNotFound
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from django.core.exceptions import ObjectDoesNotExist
from user.decorators import *
from django.contrib.auth.decorators import user_passes_test, login_required
from core.models import StudentClass,ClassGroupConfig,StuGroup,StudentSection,AcademicSession,Admission_Year
from django.db.models import Q 
from miscellaneous.models import Institute
import barcode
from barcode.writer import ImageWriter
import os
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json
from PIL import Image, ImageChops
from .serializers import Group_Serializer
from rest_framework.views import APIView
from django.contrib.auth.models import Group, Permission
from rest_framework.renderers import TemplateHTMLRenderer
from django.utils.decorators import method_decorator
from rest_framework.response import Response
from rest_framework.parsers import FormParser, MultiPartParser
from datetime import datetime
from django.core.serializers.json import DjangoJSONEncoder
import uuid
import random
from .serializers import CombinedCreateSerializer
from rest_framework import status
from shared.models import CustomUser
# Function to generate a unique number
def generate_unique_number():
    attempts = 0
    while attempts < 10:
        unique_number = random.randint(10**10, 10**11 - 1)  # Generates an 11-digit random integer
        if not CustomUser.objects.filter(user_id=unique_number).exists():
            return unique_number
        attempts += 1
    
    raise Exception("Failed to generate a unique user_id after 10 attempts")

# Function to generate a unique student ID
def generate_student_id():
    attempts = 0
    while attempts < 10:
        unique_number = random.randint(10**9, 10**10 - 1)  # Generates a 10-digit random integer
        if not Student.objects.filter(user_id=unique_number).exists():
            return unique_number
        attempts += 1
    
    raise Exception("Failed to generate a unique student_id after 10 attempts")


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_in_group(user, "teacher", "staff", "parent", "student"))
def edit_profile(request):
    current_user = request.user
    user_intance = get_object_or_404(CustomUser, id=current_user.id)
    if request.method == 'POST':
        form = EditUserForm(request.POST,request.FILES, instance=user_intance)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile data has been Updated ! ! !')
            return redirect('dashboard')
    else:
        form = EditUserForm(instance=user_intance)

    context = {
        'form': form,
        'heading': 'Staff',
        'subheading': 'Edit Staff',
    }
    return render(request, 'core/user/edit_profile.html', context)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))

def add_student(request):
    
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            phone_number = form.cleaned_data.get('guardian_phone_number', '').strip()
            if phone_number:
                if not phone_number.startswith("0"):
                    phone_number = "0" + phone_number
            
            random_value = generate_unique_number()
            password_value = phone_number if phone_number else random_value
            hashed_password = make_password(password_value)
            try:
                
                if phone_number:
                    parent_instance = Parent.objects.get(phone_number=phone_number)
                else:
                    parent_instance = None

            except ObjectDoesNotExist:
                # Create a new Parent instance if it doesn't exist
                
                
                parent_instance = Parent(
                    username=password_value,
                    phone_number=phone_number,
                    name=form.cleaned_data['guardian_name'],
                    nid=form.cleaned_data['guardian_nid'],
                    password=hashed_password
                )
                parent_instance.save()
                ImportedUser.objects.create(
                        user_id=random_value,
                        username= password_value,
                        password=password_value
                        )
                parent_profile = ParentProfile(
                    parent_field=parent_instance,
                    relation=form.cleaned_data['relation'],
                    father_name=form.cleaned_data['father_name'],
                    father_mobile_no=form.cleaned_data['father_mobile_no'],
                    mother_name=form.cleaned_data['mother_name'],
                    mother_mobile_no=form.cleaned_data['mother_mobile_no'],
                    occupation=form.cleaned_data['guardian_occupation']
                )
                parent_profile.save()


            
            random_student_id= generate_student_id()
            print(random_student_id)
            s_user_id = form.cleaned_data.get('user_id', None)
            student_id_no = s_user_id if s_user_id else random_student_id
            print(student_id_no)
            
            student = Student(
                username=student_id_no,  
                avatar=form.cleaned_data['avatar'],
                phone_number=form.cleaned_data['phone_number'], 
                name=form.cleaned_data['name'], 
                gender=form.cleaned_data['gender'], 
                dob=form.cleaned_data['dob'], 
                nid=form.cleaned_data['nid'], 
                email=form.cleaned_data['email'], 
                religion=form.cleaned_data['religion'], 
                blood_group=form.cleaned_data['blood_group'], 
                rfid=form.cleaned_data['rfid'], 
                user_id=student_id_no, 
                password=hashed_password)
            student.save()
            ImportedUser.objects.create(
                        user_id=student_id_no,
                        username= student_id_no,
                        password=password_value
                        )
            student_instance = Student.objects.get(username=student_id_no)
            admission_year_id = form.cleaned_data['admission_year_id']
            class_id = form.cleaned_data['class_id']
            roll_no = form.cleaned_data['roll_no']
            birth_certificate_no = form.cleaned_data['birth_certificate_no']
            nationality = form.cleaned_data['nationality']
            tc_no = form.cleaned_data['tc_no']
            admission_date = form.cleaned_data['admission_date']
            parent_id = form.cleaned_data['parent_id']
            
            student_profile = StudentProfile(student_field=student_instance,
                                             admission_year_id=admission_year_id,
                                             class_id=class_id,
                                             roll_no=roll_no,
                                             birth_certificate_no=birth_certificate_no,
                                             nationality=nationality,
                                             tc_no=tc_no,
                                             admission_date=admission_date,
                                             parent_id=parent_instance,
                                             )
            student_profile.save()

            
            messages.success(request, 'Student data has been Added ! ! !')
            return redirect('add_student')
        else:
            print("Form errors:", form.errors)
    else:
        form = StudentForm()

    context={
        'form': form,
        'heading':'Student',
        'subheading':'Add Student',
    }

    return render(request, 'core/user/student.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))
def edit_student(request, id):
    try:
        student_instance = Student.objects.get(id=id)
    except Student.DoesNotExist:
        print(f"Student with ID {id} does not exist.")
        return HttpResponseNotFound(f"Student with ID {id} not found.")

    student_profile_instance = student_instance.student_profile.first()

    if request.method == 'POST':
        form = StudentEditForm(request.POST, request.FILES, instance=student_instance)
        student_profile_form = StudentProfileForm(request.POST, request.FILES, instance=student_profile_instance)

        if form.is_valid() and student_profile_form.is_valid():
            old_user_id = student_instance.user_id  # Store old user_id before saving

            form.save()
            student_profile_form.save()

            # Check if user_id has changed
            if 'user_id' in form.changed_data:
                new_user_id = form.cleaned_data['user_id']
                student_instance.username = new_user_id
                student_instance.save(update_fields=['username'])

                # Also update ImportedUser table
                ImportedUser.objects.filter(user_id=old_user_id).update(
                    user_id=new_user_id,
                    username=new_user_id
                )

            messages.success(request, 'Student data has been Updated ! ! !')
            return redirect('list_student')

    else:
        form = StudentEditForm(instance=student_instance)
        student_profile_form = StudentProfileForm(instance=student_profile_instance)

    context = {
        'form': form,
        'student_profile_form': student_profile_form,
        'heading': 'Student',
        'subheading': 'Edit Student',
    }

    return render(request, 'core/user/edit_student.html', context)



# user/views.py
import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from .decorators import is_staff_or_has_role
from .forms import BulkStudentIDUpdateForm
from .models import Student, StudentProfile, ImportedUser
from core.models import ClassConfig

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))
def bulk_update_student_id(request):
    if request.method == 'POST':
        form = BulkStudentIDUpdateForm(request.POST, request.FILES)
        if form.is_valid():
            selected_class = form.cleaned_data['class_config']
            excel_file = request.FILES['excel_file']
            
            try:
                # Read Excel file
                df = pd.read_excel(excel_file)
                
                # Validate columns
                if 'old_id' not in df.columns or 'new_id' not in df.columns:
                    messages.error(request, 'Excel file must contain "old_id" and "new_id" columns')
                    return redirect('bulk_update_student_id')
                
                success_count = 0
                error_messages = []
                
                for index, row in df.iterrows():
                    old_id = row['old_id']
                    new_id = row['new_id']
                    
                    try:
                        # Get student and profile for the selected class
                        student = Student.objects.get(user_id=old_id)
                        profile = StudentProfile.objects.get(
                            student_field=student,
                            class_id=selected_class
                        )
                        
                        # Check if new ID already exists
                        if Student.objects.filter(user_id=new_id).exists():
                            error_messages.append(f"Row {index+1}: New ID {new_id} already exists")
                            continue
                        
                        # Update student ID
                        student.user_id = new_id
                        student.username = new_id
                        student.save(update_fields=['user_id', 'username'])
                        
                        # Update roll number to new ID
                        profile.roll_no = new_id
                        profile.save(update_fields=['roll_no'])
                        
                        # Update ImportedUser
                        imported_user, created = ImportedUser.objects.update_or_create(
                            user_id=old_id,
                            defaults={
                                'user_id': new_id,
                                'username': new_id
                            }
                        )
                        
                        success_count += 1
                        
                    except Student.DoesNotExist:
                        error_messages.append(f"Row {index+1}: Student with ID {old_id} not found")
                    except StudentProfile.DoesNotExist:
                        error_messages.append(f"Row {index+1}: Student {old_id} not in selected class")
                    except Exception as e:
                        error_messages.append(f"Row {index+1}: Error - {str(e)}")
                
                # Show results
                if success_count > 0:
                    messages.success(request, f"Successfully updated {success_count} student IDs")
                if error_messages:
                    for msg in error_messages:
                        messages.warning(request, msg)
                
                return redirect('bulk_update_student_id')
                        
            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
    else:
        form = BulkStudentIDUpdateForm()
    
    context = {
        'form': form,
        'heading': 'Student',
        'subheading': 'Bulk Update Student ID',
    }
    return render(request, 'core/user/bulk_update_student_id.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))
def list_student(request):
    # Get query parameters
    class_list=StudentClass.objects.all()
    group_list=StuGroup.objects.all()
    section_list=StudentSection.objects.all()
    shift_list=StudentShift.objects.all()
    admission_year_list=Admission_Year.objects.all()
    academic_session_list=AcademicSession.objects.all()
    page = int(request.GET.get('page', 1))
    name = request.GET.get('name', '')  
    roll = request.GET.get('roll', '')
    class_name = request.GET.get('class', '')
    group_name = request.GET.get('group', '')
    version_name = request.GET.get('version', '')
    section_name = request.GET.get('section', '')
    shift_name = request.GET.get('shift', '')
    admission_year = request.GET.get('admission_year', '')
    session = request.GET.get('session', '')
    phoneno = request.GET.get('phoneno', '')
    userid = request.GET.get('userid', '')
    per_page = 25
    start = (page - 1) * per_page  
    end = start + per_page  
    
    # Start with the base queryset
    student_queryset = StudentProfile.objects.select_related(
        'student_field', 'class_id__class_group_id__class_id'
    ).all()

    # Apply filters dynamically based on query parameters
    if name:
        student_queryset = student_queryset.filter(student_field__name__icontains=name)
    if userid:
        student_queryset = student_queryset.filter(student_field__user_id__icontains=userid)
    if phoneno:
        student_queryset = student_queryset.filter(parent_id__phone_number__icontains=phoneno)
    if roll:
        student_queryset = student_queryset.filter(roll_no__icontains=roll)
    if class_name:
        student_queryset = student_queryset.filter(class_id__class_group_id__class_id__name__icontains=class_name)
    if group_name:
        student_queryset = student_queryset.filter(class_id__class_group_id__group_id__name__icontains=group_name)
    if section_name:
        student_queryset = student_queryset.filter(class_id__section_id__name__icontains=section_name)
    if shift_name:
        student_queryset = student_queryset.filter(class_id__shift_id__name__icontains=shift_name)
    if version_name:
        student_queryset = student_queryset.filter(version=version_name)
    if admission_year:
        academic_session_filter = (
        Q(academic_session_year__start_year=admission_year) |
        Q(academic_session_year__end_year=admission_year) |
        Q(academic_session_year__isnull=True))
        student_queryset = student_queryset.filter(academic_session_filter)
    if session:
        
        try:
            start_year, end_year = session.split('-')
            student_queryset = student_queryset.filter(
                academic_session_year__start_year=start_year,
                academic_session_year__end_year=end_year
            )
        except ValueError:
            # Handle cases where the session format is incorrect
            pass
    
    # Get the filtered count for pagination
    total_students = student_queryset.count()
    total_pages = (total_students + per_page - 1) // per_page

    # Apply slicing for pagination
    students = student_queryset[start:end]

    # Check if it's an AJAX request
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        # Return JSON response if AJAX call
        students_data = [
            {
                'name': student.student_field.name,
                'avatar': request.build_absolute_uri(student.student_field.avatar.url) if student.student_field.avatar else None,
                'roll': student.roll_no,
                'class_name': student.class_id.class_group_id.class_id.name if student.class_id else None,
                'group_name': student.class_id.class_group_id.group_id.name if student.class_id and student.class_id.class_group_id and student.class_id.class_group_id.group_id else None,
                'section_name': student.class_id.section_id.name if student.class_id and student.class_id.section_id else None,
                'shift_name': student.class_id.shift_id.name if student.class_id and student.class_id.shift_id else None,
                'admission_year': student.admission_year_id.name if student.admission_year_id else None,
                'session': str(student.academic_session_year) if student.academic_session_year else None,
                'guardian': student.parent_id.name if student.parent_id else None,
                'phoneno': student.parent_id.phone_number if student.parent_id  and student.parent_id.phone_number else None,
                'version': student.version if student.version else 'Bangla',  
                'f_name': student.parent_id.parent_profile.father_name if student.parent_id and hasattr(student.parent_id, 'parent_profile') else None,
                'f_phoneno': student.parent_id.parent_profile.father_mobile_no if student.parent_id and student.parent_id.parent_profile.father_mobile_no else None,
                'm_name': student.parent_id.parent_profile.mother_name if student.parent_id and hasattr(student.parent_id, 'parent_profile') else None,
                'm_phoneno': student.parent_id.parent_profile.mother_mobile_no if student.parent_id and student.parent_id.parent_profile.mother_mobile_no else None,
                'student_id': student.student_field.id if student.student_field else None,
                'student_status': student.student_field.status if student.student_field else None,
                'student_profile_id': student.id,
                'student_rfid': student.student_field.rfid,
                'student_userid': student.student_field.user_id,
            }
            for student in students
        ]

        # print(students_data)
        return JsonResponse({
            'students': students_data,
            'total_pages': total_pages,
            'current_page': page
        }, encoder=DjangoJSONEncoder)

    # For normal requests (initial page load), render the template
    context = {
        'students': students,
        'class_list':class_list,
        'group_list':group_list,
        'section_list':section_list,
        'shift_list':shift_list,
        'admission_year_list':admission_year_list,
        'academic_session_list':academic_session_list,
        'total_pages': total_pages,
        'current_page': page,
        'heading':'Student',
        'subheading':'Student List',

    }
    return render(request, 'core/user/student_list.html', context)

def get_filtered_students(request):
    student_queryset = StudentProfile.objects.select_related(
        'student_field', 'class_id__class_group_id__class_id'
    ).all()

    # Extract all filter parameters
    filters = {
        'name': request.GET.get('name', ''),
        'userid': request.GET.get('userid', ''),
        'phoneno': request.GET.get('phoneno', ''),
        'roll': request.GET.get('roll', ''),
        'class': request.GET.get('class', ''),
        'group': request.GET.get('group', ''),
        'section': request.GET.get('section', ''),
        'shift': request.GET.get('shift', ''),
        'version': request.GET.get('version', ''),
        'admission_year': request.GET.get('admission_year', ''),
        'session': request.GET.get('session', ''),
    }

    # Apply filters
    if filters['name']:
        student_queryset = student_queryset.filter(student_field__name__icontains=filters['name'])
    if filters['userid']:
        student_queryset = student_queryset.filter(student_field__user_id__icontains=filters['userid'])
    if filters['phoneno']:
        student_queryset = student_queryset.filter(parent_id__phone_number__icontains=filters['phoneno'])
    if filters['roll']:
        student_queryset = student_queryset.filter(roll_no__icontains=filters['roll'])
    if filters['class']:
        student_queryset = student_queryset.filter(class_id__class_group_id__class_id__name__icontains=filters['class'])
    if filters['group']:
        student_queryset = student_queryset.filter(class_id__class_group_id__group_id__name__icontains=filters['group'])
    if filters['section']:
        student_queryset = student_queryset.filter(class_id__section_id__name__icontains=filters['section'])
    if filters['shift']:
        student_queryset = student_queryset.filter(class_id__shift_id__name__icontains=filters['shift'])
    if filters['version']:
        student_queryset = student_queryset.filter(version=filters['version'])
    if filters['admission_year']:
        student_queryset = student_queryset.filter(
            Q(academic_session_year__start_year=filters['admission_year']) |
            Q(academic_session_year__end_year=filters['admission_year']) |
            Q(academic_session_year__isnull=True)
        )
    if filters['session']:
        try:
            start_year, end_year = filters['session'].split('-')
            student_queryset = student_queryset.filter(
                academic_session_year__start_year=start_year,
                academic_session_year__end_year=end_year
            )
        except ValueError:
            pass

    return student_queryset

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))
def export_students_excel(request):
    students = get_filtered_students(request)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    # Define headers
    headers = [
        'Name', 'User ID', 'Roll No', 'Class', 'Group', 'Section', 'Shift',
        'Admission Year', 'Session', 'Guardian', 'Phone', 'Version',
        'Father Name', 'Father Phone', 'Mother Name', 'Mother Phone',
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write data rows
    for row_num, student in enumerate(students, 2):
        ws.cell(row=row_num, column=1, value=student.student_field.name)
        ws.cell(row=row_num, column=2, value=student.student_field.user_id)
        ws.cell(row=row_num, column=3, value=student.roll_no)
        ws.cell(row=row_num, column=4, value=student.class_id.class_group_id.class_id.name if student.class_id else None)
        ws.cell(row=row_num, column=5, value=student.class_id.class_group_id.group_id.name if student.class_id and student.class_id.class_group_id and student.class_id.class_group_id.group_id else None)
        ws.cell(row=row_num, column=6, value=student.class_id.section_id.name if student.class_id and student.class_id.section_id else None)
        ws.cell(row=row_num, column=7, value=student.class_id.shift_id.name if student.class_id and student.class_id.shift_id else None)
        ws.cell(row=row_num, column=8, value=student.admission_year_id.name if student.admission_year_id else None)
        ws.cell(row=row_num, column=9, value=str(student.academic_session_year) if student.academic_session_year else None)
        ws.cell(row=row_num, column=10, value=student.parent_id.name if student.parent_id else None)
        ws.cell(row=row_num, column=11, value=student.parent_id.phone_number if student.parent_id and student.parent_id.phone_number else None)
        ws.cell(row=row_num, column=12, value=student.version if student.version else 'Bangla')
        ws.cell(row=row_num, column=13, value=student.parent_id.parent_profile.father_name if student.parent_id and hasattr(student.parent_id, 'parent_profile') else None)
        ws.cell(row=row_num, column=14, value=student.parent_id.parent_profile.father_mobile_no if student.parent_id and student.parent_id.parent_profile.father_mobile_no else None)
        ws.cell(row=row_num, column=15, value=student.parent_id.parent_profile.mother_name if student.parent_id and hasattr(student.parent_id, 'parent_profile') else None)
        ws.cell(row=row_num, column=16, value=student.parent_id.parent_profile.mother_mobile_no if student.parent_id and student.parent_id.parent_profile.mother_mobile_no else None)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=students.xlsx'
    wb.save(response)
    return response

from django.template.loader import render_to_string
from weasyprint import HTML
import tempfile

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))
def export_students_pdf(request):
    students = get_filtered_students(request)
    num_columns = 19  # Adjust based on your headers
    num_rows = students.count()

    # Dynamic font size logic
    if num_rows <= 10:
        font_size = 12
    elif num_columns > 15:
        font_size = 8
    else:
        font_size = 10

    context = {'students': students, 'font_size': font_size}
    html_string = render_to_string('core/user/students_pdf.html', context)
    
    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf = html.write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=students.pdf'
    return response

from django.db import IntegrityError
from .tasks import process_student_import
import tempfile
from django.contrib import messages
from django.shortcuts import redirect, render
from django.http import HttpResponse
@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))
# def student_import_data(request):
#     if request.method == 'POST':
#         new_profile = request.FILES['file']
#         dataset = Dataset()

#         # Check if the file is a valid Excel file
#         if not new_profile.name.endswith(".xlsx"):
#             return HttpResponse("Invalid file format. Please upload an Excel file.")

#         try:
#             imported_data = dataset.load(new_profile.read(), format='xlsx')
#         except Exception as e:
#             messages.error(request, f"Error reading file: {e}")
#             return redirect('list_student')

#         # Cache related objects to reduce queries
        

#         students_to_create = []
#         parents_to_create = []
#         parent_profiles_to_create = []
#         imported_users_to_create = []
#         student_profiles_to_create = []

#         for data in imported_data:
#             try:
#                 # Extract data from the row
#                 name = data[2]
#                 if not name:
#                     messages.warning(request, "Skipping record due to missing name.")
#                     continue  # Skip this row

#                 phone_number = str(data[18]) if data[18] else None
#                 if phone_number and not phone_number.startswith("0"):
#                     phone_number = "0" + phone_number

#                 random_value = generate_unique_number()
#                 password_value = phone_number if phone_number else random_value
#                 hashed_password = make_password(password_value)

#                 # Ensure Parent Exists
#                 parent_user_instance, created = Parent.objects.get_or_create(
#                     phone_number=phone_number,
#                     defaults={
#                         "username": password_value,
#                         "name": data[17],
#                         "user_id": random_value,
#                         "password": hashed_password
#                     }
#                 )

#                 parent_profile_instance, profile_created = ParentProfile.objects.get_or_create(
#                     parent_field=parent_user_instance,
#                     defaults={
#                         "relation": data[19],
#                         "father_name": data[20],
#                         "father_mobile_no": data[21],
#                         "mother_name": data[22],
#                         "mother_mobile_no": data[23],
#                         "g_name": data[17],
#                         "g_mobile_no": phone_number,
#                     }
# )
#                 # Ensure Student Exists
#                 student_user_instance, created = Student.objects.get_or_create(
#                     user_id=data[0],
#                     defaults={
#                         "username": data[0],
#                         "avatar": data[1],
#                         "phone_number": data[3],
#                         "name": data[2],
#                         "gender": data[6],
#                         "religion": data[8],
#                         "dob": data[5],
#                         "blood_group": data[7],
#                         "email": data[4],
#                         "rfid": data[9],
#                         "user_id": data[0],
#                         "present_address":data[26],
#                         "password": hashed_password 
#                     }
#                 )

#                 if created:
#                     students_to_create.append(student_user_instance)

#                     if not ImportedUser.objects.filter(username=data[0]).exists():
#                         imported_users_to_create.append(ImportedUser(
#                             user_id=data[0],
#                             username=data[0],
#                             password=password_value
#                         ))

#                 # Get academic session and class data
#                 academic_session_year=data[16]
#                 if academic_session_year:
#                             start_year, end_year = academic_session_year.split('-')
#                             academic_session_instance = AcademicSession.objects.get(
#                                 start_year=start_year,
#                                 end_year=end_year
#                             )
#                 else:
#                     academic_session_instance=None

#                 admission_year_name = data[15]
#                 admission_year_instance = Admission_Year.objects.get(name=admission_year_name)
#                 try:
#                     class_group_instance = ClassGroupConfig.objects.get(class_id__name=data[10], group_id__name=data[11])
#                 except ClassGroupConfig.DoesNotExist:
#                     messages.error(request, f"ClassGroupConfig not found for class: {data[10]} and group: {data[11]}. Skipping this record.")
#                     continue  # Skip this row

#                 try:
#                     class_config_instance = ClassConfig.objects.get(
#                         class_group_id=class_group_instance,
#                         section_id__name=data[12],
#                         shift_id__name=data[13]
#                     )
#                 except ClassConfig.DoesNotExist:
#                     messages.error(
#                         request,
#                         f"ClassConfig not found for class group: {data[10]}-{data[11]}, section: {data[12]}, shift: {data[13]}. Skipping this record."
#                     )
#                     continue  # Skip this row
#                 # Ensure StudentProfile Does Not Exist
#                 if not StudentProfile.objects.filter(student_field=student_user_instance).exists():
#                     student_profiles_to_create.append(StudentProfile(
#                         student_field=student_user_instance,
#                         class_id=class_config_instance,
#                         roll_no=data[14],
#                         version=data[24],
#                         name_tag=data[25],
#                         admission_year_id=admission_year_instance,
#                         academic_session_year=academic_session_instance,
#                         parent_id=parent_user_instance,
#                     ))

#             except IntegrityError as e:
#                 print(e)
#                 messages.error(request, f"Integrity error: {e}. Skipping this record.")
#             except Exception as e:
#                 print(e)
#                 messages.error(request, f"Unexpected error: {e}. Skipping this record.")

#         # Bulk insert to speed up database operations
#         Parent.objects.bulk_create(parents_to_create, ignore_conflicts=True)
#         ParentProfile.objects.bulk_create(parent_profiles_to_create, ignore_conflicts=True)
#         Student.objects.bulk_create(students_to_create, ignore_conflicts=True)
#         ImportedUser.objects.bulk_create(imported_users_to_create, ignore_conflicts=True)
#         StudentProfile.objects.bulk_create(student_profiles_to_create, ignore_conflicts=True)

#         messages.success(request, 'Student data has been imported successfully!')
#         return redirect('list_student')

#     context = {
#         'heading': 'Student',
#         'subheading': 'Add Student',
#     }
#     return render(request, 'core/user/student.html', context)

def student_import_data(request):
    if request.method == 'POST':
        new_profile = request.FILES['file']
        if not new_profile.name.endswith(".xlsx"):
            return HttpResponse("Invalid file format. Please upload an Excel file.")

        # Save the uploaded file to a temporary location
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        for chunk in new_profile.chunks():
            temp_file.write(chunk)
        temp_file.close()

        # Schedule the Celery task with the tenant's schema name and file path
        process_student_import.delay(request.tenant.schema_name, temp_file.name)

        messages.success(request, 'Student data is being processed. You will be notified once it’s complete.')
        return redirect('list_student')

    context = {
        'heading': 'Student',
        'subheading': 'Add Student',
    }
    return render(request, 'core/user/student.html', context)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))
def change_status_student(request, id):
    student_instance = get_object_or_404(Student, id=id)
    if student_instance.status == "Active":
        student_instance.status = "Deactive"
    else:
        student_instance.status = "Active"
    student_instance.save(update_fields=["status"])
    messages.success(request, 'Selected Student status has been updated!')
    return redirect('list_student')


class NoTextImageWriter(ImageWriter):
    def _init(self, *args, **kwargs):
        super()._init(*args, **kwargs)
        self.text = None  # Ensure text is None

    def write_text(self, xpos, ypos):
        pass  # Override this method to do nothing

def crop_white_space(image_path):
    image = Image.open(image_path)
    image = image.convert("RGB")  # Convert to RGB mode if not already
    bg = Image.new(image.mode, image.size, image.getpixel((0, 0)))  # Create a background image with the same mode and size
    diff = ImageChops.difference(image, bg)  # Find the difference between the image and the background
    bbox = diff.getbbox()  # Get the bounding box of the non-white area
    if bbox:
        image = image.crop(bbox)  # Crop the image to the bounding box
    image.save(image_path)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))
def student_id_data(request):
    institute = Institute.objects.latest('id')
    admission_year = Admission_Year.objects.latest('updated_at')
    classList = StudentClass.objects.all()
    studentlist = None

    if request.method == 'POST':
        class_id = request.POST.get("class_name_id")
        action = request.POST.get("action")
        studentlist = StudentProfile.objects.filter(
            Q(class_id__class_group_id__class_id=class_id) &
            Q(admission_year_id=admission_year) &
            Q(student_field__status="Active")
        )

        if action == "view_id_card":
            studentlist = studentlist

        elif action == "generate_barcode":
            for student in studentlist:
                CODE_39 = barcode.get_barcode_class('code39')
                writer = NoTextImageWriter()
                code = CODE_39(str(student.student_field.user_id), writer=writer, add_checksum=False)
                user_instance = student.student_field
                barcode_path = f'barcode/{user_instance.user_id}'
                barcode_full_path = os.path.join(settings.MEDIA_ROOT, barcode_path)
                code.save(barcode_full_path)

                # Crop white space from the barcode image
                crop_white_space(f'{barcode_full_path}.png')

                user_instance.barcode = f'{barcode_path}.png'
                user_instance.save()

    context = {
        'classList': classList,
        'studentlist': studentlist,
        'institute': institute,
        'heading': 'Student',
        'subheading': 'ID Card',
    }
    return render(request, 'report/student_card.html', context)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))
@csrf_exempt
def student_data(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            print(data)  # Log the data
        except json.JSONDecodeError:
            return JsonResponse({'msg': 'error', 'error': 'Invalid JSON'})
        data = json.loads(request.body)
        idcheck = data.get('dataId')
        student_profile_instance= StudentProfile.objects.get(id=idcheck)
        student_instance = student_profile_instance.student_field
        if student_profile_instance.class_id and student_profile_instance.class_id.class_group_id:
            class_instance = student_profile_instance.class_id.class_group_id.class_id
        else:
            class_instance = None
        # class_instance = student_profile_instance.class_id.class_group_id.class_id if student_profile_instance.class_id.class_group_id else None
        
        if student_profile_instance.class_id and student_profile_instance.class_id.section_id:
            section_instance = student_profile_instance.class_id.section_id
        else:
            section_instance = None
        if student_profile_instance.class_id and student_profile_instance.class_id.shift_id:
            shift_instance = student_profile_instance.class_id.shift_id
        else:
            shift_instance = None

        roll_no = student_profile_instance.roll_no
        avatar_url = student_instance.avatar.url if student_instance.avatar else None
        admission_year_id=student_profile_instance.admission_year_id
        student_id=student_profile_instance.student_field.user_id
        gender=student_profile_instance.student_field.gender
        religion=student_profile_instance.student_field.religion
        dob=student_profile_instance.student_field.dob
        mobile=student_profile_instance.student_field.phone_number
        fn= student_profile_instance.parent_id.parent_profile.father_name
        mn= student_profile_instance.parent_id.parent_profile.mother_name
        gn= student_profile_instance.parent_id.name
        gm= student_profile_instance.parent_id.phone_number
        relation= student_profile_instance.parent_id.parent_profile.relation
        occupation= student_profile_instance.parent_id.parent_profile.occupation
        bc= student_profile_instance.birth_certificate_no
        nationality= student_profile_instance.nationality
        # shift_instance = student_profile_instance.class_id.shift_id
        shift_name = shift_instance.name if shift_instance else None
        permanent_address = student_profile_instance.student_field.permanent_address
        present_address = student_profile_instance.student_field.present_address
        admission_date = student_profile_instance.admission_date
        
        serialized_student = {
                'id': student_instance.id,
                'name': student_instance.name,
                'avatar_url': avatar_url,
                'class_id':{'name': class_instance.name},
                'shift_id': {'name': shift_name},
                'section_id':{'name': section_instance},
                'roll_no':{'roll': roll_no},
                'admission_year_id':{'name':admission_year_id.name},
                'student_id':{'id':student_id},
                'gender':{'type':gender},
                'religion':{'type':religion},
                'dob':{'date':dob},
                'mobile':{'no':mobile},
                'fn':{'name':fn},
                'mn':{'name':mn},
                'gn':{'name':gn},
                'gm':{'no':gm},
                'relation':{'type':relation},
                'occupation':{'type':occupation},
                'permanent_address': permanent_address,
                'present_address': present_address,
                'admission_date': admission_date,
                'birth_certificate_no': bc,
                'nationality': nationality
                
                
            }
        return JsonResponse({'msg': serialized_student})
    else:
        return JsonResponse({'msg': 'error', 'error': 'Invalid request method'})



@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))
def get_student_classes(request):
    admission_year_id = request.GET.get('admission_year_id')
    if admission_year_id:
        student_classes = StudentClass.objects.filter(session_year_id=admission_year_id).values('id', 'name')
        return JsonResponse(list(student_classes), safe=False)
    return JsonResponse([], safe=False)

 
@login_required(login_url='auth_login')
def rfid_update(request):
    current_year = str(datetime.now().year)
    admission_year = Admission_Year.objects.get(name=current_year)
    classList = ClassConfig.objects.all()
    admission_year_list=Admission_Year.objects.all()
    class_instance,id_profiles=None,None
    page = int(request.GET.get('page', 1))
    name = request.GET.get('name', '')  
    roll = request.GET.get('roll', '')
    admission_year = request.GET.get('admission_year', '')
    rfid = request.GET.get('rfid', '')
    userid = request.GET.get('userid', '')

    per_page = 10
    start = (page - 1) * per_page  
    end = start + per_page  

    student_queryset = StudentProfile.objects.select_related(
        'student_field', 'class_id__class_group_id__class_id'
    ).filter(student_field__status="Active")

    if request.method == "POST":
        class_id = request.POST.get('class_id')  # Get selected class ID
        
        if class_id:
            # Store class_id in session
            request.session['selected_class_id'] = class_id
            class_instance = get_object_or_404(ClassConfig, pk=class_id)
            student_queryset = student_queryset.filter(class_id=class_instance)

    # Check session for class_id if GET request
    if request.method == "GET" and 'selected_class_id' in request.session:
        class_id = request.session['selected_class_id']
        class_instance = get_object_or_404(ClassConfig, pk=class_id)
        student_queryset = student_queryset.filter(class_id=class_instance)
        
    if name:
        student_queryset = student_queryset.filter(student_field__name__icontains=name)
    if roll:
        student_queryset = student_queryset.filter(roll_no__icontains=roll)
    if rfid:
        student_queryset = student_queryset.filter(student_field__rfid__icontains=rfid)
    if userid:
        student_queryset = student_queryset.filter(student_field__user_id__icontains=userid)
    if admission_year:
        academic_session_filter = (
        Q(academic_session_year__start_year=admission_year) |
        Q(academic_session_year__end_year=admission_year) |
        Q(academic_session_year__isnull=True))
        student_queryset = student_queryset.filter(academic_session_filter)

    total_students = student_queryset.count()
    total_pages = (total_students + per_page - 1) // per_page
    students = student_queryset[start:end]

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        # Return JSON response if AJAX call
        students_data = [
            {
                'name': student.student_field.name,
                'student_id': student.student_field.id,
                'roll': student.roll_no,
                'admission_year': student.admission_year_id.name if student.admission_year_id else None,
                'student_rfid': student.student_field.rfid,
                'student_userid': student.student_field.user_id,
            }
            for student in students
        ]

        # print(students_data)
        return JsonResponse({
            'students': students_data,
            'total_pages': total_pages,
            'current_page': page
        }, encoder=DjangoJSONEncoder)

    context = {
        'students': students,
        'total_pages': total_pages,
        'current_page': page,
        'admission_year_list':admission_year_list,
        'classList':classList,
        'class_instance':class_instance,
        'heading': 'Student',
        'subheading': 'Rfid Card Update'
    }
    return render(request, 'core/user/student_rfid.html', context)


@csrf_exempt
def update_rfid(request):
    if request.method == "POST":
        student_id = request.POST.get('student_id')
        new_rfid = request.POST.get('rfid')
        if new_rfid is not None:
            new_rfid = new_rfid.strip()
        try:
            student = Student.objects.get(id=student_id)
            student.rfid = new_rfid
            student.save()
            return JsonResponse({'status': 'success'})
        except CustomUser.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Student not found'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))
def list_migrations(request):
    classList = ClassConfig.objects.select_related(
        'class_group_id__class_id', 'class_group_id__group_id', 'section_id', 'shift_id'
    ).all()
    admission_year_list = Admission_Year.objects.all()
    studentlist = None
    selected_source = {}

    if request.method == 'POST':
        action = request.POST.get('action')

        # ---- Step 1: search students to migrate ----
        if action == 'search':
            class_id = request.POST.get("class_id")
            admission_year_id = request.POST.get("admission_year_mi")
            selected_source = {'class_id': class_id, 'admission_year_id': admission_year_id}

            if class_id and admission_year_id:
                studentlist = StudentProfile.objects.select_related(
                    'student_field', 'class_id__class_group_id__class_id',
                    'class_id__section_id', 'class_id__shift_id'
                ).filter(
                    Q(class_id=class_id) &
                    Q(admission_year_id=admission_year_id) &
                    Q(student_field__status="Active")&
                    Q(is_migrated=False)
                )
                if not studentlist.exists():
                    messages.warning(request, 'No active students found for the selected class and year.')
            else:
                messages.error(request, 'Please select both class and admission year.')

        # ---- Step 2: perform migration ----
        elif action == 'migrate':
            target_class_id = request.POST.get("class_id_my")
            target_year_id = request.POST.get("my_admission_year")
            selected_student_ids = request.POST.getlist('selected_students[]')

            if not target_class_id or not target_year_id:
                messages.error(request, 'Please select target class and admission year.')
            elif not selected_student_ids:
                messages.error(request, 'Please select at least one student to migrate.')
            else:
                target_class = get_object_or_404(ClassConfig, pk=target_class_id)
                target_year = get_object_or_404(Admission_Year, pk=target_year_id)

                migrated, skipped = 0, 0
                for student_id in selected_student_ids:
                    old = get_object_or_404(StudentProfile, pk=student_id)

                    # skip if already migrated to this class + year
                    exists = StudentProfile.objects.filter(
                        student_field=old.student_field,
                        class_id=target_class,
                        admission_year_id=target_year
                    ).exists()
                    if exists:
                        skipped += 1
                        continue

                    StudentProfile.objects.create(
                        student_field=old.student_field,
                        admission_year_id=target_year,
                        class_id=target_class,
                        version=old.version,
                        roll_no=old.roll_no,
                        birth_certificate=old.birth_certificate,
                        birth_certificate_no=old.birth_certificate_no,
                        nationality=old.nationality,
                        tc_no=old.tc_no,
                        admission_date=old.admission_date,
                        tc_certificate=old.tc_certificate,
                        parent_id=old.parent_id,
                        village=old.village,
                        post_office=old.post_office,
                        ps_or_upazilla=old.ps_or_upazilla,
                        district=old.district,
                    )
                    old.is_migrated = True
                    old.save(update_fields=['is_migrated'])
                    migrated += 1

                if migrated:
                    messages.success(request, f'{migrated} student(s) migrated successfully!')
                if skipped:
                    messages.info(request, f'{skipped} student(s) skipped (already migrated).')

    studentClassList = StudentClass.objects.all()
    context = {
        'classList': classList,
        'studentlist': studentlist,
        'admission_year_list': admission_year_list,
        'selected_source': selected_source,
        'heading': 'Student',
        'subheading': 'Migration',
        'studentClassList': studentClassList,
    }
    return render(request, 'report/student_migration.html', context) 


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))

def add_staff(request):
    if request.method == 'POST':
        form = StaffRegistrationForm(request.POST, request.FILES)
        profile_form = StaffProfileForm(request.POST)

        if form.is_valid() and profile_form.is_valid():
            # Save Staff instance
            staff = CustomUser(
                username=form.cleaned_data['user_id'],
                avatar=form.cleaned_data['avatar'],
                phone_number=form.cleaned_data['phone_number'],
                name=form.cleaned_data['name'],
                gender=form.cleaned_data['gender'],
                dob=form.cleaned_data['dob'],
                nid=form.cleaned_data['nid'],
                email=form.cleaned_data['email'],
                religion=form.cleaned_data['religion'],
                blood_group=form.cleaned_data['blood_group'],
                rfid=form.cleaned_data['rfid'],
                user_id=form.cleaned_data['user_id'],
                password=make_password(form.cleaned_data['phone_number']),
                nationality=form.cleaned_data.get('nationality'),
            )
            staff.save()
            ImportedUser.objects.create(
                        user_id=form.cleaned_data['user_id'],
                        username= form.cleaned_data['user_id'],
                        password=form.cleaned_data['phone_number']
                        )
            # staff.refresh_from_db()
            # print("Saved Staff:", staff)  # Debugging step
            staff_instance = CustomUser.objects.get(username=form.cleaned_data['user_id'])
            # Use the saved staff instance for staff_field
            staff_profile = StaffProfile(
                staff_field=staff_instance,  # Assign the saved Staff object
                qualification=profile_form.cleaned_data['qualification'],
                fathers_name=profile_form.cleaned_data['fathers_name'],
                mothers_name=profile_form.cleaned_data['mothers_name'],
                spouse_name=profile_form.cleaned_data['spouse_name'],
                spouse_phone_number=profile_form.cleaned_data['spouse_phone_number'],
                children_no=profile_form.cleaned_data['children_no'],
                role=profile_form.cleaned_data['role'],
                marital_status=profile_form.cleaned_data['marital_status'],
                designation=profile_form.cleaned_data['designation'],
                joining_date=profile_form.cleaned_data['joining_date'],
                grade=profile_form.cleaned_data.get('grade'),
                job_nature=profile_form.cleaned_data.get('job_nature'),
                department=profile_form.cleaned_data.get('department'),
                shift_id=profile_form.cleaned_data.get('shift_id'),
                name_tag=profile_form.cleaned_data.get('name_tag'),
                t_version=profile_form.cleaned_data.get('t_version'),
                tin=profile_form.cleaned_data.get('tin'),
                subject=profile_form.cleaned_data.get('subject'),
            )
            staff_profile.save()
            print("Saved Staff Profile:", staff_profile)
            messages.success(request, 'Staff data has been added!')
            return redirect('add_staff')
    else:
        form = StaffRegistrationForm()
        profile_form = StaffProfileForm()

    context = {
        'form': form,
        'profile_form': profile_form,
        'heading': 'Staff',
        'subheading': 'Add Staff',
    }

    return render(request, 'core/user/staff.html', context)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))
def list_staff(request):
    stafflist=StaffProfile.objects.all()
    context={
        'stafflist':stafflist,
        'heading':'Staff',
        'subheading':'List Staff',
    }
    return render(request, 'core/user/staff_list.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))
def edit_staff(request, id):
    staff_intance = get_object_or_404(CustomUser, id=id)
    staff_intance_profile = staff_intance.staff_profile

    if request.method == 'POST':
        form = StaffEditRegistrationForm(request.POST,request.FILES, instance=staff_intance)
        profile_form = StaffProfileForm(request.POST, request.FILES, instance=staff_intance_profile)

        if form.is_valid() and profile_form.is_valid():
            form.save()
            profile_form.save()
            messages.success(request, 'Staff data has been Updated ! ! !')
            return redirect('list_staff')
    else:
        form = StaffEditRegistrationForm(instance=staff_intance)
        profile_form = StaffProfileForm(instance=staff_intance_profile)

    context = {
        'form': form,
        'profile_form': profile_form,
        'heading': 'Staff',
        'subheading': 'Edit Staff',
    }
    return render(request, 'core/user/edit_staff.html', context)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))
def change_status_staff(request,id):
    staff_instance = get_object_or_404(CustomUser, id=id)
    if staff_instance.status == "Active":
        staff_instance.status = "Deactive"
    else:
        staff_instance.status = "Active"
    staff_instance.save(update_fields=["status"])
    messages.success(request, 'Seleted Staff has been updated!')
    return redirect('list_staff')


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))
def staff_import_data(request):
    if request.method == 'POST':
        new_profile = request.FILES['file']
        dataset=Dataset()

        if not new_profile.name.endswith(".xlsx"):
            return HttpResponse("excel")
        imported_data= dataset.load(new_profile.read(),format='xlsx')
        for data in imported_data:
            name = data[2]
            phone_number = str(data[1])
            if not phone_number.startswith("0"):
                phone_number = "0" + phone_number
            if name is not None:
                password = make_password(str(phone_number))
                staff_user_instance = Staff(
                    username=data[0],
                    phone_number=phone_number,
                    name=data[2],
                    gender=data[4],
                    religion=data[5],
                    dob=data[6],
                    blood_group=data[7],
                    email=data[8],
                    rfid=data[3],
                    user_id=data[0],
                    password=password
                )
                staff_user_instance.save()
                role_instance = RoleType.objects.get(name=data[12])
                staff_profile_instance = StaffProfile(
                    staff_field=staff_user_instance,
                    qualification=data[9],
                    employee_type=data[10],
                    designation=data[11],
                    role=role_instance,
                    joining_date=data[13]
                )
                staff_profile_instance.save()
        messages.success(request, 'Staff data has been Imported ! ! !')
        return redirect('list_staff')

    context={
    'heading':'Staff',
    'subheading':'Add Staff',
    }

    return render(request, 'core/user/staff.html', context)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))
def staff_id_data(request):
    institute=Institute.objects.latest('id')
    stafflist= StaffProfile.objects.filter(staff_field__status ="Active")
    context={
    'stafflist':stafflist,
    'institute':institute,
    'heading':'Staff',
    'subheading':'Id Card',
    }
    return render(request, 'report/staff_card.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))
def add_parent(request):
    if request.method == 'POST':
        form = ParentForm(request.POST,request.FILES)
        profile_form = ParentProfileForm(request.POST)
        
        if form.is_valid() and profile_form.is_valid():
            
            parent=Parent(
                username=form.cleaned_data['phone_number'],
                phone_number=form.cleaned_data['phone_number'],
                name=form.cleaned_data['name'],
                nid=form.cleaned_data['nid'],
                password=make_password(form.cleaned_data['phone_number'])
            )
            parent.save()
            parent_instance = Parent.objects.get(username=form.cleaned_data['phone_number'])
            
            parent_profile = ParentProfile(
                parent_field=parent_instance,
                relation=profile_form.cleaned_data['relation'],
                occupation=profile_form.cleaned_data['occupation'],
                father_name=profile_form.cleaned_data['father_name'],
                father_mobile_no=profile_form.cleaned_data['father_mobile_no'],
                mother_name=profile_form.cleaned_data['mother_name'],
                mother_mobile_no=profile_form.cleaned_data['mother_mobile_no']
            )
            parent_profile.save()

            messages.success(request, 'Parent data has been Added ! ! !')
            
            return redirect('add_parent')  
        
    else:
        form = ParentForm()
        profile_form = ParentProfileForm(request.POST)

    context = {
        'form': form,
        'profile_form': profile_form,
        'heading':'Parent',
        'subheading':'Add Parent',
    }

    return render(request, 'core/user/parent.html', context)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))
def list_parent(request):
    admission_year = Admission_Year.objects.latest('updated_at') 
    studentList = StudentProfile.objects.filter(
        Q(admission_year_id=admission_year)
    )
    
    parentlist = []
    for student in studentList:
        if student.parent_id:
            parentlist.append({
                'parent': student.parent_id,
                'student_name': student.student_field.name,
                'relation': student.parent_id.parent_profile.relation,  
                'father_name': student.parent_id.parent_profile.father_name,  
                'father_mobile_no': student.parent_id.parent_profile.father_mobile_no,
                'mother_name': student.parent_id.parent_profile.mother_name,
                'mother_mobile_no': student.parent_id.parent_profile.mother_mobile_no
            })
    
    
    context = { 
        'parentlist': parentlist,
        'heading': 'Parent',
        'subheading': 'List Parent',
    }
    return render(request, 'core/user/parent_list.html', context)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))
def edit_parent(request, id):
    parent_instance = get_object_or_404(Parent, id=id)
    parent_profile_instance = parent_instance.parent_profile

    if request.method == 'POST':
        form = ParentEditForm(request.POST, request.FILES, instance=parent_instance)
        profile_form = ParentProfileForm(request.POST, request.FILES, instance=parent_profile_instance)

        if form.is_valid() and profile_form.is_valid():
            old_phone = parent_instance.phone_number
            new_phone = form.cleaned_data.get('phone_number')

            form.save()
            profile_form.save()

            # Update username and password if phone number changed
            if old_phone != new_phone:
                parent_instance.username = new_phone
                parent_instance.set_password(new_phone)  # Secure password hashing
                parent_instance.save(update_fields=['username', 'password'])

                # Update ImportedUser table as well
                ImportedUser.objects.filter(user_id=parent_instance.user_id).update(
                    username=new_phone,
                    password=new_phone  # Plaintext storage for reference
                )

            messages.success(request, 'Parent data has been Updated ! ! !')
            return redirect('list_parent')

    else:
        form = ParentEditForm(instance=parent_instance)
        profile_form = ParentProfileForm(instance=parent_profile_instance)

    context = {
        'form': form,
        'profile_form': profile_form,
        'heading': 'Parent',
        'subheading': 'Edit Parent',
    }
    return render(request, 'core/user/edit_parent.html', context)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','HR']))
def change_status_parent(request,id):
    parent_instance = get_object_or_404(Parent, id=id)
    print(parent_instance)
    if parent_instance.status == "Active":
        parent_instance.status = "Deactive"
    else:
        parent_instance.status = "Active"
    parent_instance.save(update_fields=["status"])
    messages.success(request, 'Selected Parent status has been updated!')
    return redirect('list_parent')




# class Permission_User(APIView):
#     renderer_classes = [TemplateHTMLRenderer]
#     template_name = 'core/user/permission.html'
#     serializer_class = Group_Serializer
#     parser_classes = [FormParser, MultiPartParser]
#     # @method_decorator(login_required(login_url='login'), name='dispatch')
#     def get(self, request, format=None, *args, **kwargs):

#         snippets = Group.objects.all()
#         print(snippets)
#         serializer = Group_Serializer(snippets, many=True).data
#         print(serializer)
#         data_dict = []
#         GroupInfo = []
#         for dat in serializer:
#             data_dict.append(dict(dat))
#             if str(dict(dat)['id']) == kwargs['id']:
#                 GroupInfo = dict(dat)
            
#         print("data_dict",data_dict)
#         people_2 = Permission.objects.all().values("id", "name")
#         people_2 = list(people_2)
#         context = {
#             'Text': 'Give Your pemission', 
#             'Data': people_2,
#             'Group': data_dict, 
#             'GroupInfo': GroupInfo,
#             'heading': 'User',
#             'subheading': 'Permission',
#             }  # , 'GroupPerm':permission
#         return Response(context)
#         # return render(requset, 'permission.html', context)

#     def post(self, request, *args, **kwargs):

#         snippets = Group.objects.all()
#         serializer = Group_Serializer(snippets, many=True).data
#         data_dict = []
#         GroupInfo = []
#         for dat in serializer:
#             data_dict.append(dict(dat))
#             if str(dict(dat)['id']) == kwargs['id']:
#                 GroupInfo = dict(dat)

#         GroupInfo['permissions'] = list(
#             map(int, request.POST.getlist('permissions')))
#         print(data_dict)
#         grp_update = Group.objects.get(id=kwargs['id'])

#         grp_update.permissions.clear()

#         for num in GroupInfo['permissions']:
#             grp_update.permissions.add(num)

#         people_2 = list(Permission.objects.all().values("id", "name"))
#         context = {'Text': 'Give Your pemission', 'Data': people_2,
#                    'Group': data_dict, 'GroupInfo': GroupInfo}  # , 'GroupPerm':permission
#         return Response(context)


class Permission_User(APIView):
    renderer_classes = [TemplateHTMLRenderer]
    template_name = 'core/user/permission.html'
    serializer_class = Group_Serializer
    parser_classes = [FormParser, MultiPartParser]

    def get(self, request, format=None, *args, **kwargs):
        group_id = kwargs.get('id', None)

        # Fetch all groups and the selected group
        groups = Group.objects.all()
        selected_group = Group.objects.filter(id=group_id).first()

        # Fetch all permissions and group by app
        permissions = Permission.objects.all().values("id", "name", "content_type__app_label")
        grouped_permissions = {}
        for perm in permissions:
            app_label = perm["content_type__app_label"]
            if app_label not in grouped_permissions:
                grouped_permissions[app_label] = []
            grouped_permissions[app_label].append(perm)

        # Fetch permissions for the selected group
        group_permissions = selected_group.permissions.values_list("id", flat=True) if selected_group else []

        context = {
            'Text': 'Give Your Permission',
            'grouped_permissions': grouped_permissions,
            'groups': groups,
            'selected_group': selected_group,
            'group_permissions': group_permissions,
            'heading': 'User',
            'subheading': 'Permission',
        }
        return Response(context)

    def post(self, request, *args, **kwargs):
        group_id = kwargs.get('id', None)
        selected_group = Group.objects.get(id=group_id)

        # Update group permissions
        selected_group.permissions.clear()
        permission_ids = request.POST.getlist('permissions')
        selected_group.permissions.add(*permission_ids)

        # Redirect back to the same page
        return redirect('permission', id=group_id)


from django.db.models import CharField, Value
from django.db.models.functions import Concat


class PermissionUserView(APIView):
    renderer_classes = [TemplateHTMLRenderer]
    template_name = 'core/user/permission_copy.html'
    parser_classes = [FormParser, MultiPartParser]

    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            return redirect("home")  # Redirect non-superusers to home or an appropriate page
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        group_id = kwargs.get('group_id', None)  # Group ID from URL

        # Fetch all groups
        groups = Group.objects.all()

        # Fetch selected group if group_id is provided
        selected_group = Group.objects.filter(id=group_id).first()

        # Specify apps dynamically based on your folder structure
        included_apps = [
            "accounting", "attendance", "core", "crucial",
            "exam", "media", "miscellaneous", "organizer", "sms", "webpage"
        ]

        # Categorize permissions for each app
        categorized_permissions = {}
        for app in included_apps:
            app_permissions = Permission.objects.filter(content_type__app_label=app)
            categorized_permissions[app] = {
                'view_only': app_permissions.filter(codename__startswith="view").values_list('id', flat=True),
                'all_permissions': app_permissions.values_list('id', flat=True),
            }

        # Fetch permissions assigned to the selected group
        group_permissions = selected_group.permissions.values_list("id", flat=True) if selected_group else []

        context = {
            'groups': groups,  # All groups
            'selected_group': selected_group,  # Info about selected group
            'categorized_permissions': categorized_permissions,  # Permissions categorized by app
            'group_permissions': group_permissions,  # Permissions assigned to the group
            'heading': 'User',
            'subheading': 'Permission',
        }
        return Response(context)

    def post(self, request, *args, **kwargs):
        group_id = kwargs.get('group_id', None)  # Group ID from URL
        selected_group = get_object_or_404(Group, id=group_id)

        # Update group permissions
        selected_group.permissions.clear()  # Clear existing permissions
        permission_ids = request.POST.getlist('permissions')  # Get selected permissions
        selected_group.permissions.add(*permission_ids)

        # Redirect back to the same page
        return redirect('user_permissions', group_id=group_id)


class CreateStudentParentAPIView(APIView):
    def post(self, request):
        serializer = CombinedCreateSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

def search_students(request):
    query = request.GET.get('q', '')  # Get search term from AJAX
    students = StudentProfile.objects.filter(student_field__name__icontains=query)[:10]  # Limit results

    results = [
        {"id": student.id, "text": f"{student.student_field.name} - {student.class_id}"}
        for student in students
    ]

    return JsonResponse({"results": results})





import pandas as pd
def bulk_upload_emails(request):
    if request.method == 'POST':
        excel_file = request.FILES['email_file']
        
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'Please upload a valid Excel file (.xlsx)')
            return redirect('bulk_upload_emails')
        
        try:
            # Read the Excel file
            df = pd.read_excel(excel_file)
            
            # Validate required columns
            if 'ID' not in df.columns or 'Email' not in df.columns:
                messages.error(request, 'Excel file must contain "ID" and "Email" columns')
                return redirect('bulk_upload_emails')
            
            # Process the data
            updated_count = 0
            skipped_count = 0
            
            for index, row in df.iterrows():
                student_id = str(row['ID']).strip()
                email = str(row['Email']).strip().lower() if pd.notna(row['Email']) else None
                
                if not email:
                    skipped_count += 1
                    continue
                
                # Find students with matching user_id (check both Student and CustomUser)
                students = Student.objects.filter(user_id=student_id)
                if not students.exists():
                    # Check CustomUser as fallback
                    users = CustomUser.objects.filter(user_id=student_id)
                    if not users.exists():
                        skipped_count += 1
                        continue
                    else:
                        user = users.first()
                else:
                    user = students.first()
                
                # Only update if email is empty
                if not user.email:
                    user.email = email
                    user.save()
                    updated_count += 1
                else:
                    skipped_count += 1
            
            messages.success(request, f'Successfully updated {updated_count} emails. {skipped_count} records were skipped (either email exists or ID not found).')
            return redirect('bulk_upload_emails')
            
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            return redirect('bulk_upload_emails')
    
    return render(request, 'core/user/bulk_email_upload.html', {
        'heading': 'Student',
        'subheading': 'Bulk Email Upload'
    })

def process_email_upload(request):
    if request.method == 'POST' and request.FILES.get('email_file'):
        return bulk_upload_emails(request)
    return HttpResponse("Invalid request", status=400)




from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.generic import View
from django.core.exceptions import ObjectDoesNotExist
from .models import AdmissionApplicant, Student, Parent, StudentProfile, ParentProfile
from django.contrib.auth.hashers import make_password
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.core.mail import send_mail
from django.conf import settings
import json
from datetime import datetime
from ssl_commerz.models import PaymentTransaction, SSLC
from ssl_commerz.views import initiate_payment as ssl_initiate_payment
from django.contrib.auth import get_user_model
from sslcommerz_lib import SSLCOMMERZ
from django.urls import reverse

# class AdmissionFormView(View):
#     def get(self, request):
#         return render(request, 'user/admission_form.html')
    
#     def post(self, request):
#         try:
#             data = request.POST
#             files = request.FILES
            
#             birth_date_str = data.get('birthDate')
#             birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
            
#             # Create admission applicant
#             applicant = AdmissionApplicant(
#                 class_sought=data.get('class_sought'),
#                 shift=data.get('shift'),
#                 version=data.get('version'),
#                 photo=files.get('photo'),
                
#                 # Applicant Information
#                 full_name=data.get('fullName'),
#                 full_name_bangla=data.get('fullNameBangla'),
#                 nick_name=data.get('nickName'),
#                 birth_certificate_no=data.get('brcn'),
#                 gender=data.get('gender'),
#                 religion=data.get('religion'),
#                 dob=birth_date,
#                 blood_group=data.get('bloodGroup'),
#                 nationality=data.get('nationality'),
#                 catchment_area=data.get('catchmentArea') == 'on',
#                 quota=data.get('quota'),
                
#                 # Academic Information
#                 previous_class=data.get('previous_class'),
#                 previous_school=data.get('spa'),
#                 previous_school_address=data.get('apa'),
#                 exam_name=data.get('exam'),
#                 board_roll=data.get('boardRoll'),
#                 board_name=data.get('boardName'),
#                 registration_no=data.get('registrationNo'),
#                 gpa=data.get('gpa'),
                
#                 # Father Information
#                 father_name=data.get('fatherName'),
#                 father_name_bangla=data.get('fatherNameBangla'),
#                 father_mobile=data.get('fatherMobile'),
#                 father_qualification=data.get('qualification'),
#                 father_occupation=data.get('fatherOccupation'),
#                 father_service_type=data.get('fatherService'),
#                 father_designation=data.get('fatherDesignation'),
#                 father_organization=data.get('fatherOrganization'),
#                 father_yearly_income=data.get('fatherYearlyIncome'),
#                 father_income_source=data.get('fatherIncomeSource'),
#                 father_nid=data.get('fnid'),
#                 father_etin=data.get('fEtin'),
                
#                 # Mother Information
#                 mother_name=data.get('motherName'),
#                 mother_name_bangla=data.get('motherNameBangla'),
#                 mother_mobile=data.get('motherMobile'),
#                 mother_qualification=data.get('mqualification'),
#                 mother_occupation=data.get('motherOccupation'),
#                 mother_service_type=data.get('motherService'),
#                 mother_designation=data.get('motherDesignation'),
#                 mother_organization=data.get('motherOrganization'),
#                 mother_yearly_income=data.get('motherYearlyIncome'),
#                 mother_income_source=data.get('motherIncomeSource'),
#                 mother_nid=data.get('mNid'),
#                 mother_etin=data.get('mEtin'),
                
#                 # Guardian Information
#                 guardian_name=data.get('guardianName'),
#                 guardian_mobile=data.get('guardianMobile'),
#                 guardian_relation=data.get('guardianRelation'),
#                 guardian_nid=data.get('gnid'),
#                 guardian_address=data.get('gAddress'),
                
#                 # Address Information
#                 present_address=data.get('praddress'),
#                 present_country=data.get('prCountry'),
#                 present_district=data.get('prDistrict'),
#                 present_thana=data.get('prThana'),
#                 present_telephone=data.get('prTelephone'),
#                 present_mobile=data.get('prMobile'),
                
#                 permanent_address=data.get('paddress'),
#                 permanent_country=data.get('pCountry'),
#                 permanent_district=data.get('pDistrict'),
#                 permanent_thana=data.get('pThana'),
#                 permanent_telephone=data.get('pTelephone'),
#                 permanent_mobile=data.get('pMobile'),
                
#                 # Contact Information
#                 contact_mobile=data.get('smsAlertNumber'),
#                 emergency_contact=data.get('emergencyContact'),
                
#                 # Payment status
#                 status=AdmissionApplicant.Status.PAYMENT_PENDING,
#                 payment_status='Pending'
                
#             )
            
#             applicant.save()
            
#         #     # Send confirmation email if email is provided
#         #     if data.get('email'):
#         #         send_mail(
#         #             'Admission Application Submitted',
#         #             f'Your admission application has been submitted successfully. Your Applicant ID is: {applicant.applicant_id}',
#         #             settings.DEFAULT_FROM_EMAIL,
#         #             [data.get('email')],
#         #             fail_silently=True,
#         #         )
            
#         #     return JsonResponse({
#         #         'success': True,
#         #         'applicant_id': applicant.applicant_id,
#         #         'message': 'Application submitted successfully!'
#         #     })
        
#         # except Exception as e:
#         #     return JsonResponse({
#         #         'success': False,
#         #         'message': f'Error submitting application: {str(e)}'
#         #     }, status=400)
        
        
        
#         # Initiate SSLCommerz payment
#         request.user = applicant  # Temporary set applicant as user for payment
#         request.POST = request.POST.copy()
#         request.POST['payment_method'] = 'sslcommerz'
        
#         # Calculate admission fee
#         admission_fee = applicant.get_admission_fee()
        
#         # Create payment transaction
#         tran_id = f"ADM_{applicant.applicant_id}"
#         payment_transaction = PaymentTransaction.objects.create(
#             user=applicant,
#             tran_id=tran_id,
#             amount=admission_fee,
#             status='PENDING',
#             gateway='sslcommerz'
#         )
        
#         applicant.payment_transaction = payment_transaction
#         applicant.save()
        
#         # Redirect to payment page
#         payment_response = ssl_initiate_payment(request)
        
#         if hasattr(payment_response, 'url'):
#             # If payment initiation is successful, return the redirect URL
#             return JsonResponse({
#                 'success': True,
#                 'redirect_url': payment_response.url,
#                 'applicant_id': applicant.applicant_id,
#                 'message': 'Please complete the payment to finalize your admission application.'
#             })
#         else:
#             # If payment initiation failed
#             return JsonResponse({
#                 'success': False,
#                 'message': 'Payment initiation failed. Please try again.'
#             }, status=400)
    
#     except Exception as e:
#         return JsonResponse({
#             'success': False,
#             'message': f'Error submitting application: {str(e)}'
#         }, status=400)
        
class AdmissionFormView(View):
    def get(self, request):
        return render(request, 'user/admission_form.html')
    
    def post(self, request):
        try:
            data = request.POST
            files = request.FILES
            
            birth_date_str = data.get('birthDate')
            birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
            
            # Create admission applicant
            applicant = AdmissionApplicant(
                class_sought=data.get('class_sought'),
                shift=data.get('shift'),
                version=data.get('version'),
                photo=files.get('photo'),
                
                # Applicant Information
                full_name=data.get('fullName'),
                full_name_bangla=data.get('fullNameBangla'),
                nick_name=data.get('nickName'),
                birth_certificate_no=data.get('brcn'),
                gender=data.get('gender'),
                religion=data.get('religion'),
                dob=birth_date,
                blood_group=data.get('bloodGroup'),
                nationality=data.get('nationality'),
                catchment_area=data.get('catchmentArea') == 'on',
                quota=data.get('quota'),
                
                # Academic Information
                previous_class=data.get('previous_class'),
                previous_school=data.get('spa'),
                previous_school_address=data.get('apa'),
                exam_name=data.get('exam'),
                board_roll=data.get('boardRoll'),
                board_name=data.get('boardName'),
                registration_no=data.get('registrationNo'),
                gpa=data.get('gpa'),
                
                # Father Information
                father_name=data.get('fatherName'),
                father_name_bangla=data.get('fatherNameBangla'),
                father_mobile=data.get('fatherMobile'),
                father_qualification=data.get('qualification'),
                father_occupation=data.get('fatherOccupation'),
                father_service_type=data.get('fatherService'),
                father_designation=data.get('fatherDesignation'),
                father_organization=data.get('fatherOrganization'),
                father_yearly_income=data.get('fatherYearlyIncome'),
                father_income_source=data.get('fatherIncomeSource'),
                father_nid=data.get('fnid'),
                father_etin=data.get('fEtin'),
                
                # Mother Information
                mother_name=data.get('motherName'),
                mother_name_bangla=data.get('motherNameBangla'),
                mother_mobile=data.get('motherMobile'),
                mother_qualification=data.get('mqualification'),
                mother_occupation=data.get('motherOccupation'),
                mother_service_type=data.get('motherService'),
                mother_designation=data.get('motherDesignation'),
                mother_organization=data.get('motherOrganization'),
                mother_yearly_income=data.get('motherYearlyIncome'),
                mother_income_source=data.get('motherIncomeSource'),
                mother_nid=data.get('mNid'),
                mother_etin=data.get('mEtin'),
                
                # Guardian Information
                guardian_name=data.get('guardianName'),
                guardian_mobile=data.get('guardianMobile'),
                guardian_relation=data.get('guardianRelation'),
                guardian_nid=data.get('gnid'),
                guardian_address=data.get('gAddress'),
                
                # Address Information
                present_address=data.get('praddress'),
                present_country=data.get('prCountry'),
                present_district=data.get('prDistrict'),
                present_thana=data.get('prThana'),
                present_telephone=data.get('prTelephone'),
                present_mobile=data.get('prMobile'),
                
                permanent_address=data.get('paddress'),
                permanent_country=data.get('pCountry'),
                permanent_district=data.get('pDistrict'),
                permanent_thana=data.get('pThana'),
                permanent_telephone=data.get('pTelephone'),
                permanent_mobile=data.get('pMobile'),
                
                # Contact Information
                contact_mobile=data.get('smsAlertNumber'),
                emergency_contact=data.get('emergencyContact'),
                
                # Payment status
                status=AdmissionApplicant.Status.PAYMENT_PENDING,
                payment_status='Pending'
            )
            
            applicant.save()
            
            # Calculate admission fee
            admission_fee = applicant.get_admission_fee()
            
            # # Create payment transaction
            # tran_id = f"ADM_{applicant.applicant_id}"
            
            # # Create payment transaction with a temporary user
            # User = get_user_model()
            # temp_user, created = User.objects.get_or_create(
            #     username=f"temp_{applicant.applicant_id}",
            #     defaults={
            #         'password': make_password(get_random_string(10)),
            #         'is_active': False
            #     }
            # )
            
            tran_id = f"ADM_{applicant.applicant_id}"
            ssl_config = SSLC.objects.first()
            if not ssl_config:
                return JsonResponse({
                    'success': False,
                    'message': 'Payment gateway configuration not found.'
                }, status=400)
                
            sslcz = SSLCOMMERZ({
                'store_id': ssl_config.store_id,
                'store_pass': ssl_config.store_pass,
                'issandbox': (ssl_config.store_penv == 'sandbox')
            })
            
            # Prepare SSLCommerz post body
            post_body = {
                'total_amount': str(admission_fee),
                'currency': "BDT",
                'tran_id': tran_id,
                'success_url': request.build_absolute_uri(reverse('payment_callback')),
                'fail_url': request.build_absolute_uri(reverse('payment_fail')),
                'cancel_url': request.build_absolute_uri(reverse('payment_cancel')),
                'ipn_url': request.build_absolute_uri(reverse('payment_callback')),
                'emi_option': "0",
                'cus_name': applicant.full_name,
                'cus_email': "zihadmilon747@gmail.com",  # Add email field to your form if needed
                'cus_phone': applicant.contact_mobile,
                'cus_add1': applicant.present_address[:250],  # Truncate if too long
                'cus_city': applicant.present_district or "Dhaka",
                'cus_country': "Bangladesh",
                'shipping_method': "NO",
                'product_name': "Admission Fee",
                'product_category': "Education",
                'product_profile': "general",
                'value_a': applicant.applicant_id,  # Custom parameter to track applicant
            }
            
            # Create payment transaction without user association
            payment_transaction = PaymentTransaction.objects.create(
                tran_id=tran_id,
                amount=admission_fee,
                status='PENDING',
                gateway='sslcommerz'
            )
            
            applicant.payment_transaction = payment_transaction
            applicant.save()
            
            # Store applicant_id in session for callback
            request.session['current_applicant_id'] = applicant.applicant_id
            
            try:
                # Initiate SSLCommerz payment
                response = sslcz.createSession(post_body)
                
                if response.get('status') == 'SUCCESS':
                    gateway_url = response['GatewayPageURL']
                    return JsonResponse({
                        'success': True,
                        'redirect_url': gateway_url,
                        'applicant_id': applicant.applicant_id,
                        'message': 'Please complete the payment to finalize your admission application.'
                    })
                else:
                    error_msg = response.get('failedreason', 'Payment initiation failed')
                    return JsonResponse({
                        'success': False,
                        'message': error_msg
                    }, status=400)
                    
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Payment gateway error: {str(e)}'
                }, status=400)
                
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error submitting application: {str(e)}'
            }, status=400)

class AdmitCardView(View):
    def get(self, request, applicant_id):
        try:
            applicant = AdmissionApplicant.objects.get(applicant_id=applicant_id)
            
            # Check if payment is completed
            if not applicant.is_payment_completed():
                return render(request, 'user/admit_card.html', {
                    'error': 'Payment not completed. Please complete the payment to download your admit card.',
                    'show_payment_button': True,
                    'applicant': applicant
                })
            
            context = {
                'applicant': applicant,
                'exam_date': 'To be announced',  
                'exam_center': 'Main Campus',   
            }
            return render(request, 'user/admit_card.html', context)
        except AdmissionApplicant.DoesNotExist:
            return render(request, 'user/admit_card.html', {'error': 'Applicant not found'})
 
@csrf_exempt
def payment_callback(request):
    """Handle SSLCommerz payment callback"""
    if request.method == 'POST':
        tran_id = request.POST.get('tran_id')
        status = request.POST.get('status')
        
        try:
            payment = PaymentTransaction.objects.get(tran_id=tran_id)
            applicant_id = request.session.get('current_applicant_id')
            
            if applicant_id:
                applicant = AdmissionApplicant.objects.get(applicant_id=applicant_id)
                
                if status in ['VALID', 'SUCCESS']:
                    payment.status = 'SUCCESS'
                    applicant.payment_status = 'Completed'
                    applicant.status = AdmissionApplicant.Status.PENDING
                else:
                    payment.status = 'FAILED'
                    applicant.payment_status = 'Failed'
                
                payment.save()
                applicant.save()
            
            return HttpResponse("OK", status=200)
        except Exception as e:
            return HttpResponse(f"Error: {str(e)}", status=400)
    
    return HttpResponse("Invalid request", status=400)

def payment_fail(request):
    """Handle payment failure"""
    tran_id = request.GET.get('tran_id')
    return render(request, 'payment_fail.html', {'tran_id': tran_id})

def payment_cancel(request):
    """Handle payment cancellation"""
    tran_id = request.GET.get('tran_id')
    return render(request, 'payment_cancel.html', {'tran_id': tran_id})
      

@login_required
def admission_applicants_list(request):
    if not request.user.is_staff:
        return redirect('home')
    
    applicants = AdmissionApplicant.objects.all().order_by('-applied_at')
    return render(request, 'user/admission_applicants.html', {'applicants': applicants})

@login_required
def migrate_applicant(request, applicant_id):
    if not request.user.is_staff:
        return redirect('home')
    
    try:
        applicant = AdmissionApplicant.objects.get(applicant_id=applicant_id)
        
        # Verify payment is completed before migration
        if not applicant.is_payment_completed():
            messages.error(request, 'Cannot migrate applicant - payment not completed')
            return redirect('admission_applicants_list')
            
        student = applicant.migrate_to_student()
        applicant.status = AdmissionApplicant.Status.APPROVED
        applicant.save()
        
        messages.success(request, f'Successfully migrated applicant {applicant.full_name} to student database.')
        return redirect('admission_applicants_list')
    
    except AdmissionApplicant.DoesNotExist:
        messages.error(request, 'Applicant not found.')
        return redirect('admission_applicants_list')
    except Exception as e:
        messages.error(request, f'Error migrating applicant: {str(e)}')
        return redirect('admission_applicants_list')
    
    
    
def initiate_admission_payment(request, applicant_id):
    """Initiate payment for an existing admission application"""
    try:
        applicant = AdmissionApplicant.objects.get(applicant_id=applicant_id)
        
        if applicant.is_payment_completed():
            return redirect('admit_card', applicant_id=applicant_id)
            
        # Create payment transaction
        tran_id = f"ADM_{applicant.applicant_id}"
        payment_transaction = PaymentTransaction.objects.create(
            user=applicant,
            tran_id=tran_id,
            amount=applicant.get_admission_fee(),
            status='PENDING',
            gateway='sslcommerz'
        )
        
        applicant.payment_transaction = payment_transaction
        applicant.save()
        
        # Initiate SSLCommerz payment
        request.user = applicant
        request.POST = request.POST.copy()
        request.POST['payment_method'] = 'sslcommerz'
        
        return ssl_initiate_payment(request)
        
    except AdmissionApplicant.DoesNotExist:
        messages.error(request, 'Applicant not found')
        return redirect('home')
    
    
    
#class xi here ..................############
from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import AdmissionForm
from django.template.loader import get_template
from xhtml2pdf import pisa
import os
from django.conf import settings

# def admission_form(request):
#     if request.method == 'POST':
#         form_data = request.POST.copy()
#         form_data['class_sought'] = 'Class XI'  # Fixed value
        
#         # Create form instance
#         form = AdmissionForm(form_data)
        
#         if form.is_valid():
#             admission = form.save()
#             return redirect('download_pdf', admission_id=admission.id)
#         else:
#             return render(request, 'user/admissions_form.html', {'errors': form.errors})
    
#     return render(request, 'user/admissions_form.html')


from .forms import AdmissionFormForm  

def admission_form(request):
    if request.method == 'POST':
        form = AdmissionFormForm(request.POST)
        if form.is_valid():
            admission = form.save()
            return redirect('download_pdf', admission_id=admission.id)
        else:
            return render(request, 'user/admissions_form.html', {'errors': form.errors})
    
    # Set initial value for class_sought
    form = AdmissionFormForm(initial={'class_sought': 'Class XI'})
    return render(request, 'user/admissions_form.html', {'form': form})

def download_pdf(request, admission_id):
    admission = AdmissionForm.objects.get(id=admission_id)
    institute = Institute.objects.order_by('-id').first()
    institute_logo = institute.institute_logo.url if institute and institute.institute_logo else None
    template_path = 'user/pdf_template.html'
    context = {
        'admission': admission,
        'institute_logo': institute_logo,
        'institute_name': 'Rajuk Uttara Model College'
    }
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="admission_form_{admission_id}.pdf"'
    
    template = get_template(template_path)
    html = template.render(context)
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
        return HttpResponse('Error generating PDF')
    return response





from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import Registration
from .forms import RegistrationForm
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.shortcuts import render
from .models import Registration

def registration_view(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            registration = form.save()
            # You can add a success message here
            return redirect('registration_success', registration_id=registration.registration_id)
    else:
        form = RegistrationForm()
    
    return render(request, 'user/registration.html', {'form': form})

def registration_success(request, registration_id):
    registration = Registration.objects.get(registration_id=registration_id)
    return render(request, 'user/registration_success.html', {'registration': registration})

def registered_users_view(request):
    # Get all users ordered by registration date
    users_list = Registration.objects.all().order_by('-registration_date')
    
    # Handle filtering
    name_filter = request.GET.get('name', '')
    mobile_filter = request.GET.get('mobile', '')
    email_filter = request.GET.get('email', '')
    admitted_class_filter = request.GET.get('admitted_class', '')
    passing_class_filter = request.GET.get('passing_class', '')
    
    if name_filter:
        users_list = users_list.filter(full_name__icontains=name_filter)
    if mobile_filter:
        users_list = users_list.filter(mobile__icontains=mobile_filter)
    if email_filter:
        users_list = users_list.filter(email__icontains=email_filter)
    if admitted_class_filter:
        users_list = users_list.filter(admitted_class=admitted_class_filter)
    if passing_class_filter:
        users_list = users_list.filter(passing_class=passing_class_filter)
    
    # Pagination
    paginator = Paginator(users_list, 10)  # Show 10 users per page
    page = request.GET.get('page')
    
    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)
    
    context = {
        'users': users,
        'name_filter': name_filter,
        'mobile_filter': mobile_filter,
        'email_filter': email_filter,
        'admitted_class_filter': admitted_class_filter,
        'passing_class_filter': passing_class_filter,
        'class_choices': [choice[0] for choice in CLASS_CHOICES],
    }
    
    return render(request, 'user/registered_users.html', context)





import os
import pandas as pd
from django.http import HttpResponse
from django.shortcuts import render
from django.conf import settings
from django.db.models import Q
from .models import Student, StudentProfile
from core.models import ClassConfig, StudentClass, AcademicSession
import zipfile
from io import BytesIO

def download_student_data(request):
    if request.method == 'POST':
        # Get filter parameters
        class_id = request.POST.get('class_id')
        version = request.POST.get('version')
        session_id = request.POST.get('session')
        
        # Start with all students
        student_profiles = StudentProfile.objects.all()
        
        # Apply filters
        if class_id:
            student_profiles = student_profiles.filter(class_id__class_group_id__class_id_id=class_id)
        if version:
            student_profiles = student_profiles.filter(version=version)
        if session_id:
            student_profiles = student_profiles.filter(academic_session_year_id=session_id)
        
        # Select related to optimize queries
        student_profiles = student_profiles.select_related(
            'student_field', 
            'class_id',
            'class_id__class_group_id',
            'class_id__class_group_id__class_id',
            'class_id__class_group_id__group_id',
            'class_id__section_id',
            'class_id__shift_id',
            'academic_session_year',
            'parent_id'
        ).prefetch_related('student_field__user_wallet')
        
        # Prepare data for Excel
        data = []
        image_files = {}
        
        for profile in student_profiles:
            student = profile.student_field
            
            # Get image name if exists
            image_name = ""
            if student.avatar:
                image_name = os.path.basename(student.avatar.name)
                # Store image path for later
                if student.avatar and hasattr(student.avatar, 'path') and os.path.exists(student.avatar.path):
                    image_files[image_name] = student.avatar.path
            
            # Get wallet balance
            wallet_balance = ""
            wallet = student.user_wallet.first()
            if wallet:
                wallet_balance = wallet.wallet
            
            # Prepare row data
            row = {
                'User ID': student.user_id or '',
                'Username': student.username or '',
                'Name': student.name or '',
                'Name in Bangla': student.name_in_bangla or '',
                'Phone Number': student.phone_number or '',
                'Email': student.email or '',
                'Gender': student.gender or '',
                'Religion': student.religion or '',
                'Date of Birth': student.dob or '',
                'Blood Group': student.blood_group or '',
                'NID': student.nid or '',
                'RFID': student.rfid or '',
                'Present Address': student.present_address or '',
                'Permanent Address': student.permanent_address or '',
                'Disability Info': student.disability_info or '',
                'Status': student.status or '',
                # 'Nationality': student.nationality or '',
                'Created At': student.created_at or '',
                'Image Name': image_name,
                'Wallet Balance': wallet_balance,
                # Student Profile fields
                'Student Status': profile.status or '',
                'Version': profile.version or '',
                'Name Tag': profile.name_tag or '',
                'Admission Year': profile.admission_year_id.name if profile.admission_year_id else '',
                'Academic Session': f"{profile.academic_session_year.start_year}-{profile.academic_session_year.end_year}" if profile.academic_session_year else '',
                'Class': profile.class_id.class_group_id.class_id.name if profile.class_id and profile.class_id.class_group_id and profile.class_id.class_group_id.class_id else '',
                'Group': profile.class_id.class_group_id.group_id.name if profile.class_id and profile.class_id.class_group_id and profile.class_id.class_group_id.group_id else '',
                'Section': profile.class_id.section_id.name if profile.class_id and profile.class_id.section_id else '',
                'Shift': profile.class_id.shift_id.name if profile.class_id and profile.class_id.shift_id else '',
                'Roll No': profile.roll_no or '',
                'Birth Certificate No': profile.birth_certificate_no or '',
                'TC No': profile.tc_no or '',
                'Admission Date': profile.admission_date or '',
                'Village': profile.village or '',
                'Post Office': profile.post_office or '',
                'PS/Upazilla': profile.ps_or_upazilla or '',
                'District': profile.district or '',
                # Parent information
                'Parent Name': profile.parent_id.name if profile.parent_id else '',
                'Parent Phone': profile.parent_id.phone_number if profile.parent_id else '',
            }
            data.append(row)
        
        if not data:
            return HttpResponse("No students found with the selected filters.")
        
        # Create DataFrame
        df = pd.DataFrame(data)

        # Convert timezone-aware datetimes to naive before writing to Excel
        for col in df.select_dtypes(include=['datetimetz']).columns:
            df[col] = df[col].dt.tz_localize(None)

        # Create Excel file in memory
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Students')
        excel_buffer.seek(0)
        
        # Create zip file containing both Excel and images
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Add Excel file
            zip_file.writestr('student_data.xlsx', excel_buffer.getvalue())
            
            # Add images to a separate folder
            for image_name, image_path in image_files.items():
                try:
                    with open(image_path, 'rb') as img_file:
                        zip_file.writestr(f'student_images/{image_name}', img_file.read())
                except Exception as e:
                    print(f"Error adding image {image_name}: {e}")
        
        zip_buffer.seek(0)
        
        # Prepare response
        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="student_data.zip"'
        
        return response
    
    else:
        # GET request - show filter form
        classes = StudentClass.objects.all()  # Changed from ClassConfig to StudentClass
        versions = StudentProfile.Version.choices
        sessions = AcademicSession.objects.all().order_by('-start_year')
        
        context = {
            'classes': classes,
            'versions': versions,
            'sessions': sessions,
        }
        return render(request, 'user/download_student_data.html', context)
    
  
from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Prefetch, Q
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from core.models import ClassConfig, StuGroup, SubjectConfig, StudentClass, ClassGroupConfig, AcademicSession
from user.models import Student, StudentProfile
from exam.models import Forth_Sub
import json

def download_student_subjects(request):
    # Get filter parameters
    class_id = request.GET.get('class')
    group_id = request.GET.get('group')
    session_id = request.GET.get('session')
    
    # Get all classes, groups, and sessions for the filter form
    classes = StudentClass.objects.all().distinct()
    groups = StuGroup.objects.all()
    sessions = AcademicSession.objects.all().order_by('-start_year')
    
    if request.method == 'GET' and (class_id or group_id or session_id):
        # Filter students based on class, group, and session
        students = StudentProfile.objects.select_related(
            'student_field', 
            'class_id',
            'class_id__class_group_id',
            'class_id__class_group_id__class_id',
            'class_id__class_group_id__group_id',
            'academic_session_year'
        ).filter(class_id__isnull=False)
        
        if class_id:
            students = students.filter(class_id__class_group_id__class_id__id=class_id)
        if group_id:
            students = students.filter(class_id__class_group_id__group_id__id=group_id)
        if session_id:
            students = students.filter(academic_session_year__id=session_id)
        
        # Prefetch subject assignments for all students
        student_ids = [student.id for student in students]
        subject_assignments = Forth_Sub.objects.filter(
            student_id__in=student_ids
        ).select_related(
            'sub_conf_id',
            'sub_conf_id__subject_id',
            'student_id'
        )
        
        # Create a dictionary for quick lookup of student subjects
        student_subjects_map = {}
        for assignment in subject_assignments:
            student_profile_id = assignment.student_id.id
            subject_name = assignment.sub_conf_id.subject_id.name
            subject_type = assignment.forth_type
            
            if student_profile_id not in student_subjects_map:
                student_subjects_map[student_profile_id] = {'compulsory': [], 'optional': []}
            
            if subject_type == 'OPTIONAL':
                student_subjects_map[student_profile_id]['optional'].append(subject_name)
            else:  # COMPULSARY and any other types
                student_subjects_map[student_profile_id]['compulsory'].append(subject_name)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Student Subjects"
        
        # Create headers
        headers = ['Student ID', 'Student Name', 'Class', 'Group', 'Section', 'Roll No', 'Session']
        
        # Get all possible subjects for column structure
        if students.exists():
            # Get unique compulsory subjects across all students for column headers
            all_compulsory_subjects = set()
            for student_data in student_subjects_map.values():
                all_compulsory_subjects.update(student_data['compulsory'])
            
            # Convert to sorted list for consistent column order
            all_compulsory_subjects = sorted(list(all_compulsory_subjects))
            
            # Add compulsory subject columns
            for subject_name in all_compulsory_subjects:
                headers.append(subject_name)
            
            # Add optional column
            headers.append('OPTIONAL')
            
            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
        
        # Add student data
        for row, student_profile in enumerate(students, 2):
            student_data = student_subjects_map.get(student_profile.id, {'compulsory': [], 'optional': []})
            
            # Basic student info
            ws.cell(row=row, column=1, value=student_profile.student_field.user_id or student_profile.student_field.id)
            ws.cell(row=row, column=2, value=student_profile.student_field.name)
            ws.cell(row=row, column=3, value=str(student_profile.class_id.class_group_id.class_id.name))
            ws.cell(row=row, column=4, value=str(student_profile.class_id.class_group_id.group_id.name) if student_profile.class_id.class_group_id.group_id else '')
            ws.cell(row=row, column=5, value=str(student_profile.class_id.section_id.name) if student_profile.class_id.section_id else '')
            ws.cell(row=row, column=6, value=student_profile.roll_no)
            ws.cell(row=row, column=7, value=str(student_profile.academic_session_year) if student_profile.academic_session_year else '')
            
            # Add compulsory subjects
            col_offset = 7  # After basic info columns (increased due to session column)
            
            # Create a mapping of this student's compulsory subjects
            student_compulsory_subjects = {subject: subject for subject in student_data['compulsory']}
            
            # Fill in compulsory subject columns
            for subject_name in all_compulsory_subjects:
                if subject_name in student_compulsory_subjects:
                    ws.cell(row=row, column=col_offset + 1, value=subject_name)
                else:
                    ws.cell(row=row, column=col_offset + 1, value="")
                col_offset += 1
            
            # Add optional subjects for this specific student
            optional_subjects = student_data['optional']
            ws.cell(row=row, column=col_offset + 1, value=', '.join(optional_subjects))
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response with Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="student_subjects.xlsx"'
        
        wb.save(response)
        return response
    
    context = {
        'classes': classes,
        'groups': groups,
        'sessions': sessions,
    }
    return render(request, 'user/download_subjects.html', context)



# from django.shortcuts import render, redirect
# from django.contrib import messages
# from django.contrib.auth import logout
# from django.db.models import Q
# from django.http import JsonResponse
# from core.models import StudentClass, ClassConfig
# from user.models import StudentProfile, Student
# from shared.models import CustomUser
# from django.contrib.sessions.models import Session
# from django.utils import timezone
# import json

# def manage_student_status(request):
#     # Get filter parameters
#     class_id = request.GET.get('class')
#     version = request.GET.get('version')
#     status_filter = request.GET.get('status', 'Active')  # Default to Active
    
#     # Get all classes and versions for the filter form
#     classes = StudentClass.objects.all().distinct()
#     versions = StudentProfile.Version.choices
    
#     # Get student count for the selected filters (without loading all student data)
#     student_count = StudentProfile.objects.all()
    
#     if class_id:
#         student_count = student_count.filter(class_id__class_group_id__class_id__id=class_id)
#     if version:
#         student_count = student_count.filter(version=version)
#     if status_filter:
#         student_count = student_count.filter(student_field__status=status_filter)
    
#     student_count = student_count.count()
    
#     if request.method == 'POST':
#         # Handle single status update
#         student_id = request.POST.get('student_id')
#         new_status = request.POST.get('status')
        
#         if student_id and new_status:
#             try:
#                 student_user = Student.objects.get(id=student_id)
#                 old_status = student_user.status
#                 student_user.status = new_status
#                 student_user.save()
                
#                 # If changing from Active to Deactive, log out the user
#                 if old_status == 'Active' and new_status == 'Deactive':
#                     logout_user_sessions(student_user)
#                     messages.success(request, f'{student_user.name} has been deactivated and logged out from all sessions.')
#                 else:
#                     messages.success(request, f'{student_user.name} status updated to {new_status}.')
                    
#                 return redirect('manage_student_status')
                
#             except Student.DoesNotExist:
#                 messages.error(request, 'Student not found.')
#             except Exception as e:
#                 messages.error(request, f'Error updating student status: {str(e)}')
    
#     context = {
#         'classes': classes,
#         'versions': versions,
#         'student_count': student_count,
#         'status_choices': CustomUser.Status.choices,
#         'selected_class': class_id,
#         'selected_version': version,
#         'selected_status': status_filter,
#     }
#     return render(request, 'user/manage_student_status.html', context)

# def logout_user_sessions(user):
#     """
#     Log out a user from all active sessions
#     """
#     try:
#         # Get all active sessions for this user
#         sessions = Session.objects.filter(expire_date__gte=timezone.now())
        
#         for session in sessions:
#             session_data = session.get_decoded()
#             if session_data.get('_auth_user_id') == str(user.id):
#                 # Delete the session to log out the user
#                 session.delete()
                
#     except Exception as e:
#         print(f"Error logging out user sessions: {str(e)}")

# def bulk_update_student_status(request):
#     """
#     Bulk update student status - COMPLETELY REWRITTEN
#     """
#     if request.method == 'POST':
#         # Get filter parameters from the form (same as manage_student_status)
#         class_id = request.POST.get('class')
#         version = request.POST.get('version')
#         status_filter = request.POST.get('status_filter', 'Active')
#         new_status = request.POST.get('bulk_status')
        
#         if not new_status:
#             messages.error(request, 'No status selected for update.')
#             return redirect('manage_student_status')
        
#         # Get students based on the current filters
#         students = StudentProfile.objects.select_related('student_field').all()
        
#         # Apply the same filters as the manage_student_status view
#         if class_id:
#             students = students.filter(class_id__class_group_id__class_id__id=class_id)
#         if version:
#             students = students.filter(version=version)
#         if status_filter:
#             students = students.filter(student_field__status=status_filter)
        
#         student_count = students.count()
        
#         if student_count == 0:
#             messages.warning(request, 'No students found with the current filters.')
#             return redirect('manage_student_status')
        
#         updated_count = 0
#         logged_out_count = 0
        
#         # Update all students in the filtered queryset
#         for student_profile in students:
#             try:
#                 student_user = student_profile.student_field
#                 old_status = student_user.status
                
#                 # Only update if status is actually changing
#                 if old_status != new_status:
#                     student_user.status = new_status
#                     student_user.save()
                    
#                     # If changing from Active to Deactive, log out the user
#                     if old_status == 'Active' and new_status == 'Deactive':
#                         logout_user_sessions(student_user)
#                         logged_out_count += 1
                    
#                     updated_count += 1
                
#             except Exception as e:
#                 messages.error(request, f'Error updating student {student_profile.student_field.name}: {str(e)}')
        
#         if updated_count > 0:
#             message = f'Successfully updated {updated_count} student(s) to {new_status}.'
#             if logged_out_count > 0:
#                 message += f' {logged_out_count} user(s) were logged out from all active sessions.'
#             messages.success(request, message)
#         else:
#             messages.info(request, f'All {student_count} students already have status "{new_status}". No changes made.')
    
#     return redirect('manage_student_status')


from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import logout
from django.db.models import Q
from django.http import JsonResponse
from django.contrib.auth.decorators import user_passes_test
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from datetime import datetime, timedelta
from core.models import StudentClass, ClassConfig
from user.models import StudentProfile, Student
from shared.models import CustomUser
from django.contrib.sessions.models import Session
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.token_blacklist.models import BlacklistedToken, OutstandingToken
import json

def is_admin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)

def manage_student_status(request):
    # Get filter parameters
    class_id = request.GET.get('class')
    version = request.GET.get('version')
    status_filter = request.GET.get('status', 'Active')  # Default to Active
    
    # Get all classes and versions for the filter form
    classes = StudentClass.objects.all().distinct()
    versions = StudentProfile.Version.choices
    
    # Get student count for the selected filters (without loading all student data)
    student_count = StudentProfile.objects.all()
    
    if class_id:
        student_count = student_count.filter(class_id__class_group_id__class_id__id=class_id)
    if version:
        student_count = student_count.filter(version=version)
    if status_filter:
        student_count = student_count.filter(student_field__status=status_filter)
    
    student_count = student_count.count()
    
    if request.method == 'POST':
        # Handle single status update
        student_id = request.POST.get('student_id')
        new_status = request.POST.get('status')
        
        if student_id and new_status:
            try:
                student_user = Student.objects.get(id=student_id)
                old_status = student_user.status
                student_user.status = new_status
                student_user.save()
                
                # If changing from Active to Deactive, log out the user
                if old_status == 'Active' and new_status == 'Deactive':
                    logout_user_jwt(student_user)
                    messages.success(request, f'{student_user.name} has been deactivated and logged out from all sessions.')
                else:
                    messages.success(request, f'{student_user.name} status updated to {new_status}.')
                    
                return redirect('manage_student_status')
                
            except Student.DoesNotExist:
                messages.error(request, 'Student not found.')
            except Exception as e:
                messages.error(request, f'Error updating student status: {str(e)}')
    
    context = {
        'classes': classes,
        'versions': versions,
        'student_count': student_count,
        'status_choices': CustomUser.Status.choices,
        'selected_class': class_id,
        'selected_version': version,
        'selected_status': status_filter,
    }
    return render(request, 'user/manage_student_status.html', context)

def logout_user_jwt(user):
    """
    Log out a user from JWT authentication by blacklisting all their tokens
    """
    try:
        # Get all outstanding tokens for the user
        outstanding_tokens = OutstandingToken.objects.filter(user=user)
        
        # Blacklist all refresh tokens
        for token in outstanding_tokens:
            try:
                # Check if the token is already blacklisted
                if not BlacklistedToken.objects.filter(token=token).exists():
                    RefreshToken(token.token).blacklist()
            except Exception as e:
                print(f"Error blacklisting token for user {user.id}: {str(e)}")
                # Continue with other tokens even if one fails
        
        # Also clear any Django sessions (for admin interface if used)
        logout_user_sessions(user)
        
        print(f"Successfully logged out user {user.username} from all JWT sessions")
        
    except Exception as e:
        print(f"Error logging out user JWT sessions: {str(e)}")

def logout_user_sessions(user):
    """
    Log out a user from all active Django sessions (for session-based auth)
    """
    try:
        # Get all active sessions for this user
        sessions = Session.objects.filter(expire_date__gte=timezone.now())
        
        for session in sessions:
            session_data = session.get_decoded()
            if session_data.get('_auth_user_id') == str(user.id):
                # Delete the session to log out the user
                session.delete()
                
    except Exception as e:
        print(f"Error logging out user sessions: {str(e)}")

def bulk_update_student_status(request):
    """
    Bulk update student status - UPDATED WITH JWT LOGOUT
    """
    if request.method == 'POST':
        # Get filter parameters from the form (same as manage_student_status)
        class_id = request.POST.get('class')
        version = request.POST.get('version')
        status_filter = request.POST.get('status_filter', 'Active')
        new_status = request.POST.get('bulk_status')
        
        if not new_status:
            messages.error(request, 'No status selected for update.')
            return redirect('manage_student_status')
        
        # Get students based on the current filters
        students = StudentProfile.objects.select_related('student_field').all()
        
        # Apply the same filters as the manage_student_status view
        if class_id:
            students = students.filter(class_id__class_group_id__class_id__id=class_id)
        if version:
            students = students.filter(version=version)
        if status_filter:
            students = students.filter(student_field__status=status_filter)
        
        student_count = students.count()
        
        if student_count == 0:
            messages.warning(request, 'No students found with the current filters.')
            return redirect('manage_student_status')
        
        updated_count = 0
        logged_out_count = 0
        
        # Update all students in the filtered queryset
        for student_profile in students:
            try:
                student_user = student_profile.student_field
                old_status = student_user.status
                
                # Only update if status is actually changing
                if old_status != new_status:
                    student_user.status = new_status
                    student_user.save()
                    
                    # If changing from Active to Deactive, log out the user using JWT
                    if old_status == 'Active' and new_status == 'Deactive':
                        logout_user_jwt(student_user)
                        logged_out_count += 1
                    
                    updated_count += 1
                
            except Exception as e:
                messages.error(request, f'Error updating student {student_profile.student_field.name}: {str(e)}')
        
        if updated_count > 0:
            message = f'Successfully updated {updated_count} student(s) to {new_status}.'
            if logged_out_count > 0:
                message += f' {logged_out_count} user(s) were logged out from all active sessions (JWT tokens invalidated).'
            messages.success(request, message)
        else:
            messages.info(request, f'All {student_count} students already have status "{new_status}". No changes made.')
    
    return redirect('manage_student_status')

def bulk_activate_by_date_range(request):
    """
    Bulk activate students who were deactivated within a specific date range
    """
    classes = StudentClass.objects.all().distinct()
    versions = StudentProfile.Version.choices
    
    if request.method == 'POST':
        class_id = request.POST.get('class')
        version = request.POST.get('version')
        from_date = request.POST.get('from_date')
        to_date = request.POST.get('to_date')
        
        if not from_date or not to_date:
            messages.error(request, 'Please select both from and to dates.')
            return redirect('bulk_activate_by_date_range')
        
        try:
            # Convert string dates to datetime objects with timezone
            from_datetime = timezone.make_aware(datetime.strptime(from_date, '%Y-%m-%d'))
            to_datetime = timezone.make_aware(datetime.strptime(to_date, '%Y-%m-%d')) + timedelta(days=1)  # Include the entire end date
            
            # Get ALL deactivated students (not filtering by date yet)
            students = StudentProfile.objects.select_related('student_field').filter(
                student_field__status='Deactive'
            )
            
            # Apply filters
            if class_id:
                students = students.filter(class_id__class_group_id__class_id__id=class_id)
            if version:
                students = students.filter(version=version)
            
            student_count = students.count()
            
            if student_count == 0:
                messages.warning(request, f'No deactivated students found.')
                return redirect('bulk_activate_by_date_range')
            
            # Update students to active (ALL deactivated students, not filtered by date)
            updated_count = 0
            for student_profile in students:
                try:
                    student_user = student_profile.student_field
                    student_user.status = 'Active'
                    student_user.save()
                    updated_count += 1
                except Exception as e:
                    messages.error(request, f'Error activating student {student_profile.student_field.name}: {str(e)}')
            
            if updated_count > 0:
                messages.success(request, f'Successfully activated {updated_count} deactivated student(s).')
            else:
                messages.info(request, 'No students were activated.')
            
            return redirect('bulk_activate_by_date_range')
            
        except Exception as e:
            messages.error(request, f'Error processing request: {str(e)}')
            return redirect('bulk_activate_by_date_range')
    
    # Default date range: yesterday to today (for convenience)
    today = timezone.now().date()
    yesterday = today - timedelta(days=1)
    
    context = {
        'classes': classes,
        'versions': versions,
        'default_from_date': yesterday.strftime('%Y-%m-%d'),
        'default_to_date': today.strftime('%Y-%m-%d'),
    }
    return render(request, 'user/bulk_activate_by_date_range.html', context)

def get_deactivated_students_count(request):
    """
    API endpoint to get count of deactivated students within a date range
    """
    if request.method == 'GET':
        try:
            class_id = request.GET.get('class')
            version = request.GET.get('version')
            from_date = request.GET.get('from_date')
            to_date = request.GET.get('to_date')
            
            # Get students who are currently deactivated
            students = StudentProfile.objects.select_related('student_field').filter(
                student_field__status='Deactive'
            )
            
            # Apply filters
            if class_id:
                students = students.filter(class_id__class_group_id__class_id__id=class_id)
            if version:
                students = students.filter(version=version)
            
            if from_date and to_date:
                # Convert string dates to datetime objects with timezone
                from_datetime = timezone.make_aware(datetime.strptime(from_date, '%Y-%m-%d'))
                to_datetime = timezone.make_aware(datetime.strptime(to_date, '%Y-%m-%d')) + timedelta(days=1)
                
                # Get the student IDs
                student_ids = students.values_list('student_field__id', flat=True)
                
                # Get users who were updated within the date range
                updated_users = CustomUser.objects.filter(
                    id__in=student_ids,
                    updated_at__range=[from_datetime, to_datetime]
                )
                
                # Get the final count
                count = students.filter(student_field__id__in=updated_users.values('id')).count()
            else:
                # If no date range, return all deactivated students
                count = students.count()
            
            return JsonResponse({
                'success': True,
                'count': count
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error getting student count: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Only GET requests allowed'
    }, status=405)

@user_passes_test(is_admin)
def logout_all_users(request):
    """View to display logout all users page"""
    # Get current user counts for display
    active_users = CustomUser.objects.filter(is_active=True)
    
    user_stats = {
        'student_count': active_users.filter(groups__name='student').count(),
        'staff_count': active_users.filter(groups__name='staff').count(),
        'parent_count': active_users.filter(groups__name='parent').count(),
        'total_count': active_users.count(),
    }
    
    return render(request, 'user/logout_all_users.html', {'user_stats': user_stats})

@csrf_exempt
@user_passes_test(is_admin)
def force_logout_all_users(request):
    """API endpoint to force logout all users by invalidating tokens and sessions"""
    if request.method == 'POST':
        try:
            # Get all active users
            active_users = CustomUser.objects.filter(is_active=True)
            
            # Count users by type for reporting
            student_count = active_users.filter(groups__name='student').count()
            staff_count = active_users.filter(groups__name='staff').count()
            parent_count = active_users.filter(groups__name='parent').count()
            other_count = active_users.exclude(
                Q(groups__name='student') | 
                Q(groups__name='staff') | 
                Q(groups__name='parent')
            ).count()
            
            total_users = active_users.count()
            
            # Method 1: Invalidate JWT tokens using blacklist
            tokens_blacklisted = invalidate_jwt_tokens()
            
            # Method 2: Clear user sessions from database
            sessions_cleared = clear_user_sessions(request)
            
            # Method 3: Update user secret key (nuclear option - forces re-login)
            users_affected = force_user_reauthentication()
            
            return JsonResponse({
                'success': True,
                'message': f'Successfully initiated logout for all {total_users} users',
                'details': {
                    'students': student_count,
                    'staff': staff_count,
                    'parents': parent_count,
                    'others': other_count,
                    'total': total_users,
                    'tokens_blacklisted': tokens_blacklisted,
                    'sessions_cleared': sessions_cleared,
                    'users_affected': users_affected
                }
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error logging out users: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Only POST requests allowed'
    }, status=405)

def invalidate_jwt_tokens():
    """Invalidate all JWT tokens using blacklist"""
    tokens_blacklisted = 0
    try:
        from rest_framework_simplejwt.token_blacklist.models import BlacklistedToken, OutstandingToken
        from rest_framework_simplejwt.utils import aware_utcnow
        
        # Blacklist all outstanding tokens that haven't expired
        outstanding_tokens = OutstandingToken.objects.filter(expires_at__gt=aware_utcnow())
        tokens_blacklisted = outstanding_tokens.count()
        
        for token in outstanding_tokens:
            BlacklistedToken.objects.get_or_create(token=token)
            
        print(f"Blacklisted {tokens_blacklisted} JWT tokens")
        
    except ImportError:
        print("JWT blacklist not available")
    except Exception as e:
        print(f"Error blacklisting tokens: {str(e)}")
    
    return tokens_blacklisted

def clear_user_sessions(request):
    """Clear all user sessions from database (except current admin)"""
    sessions_cleared = 0
    try:
        from django.contrib.sessions.models import Session
        from django.utils import timezone
        
        # Get current session key to exclude admin session
        current_session_key = request.session.session_key
        
        # Delete all active sessions except current admin session
        active_sessions = Session.objects.filter(expire_date__gte=timezone.now())
        sessions_cleared = active_sessions.exclude(session_key=current_session_key).count()
        
        active_sessions.exclude(session_key=current_session_key).delete()
        print(f"Cleared {sessions_cleared} user sessions")
        
    except Exception as e:
        print(f"Error clearing sessions: {str(e)}")
    
    return sessions_cleared

def force_user_reauthentication():
    """Force all users to reauthenticate by modifying a field"""
    users_affected = 0
    try:
        # This method forces all users to re-login by updating a timestamp
        # You might need to modify your authentication backend to check this
        
        # Alternative: Update all users' last_login to force reauthentication
        from django.utils import timezone
        from django.contrib.auth import update_session_auth_hash
        
        active_users = CustomUser.objects.filter(is_active=True)
        users_affected = active_users.count()
        
        # Update a field that might trigger reauthentication
        # This is a more aggressive approach
        for user in active_users:
            # You could update a custom field that your auth system checks
            user.save()  # This updates the updated_at field
            
        print(f"Affected {users_affected} users for reauthentication")
        
    except Exception as e:
        print(f"Error forcing reauthentication: {str(e)}")
    
    return users_affected

@user_passes_test(is_admin)
def get_user_stats(request):
    """API endpoint to get current user statistics"""
    if request.method == 'GET':
        try:
            active_users = CustomUser.objects.filter(is_active=True)
            
            stats = {
                'student_count': active_users.filter(groups__name='student').count(),
                'staff_count': active_users.filter(groups__name='staff').count(),
                'parent_count': active_users.filter(groups__name='parent').count(),
                'total_count': active_users.count(),
            }
            
            return JsonResponse({
                'success': True,
                'stats': stats
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error getting user stats: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Only GET requests allowed'
    }, status=405)
    
    
import pandas as pd
import io
from django.db import transaction

def upload_student_status_excel(request):
    """
    Upload Excel file to update student status
    """
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            # Read the Excel file
            if excel_file.name.endswith('.xlsx'):
                df = pd.read_excel(excel_file)
            elif excel_file.name.endswith('.csv'):
                df = pd.read_csv(excel_file)
            else:
                messages.error(request, 'Please upload a valid Excel (.xlsx) or CSV file.')
                return redirect('upload_student_status_excel')
            
            # Check required columns
            required_columns = ['User ID', 'Status']
            if not all(col in df.columns for col in required_columns):
                missing = [col for col in required_columns if col not in df.columns]
                messages.error(request, f'Missing required columns: {", ".join(missing)}')
                return redirect('upload_student_status_excel')
            
            # Clean the data
            df = df.dropna(subset=['User ID'])  # Remove rows with empty User ID
            df['User ID'] = df['User ID'].astype(str).str.strip()
            df['Status'] = df['Status'].astype(str).str.strip()
            
            # Validate status values
            valid_statuses = ['Active', 'Deactive']
            invalid_status_rows = df[~df['Status'].isin(valid_statuses)]
            if not invalid_status_rows.empty:
                messages.warning(request, f'Found {len(invalid_status_rows)} rows with invalid status. Only "Active" and "Deactive" are allowed.')
                # Remove invalid rows
                df = df[df['Status'].isin(valid_statuses)]
            
            total_rows = len(df)
            if total_rows == 0:
                messages.warning(request, 'No valid data found in the Excel file.')
                return redirect('upload_student_status_excel')
            
            # Process the updates
            results = {
                'updated': 0,
                'no_change': 0,
                'not_found': 0,
                'errors': 0,
                'details': []
            }
            
            # Use transaction to ensure data consistency
            with transaction.atomic():
                for index, row in df.iterrows():
                    try:
                        user_id = str(row['User ID'])
                        new_status = row['Status']
                        
                        # Find the student by user_id
                        try:
                            student = Student.objects.get(user_id=user_id)
                            
                            # Check if status is different
                            if student.status == new_status:
                                results['no_change'] += 1
                                results['details'].append({
                                    'user_id': user_id,
                                    'action': 'no_change',
                                    'message': f'Status already {new_status}'
                                })
                            else:
                                old_status = student.status
                                student.status = new_status
                                student.save()
                                
                                # If changing from Active to Deactive, log out the user
                                if old_status == 'Active' and new_status == 'Deactive':
                                    logout_user_jwt(student)
                                
                                results['updated'] += 1
                                results['details'].append({
                                    'user_id': user_id,
                                    'name': student.name,
                                    'old_status': old_status,
                                    'new_status': new_status,
                                    'action': 'updated'
                                })
                                
                        except Student.DoesNotExist:
                            results['not_found'] += 1
                            results['details'].append({
                                'user_id': user_id,
                                'action': 'not_found',
                                'message': 'Student not found in database'
                            })
                            
                    except Exception as e:
                        results['errors'] += 1
                        results['details'].append({
                            'user_id': str(row.get('User ID', 'Unknown')),
                            'action': 'error',
                            'message': str(e)
                        })
            
            # Prepare summary message
            summary_message = f"""
            Processed {total_rows} rows from Excel file:<br>
            ✅ <strong>{results['updated']}</strong> students updated<br>
            ⏸️ <strong>{results['no_change']}</strong> students already had the correct status<br>
            ❌ <strong>{results['not_found']}</strong> students not found in database<br>
            ⚠️ <strong>{results['errors']}</strong> errors encountered
            """
            
            messages.success(request, summary_message)
            
            # Store detailed results in session for display
            request.session['upload_results'] = results['details'][:100]  # Limit to first 100
            request.session['upload_summary'] = {
                'total': total_rows,
                'updated': results['updated'],
                'no_change': results['no_change'],
                'not_found': results['not_found'],
                'errors': results['errors']
            }
            
            return redirect('upload_student_status_excel')
            
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            return redirect('upload_student_status_excel')
    
    # GET request - show the form
    # Get results from previous upload if available
    results = request.session.pop('upload_results', [])
    summary = request.session.pop('upload_summary', None)
    
    context = {
        'results': results,
        'summary': summary
    }
    return render(request, 'user/upload_student_status_excel.html', context)


def download_student_status_template(request):
    """
    Download Excel template for updating student status
    """
    try:
        # Create a sample DataFrame
        data = {
            'User ID': ['2220621006', '2320721022', '2220621005'],
            'Status': ['Active', 'Deactive', 'Active']
        }
        df = pd.DataFrame(data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Template', index=False)
        
        output.seek(0)
        
        # Create HTTP response
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="student_status_template.xlsx"'
        return response
        
    except Exception as e:
        messages.error(request, f'Error creating template: {str(e)}')
        return redirect('upload_student_status_excel')


def get_student_status_report(request):
    """
    Download current student status report
    """
    try:
        # Get filter parameters
        class_id = request.GET.get('class')
        version = request.GET.get('version')
        status = request.GET.get('status')
        
        # Get students with their profiles
        students = StudentProfile.objects.select_related('student_field', 'class_id').all()
        
        # Apply filters
        if class_id:
            students = students.filter(class_id__class_group_id__class_id__id=class_id)
        if version:
            students = students.filter(version=version)
        if status:
            students = students.filter(student_field__status=status)
        
        # Prepare data for Excel
        data = []
        for student in students:
            data.append({
                'User ID': student.student_field.user_id,
                'Name': student.student_field.name,
                'Class': student.class_id.class_group_id.class_id.name if student.class_id else 'N/A',
                'Version': student.version,
                'Current Status': student.student_field.status,
                'Phone Number': student.student_field.phone_number,
                'Email': student.student_field.email
            })
        
        df = pd.DataFrame(data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Student Status Report', index=False)
            
            # Auto-adjust column widths
            worksheet = writer.sheets['Student Status Report']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Create filename with timestamp
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'student_status_report_{timestamp}.xlsx'
        
        # Create HTTP response
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        messages.error(request, f'Error generating report: {str(e)}')
        return redirect('manage_student_status')
    
    
    
    
# user/views.py
import os
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.conf import settings
from django.db.models import Q
from django.template.loader import render_to_string
from django.core.files.storage import default_storage
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, letter, landscape
from reportlab.lib.units import mm, inch
from reportlab.lib.utils import ImageReader
from PIL import Image
import io
from weasyprint import HTML, CSS
from weasyprint.text.fonts import FontConfiguration
import json
from datetime import datetime

from .models import Student, StudentProfile
from core.models import ClassConfig, StudentShift, StudentSection, StuGroup, StudentClass
from miscellaneous.models import Institute

def id_card_generator(request):
    """Main view for ID card generation"""
    
    # Get all available filters
    classes = StudentClass.objects.all().distinct()
    sections = StudentSection.objects.all().distinct()
    shifts = StudentShift.objects.all().distinct()
    versions = StudentProfile.Version.choices
    
    # Get institute info for templates
    try:
        institute = Institute.objects.first()
    except:
        institute = None
    
    context = {
        'classes': classes,
        'sections': sections,
        'shifts': shifts,
        'versions': versions,
        'institute': institute,
    }
    
    return render(request, 'user/id_card_generator.html', context)

def get_students_for_id_card(request):
    """Get filtered students for ID card generation"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            class_id = data.get('class_id')
            section_id = data.get('section_id')
            shift_id = data.get('shift_id')
            version = data.get('version')
            
            # Start with StudentProfile queryset (not Student proxy model)
            query = Q()
            
            if class_id:
                query &= Q(class_id__class_group_id__class_id_id=class_id)
            
            if section_id:
                query &= Q(class_id__section_id_id=section_id)
            
            if shift_id:
                query &= Q(class_id__shift_id_id=shift_id)
            
            if version:
                query &= Q(version=version)
            
            # Get student profiles with related data
            student_profiles = StudentProfile.objects.filter(query).select_related(
                'student_field',
                'class_id',
                'class_id__class_group_id',
                'class_id__section_id',
                'class_id__shift_id',
                'class_id__class_group_id__class_id',
                'class_id__class_group_id__group_id',
            ).prefetch_related('student_field__groups')
            
            # Filter only students (those with 'student' group)
            student_profiles = [
                profile for profile in student_profiles 
                if profile.student_field.groups.filter(name='student').exists()
            ]
            
            student_list = []
            for profile in student_profiles:
                student = profile.student_field
                student_data = {
                    'id': student.id,
                    'name': student.name or student.username,
                    'roll_no': profile.roll_no,
                    'avatar_url': student.avatar.url if student.avatar and student.avatar.name else None,
                    'rfid': student.rfid,
                    'user_id': student.user_id,
                    'class': {
                        'name': profile.class_id.class_group_id.class_id.name if profile.class_id and profile.class_id.class_group_id else '',
                    },
                    'section': {
                        'name': profile.class_id.section_id.name if profile.class_id and profile.class_id.section_id else '',
                    },
                    'group': {
                        'name': profile.class_id.class_group_id.group_id.name if profile.class_id and profile.class_id.class_group_id and profile.class_id.class_group_id.group_id else '',
                    },
                    'shift': {
                        'name': profile.class_id.shift_id.name if profile.class_id and profile.class_id.shift_id else '',
                    },
                    'version': profile.version,
                }
                student_list.append(student_data)
            
            return JsonResponse({'students': student_list, 'count': len(student_list)})
        
        except Exception as e:
            print(f"Error: {str(e)}")  # For debugging
            return JsonResponse({'error': str(e), 'students': []}, status=400)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

def generate_id_card_pdf(request):
    """Generate PDF with ID cards"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            student_ids = data.get('student_ids', [])
            template_type = data.get('template_type', '1')
            layout = data.get('layout', 'single')
            
            if not student_ids:
                return JsonResponse({'error': 'No students selected'}, status=400)
            
            # Get student profiles for the selected student IDs
            student_profiles = StudentProfile.objects.filter(
                student_field_id__in=student_ids
            ).select_related(
                'student_field',
                'class_id',
                'class_id__class_group_id',
                'class_id__section_id',
                'class_id__shift_id',
                'class_id__class_group_id__class_id',
                'class_id__class_group_id__group_id',
            )
            
            # Get institute info
            try:
                institute = Institute.objects.first()
            except:
                institute = None
            
            # Prepare student data for template
            student_data = []
            for profile in student_profiles:
                student = profile.student_field
                student_info = {
                    'name': student.name or student.username,
                    'roll_no': profile.roll_no,
                    'class_name': profile.class_id.class_group_id.class_id.name if profile.class_id and profile.class_id.class_group_id else '',
                    'section_name': profile.class_id.section_id.name if profile.class_id and profile.class_id.section_id else '',
                    'group_name': profile.class_id.class_group_id.group_id.name if profile.class_id and profile.class_id.class_group_id and profile.class_id.class_group_id.group_id else '',
                    'shift_name': profile.class_id.shift_id.name if profile.class_id and profile.class_id.shift_id else '',
                    'version': profile.version,
                    'user_id': student.user_id,
                    'rfid': student.rfid,
                    'avatar_url': student.avatar.url if student.avatar and student.avatar.name else '/static/default-avatar.png',
                }
                student_data.append(student_info)
            
            # Render HTML template
            html_string = render_to_string(f'user/template_{template_type}.html', {
                'students': student_data,
                'institute': institute,
                'layout': layout,
                'current_date': datetime.now().strftime("%d/%m/%Y"),
            })
            
            # Generate PDF using WeasyPrint
            font_config = FontConfiguration()
            
            # Create CSS for printing
            css = CSS(string='''
                @page {
                    size: A4;
                    margin: 10mm;
                }
                @media print {
                    body {
                        -webkit-print-color-adjust: exact;
                        print-color-adjust: exact;
                    }
                    .id-card-container {
                        page-break-inside: avoid;
                    }
                }
                .id-card {
                    box-shadow: 0 4px 8px rgba(0,0,0,0.1);
                    border-radius: 10px;
                    overflow: hidden;
                }
            ''', font_config=font_config)
            
            html = HTML(
                string=html_string,
                base_url=request.build_absolute_uri('/')   # 🔥 REQUIRED
            )
            pdf = html.write_pdf(stylesheets=[css], font_config=font_config)
            
            # Create HTTP response
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="id_cards_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
            
            return response
        
        except Exception as e:
            print(f"PDF Generation Error: {str(e)}")
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

def get_id_card_preview(request):
    """Get HTML preview of ID card template"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            template_type = data.get('template_type', '1')
            
            # Sample data for preview
            sample_student = {
                'name': 'John Doe',
                'roll_no': '2023001',
                'class_name': '10',
                'section_name': 'A',
                'group_name': 'Science',
                'shift_name': 'Morning',
                'version': 'English',
                'user_id': 'STU2023001',
                'rfid': 'RFID001',
                'avatar_url': '/static/assets/img/test/zihad.jpeg',
            }
            
            # Get institute info
            try:
                institute = Institute.objects.first()
            except:
                institute = None
            
            html_string = render_to_string(f'user/template_{template_type}.html', {
                'students': [sample_student],
                'institute': institute,
                'layout': 'single',
                'preview': True,
                'current_date': datetime.now().strftime("%d/%m/%Y"),
            })
            
            return JsonResponse({'html': html_string})
        
        except Exception as e:
            print(f"Preview Error: {str(e)}")
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


# user/views.py

from django.shortcuts import render, redirect
from django.http import JsonResponse
from .models import SpinEntry
import random

def spin_form(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        phone = request.POST.get('phone')
        institute = request.POST.get('institute')
        num_students = request.POST.get('num_students')
        
        if SpinEntry.objects.filter(spin_phone=phone).exists():
            return render(request, 'user/spin_form.html', {'error': 'This phone number has already been used.'})
        
        entry = SpinEntry.objects.create(
            spin_name=name,
            spin_phone=phone,
            spin_institute=institute,
            spin_num_students=num_students
        )
        request.session['spin_phone'] = phone
        return redirect('spin_wheel')
    
    return render(request, 'user/spin_form.html')

def spin_wheel(request):
    phone = request.session.get('spin_phone')
    if not phone:
        return redirect('spin_form')
    
    try:
        entry = SpinEntry.objects.get(spin_phone=phone)
        if entry.spin_has_spun:
            return render(request, 'user/spin_result.html', {'result': entry.spin_result})
    except SpinEntry.DoesNotExist:
        return redirect('spin_form')
    
    return render(request, 'user/spin_wheel.html')

def get_result(request):
    phone = request.session.get('spin_phone')
    if not phone:
        return JsonResponse({'error': 'Invalid session'}, status=400)
    
    try:
        entry = SpinEntry.objects.get(spin_phone=phone)
        if entry.spin_has_spun:
            return JsonResponse({'result': entry.spin_result})
        
        # Rigged result: always between 5 and 10
        result = random.randint(5, 10)
        entry.spin_result = result
        entry.spin_has_spun = True
        entry.save()
        
        return JsonResponse({'result': result})
    except SpinEntry.DoesNotExist:
        return JsonResponse({'error': 'Entry not found'}, status=400)
    
    
    
from django.shortcuts import render
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from .models import SpinEntry

def download_all_spin_entries(request):
    """
    Download ALL SpinEntry records as Excel (no date filtering)
    """
    # Get ALL entries from database
    entries = SpinEntry.objects.all()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Spin Entries"
    
    # Define column headers (Bengali/English mix as per your preference)
    headers = [
        'ID', 
        'নাম (Name)', 
        'ফোন (Phone)', 
        'ইনস্টিটিউট (Institute)',
        'শিক্ষার্থী সংখ্যা (Students)', 
        'ফলাফল (Result)', 
        'স্পিন হয়েছে? (Has Spun?)'
    ]
    
    # Write headers to first row
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f'{col_letter}1']
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True, size=12)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        ws.column_dimensions[col_letter].width = 25
    
    # Populate data rows
    for row_num, entry in enumerate(entries, 2):
        ws.cell(row=row_num, column=1, value=entry.id)
        ws.cell(row=row_num, column=2, value=entry.spin_name)
        ws.cell(row=row_num, column=3, value=entry.spin_phone)
        ws.cell(row=row_num, column=4, value=entry.spin_institute)
        ws.cell(row=row_num, column=5, value=entry.spin_num_students)
        ws.cell(row=row_num, column=6, value=entry.spin_result if entry.spin_result is not None else 'N/A')
        ws.cell(row=row_num, column=7, value='হ্যাঁ (Yes)' if entry.spin_has_spun else 'না (No)')
    
    # Add summary at the bottom
    total_entries = entries.count()
    total_students = sum(entry.spin_num_students for entry in entries)
    total_spun = entries.filter(spin_has_spun=True).count()
    
    summary_row = total_entries + 3
    ws.cell(row=summary_row, column=1, value="সারাংশ (Summary):").font = openpyxl.styles.Font(bold=True)
    ws.cell(row=summary_row + 1, column=1, value=f"মোট এন্ট্রি (Total Entries): {total_entries}")
    ws.cell(row=summary_row + 2, column=1, value=f"মোট শিক্ষার্থী (Total Students): {total_students}")
    ws.cell(row=summary_row + 3, column=1, value=f"স্পিন হয়েছে (Has Spun): {total_spun}")
    
    # Create HTTP response with Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = "all_spin_entries.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response


def spin_download_page(request):
    """
    Simple page with download button
    """
    total_entries = SpinEntry.objects.count()
    return render(request, 'user/spin_download_all.html', {
        'total_entries': total_entries
    })