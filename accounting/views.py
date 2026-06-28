# Python Standard Library
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from io import BytesIO
import tempfile
import logging

# Django Modules
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.template.loader import get_template, render_to_string
from django.utils.timezone import now, timezone
from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction, IntegrityError
from django.db.models import (
    Sum, F, Q, Value, CharField, DecimalField
)
from django.db.models.functions import Coalesce
from django.urls import reverse
from django.views import View
from django.contrib import messages
from django.utils.dateparse import parse_date
from django.utils.dateformat import format
from django.http import HttpResponse, JsonResponse, FileResponse, HttpResponseRedirect
from django.contrib.auth.decorators import user_passes_test, login_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

# Third-Party Libraries
from xhtml2pdf import pisa
from reportlab.pdfgen import canvas
from weasyprint import HTML

# Application-Specific Modules
from crucial.models import StudentProfile, FeeHead
from user.decorators import is_staff_or_has_role
from .models import *
from .forms import *






#________________ Accounting Part ____________________

def ledger_list(request):
    categories = LedgerCategory.objects.all()
    ledgers = Ledger.objects.all().select_related('category')
    ledger_queryset = Ledger.objects.all().select_related('category').order_by('code')
    
    page = request.GET.get('page', 1)
    paginator = Paginator(ledger_queryset, 15) 
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
        
    return render(request, 'accounting/ledger_form.html', {
    'categories': categories,
    'ledgers': ledgers,
    'page_obj': page_obj
})

def create_ledger(request):
    if request.method == 'POST':
        try:
            # Get form data
            category = get_object_or_404(LedgerCategory, id=request.POST.get('category'))
            name = request.POST.get('name')
            description = request.POST.get('description', '')
            is_active = 'is_active' in request.POST

            if not name:
                return JsonResponse({'error': 'Ledger name is required'}, status=400)

            # Generate code based on category
            category_code = category.code
            existing_codes = Ledger.objects.filter(code__startswith=f"{category_code}-").values_list('code', flat=True)
            
            max_number = 0
            for code_str in existing_codes:
                if not code_str.startswith(f"{category_code}-"):
                    continue
                suffix = code_str[len(category_code) + 1:]
                if len(suffix) == 3 and suffix.isdigit():
                    try:
                        num = int(suffix)
                        max_number = max(max_number, num)
                    except ValueError:
                        continue

            new_code = f"{category_code}-{max_number + 1:03d}"

            balance_type_map = {
                'Asset': 'Debit',
                'Liability': 'Credit',
                'Equity': 'Credit',
                'Income': 'Credit',
                'Expense': 'Debit'
            }
            balance_type = balance_type_map.get(category.name, 'Debit')

            ledger = Ledger.objects.create(
                category=category,
                name=name,
                code=new_code,
                opening_balance=Decimal('0.00'),  
                balance_type=balance_type,
                description=description,
                is_active=is_active
            )

            return JsonResponse({
                'id': ledger.id,
                'category': ledger.category.name,
                'name': ledger.name,
                'code': ledger.code,
                'balance_type': ledger.balance_type,
                'description': ledger.description,
                'opening_balance': '0.00'  
            })

        except IntegrityError as e:
            return JsonResponse({'error': 'Ledger with this name or code already exists'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid request method'}, status=405)



# def delete_ledger(request, ledger_id):
#     if request.method == 'POST':
#         try:
#             ledger = get_object_or_404(Ledger, id=ledger_id)
#             ledger.delete()
#             return JsonResponse({'success': True})
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=400)
#     return JsonResponse({'error': 'Invalid request method'}, status=405)

def delete_ledger(request, ledger_id):
    if request.method == 'POST':
        try:
            with transaction.atomic():  # Wrap in atomic transaction
                ledger = get_object_or_404(Ledger, id=ledger_id)
                ledger.delete()
                return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid request method'}, status=405)



def ledger_entry_list(request):
    ledgers = Ledger.objects.all()
    entries = LedgerEntry.objects.all().select_related('ledger')
    return render(request, 'accounting/ledger_entry_form.html', {
        'ledgers': ledgers,
        'entries': entries
    })

def create_ledger_entry(request):
    if request.method == 'POST':
        try:
            ledger = get_object_or_404(Ledger, id=request.POST.get('ledger'))
            entry_type = request.POST.get('entry_type')
            amount = request.POST.get('amount')
            date_str = request.POST.get('date')
            description = request.POST.get('description', '')

            try:
                amount = Decimal(amount)
                if amount <= 0:
                    return JsonResponse({'error': 'Amount must be greater than zero'}, status=400)
            except InvalidOperation:
                return JsonResponse({'error': 'Invalid amount format'}, status=400)

            try:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({'error': 'Invalid date format'}, status=400)

            with transaction.atomic():
                entry = LedgerEntry.objects.create(
                    ledger=ledger,
                    entry_type=entry_type,
                    amount=amount,
                    date=date_obj,
                    description=description
                )

                corresponding_entry = determine_corresponding_entry(ledger, entry_type, amount, date_obj, description)

                if corresponding_entry:
                    LedgerEntry.objects.create(
                        ledger=corresponding_entry['ledger'],
                        entry_type=corresponding_entry['entry_type'],
                        amount=amount,
                        date=date_obj,
                        description=corresponding_entry['description']
                    )

            return JsonResponse({
                'id': entry.id,
                'ledger': entry.ledger.name,
                'entry_type': entry.entry_type,
                'amount': str(entry.amount),
                'date': entry.date.strftime('%Y-%m-%d'),
                'description': entry.description or "N/A"
            })

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid request method'}, status=405)

def determine_corresponding_entry(ledger, entry_type, amount, date_obj, description):

    try:
        category_name = ledger.category.name

        if category_name == "Expense":
            if entry_type == "Debit":
                corresponding_ledger = Ledger.objects.get(category__name="Asset", name="Cash in Hand")
                return {
                    'ledger': corresponding_ledger,
                    'entry_type': "Credit",
                    'description': f"Cash disbursement for {ledger.name}"
                }

        elif category_name == "Income":
            if entry_type == "Credit":
                corresponding_ledger = Ledger.objects.get(category__name="Asset", name="Cash in Hand")
                return {
                    'ledger': corresponding_ledger,
                    'entry_type': "Debit",
                    'description': f"Cash receipt from {ledger.name}"
                }

        elif category_name == "Asset":
            if entry_type == "Debit":
                corresponding_ledger = Ledger.objects.filter(category__name="Liability").first()
                if corresponding_ledger:
                    return {
                        'ledger': corresponding_ledger,
                        'entry_type': "Credit",
                        'description': f"Increase in {ledger.name}"
                    }
            elif entry_type == "Credit":
                corresponding_ledger = Ledger.objects.filter(category__name="Expense").first()
                if corresponding_ledger:
                    return {
                        'ledger': corresponding_ledger,
                        'entry_type': "Debit",
                        'description': f"Decrease in {ledger.name} due to {corresponding_ledger.name}"
                    }

        elif category_name in ["Liability", "Equity"]:
            if entry_type == "Credit":
                corresponding_ledger = Ledger.objects.get(category__name="Asset", name="Cash in Hand")
                return {
                    'ledger': corresponding_ledger,
                    'entry_type': "Debit",
                    'description': f"Cash withdrawal for {ledger.name}"
                }

    except Ledger.DoesNotExist:
        return None

    return None


def delete_ledger_entry(request, entry_id):
    if request.method == 'POST':
        try:
            entry = get_object_or_404(LedgerEntry, id=entry_id)
            entry.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid request method'}, status=405)

def receive_management(request):
    if request.method == 'POST':
        try:
            required_fields = ['date', 'amount', 'income_ledger', 'description']
            missing = [field for field in required_fields if not request.POST.get(field)]
            if missing:
                return JsonResponse({'error': f'Missing fields: {", ".join(missing)}'}, status=400)

            voucher_no = request.POST.get('voucher_no')
            if not voucher_no:
                current_time = datetime.now()
                voucher_no = current_time.strftime("REC%Y%m%d%H%M%S")

            date_str = request.POST['date']
            amount = request.POST['amount']
            income_ledger_id = request.POST['income_ledger']
            description = request.POST['description']

            try:
                date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({'error': 'Invalid date format (YYYY-MM-DD required)'}, status=400)

            try:
                amount = Decimal(amount)
                if amount <= 0:
                    return JsonResponse({'error': 'Amount must be greater than zero'}, status=400)
            except InvalidOperation:
                return JsonResponse({'error': 'Invalid amount format'}, status=400)

            income_ledger = get_object_or_404(Ledger, id=income_ledger_id)
            if income_ledger.category.name != 'Income':
                return JsonResponse({'error': 'Selected ledger must be an Income account'}, status=400)

            cash_ledger, created = Ledger.objects.get_or_create(name='Cash in Hand', defaults={'category_id': 1})

            student = None
            if student_id := request.POST.get('student'):
                student = get_object_or_404(StudentProfile, id=student_id)

            fee_head = None
            if fee_head_id := request.POST.get('fee_head'):
                fee_head = get_object_or_404(FeeHead, id=fee_head_id)

            with transaction.atomic():
                receive = Receive.objects.create(
                    voucher_no=voucher_no,
                    date=date,
                    amount=amount,
                    student=student,
                    fee_head=fee_head,
                    cash_ledger=cash_ledger,
                    income_ledger=income_ledger,
                    description=description,
                )

            return JsonResponse({
                'success': True,
                'voucher_no': receive.voucher_no,
                'date': receive.date.strftime('%Y-%m-%d'),
                'amount': str(receive.amount),
                'cash_ledger': cash_ledger.name,
                'income_ledger': income_ledger.name,
                'description': description,
                'created_at': receive.created_at.strftime('%Y-%m-%d %H:%M:%S')
            })

        except IntegrityError:
            return JsonResponse({'error': 'Voucher number already exists'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
        
    receives = Receive.objects.all().order_by('-id')
    paginator = Paginator(receives, 15)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    income_ledgers = Ledger.objects.filter(category__name='Income')
    return render(request, 'accounting/receive_list.html', {
        'receives': page_obj,
        'income_ledgers': income_ledgers
    })

    receives = Receive.objects.all()
    income_ledgers = Ledger.objects.filter(category__name='Income')
    return render(request, 'accounting/receive_list.html', {'receives': receives, 'income_ledgers': income_ledgers})


def delete_receive(request, receive_id):
    if request.method == 'POST':
        try:
            receive = get_object_or_404(Receive, id=receive_id)
            receive.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid request method'}, status=405)

def payment_management(request):
    if request.method == 'POST':
        try:
            required_fields = ['date', 'amount', 'expense_ledger', 'description']
            missing = [field for field in required_fields if not request.POST.get(field)]
            if missing:
                return JsonResponse({'error': f'Missing fields: {", ".join(missing)}'}, status=400)

            voucher_no = request.POST.get('voucher_no')
            if not voucher_no:
                current_time = datetime.now()
                voucher_no = current_time.strftime("PAY%Y%m%d%H%M%S")

            date_str = request.POST['date']
            amount = request.POST['amount']
            expense_ledger_id = request.POST['expense_ledger']
            description = request.POST['description']

            try:
                date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({'error': 'Invalid date format (YYYY-MM-DD required)'}, status=400)

            try:
                amount = Decimal(amount)
                if amount <= 0:
                    return JsonResponse({'error': 'Amount must be greater than zero'}, status=400)
            except InvalidOperation:
                return JsonResponse({'error': 'Invalid amount format'}, status=400)

            expense_ledger = get_object_or_404(Ledger, id=expense_ledger_id)
            if expense_ledger.category.name != 'Expense':
                return JsonResponse({'error': 'Selected ledger must be an Expense account'}, status=400)

            cash_ledger, created = Ledger.objects.get_or_create(name='Cash in Hand', defaults={'category_id': 1})

            staff = None
            if staff_id := request.POST.get('staff'):
                staff = get_object_or_404(StaffProfile, id=staff_id)

            with transaction.atomic():
                payment = Payment.objects.create(
                    voucher_no=voucher_no,
                    date=date,
                    amount=amount,
                    staff=staff,
                    cash_ledger=cash_ledger,
                    expense_ledger=expense_ledger,
                    description=description,
                )

                main_balance, created = MainBalance.objects.get_or_create(cash_ledger=cash_ledger)
                main_balance.balance = cash_ledger.current_balance
                main_balance.save(update_fields=['balance'])

            return JsonResponse({
                'success': True,
                'voucher_no': payment.voucher_no,
                'date': payment.date.strftime('%Y-%m-%d'),
                'amount': str(payment.amount),
                'cash_ledger': cash_ledger.name,
                'expense_ledger': expense_ledger.name,
                'description': description,
                'created_at': payment.created_at.strftime('%Y-%m-%d %H:%M:%S')
            })

        except IntegrityError:
            return JsonResponse({'error': 'Voucher number already exists'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    # GET request handling with pagination
    payments = Payment.objects.all().order_by('-id')
    paginator = Paginator(payments, 15)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    expense_ledgers = Ledger.objects.filter(category__name='Expense')
    return render(request, 'accounting/payment_list.html', {
        'payments': page_obj,
        'expense_ledgers': expense_ledgers
    })

# ------------------------------ Accounting Report Part start --------------------------



@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','Accountant']))
def cash_summary(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if not from_date or not to_date:
        today = datetime.today()
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
        to_date = today.strftime('%Y-%m-%d')

    cash_ledgers = Ledger.objects.filter(category__name="Asset", name__icontains="Cash")

    total_opening_balance = cash_ledgers.aggregate(total=Sum('opening_balance'))['total'] or 0

    cash_inflows_queryset = LedgerEntry.objects.filter(
        ledger__in=cash_ledgers,
        entry_type="Debit",
        date__range=[from_date, to_date]
    ).values('ledger__name', 'description').annotate(total_amount=Sum('amount'))

    cash_inflows_list = list(cash_inflows_queryset)
    total_cash_inflows = sum(item['total_amount'] for item in cash_inflows_list)
    paginator_inflow = Paginator(cash_inflows_list, 15)
    page_inflow = paginator_inflow.get_page(request.GET.get('inflow_page'))

    cash_outflows_queryset = LedgerEntry.objects.filter(
        ledger__in=cash_ledgers,
        entry_type="Credit",
        date__range=[from_date, to_date]
    ).values('ledger__name', 'description').annotate(total_amount=Sum('amount'))

    cash_outflows_list = list(cash_outflows_queryset)
    total_cash_outflows = sum(item['total_amount'] for item in cash_outflows_list)
    paginator_outflow = Paginator(cash_outflows_list, 15)
    page_outflow = paginator_outflow.get_page(request.GET.get('outflow_page'))

    total_closing_balance = total_opening_balance + total_cash_inflows - total_cash_outflows

    context = {
        'from_date': from_date,
        'to_date': to_date,
        'total_opening_balance': total_opening_balance,
        'cash_inflows': page_inflow.object_list,
        'total_cash_inflows': total_cash_inflows,
        'cash_outflows': page_outflow.object_list,
        'total_cash_outflows': total_cash_outflows,
        'total_closing_balance': total_closing_balance,
        'page_inflow': page_inflow,
        'page_outflow': page_outflow,
    }

    return render(request, 'accounting/cash_summary.html', context)


from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from io import BytesIO
from django.db.models import Sum
from datetime import datetime

def cash_summary_pdf(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if not from_date or not to_date:
        today = datetime.today()
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
        to_date = today.strftime('%Y-%m-%d')

    cash_ledgers = Ledger.objects.filter(category__name="Asset", name__icontains="Cash")

    total_opening_balance = cash_ledgers.aggregate(total=Sum('opening_balance'))['total'] or 0

    # Get all cash inflows without pagination
    cash_inflows = LedgerEntry.objects.filter(
        ledger__in=cash_ledgers,
        entry_type="Debit",
        date__range=[from_date, to_date]
    ).values('ledger__name', 'description').annotate(total_amount=Sum('amount'))

    total_cash_inflows = sum(item['total_amount'] for item in cash_inflows)

    # Get all cash outflows without pagination
    cash_outflows = LedgerEntry.objects.filter(
        ledger__in=cash_ledgers,
        entry_type="Credit",
        date__range=[from_date, to_date]
    ).values('ledger__name', 'description').annotate(total_amount=Sum('amount'))

    total_cash_outflows = sum(item['total_amount'] for item in cash_outflows)

    total_closing_balance = total_opening_balance + total_cash_inflows - total_cash_outflows

    context = {
        'from_date': from_date,
        'to_date': to_date,
        'total_opening_balance': total_opening_balance,
        'cash_inflows': cash_inflows,
        'total_cash_inflows': total_cash_inflows,
        'cash_outflows': cash_outflows,
        'total_cash_outflows': total_cash_outflows,
        'total_closing_balance': total_closing_balance,
    }

    # Render HTML template
    html_string = render_to_string('accounting/cash_summary_pdf.html', context)
    
    # Create PDF
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_string.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        filename = f"cash_summary_{from_date}_to_{to_date}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    return HttpResponse('Error generating PDF', status=400)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','Accountant']))
# def trial_balance(request):
#     from_date = request.GET.get('from_date')
#     to_date = request.GET.get('to_date')

#     if not from_date or not to_date:
#         today = datetime.today()
#         from_date = today.replace(day=1).strftime('%Y-%m-%d')
#         to_date = today.strftime('%Y-%m-%d')

#     categories = LedgerCategory.objects.all()
#     trial_balance = {}
#     total_debit = 0
#     total_credit = 0

#     trial_balance_list = []
#     for category in categories:
#         ledgers = Ledger.objects.filter(category=category, is_active=True)
#         category_data = []
        
#         for ledger in ledgers:
#             if ledger.balance_type == 'Debit':
#                 debit = ledger.current_balance
#                 credit = 0
#             else:
#                 credit = ledger.current_balance
#                 debit = 0
                
#             category_data.append({
#                 'category': category.name,
#                 'ledger': ledger.name,
#                 'debit': debit,
#                 'credit': credit
#             })
            
#             total_debit += debit
#             total_credit += credit
        
#         if category_data:
#             trial_balance_list.extend(category_data)

#     paginator = Paginator(trial_balance_list, 15)
#     page_number = request.GET.get('page')
#     page_obj = paginator.get_page(page_number)

#     balance_status = "Balanced" if total_debit == total_credit else "Not Balanced"

#     context = {
#         'page_obj': page_obj,
#         'from_date': from_date,
#         'to_date': to_date,
#         'total_debit': total_debit,
#         'total_credit': total_credit,
#         'balance_status': balance_status,
#     }

#     if request.GET.get('download') == 'pdf':
#         return generate_pdf('accounting/trial_balance_pdf.html', context)

#     return render(request, 'accounting/trial_balance.html', context)


def trial_balance(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if not from_date or not to_date:
        today = datetime.today()
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
        to_date = today.strftime('%Y-%m-%d')

    # Check Main Balance > 0
    try:
        cash_ledger = Ledger.objects.get(name='Cash in Hand')
        main_balance = MainBalance.objects.get(cash_ledger=cash_ledger)
    except (Ledger.DoesNotExist, MainBalance.DoesNotExist):
        main_balance = None

    if not main_balance or main_balance.balance <= 0:
        context = {
            'error': "Trial Balance can only be generated when Main Balance is positive",
            'from_date': from_date,
            'to_date': to_date
        }
        return render(request, 'accounting/trial_balance.html', context)

    # Build full dataset once (used by both UI + PDF)
    categories = LedgerCategory.objects.all()
    trial_balance_list = []
    total_debit = 0.0
    total_credit = 0.0

    for category in categories:
        ledgers = Ledger.objects.filter(category=category, is_active=True)
        for ledger in ledgers:
            if ledger.balance_type == 'Debit':
                debit = float(ledger.current_balance or 0)
                credit = 0.0
            else:
                credit = float(ledger.current_balance or 0)
                debit = 0.0

            trial_balance_list.append({
                'category': category.name,
                'ledger': ledger.name,
                'debit': debit,
                'credit': credit,
            })
            total_debit += debit
            total_credit += credit

    balance_status = "Balanced" if round(total_debit, 2) == round(total_credit, 2) else "Not Balanced"

    # If PDF requested -> no pagination; send full list to the template
    if request.GET.get('download') == 'pdf':
        context = {
            'from_date': from_date,
            'to_date': to_date,
            'total_debit': total_debit,
            'total_credit': total_credit,
            'balance_status': balance_status,
            'trial_balance_list': trial_balance_list,  # for regroup in template
        }
        filename = f"trial_balance_{from_date}_to_{to_date}.pdf"
        return render_to_pdf('accounting/trial_balance_pdf.html', context, filename)

    # UI path -> paginate only for on-screen table
    paginator = Paginator(trial_balance_list, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'from_date': from_date,
        'to_date': to_date,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'balance_status': balance_status,
    }
    return render(request, 'accounting/trial_balance.html', context)


def render_to_pdf(template_src, context_dict, filename=None):
    now = datetime.now()
    context_dict['generated_date'] = now.strftime("%Y-%m-%d")
    context_dict['generated_time'] = now.strftime("%H:%M:%S")

    template = get_template(template_src)
    html = template.render(context_dict)

    pdf_buffer = BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=pdf_buffer)
    if pisa_status.err:
        return HttpResponse("Error generating PDF", status=500)

    pdf_buffer.seek(0)

    fname = filename or f'trial-balance_{slugify(context_dict.get("from_date", ""))}_to_{slugify(context_dict.get("to_date", ""))}.pdf'

    resp = FileResponse(pdf_buffer, as_attachment=True, filename=fname)
    resp['Content-Type'] = 'application/pdf'
    resp['X-Content-Type-Options'] = 'nosniff'
    resp['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp['Pragma'] = 'no-cache'
    resp['Content-Transfer-Encoding'] = 'binary'
    return resp





@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','Accountant']))
def balance_sheet(request):
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    if not from_date_str or not to_date_str:
        today = now().date()
        from_date = today.replace(day=1) 
        to_date = today
    else:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()

    def calculate_balances(ledgers):
        ledger_data = []
        category_total = 0
        
        for ledger in ledgers:
            debit_sum = ledger.entries.filter(
                entry_type='Debit', 
                date__lte=to_date
            ).aggregate(total=Sum('amount'))['total'] or 0
            
            credit_sum = ledger.entries.filter(
                entry_type='Credit', 
                date__lte=to_date
            ).aggregate(total=Sum('amount'))['total'] or 0

            if ledger.balance_type == 'Debit':
                balance = ledger.opening_balance + debit_sum - credit_sum
            else:
                balance = ledger.opening_balance + credit_sum - debit_sum

            ledger_data.append({
                'name': ledger.name,
                'balance': balance
            })
            category_total += balance
            
        return ledger_data, category_total

    assets_data, total_assets = calculate_balances(Ledger.objects.filter(category__name='Asset', is_active=True))
    liabilities_data, total_liabilities = calculate_balances(Ledger.objects.filter(category__name='Liability', is_active=True))
    equity_data, total_equity = calculate_balances(Ledger.objects.filter(category__name='Equity', is_active=True))

    accounting_equation = total_assets == (total_liabilities + total_equity)

    if request.GET.get('download') == 'pdf':
        context = {
            'from_date': from_date.strftime('%Y-%m-%d'),
            'to_date': to_date.strftime('%Y-%m-%d'),
            'assets': assets_data,
            'liabilities': liabilities_data,
            'equity': equity_data,
            'total_assets': total_assets,
            'total_liabilities': total_liabilities,
            'total_equity': total_equity,
            'balance_status': "Balanced" if accounting_equation else "Not Balanced",
        }
        return generate_pdf('accounting/balance_sheet_pdf.html', context)

    paginator_assets = Paginator(assets_data, 15)
    page_assets = paginator_assets.get_page(request.GET.get('assets_page'))

    paginator_liabilities = Paginator(liabilities_data, 15)
    page_liabilities = paginator_liabilities.get_page(request.GET.get('liabilities_page'))

    paginator_equity = Paginator(equity_data, 15)
    page_equity = paginator_equity.get_page(request.GET.get('equity_page'))

    context = {
        'from_date': from_date.strftime('%Y-%m-%d'),
        'to_date': to_date.strftime('%Y-%m-%d'),
        'page_assets': page_assets,
        'page_liabilities': page_liabilities,
        'page_equity': page_equity,
        'total_assets': total_assets,
        'total_liabilities': total_liabilities,
        'total_equity': total_equity,
        'balance_status': "Balanced" if accounting_equation else "Not Balanced",
    }

    return render(request, 'accounting/balance_sheet.html', context)

from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from io import BytesIO

def generate_pdf(template_src, context_dict):
    html = render_to_string(template_src, context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None

def balance_sheet_pdf(request):
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    if not from_date_str or not to_date_str:
        today = now().date()
        from_date = today.replace(day=1) 
        to_date = today
    else:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()

    def calculate_balances(ledgers):
        ledger_data = []
        category_total = 0
        
        for ledger in ledgers:
            debit_sum = ledger.entries.filter(
                entry_type='Debit', 
                date__lte=to_date
            ).aggregate(total=Sum('amount'))['total'] or 0
            
            credit_sum = ledger.entries.filter(
                entry_type='Credit', 
                date__lte=to_date
            ).aggregate(total=Sum('amount'))['total'] or 0

            if ledger.balance_type == 'Debit':
                balance = ledger.opening_balance + debit_sum - credit_sum
            else:
                balance = ledger.opening_balance + credit_sum - debit_sum

            ledger_data.append({
                'name': ledger.name,
                'balance': balance
            })
            category_total += balance
            
        return ledger_data, category_total

    assets_data, total_assets = calculate_balances(Ledger.objects.filter(category__name='Asset', is_active=True))
    liabilities_data, total_liabilities = calculate_balances(Ledger.objects.filter(category__name='Liability', is_active=True))
    equity_data, total_equity = calculate_balances(Ledger.objects.filter(category__name='Equity', is_active=True))

    accounting_equation = total_assets == (total_liabilities + total_equity)

    context = {
        'from_date': from_date.strftime('%Y-%m-%d'),
        'to_date': to_date.strftime('%Y-%m-%d'),
        'assets': assets_data,
        'liabilities': liabilities_data,
        'equity': equity_data,
        'total_assets': total_assets,
        'total_liabilities': total_liabilities,
        'total_equity': total_equity,
        'balance_status': "Balanced" if accounting_equation else "Not Balanced",
    }
    
    response = generate_pdf('accounting/balance_sheet_pdf.html', context)
    if response:
        filename = f"balance_sheet_{from_date}_to_{to_date}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    return HttpResponse('Error generating PDF', status=400)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager','Accountant']))
def income_statement(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if not from_date or not to_date:
        today = datetime.today()
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
        to_date = today.strftime('%Y-%m-%d')

    income_ledgers = Ledger.objects.filter(category__name="Income", is_active=True)
    income_details = []
    total_income = 0
    for ledger in income_ledgers:
        income_amount = LedgerEntry.objects.filter(
            ledger=ledger, entry_type="Credit", date__range=[from_date, to_date]
        ).aggregate(total=Sum('amount'))['total'] or 0
        income_details.append({'name': ledger.name, 'amount': income_amount})
        total_income += income_amount

    income_paginator = Paginator(income_details, 15)
    income_page = income_paginator.get_page(request.GET.get('income_page'))

    expense_ledgers = Ledger.objects.filter(category__name="Expense", is_active=True)
    expense_details = []
    total_expenses = 0
    for ledger in expense_ledgers:
        expense_amount = LedgerEntry.objects.filter(
            ledger=ledger, entry_type="Debit", date__range=[from_date, to_date]
        ).aggregate(total=Sum('amount'))['total'] or 0
        expense_details.append({'name': ledger.name, 'amount': expense_amount})
        total_expenses += expense_amount

    expense_paginator = Paginator(expense_details, 15)
    expense_page = expense_paginator.get_page(request.GET.get('expense_page'))

    net_profit = total_income - total_expenses

    context = {
        'from_date': from_date,
        'to_date': to_date,
        'income_page': income_page,
        'expense_page': expense_page,
        'total_income': total_income,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
        'status': 'Profit' if net_profit >= 0 else 'Loss'
    }

    return render(request, 'accounting/income_statement.html', context)


from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from io import BytesIO
from django.db.models import Sum
from .models import Ledger, LedgerEntry
from datetime import datetime

def generate_pdf(template_src, context_dict, filename):
    html = render_to_string(template_src, context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    return None

def income_statement_pdf(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if not from_date or not to_date:
        today = datetime.today()
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
        to_date = today.strftime('%Y-%m-%d')

    # Get income data and filter out zero amounts
    income_ledgers = Ledger.objects.filter(category__name="Income", is_active=True)
    income_details = []
    total_income = 0
    for ledger in income_ledgers:
        income_amount = LedgerEntry.objects.filter(
            ledger=ledger, entry_type="Credit", date__range=[from_date, to_date]
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        # Only include if amount is not zero
        if income_amount != 0:
            income_details.append({'name': ledger.name, 'amount': income_amount})
            total_income += income_amount

    # Get expense data and filter out zero amounts
    expense_ledgers = Ledger.objects.filter(category__name="Expense", is_active=True)
    expense_details = []
    total_expenses = 0
    for ledger in expense_ledgers:
        expense_amount = LedgerEntry.objects.filter(
            ledger=ledger, entry_type="Debit", date__range=[from_date, to_date]
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        # Only include if amount is not zero
        if expense_amount != 0:
            expense_details.append({'name': ledger.name, 'amount': expense_amount})
            total_expenses += expense_amount

    net_profit = total_income - total_expenses

    context = {
        'from_date': from_date,
        'to_date': to_date,
        'income_details': income_details,
        'expense_details': expense_details,
        'total_income': total_income,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
        'status': 'Profit' if net_profit >= 0 else 'Loss'
    }
    
    filename = f"income_statement_{from_date}_to_{to_date}.pdf"
    return generate_pdf('accounting/income_statement_pdf.html', context, filename)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def bank_book(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if not from_date or not to_date:
        today = datetime.today()
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
        to_date = today.strftime('%Y-%m-%d')

    bank_ledgers = Ledger.objects.filter(category__name="Asset", name__icontains="Bank")

    total_opening_balance = bank_ledgers.aggregate(total=Sum('opening_balance'))['total'] or 0

    bank_entries = LedgerEntry.objects.filter(
        ledger__in=bank_ledgers,
        date__range=[from_date, to_date]
    ).order_by('date')

    total_debit = bank_entries.filter(entry_type="Debit").aggregate(total=Sum('amount'))['total'] or 0
    total_credit = bank_entries.filter(entry_type="Credit").aggregate(total=Sum('amount'))['total'] or 0

    closing_balance = total_opening_balance + total_debit - total_credit

    context = {
        'from_date': from_date,
        'to_date': to_date,
        'bank_entries': bank_entries,
        'total_opening_balance': total_opening_balance,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'closing_balance': closing_balance
    }

    return render(request, 'accounting/bank_book.html', context)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def outstanding_receivables(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if not from_date or not to_date:
        today = datetime.today()
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
        to_date = today.strftime('%Y-%m-%d')

    income_ledgers = Ledger.objects.filter(category__name="Income", is_active=True)

    outstanding_data = []

    for ledger in income_ledgers:
        total_income = ledger.entries.filter(entry_type='Credit', date__range=[from_date, to_date]).aggregate(
            total=Sum('amount'))['total'] or 0

        total_received = Receive.objects.filter(income_ledger=ledger, date__range=[from_date, to_date]).aggregate(
            total=Sum('amount'))['total'] or 0

        outstanding_data.append({
            'ledger_name': ledger.name,
            'total_income': total_income,
            'total_received': total_received,
            'outstanding_balance': total_income - total_received
        })

    context = {
        'from_date': from_date,
        'to_date': to_date,
        'outstanding_data': outstanding_data,
    }

    return render(request, 'accounting/outstanding_receivables.html', context)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def payment_summary(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if not from_date or not to_date:
        today = datetime.today()
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
        to_date = today.strftime('%Y-%m-%d')

    payments = Payment.objects.filter(date__range=[from_date, to_date]).order_by('-date')

    total_payments = payments.aggregate(total=Sum('amount'))['total'] or 0

    context = {
        'from_date': from_date,
        'to_date': to_date,
        'payments': payments,
        'total_payments': total_payments,
    }

    return render(request, 'accounting/payment_summary.html', context)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))

def funds_flow_report(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if not from_date or not to_date:
        today = datetime.today()
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
        to_date = today.strftime('%Y-%m-%d')

    fund_sources = Ledger.objects.filter(category__name__in=["Liability", "Income"], is_active=True)
    fund_sources_data = []
    total_fund_sources = 0

    for ledger in fund_sources:
        credit_total = ledger.entries.filter(entry_type="Credit", date__range=[from_date, to_date]).aggregate(
            total=Sum('amount')
        )['total'] or 0

        fund_sources_data.append({
            'category': ledger.category.name,
            'ledger_name': ledger.name,
            'amount': credit_total
        })
        total_fund_sources += credit_total

    fund_applications = Ledger.objects.filter(category__name__in=["Asset", "Expense"], is_active=True)
    fund_applications_data = []
    total_fund_applications = 0

    for ledger in fund_applications:
        debit_total = ledger.entries.filter(entry_type="Debit", date__range=[from_date, to_date]).aggregate(
            total=Sum('amount')
        )['total'] or 0

        fund_applications_data.append({
            'category': ledger.category.name,
            'ledger_name': ledger.name,
            'amount': debit_total
        })
        total_fund_applications += debit_total

    context = {
        'from_date': from_date,
        'to_date': to_date,
        'fund_sources': fund_sources_data,
        'fund_applications': fund_applications_data,
        'total_fund_sources': total_fund_sources,
        'total_fund_applications': total_fund_applications,
        'status': "Balanced" if total_fund_sources == total_fund_applications else "Unbalanced"
    }

    return render(request, 'accounting/funds_flow_report.html', context)

from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from io import BytesIO
from django.db.models import Sum
from .models import Ledger, LedgerEntry

def generate_pdf(template_src, context_dict, filename):
    html = render_to_string(template_src, context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    return None

def funds_flow_report_pdf(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if not from_date or not to_date:
        today = datetime.today()
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
        to_date = today.strftime('%Y-%m-%d')

    # Get fund sources data and filter out zero amounts
    fund_sources = Ledger.objects.filter(category__name__in=["Liability", "Income"], is_active=True)
    fund_sources_data = []
    total_fund_sources = 0

    for ledger in fund_sources:
        credit_total = ledger.entries.filter(entry_type="Credit", date__range=[from_date, to_date]).aggregate(
            total=Sum('amount')
        )['total'] or 0

        # Only include if amount is not zero
        if credit_total != 0:
            fund_sources_data.append({
                'category': ledger.category.name,
                'ledger_name': ledger.name,
                'amount': credit_total
            })
            total_fund_sources += credit_total

    # Get fund applications data and filter out zero amounts
    fund_applications = Ledger.objects.filter(category__name__in=["Asset", "Expense"], is_active=True)
    fund_applications_data = []
    total_fund_applications = 0

    for ledger in fund_applications:
        debit_total = ledger.entries.filter(entry_type="Debit", date__range=[from_date, to_date]).aggregate(
            total=Sum('amount')
        )['total'] or 0

        # Only include if amount is not zero
        if debit_total != 0:
            fund_applications_data.append({
                'category': ledger.category.name,
                'ledger_name': ledger.name,
                'amount': debit_total
            })
            total_fund_applications += debit_total

    context = {
        'from_date': from_date,
        'to_date': to_date,
        'fund_sources': fund_sources_data,
        'fund_applications': fund_applications_data,
        'total_fund_sources': total_fund_sources,
        'total_fund_applications': total_fund_applications,
        'net_funds_flow': total_fund_sources - total_fund_applications,
        'status': "Balanced" if total_fund_sources == total_fund_applications else "Unbalanced"
    }
    
    filename = f"funds_flow_report_{from_date}_to_{to_date}.pdf"
    return generate_pdf('accounting/funds_flow_report_pdf.html', context, filename)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def voucher_list(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    transaction_type = request.GET.get('transaction_type')

    vouchers = []

    if from_date and to_date:
        if transaction_type == "Receive":
            receive_vouchers = Receive.objects.filter(date__range=[from_date, to_date])
            for voucher in receive_vouchers:
                voucher.transaction_type = "Receive"
            vouchers.extend(receive_vouchers)

        elif transaction_type == "Payment":
            payment_vouchers = Payment.objects.filter(date__range=[from_date, to_date])
            for voucher in payment_vouchers:
                voucher.transaction_type = "Payment"
            vouchers.extend(payment_vouchers)

        elif transaction_type == "Contra":
            contra_vouchers = Contra.objects.filter(date__range=[from_date, to_date])
            for voucher in contra_vouchers:
                voucher.transaction_type = "Contra Transaction"
            vouchers.extend(contra_vouchers)

        elif transaction_type == "Journal":
            journal_vouchers = Journal.objects.filter(date__range=[from_date, to_date])
            for voucher in journal_vouchers:
                voucher.transaction_type = "Journal Transaction"
            vouchers.extend(journal_vouchers)
        
        paginator = Paginator(vouchers, 15)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
    else:
        page_obj = None

    context = {
        'from_date': from_date,
        'to_date': to_date,
        'transaction_type': transaction_type,
        'vouchers': vouchers,
        'page_obj': page_obj,
    }

    return render(request, 'accounting/voucher_list.html', context)


from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from io import BytesIO
from .models import Receive, Payment, Contra, Journal

def generate_pdf(template_src, context_dict, filename):
    html = render_to_string(template_src, context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    return None

def voucher_list_pdf(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    transaction_type = request.GET.get('transaction_type')

    vouchers = []

    if from_date and to_date:
        if transaction_type == "Receive":
            receive_vouchers = Receive.objects.filter(date__range=[from_date, to_date])
            for voucher in receive_vouchers:
                voucher.transaction_type = "Receive"
            vouchers.extend(receive_vouchers)

        elif transaction_type == "Payment":
            payment_vouchers = Payment.objects.filter(date__range=[from_date, to_date])
            for voucher in payment_vouchers:
                voucher.transaction_type = "Payment"
            vouchers.extend(payment_vouchers)

        elif transaction_type == "Contra":
            contra_vouchers = Contra.objects.filter(date__range=[from_date, to_date])
            for voucher in contra_vouchers:
                voucher.transaction_type = "Contra Transaction"
            vouchers.extend(contra_vouchers)

        elif transaction_type == "Journal":
            journal_vouchers = Journal.objects.filter(date__range=[from_date, to_date])
            for voucher in journal_vouchers:
                voucher.transaction_type = "Journal Transaction"
            vouchers.extend(journal_vouchers)
        
        # For PDF, we don't paginate - we include all results
        total_amount = sum(voucher.amount for voucher in vouchers)
    else:
        vouchers = []
        total_amount = 0

    context = {
        'from_date': from_date,
        'to_date': to_date,
        'transaction_type': transaction_type,
        'vouchers': vouchers,
        'total_amount': total_amount,
    }
    
    filename = f"voucher_list_{from_date}_to_{to_date}_{transaction_type or 'all'}.pdf"
    return generate_pdf('accounting/voucher_list_pdf.html', context, filename)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def journal_report(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if not from_date or not to_date:
        today = datetime.today()
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
        to_date = today.strftime('%Y-%m-%d')

    journals = Journal.objects.filter(date__range=[from_date, to_date]).prefetch_related('entries')

    paginator = Paginator(journals, 15)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    journal_data = []
    for journal in page_obj:
        entries = journal.entries.all()
        journal_data.append({
            "voucher_no": journal.voucher_no,
            "date": journal.date,
            "description": journal.description,
            "debits": [e for e in entries if e.entry_type == "Debit"],
            "credits": [e for e in entries if e.entry_type == "Credit"],
        })

    context = {
        "from_date": from_date,
        "to_date": to_date,
        "journal_data": journal_data,
        "page_obj": page_obj,
    }

    return render(request, 'accounting/journal_report.html', context)

from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from io import BytesIO
from .models import Journal

def generate_pdf(template_src, context_dict, filename):
    html = render_to_string(template_src, context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    return None

def journal_report_pdf(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if not from_date or not to_date:
        today = datetime.today()
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
        to_date = today.strftime('%Y-%m-%d')

    # Get all journals without pagination for PDF
    journals = Journal.objects.filter(date__range=[from_date, to_date]).prefetch_related('entries')

    journal_data = []
    total_debit = 0
    total_credit = 0
    
    for journal in journals:
        entries = journal.entries.all()
        journal_debits = [e for e in entries if e.entry_type == "Debit"]
        journal_credits = [e for e in entries if e.entry_type == "Credit"]
        
        journal_debit_total = sum(debit.amount for debit in journal_debits)
        journal_credit_total = sum(credit.amount for credit in journal_credits)
        
        total_debit += journal_debit_total
        total_credit += journal_credit_total
        
        journal_data.append({
            "voucher_no": journal.voucher_no,
            "date": journal.date,
            "description": journal.description,
            "debits": journal_debits,
            "credits": journal_credits,
            "debit_total": journal_debit_total,
            "credit_total": journal_credit_total,
        })

    context = {
        "from_date": from_date,
        "to_date": to_date,
        "journal_data": journal_data,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "is_balanced": total_debit == total_credit,
    }
    
    filename = f"journal_report_{from_date}_to_{to_date}.pdf"
    return generate_pdf('accounting/journal_report_pdf.html', context, filename)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def ledger_summary(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    category_id = request.GET.get('category_id')

    if not from_date or not to_date:
        today = datetime.today()
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
        to_date = today.strftime('%Y-%m-%d')

    categories = LedgerCategory.objects.all()

    if category_id:
        ledgers = Ledger.objects.filter(category_id=category_id, is_active=True)
    else:
        ledgers = Ledger.objects.filter(is_active=True)

    ledger_data = []

    for ledger in ledgers:
        debit_total = ledger.entries.filter(entry_type='Debit', date__range=[from_date, to_date]).aggregate(total=Sum('amount'))['total'] or 0
        credit_total = ledger.entries.filter(entry_type='Credit', date__range=[from_date, to_date]).aggregate(total=Sum('amount'))['total'] or 0

        ledger_data.append({
            'name': ledger.name,
            'category': ledger.category.name,
            'opening_balance': ledger.opening_balance,
            'debit_total': debit_total,
            'credit_total': credit_total,
            'closing_balance': ledger.opening_balance + debit_total - credit_total
        })

    context = {
        'from_date': from_date,
        'to_date': to_date,
        'categories': categories,
        'selected_category': category_id,
        'ledger_data': ledger_data,
    }

    return render(request, 'accounting/ledger_summary.html', context)


from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from io import BytesIO
from django.db.models import Sum
from .models import Ledger, LedgerCategory, LedgerEntry

def generate_pdf(template_src, context_dict, filename):
    html = render_to_string(template_src, context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    return None

def ledger_summary_pdf(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    category_id = request.GET.get('category_id')

    if not from_date or not to_date:
        today = datetime.today()
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
        to_date = today.strftime('%Y-%m-%d')

    categories = LedgerCategory.objects.all()

    if category_id:
        ledgers = Ledger.objects.filter(category_id=category_id, is_active=True)
        selected_category = LedgerCategory.objects.get(id=category_id)
    else:
        ledgers = Ledger.objects.filter(is_active=True)
        selected_category = None

    ledger_data = []
    total_opening = 0
    total_debit = 0
    total_credit = 0
    total_closing = 0

    for ledger in ledgers:
        debit_total = ledger.entries.filter(entry_type='Debit', date__range=[from_date, to_date]).aggregate(total=Sum('amount'))['total'] or 0
        credit_total = ledger.entries.filter(entry_type='Credit', date__range=[from_date, to_date]).aggregate(total=Sum('amount'))['total'] or 0
        closing_balance = ledger.opening_balance + debit_total - credit_total

        ledger_data.append({
            'name': ledger.name,
            'category': ledger.category.name,
            'opening_balance': ledger.opening_balance,
            'debit_total': debit_total,
            'credit_total': credit_total,
            'closing_balance': closing_balance
        })
        
        total_opening += ledger.opening_balance
        total_debit += debit_total
        total_credit += credit_total
        total_closing += closing_balance

    context = {
        'from_date': from_date,
        'to_date': to_date,
        'categories': categories,
        'selected_category': selected_category,
        'ledger_data': ledger_data,
        'total_opening': total_opening,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'total_closing': total_closing,
    }
    
    if selected_category:
        filename = f"ledger_summary_{selected_category.name}_{from_date}_to_{to_date}.pdf"
    else:
        filename = f"ledger_summary_all_{from_date}_to_{to_date}.pdf"
    
    return generate_pdf('accounting/ledger_summary_pdf.html', context, filename)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def category_summary(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    selected_category_id = request.GET.get('ledger_category')

    if not from_date or not to_date:
        today = datetime.today()
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
        to_date = today.strftime('%Y-%m-%d')

    categories = LedgerCategory.objects.all()
    selected_category = None
    ledgers = []
    total_debit = 0
    total_credit = 0

    if selected_category_id:
        selected_category = LedgerCategory.objects.filter(id=selected_category_id).first()
        if selected_category:
            ledgers = Ledger.objects.filter(category=selected_category, is_active=True)

            ledger_details = []
            for ledger in ledgers:
                debit_total = LedgerEntry.objects.filter(
                    ledger=ledger, entry_type="Debit", date__range=[from_date, to_date]
                ).aggregate(total=Sum('amount'))['total'] or 0

                credit_total = LedgerEntry.objects.filter(
                    ledger=ledger, entry_type="Credit", date__range=[from_date, to_date]
                ).aggregate(total=Sum('amount'))['total'] or 0

                ledger_details.append({
                    'name': ledger.name,
                    'debit': debit_total,
                    'credit': credit_total
                })

                total_debit += debit_total
                total_credit += credit_total

    context = {
        'from_date': from_date,
        'to_date': to_date,
        'categories': categories,
        'selected_category': selected_category,
        'ledger_details': ledger_details if selected_category else [],
        'total_debit': total_debit,
        'total_credit': total_credit,
    }

    return render(request, 'accounting/category_summary.html', context)

@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))
def ledger_book(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    ledger_id = request.GET.get('ledger_id')

    ledgers = Ledger.objects.filter(is_active=True)

    entries = []
    selected_ledger = None
    total_debit = 0
    total_credit = 0

    if ledger_id:
        selected_ledger = Ledger.objects.filter(id=ledger_id).first()
        if selected_ledger:
            entries = LedgerEntry.objects.filter(
                ledger=selected_ledger,
                date__range=[from_date, to_date]
            ).order_by('date')

            total_debit = entries.filter(entry_type="Debit").aggregate(total=Sum('amount'))['total'] or 0
            total_credit = entries.filter(entry_type="Credit").aggregate(total=Sum('amount'))['total'] or 0

    context = {
        'from_date': from_date,
        'to_date': to_date,
        'ledgers': ledgers,
        'selected_ledger': selected_ledger,
        'entries': entries,
        'total_debit': total_debit,
        'total_credit': total_credit,
    }

    return render(request, 'accounting/ledger_book.html', context)


@login_required(login_url='login')
@user_passes_test(lambda user: is_staff_or_has_role(user, roles=['Manager', 'Accountant']))

def cash_book(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if not from_date or not to_date:
        today = datetime.today()
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
        to_date = today.strftime('%Y-%m-%d')

    cash_ledgers = Ledger.objects.filter(category__name="Asset", name__icontains="Cash")

    total_opening_balance = cash_ledgers.aggregate(total=Sum('opening_balance'))['total'] or 0

    cash_entries = LedgerEntry.objects.filter(
        ledger__in=cash_ledgers,
        date__range=[from_date, to_date]
    ).order_by('date')

    total_debit = cash_entries.filter(entry_type="Debit").aggregate(total=Sum('amount'))['total'] or 0
    total_credit = cash_entries.filter(entry_type="Credit").aggregate(total=Sum('amount'))['total'] or 0

    closing_balance = total_opening_balance + total_debit - total_credit

    context = {
        'from_date': from_date,
        'to_date': to_date,
        'cash_entries': cash_entries,
        'total_opening_balance': total_opening_balance,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'closing_balance': closing_balance
    }

    return render(request, 'accounting/cash_book.html', context)




from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook

def fee_report(request):
    # Filter receives with voucher numbers starting with FEE-
    receives = Receive.objects.filter(voucher_no__startswith='FEE-').select_related(
        'student__student_field',
        'student__class_id__class_group_id__class_id',
        'student__class_id__section_id',
        'student__class_id__shift_id',
        'student__class_id__class_group_id__group_id',
    ).order_by('-date')

    # Excel Export Handling
    if 'export' in request.GET:
        wb = Workbook()
        ws = wb.active
        headers = ['Student Name', 'Roll Number', 'Class', 'Section', 'Group', 'Version', 'Shift', 'Amount', 'Date']
        ws.append(headers)
        
        for receive in receives:
            student = receive.student
            student_name = student.student_field.name if student and student.student_field else '-'
            roll_no = student.roll_no if student and student.roll_no is not None else '-'
            
            class_config = student.class_id if student else None
            class_group_config = class_config.class_group_id if class_config else None
            
            # Class and Group
            class_name = class_group_config.class_id.name if class_group_config and class_group_config.class_id else '-'
            group = class_group_config.group_id.name if class_group_config and class_group_config.group_id else '-'
            
            # Section and Shift
            section = class_config.section_id.name if class_config and class_config.section_id else '-'
            shift = class_config.shift_id.name if class_config and class_config.shift_id else '-'
            
            # Version
            version = student.get_version_display() if student and student.version else '-'
            
            # Amount and Date
            amount = receive.amount
            date = receive.date.strftime('%Y-%m-%d') if receive.date else '-'
            
            ws.append([
                student_name,
                roll_no,
                class_name,
                section,
                group,
                version,
                shift,
                amount,
                date
            ])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="fee_report.xlsx"'
        wb.save(response)
        return response

    # Pagination for HTML View
    paginator = Paginator(receives, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    report_data = []
    for receive in page_obj:
        student = receive.student
        student_name = student.student_field.name if student and student.student_field else '-'
        roll_no = student.roll_no if student and student.roll_no is not None else '-'
        
        class_config = student.class_id if student else None
        class_group_config = class_config.class_group_id if class_config else None
        
        # Class and Group
        class_name = class_group_config.class_id.name if class_group_config and class_group_config.class_id else '-'
        group = class_group_config.group_id.name if class_group_config and class_group_config.group_id else '-'
        
        # Section and Shift
        section = class_config.section_id.name if class_config and class_config.section_id else '-'
        shift = class_config.shift_id.name if class_config and class_config.shift_id else '-'
        
        # Version
        version = student.get_version_display() if student and student.version else '-'
        
        report_data.append({
            'student_name': student_name,
            'roll_no': roll_no,
            'class': class_name,
            'section': section,
            'group': group,
            'version': version,
            'shift': shift,
            'amount': receive.amount,
            'date': receive.date.strftime('%Y-%m-%d') if receive.date else '-'
        })

    return render(request, 'accounting/fee_report.html', {
        'report_data': report_data,
        'page_obj': page_obj
    })


def fee_report_ui(request):
    # Filter receives with voucher numbers starting with FEE-
    receives = Receive.objects.filter(voucher_no__startswith='FEE-').select_related(
        'student__student_field',
        'student__class_id__class_group_id__class_id',
        'student__class_id__section_id',
        'student__class_id__shift_id',
        'student__class_id__class_group_id__group_id',
    ).order_by('-date')

    # Pagination for HTML View (20 rows per page)
    paginator = Paginator(receives, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    report_data = []
    for receive in page_obj:
        student = receive.student
        student_name = student.student_field.name if student and student.student_field else '-'
        roll_no = student.roll_no if student and student.roll_no is not None else '-'
        
        class_config = student.class_id if student else None
        class_group_config = class_config.class_group_id if class_config else None
        
        # Class and Group
        class_name = class_group_config.class_id.name if class_group_config and class_group_config.class_id else '-'
        group = class_group_config.group_id.name if class_group_config and class_group_config.group_id else '-'
        
        # Section and Shift
        section = class_config.section_id.name if class_config and class_config.section_id else '-'
        shift = class_config.shift_id.name if class_config and class_config.shift_id else '-'
        
        # Version
        version = student.get_version_display() if student and student.version else '-'
        
        report_data.append({
            'student_name': student_name,
            'roll_no': roll_no,
            'class': class_name,
            'section': section,
            'group': group,
            'version': version,
            'shift': shift,
            'amount': receive.amount,
            'date': receive.date.strftime('%Y-%m-%d') if receive.date else '-'
        })

    return render(request, 'accounting/fee_report_ui.html', {
        'report_data': report_data,
        'page_obj': page_obj
    })
    

    
   
from django.shortcuts import render, HttpResponse
from django.core.paginator import Paginator
from django.template.loader import render_to_string
from django.db.models import Sum
from datetime import datetime
from weasyprint import HTML
import tempfile
from .models import Payment, Ledger

def expense_report(request):
    # Date validation and filtering
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    if not from_date or not to_date:
        return render(request, 'accounting/expense_report.html', {'error': 'Please select both from and to dates'})
    
    try:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
    except ValueError:
        return render(request, 'accounting/expense_report.html', {'error': 'Invalid date format'})

    # Get filtered payments
    payments = Payment.objects.filter(
        date__range=[from_date, to_date]
    ).select_related('expense_ledger', 'staff__staff_field').order_by('date')

    # Group by ledger and process data
    ledgers = Ledger.objects.filter(
        expense_payments__in=payments
    ).distinct().annotate(
        total_amount=Sum('expense_payments__amount')
    ).order_by('code')

    ledger_data = []
    grand_total = 0
    entries_per_page = 6  

    for ledger in ledgers:
        ledger_payments = payments.filter(expense_ledger=ledger)
        total = ledger.total_amount or 0
        grand_total += total
        
        paginator = Paginator(ledger_payments, entries_per_page)
        
        for page_num in paginator.page_range:
            page = paginator.page(page_num)
            ledger_data.append({
                'ledger': ledger,
                'entries': page.object_list,
                'total': total,
                'page_number': page_num,
                'total_pages': paginator.num_pages,
            })

    context = {
        'ledger_data': ledger_data,
        'grand_total': grand_total,
        'from_date': from_date.strftime('%Y-%m-%d'),
        'to_date': to_date.strftime('%Y-%m-%d'),
    }
    return render(request, 'accounting/expense_report.html', context)

from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from io import BytesIO
from django.db.models import Sum
from .models import Payment, Ledger

def generate_pdf(template_src, context_dict, filename):
    html = render_to_string(template_src, context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    return None

def expense_report_pdf(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if not from_date or not to_date:
        today = datetime.today()
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
        to_date = today.strftime('%Y-%m-%d')

    try:
        from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()
    except:
        return HttpResponse('Invalid date parameters')

    # Get filtered payments
    payments = Payment.objects.filter(
        date__range=[from_date_obj, to_date_obj]
    ).select_related('expense_ledger', 'staff__staff_field').order_by('date')

    # Group by ledger and process data
    ledgers = Ledger.objects.filter(
        expense_payments__in=payments
    ).distinct().annotate(
        total_amount=Sum('expense_payments__amount')
    ).order_by('code')

    ledger_data = []
    grand_total = 0

    for ledger in ledgers:
        ledger_payments = payments.filter(expense_ledger=ledger)
        total = ledger.total_amount or 0
        grand_total += total
        
        # For PDF, we don't paginate - include all entries
        ledger_data.append({
            'ledger': ledger,
            'entries': ledger_payments,
            'total': total,
            'page_number': 1,
            'total_pages': 1,
        })

    context = {
        'ledger_data': ledger_data,
        'grand_total': grand_total,
        'from_date': from_date_obj,
        'to_date': to_date_obj,
    }
    
    filename = f"expense_report_{from_date}_to_{to_date}.pdf"
    return generate_pdf('accounting/expense_report_pdf.html', context, filename)

from django.shortcuts import render, HttpResponse
from django.core.paginator import Paginator
from django.template.loader import render_to_string
from django.db.models import Sum
from datetime import datetime
from weasyprint import HTML
from .models import Receive, Ledger

# def income_report(request):
#     # Date validation and filtering
#     from_date = request.GET.get('from_date')
#     to_date = request.GET.get('to_date')
    
#     if not from_date or not to_date:
#         return render(request, 'accounting/income_report.html', {'error': 'Please select both from and to dates'})
    
#     try:
#         from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
#         to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
#     except ValueError:
#         return render(request, 'accounting/income_report.html', {'error': 'Invalid date format'})

#     # Get filtered receipts
#     receipts = Receive.objects.filter(
#         date__range=[from_date, to_date]
#     ).select_related('income_ledger').order_by('date')

#     # Group by ledger and process data
#     ledgers = Ledger.objects.filter(
#         income_receipts__in=receipts
#     ).distinct().annotate(
#         total_amount=Sum('income_receipts__amount')
#     ).order_by('code')

#     ledger_data = []
#     grand_total = 0
#     entries_per_page = 10  

#     for ledger in ledgers:
#         ledger_receipts = receipts.filter(income_ledger=ledger)
#         total = ledger.total_amount or 0
#         grand_total += total
        
#         paginator = Paginator(ledger_receipts, entries_per_page)
        
#         for page_num in paginator.page_range:
#             page = paginator.page(page_num)
#             ledger_data.append({
#                 'ledger': ledger,
#                 'entries': page.object_list,
#                 'total': total,
#                 'page_number': page_num,
#                 'total_pages': paginator.num_pages,
#             })

#     context = {
#         'ledger_data': ledger_data,
#         'grand_total': grand_total,
#         'from_date': from_date.strftime('%Y-%m-%d'),
#         'to_date': to_date.strftime('%Y-%m-%d'),
#     }
#     return render(request, 'accounting/income_report.html', context)


from datetime import datetime
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Sum
from django.shortcuts import render
from .models import Receive, Ledger

def income_report(request):
    # Date validation and filtering
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    if not from_date or not to_date:
        return render(request, 'accounting/income_report.html', {'error': 'Please select both from and to dates'})
    
    try:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
    except ValueError:
        return render(request, 'accounting/income_report.html', {'error': 'Invalid date format'})

    # Get filtered receipts
    receipts = Receive.objects.filter(
        date__range=[from_date, to_date]
    ).select_related('income_ledger').order_by('date')

    # Group by ledger and process data
    ledgers = Ledger.objects.filter(
        income_receipts__in=receipts
    ).distinct().annotate(
        total_amount=Sum('income_receipts__amount')
    ).order_by('code')

    ledger_data = []
    grand_total = 0
    entries_per_page = 10  # Items per page per ledger
    page = request.GET.get('page', 1)

    for ledger in ledgers:
        ledger_receipts = receipts.filter(income_ledger=ledger)
        total = ledger.total_amount or 0
        grand_total += total
        
        paginator = Paginator(ledger_receipts, entries_per_page)
        try:
            current_page = paginator.page(page)
        except PageNotAnInteger:
            current_page = paginator.page(1)
        except EmptyPage:
            current_page = paginator.page(paginator.num_pages)

        ledger_data.append({
            'ledger': ledger,
            'entries': current_page.object_list,
            'total': total,
            'page': current_page,
            'has_pages': paginator.num_pages > 1
        })

    context = {
        'ledger_data': ledger_data,
        'grand_total': grand_total,
        'from_date': from_date.strftime('%Y-%m-%d'),
        'to_date': to_date.strftime('%Y-%m-%d'),
        'current_page': page
    }
    return render(request, 'accounting/income_report.html', context)


def income_report_pdf(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    try:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
    except:
        return HttpResponse('Invalid date parameters')

    receipts = Receive.objects.filter(
        date__range=[from_date, to_date]
    ).select_related('income_ledger').order_by('date')

    ledgers = Ledger.objects.filter(
        income_receipts__in=receipts
    ).distinct().annotate(
        total_amount=Sum('income_receipts__amount')
    ).order_by('code')

    ledger_data = []
    grand_total = 0
    entries_per_page = 6

    for ledger in ledgers:
        ledger_receipts = receipts.filter(income_ledger=ledger)
        total = ledger.total_amount or 0
        grand_total += total
        
        paginator = Paginator(ledger_receipts, entries_per_page)
        
        for page_num in paginator.page_range:
            page = paginator.page(page_num)
            ledger_data.append({
                'ledger': ledger,
                'entries': page.object_list,
                'total': total,
                'page_number': page_num,
                'total_pages': paginator.num_pages,
            })

    html_string = render_to_string('accounting/income_report_pdf.html', {
        'ledger_data': ledger_data,
        'grand_total': grand_total,
        'from_date': from_date,
        'to_date': to_date,
    })

    html = HTML(string=html_string)
    result = html.write_pdf()

    response = HttpResponse(result, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="income_report_{from_date}_to_{to_date}.pdf"'
    return response


from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Q
from openpyxl import Workbook
from .models import Receive

def voucher_report(request):
    # EXCLUDE records starting with these prefixes
    queryset = Receive.objects.exclude(
        Q(voucher_no__startswith='RECV') |
        Q(voucher_no__startswith='FEE')
    ).select_related('student__student_field', 'fee_head')

    # Handle Excel export
    if 'export' in request.GET:
        wb = Workbook()
        ws = wb.active
        ws.append([
            'Voucher No', 'Date', 'Student', 'Fee Head', 'Amount',
            'Student Name', 'Student Roll Number'
        ])

        for rec in queryset:
            student_name = rec.student.student_field.name if rec.student and rec.student.student_field else ''
            roll_no = rec.student.roll_no if rec.student else ''
            fee_head = rec.fee_head.name if rec.fee_head else ''

            ws.append([
                rec.voucher_no,
                rec.date,
                str(rec.student) if rec.student else '',
                fee_head,
                rec.amount,
                student_name,
                roll_no
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="voucher_report.xlsx"'
        wb.save(response)
        return response

    # Prepare context for HTML view
    context = []
    for rec in queryset:
        student_name = rec.student.student_field.name if rec.student and rec.student.student_field else ''
        roll_no = rec.student.roll_no if rec.student else ''
        fee_head = rec.fee_head.name if rec.fee_head else ''

        context.append({
            'voucher_no': rec.voucher_no,
            'date': rec.date,
            'student': str(rec.student) if rec.student else '',
            'fee_head': fee_head,
            'amount': rec.amount,
            'student_name': student_name,
            'student_roll': roll_no,
        })

    return render(request, 'accounting/voucher_report.html', {'vouchers': context})

import json
def journal_management(request):
    if request.method == 'POST':
        try:
            required_fields = ['date', 'description']
            missing = [field for field in required_fields if not request.POST.get(field)]
            if missing:
                return JsonResponse({'error': f'Missing fields: {", ".join(missing)}'}, status=400)

            voucher_no = request.POST.get('voucher_no')
            if not voucher_no:
                current_time = datetime.now()
                voucher_no = current_time.strftime("JRNL%Y%m%d%H%M%S")

            date_str = request.POST['date']
            description = request.POST['description']
            entries = json.loads(request.POST.get('entries', '[]'))

            if len(entries) < 2:
                return JsonResponse({'error': 'At least two entries required (one debit and one credit)'}, status=400)

            total_debit = sum(Decimal(entry['amount']) for entry in entries if entry['entry_type'] == 'Debit')
            total_credit = sum(Decimal(entry['amount']) for entry in entries if entry['entry_type'] == 'Credit')

            if total_debit != total_credit:
                return JsonResponse({'error': 'Total debits must equal total credits'}, status=400)

            try:
                date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({'error': 'Invalid date format (YYYY-MM-DD required)'}, status=400)

            with transaction.atomic():
                journal = Journal.objects.create(
                    voucher_no=voucher_no,
                    date=date,
                    description=description,
                    created_by = getattr(request.user, 'staffprofile', None)
                )

                for entry in entries:
                    ledger = get_object_or_404(Ledger, id=entry['ledger_id'])
                    JournalEntry.objects.create(
                        journal=journal,
                        ledger=ledger,
                        entry_type=entry['entry_type'],
                        amount = Decimal(str(entry['amount'])),
                        description=entry['description'] or ''
                    )

                return JsonResponse({
                    'success': True,
                    'voucher_no': journal.voucher_no,
                    'date': journal.date.strftime('%Y-%m-%d'),
                    'description': journal.description,
                    'entries': [{
                        'ledger': entry.ledger.name,
                        'entry_type': entry.entry_type,
                        'amount': str(entry.amount),
                        'description': entry.description
                    } for entry in journal.entries.all()]
                })

        except IntegrityError:
            return JsonResponse({'error': 'Voucher number already exists'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    journals = Journal.objects.all().prefetch_related('entries').order_by('-date')
    paginator = Paginator(journals, 15)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    
    
    ledgers = Ledger.objects.filter(is_active=True)
    return render(request, 'accounting/journal_entry.html', {
        'journals': page_obj,
        'ledgers': ledgers
    })
    
    
    
def contra_management(request):
    if request.method == 'POST':
        try:
            required_fields = ['date', 'amount', 'from_ledger', 'to_ledger']
            missing = [field for field in required_fields if not request.POST.get(field)]
            if missing:
                return JsonResponse({'error': f'Missing fields: {", ".join(missing)}'}, status=400)

            voucher_no = request.POST.get('voucher_no')
            if not voucher_no:
                current_time = datetime.now()
                voucher_no = current_time.strftime("CONTRA%Y%m%d%H%M%S%f")

            date_str = request.POST['date']
            amount = request.POST['amount']
            from_ledger_id = request.POST['from_ledger']
            to_ledger_id = request.POST['to_ledger']
            description = request.POST.get('description', '')

            try:
                date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({'error': 'Invalid date format (YYYY-MM-DD required)'}, status=400)

            try:
                amount = Decimal(amount)
                if amount <= 0:
                    return JsonResponse({'error': 'Amount must be greater than zero'}, status=400)
            except InvalidOperation:
                return JsonResponse({'error': 'Invalid amount format'}, status=400)

            from_ledger = get_object_or_404(Ledger, id=from_ledger_id)
            to_ledger = get_object_or_404(Ledger, id=to_ledger_id)

            with transaction.atomic():
                contra = Contra(
                    voucher_no=voucher_no,
                    date=date,
                    amount=amount,
                    from_ledger=from_ledger,
                    to_ledger=to_ledger,
                    description=description,
                    created_by=getattr(request.user, 'staffprofile', None)  # Added comma here
                )
                # Validation and save moved outside constructor
                contra.full_clean()
                contra.save()

            return JsonResponse({
                'success': True,
                'voucher_no': contra.voucher_no,
                'date': contra.date.strftime('%Y-%m-%d'),
                'amount': str(contra.amount),
                'from_ledger': from_ledger.name,
                'to_ledger': to_ledger.name,
                'description': description,
                'created_at': contra.created_at.strftime('%Y-%m-%d %H:%M:%S')
            })

        except ValidationError as e:
            return JsonResponse({'error': str(e)}, status=400)
        except IntegrityError:
            return JsonResponse({'error': 'Voucher number already exists'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    contra_entries = Contra.objects.all().order_by('-date')
    paginator = Paginator(contra_entries, 15)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    asset_ledgers = Ledger.objects.filter(category__name__iexact='Asset', is_active=True)
    return render(request, 'accounting/contra_list.html', {
        'contra_entries': page_obj,
        'asset_ledgers': asset_ledgers.order_by('name')
    })
    
def delete_contra(request, contra_id):
    if request.method == 'POST':
        try:
            contra = get_object_or_404(Contra, id=contra_id)
            contra.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid request method'}, status=405)


from django.shortcuts import render
from django.db.models import Sum, Q
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime
import io
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
import xlsxwriter

def student_fee_report(request):
    """Display the report form"""
    try:
        return render(request, 'accounting/student_fee_report.html')
    except:
        # Try alternative template path
        return render(request, 'student_fee_report.html')

def generate_student_fee_report(request):
    """Generate the report based on date range"""
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        report_type = request.POST.get('report_type')  # 'excel' or 'pdf'
        
        # Validate dates
        if not start_date or not end_date:
            try:
                return render(request, 'accounting/student_fee_report.html', 
                             {'error': 'Please select both start and end dates'})
            except:
                return render(request, 'student_fee_report.html', 
                             {'error': 'Please select both start and end dates'})
        
        # Import models inside the function to avoid circular imports
        from .models import Receive
        
        try:
            # Filter Receive records based on date range and exclude FEE voucher_no
            receives = Receive.objects.filter(
                date__range=[start_date, end_date]
            ).exclude(
                voucher_no__startswith='FEE'
            ).select_related(
                'student__student_field',
                'student__class_id__class_group_id__class_id',
                'student__class_id__class_group_id__group_id',
                'student__class_id__section_id',
                'student__class_id__shift_id',
                'student__academic_session_year'
            )
            
            # Get unique students first
            student_ids = receives.filter(student__isnull=False).values_list('student', flat=True).distinct()
            
            report_data = []
            total_amount = 0
            
            for student_id in student_ids:
                student_receives = receives.filter(student_id=student_id)
                student_total = student_receives.aggregate(total=Sum('amount'))['total'] or 0
                
                if student_total > 0:
                    # Get student details from first receive
                    student_receive = student_receives.first()
                    student_profile = student_receive.student
                    
                    if not student_profile:
                        continue
                    
                    # Safely access related fields
                    class_name = ""
                    if student_profile.class_id and student_profile.class_id.class_group_id:
                        class_name = student_profile.class_id.class_group_id.class_id.name
                    
                    group_name = ""
                    if student_profile.class_id and student_profile.class_id.class_group_id and student_profile.class_id.class_group_id.group_id:
                        group_name = student_profile.class_id.class_group_id.group_id.name
                    
                    section_name = ""
                    if student_profile.class_id and student_profile.class_id.section_id:
                        section_name = student_profile.class_id.section_id.name
                    
                    shift_name = ""
                    if student_profile.class_id and student_profile.class_id.shift_id:
                        shift_name = student_profile.class_id.shift_id.name
                    
                    session_name = ""
                    if student_profile.academic_session_year:
                        session_name = f"{student_profile.academic_session_year.start_year}-{student_profile.academic_session_year.end_year}"
                    
                    report_data.append({
                        'student_name': student_profile.student_field.name if student_profile.student_field else "N/A",
                        'roll': student_profile.roll_no or "N/A",
                        'class': class_name,
                        'version': student_profile.version,
                        'group': group_name,
                        'section': section_name,
                        'shift': shift_name,
                        'session': session_name,
                        'amount': float(student_total)
                    })
                    total_amount += float(student_total)
            
            # Sort by class and roll number
            report_data.sort(key=lambda x: (
                x['class'], 
                x['roll'] if isinstance(x.get('roll'), (int, float)) else 0
            ))
            
            # Add total row
            report_data.append({
                'student_name': 'TOTAL',
                'roll': '',
                'class': '',
                'version': '',
                'group': '',
                'section': '',
                'shift': '',
                'session': '',
                'amount': total_amount
            })
            
            if report_type == 'excel':
                return generate_excel_report(report_data, start_date, end_date)
            else:
                return generate_pdf_report(report_data, start_date, end_date)
                
        except Exception as e:
            error_message = f"Error generating report: {str(e)}"
            print(f"Error: {error_message}")  # For debugging
            try:
                return render(request, 'accounting/student_fee_report.html', 
                             {'error': error_message})
            except:
                return render(request, 'student_fee_report.html', 
                             {'error': error_message})
    
    # If not POST, return to form
    try:
        return render(request, 'accounting/student_fee_report.html')
    except:
        return render(request, 'student_fee_report.html')

def generate_excel_report(data, start_date, end_date):
    """Generate Excel report using xlsxwriter"""
    try:
        output = io.BytesIO()
        
        # Create workbook and worksheet
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Student Fee Report')
        
        # Define formats - FIXED: set font size in the format dictionary
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4472C4',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'font_size': 11  # Added font size here
        })
        
        cell_format = workbook.add_format({
            'border': 1,
            'align': 'left',
            'valign': 'vcenter',
            'font_size': 10
        })
        
        amount_format = workbook.add_format({
            'border': 1,
            'align': 'right',
            'valign': 'vcenter',
            'num_format': '#,##0.00',
            'font_size': 10
        })
        
        total_format = workbook.add_format({
            'bold': True,
            'bg_color': '#F2F2F2',
            'border': 1,
            'align': 'right',
            'valign': 'vcenter',
            'num_format': '#,##0.00',
            'font_size': 11
        })
        
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 14,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        # Set column widths - make them wider
        column_widths = [35, 15, 15, 15, 15, 15, 15, 15, 20]
        for i, width in enumerate(column_widths):
            worksheet.set_column(i, i, width)
        
        # Write title
        title = f"Student Fee Report ({start_date} to {end_date})"
        worksheet.merge_range('A1:I1', title, title_format)
        
        # Write headers (starting from row 2)
        headers = ['Student Name', 'Roll', 'Class', 'Version', 'Group', 'Section', 'Shift', 'Session', 'Amount']
        for col, header in enumerate(headers):
            worksheet.write(1, col, header, header_format)
        
        # Write data
        for row, item in enumerate(data[:-1], start=2):  # Start from row 2, exclude total row for now
            worksheet.write(row, 0, item['student_name'], cell_format)
            worksheet.write(row, 1, item['roll'] if item['roll'] != "N/A" else "", cell_format)
            worksheet.write(row, 2, item['class'], cell_format)
            worksheet.write(row, 3, item['version'], cell_format)
            worksheet.write(row, 4, item['group'], cell_format)
            worksheet.write(row, 5, item['section'], cell_format)
            worksheet.write(row, 6, item['shift'], cell_format)
            worksheet.write(row, 7, item['session'], cell_format)
            worksheet.write(row, 8, item['amount'], amount_format)
        
        # Write total row
        total_row = len(data) + 1  # +1 because we started from row 2
        total_item = data[-1]
        
        # Create a format for the total row label
        total_label_format = workbook.add_format({
            'bold': True,
            'bg_color': '#F2F2F2',
            'border': 1,
            'align': 'left',
            'valign': 'vcenter',
            'font_size': 11
        })
        
        worksheet.write(total_row, 0, total_item['student_name'], total_label_format)
        # Merge cells B to H for the total row
        for col in range(1, 8):
            worksheet.write_blank(total_row, col, None, total_label_format)
        worksheet.write(total_row, 8, total_item['amount'], total_format)
        
        workbook.close()
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'student_fee_report_{start_date}_to_{end_date}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    except Exception as e:
        # Return error response if Excel generation fails
        import traceback
        error_details = traceback.format_exc()
        print(f"Excel Generation Error: {str(e)}")
        print(f"Traceback: {error_details}")
        
        # Create a simple error Excel file
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Error')
        worksheet.write(0, 0, f"Error generating report: {str(e)}")
        workbook.close()
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="error_report.xlsx"'
        return response
    
    
def generate_pdf_report(data, start_date, end_date):
    """Generate PDF report using ReportLab - MAXIMUM SIZE with wide columns"""
    try:
        buffer = io.BytesIO()
        
        # Use A3 landscape for maximum space (larger than A4)
        from reportlab.lib.pagesizes import A3
        page_width, page_height = landscape(A3)
        
        # Calculate available width after margins
        left_margin = 0.5*inch
        right_margin = 0.5*inch
        available_width = page_width - left_margin - right_margin
        
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A3),  # A3 is much larger than A4
            rightMargin=right_margin,
            leftMargin=left_margin,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Add title with very large font
        title_text = f"<para align=center><font size=20><b>STUDENT FEE REPORT</b></font><br/><font size=14>{start_date} to {end_date}</font></para>"
        title = Paragraph(title_text, styles['Title'])
        elements.append(title)
        elements.append(Paragraph("<br/><br/>", styles['Normal']))  # Add more space
        
        # Prepare table data
        table_data = []
        
        # Headers - use very large, bold font
        headers = ['Student Name', 'Roll', 'Class', 'Version', 'Group', 'Section', 'Shift', 'Session', 'Amount']
        table_data.append(headers)
        
        # Add data rows
        for item in data[:-1]:  # Exclude total row
            # Don't truncate names - we have wide columns now
            student_name = item['student_name']
                
            row = [
                student_name,
                str(item['roll']),
                item['class'],
                item['version'],
                item['group'],
                item['section'],
                item['shift'],
                item['session'],
                f"{item['amount']:,.2f}"
            ]
            table_data.append(row)
        
        # Add total row
        total_item = data[-1]
        total_row = [
            total_item['student_name'],
            '',
            '',
            '',
            '',
            '',
            '',
            'TOTAL:',
            f"{total_item['amount']:,.2f}"
        ]
        table_data.append(total_row)
        
        # Create table with MAXIMUM WIDTH columns
        # Calculate widths to use almost all available space
        col_widths = [
            3.5*inch,  # Student Name - VERY WIDE
            1.0*inch,  # Roll
            1.2*inch,  # Class
            1.2*inch,  # Version
            1.2*inch,  # Group
            1.2*inch,  # Section
            1.2*inch,  # Shift
            1.5*inch,  # Session - WIDE
            1.8*inch,  # Amount - VERY WIDE
        ]
        
        # Verify total width doesn't exceed available width
        total_table_width = sum(col_widths)
        if total_table_width > available_width:
            # Scale down proportionally
            scale_factor = available_width / total_table_width
            col_widths = [w * scale_factor for w in col_widths]
        
        table = Table(table_data, colWidths=col_widths, repeatRows=1)  # Repeat headers on each page
        
        # Style the table - MAXIMUM VISIBILITY with large fonts
        style = TableStyle([
            # Header style - VERY LARGE, BOLD, with thick border
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),  # Very dark blue
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),  # VERY LARGE header font
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # VERY THICK grid lines
            ('GRID', (0, 0), (-1, -1), 1.5, colors.black),
            
            # Data rows alignment
            ('ALIGN', (0, 1), (-2, -2), 'LEFT'),
            ('ALIGN', (-1, 1), (-1, -2), 'RIGHT'),
            
            # Font for data rows - LARGE for readability
            ('FONTSIZE', (0, 1), (-1, -2), 12),  # Large data font
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            
            # Alternate row colors for maximum readability
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), 
             [colors.HexColor('#F8F9FA'), colors.HexColor('#E9ECEF')]),
            
            # Generous padding for data rows
            ('TOPPADDING', (0, 1), (-1, -2), 10),
            ('BOTTOMPADDING', (0, 1), (-1, -2), 10),
            ('LEFTPADDING', (0, 1), (-1, -2), 8),
            ('RIGHTPADDING', (0, 1), (-1, -2), 8),
            
            # Total row style - VERY PROMINENT
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#1A1A1A')),  # Very dark gray
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 16),  # Extra large for total
            ('ALIGN', (-2, -1), (-2, -1), 'RIGHT'),
            
            # EXTRA THICK border for total row
            ('LINEABOVE', (0, -1), (-1, -1), 3, colors.black),
            
            # Vertical alignment for all cells
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Center align specific columns
            ('ALIGN', (1, 1), (1, -2), 'CENTER'),  # Roll
            ('ALIGN', (2, 1), (2, -2), 'CENTER'),  # Class
            ('ALIGN', (3, 1), (3, -2), 'CENTER'),  # Version
            ('ALIGN', (4, 1), (4, -2), 'CENTER'),  # Group
            ('ALIGN', (5, 1), (5, -2), 'CENTER'),  # Section
            ('ALIGN', (6, 1), (6, -2), 'CENTER'),  # Shift
            ('ALIGN', (7, 1), (7, -2), 'CENTER'),  # Session
            
            # Make column separators more visible
            ('LINEBEFORE', (1, 0), (1, -1), 1, colors.HexColor('#666666')),
            ('LINEBEFORE', (2, 0), (2, -1), 1, colors.HexColor('#666666')),
            ('LINEBEFORE', (3, 0), (3, -1), 1, colors.HexColor('#666666')),
            ('LINEBEFORE', (4, 0), (4, -1), 1, colors.HexColor('#666666')),
            ('LINEBEFORE', (5, 0), (5, -1), 1, colors.HexColor('#666666')),
            ('LINEBEFORE', (6, 0), (6, -1), 1, colors.HexColor('#666666')),
            ('LINEBEFORE', (7, 0), (7, -1), 1, colors.HexColor('#666666')),
            ('LINEBEFORE', (8, 0), (8, -1), 1, colors.HexColor('#666666')),
        ])
        
        table.setStyle(style)
        elements.append(table)
        
        # Add footer with larger font
        elements.append(Paragraph("<br/><br/>", styles['Normal']))
        footer_text = f"<para align=right><font size=10><b>Generated on {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}</b></font></para>"
        footer = Paragraph(footer_text, styles['Normal'])
        elements.append(footer)
        
        # Add page numbers if needed (for multi-page reports)
        def add_page_numbers(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 9)
            page_num = canvas.getPageNumber()
            text = f"Page {page_num}"
            canvas.drawRightString(page_width - right_margin, 0.4*inch, text)
            canvas.restoreState()
        
        # Build PDF
        doc.build(elements, onFirstPage=add_page_numbers, onLaterPages=add_page_numbers)
        
        buffer.seek(0)
        
        # Create response
        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f'student_fee_report_{start_date}_to_{end_date}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    except Exception as e:
        # Return error response if PDF generation fails
        import traceback
        error_details = traceback.format_exc()
        print(f"PDF Generation Error: {str(e)}")
        print(f"Traceback: {error_details}")
        
        # Create a simple error PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A3))
        elements = []
        styles = getSampleStyleSheet()
        error_style = styles['Title']
        error_style.fontSize = 16
        elements.append(Paragraph(f"<para align=center><b>Error generating PDF report</b><br/>{str(e)}</para>", error_style))
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="error_report.pdf"'
        return response


